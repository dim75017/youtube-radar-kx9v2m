#!/usr/bin/env python3
"""Build the small, first-paint livestream snapshot used by the dashboard."""

from __future__ import annotations

import argparse
import io
import json
import math
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


SHEET_ID = "1XE_M9pQWn8w2Qu83vV_tv9sEFDFQ13fTePseG6mh1vI"
XLSX_URL = (
    "https://docs.google.com/spreadsheets/d/"
    + SHEET_ID
    + "/export?format=xlsx"
)
VIDEO_ID = re.compile(r"^[\w-]{11}$")
TITLE_SCAN_SUFFIX = re.compile(
    r"\s*\d{4}-\d{2}-\d{2}[ T]?\d{2}:\d{2}(?::\d{2})?\s*$"
)
DAY_MS = 86_400_000
OFFICIAL_CHANNEL_ID = "UCSJ4gkVC6NrvII8umztf0Ow"
OFFICIAL_STREAMS_URL = "https://www.youtube.com/@LofiGirl/streams"
OFFICIAL_UPLOADS_URL = "https://www.youtube.com/playlist?list=UUSJ4gkVC6NrvII8umztf0Ow"
OFFICIAL_LISTING_LIMIT = 100
NEXT_ENDPOINT = "https://www.youtube.com/youtubei/v1/next?prettyPrint=false"
NEXT_CLIENT_VERSION = "2.20260114.08.00"
NEXT_MAX_JSON_BYTES = 2 * 1024 * 1024
NEXT_TRANSIENT_HTTP_CODES = {408, 409, 425, 429, 500, 502, 503, 504}
WATCH_ENDPOINT = "https://www.youtube.com/watch"
WATCH_MAX_HTML_BYTES = 4 * 1024 * 1024
WATCH_PLAYER_ASSIGNMENT = re.compile(r"(?:var\s+)?ytInitialPlayerResponse\s*=\s*")
WATCH_TRANSIENT_HTTP_CODES = NEXT_TRANSIENT_HTTP_CODES
ENDED_UNPLAYABLE_REASON = "This live stream recording is not available."


def utc_now_ms() -> int:
    return int(datetime.now(timezone.utc).timestamp() * 1000)


def to_ms(value: object) -> int | None:
    if value is None or value == "":
        return None
    parsed: datetime
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
    elif isinstance(value, (int, float)):
        parsed = from_excel(value)
    else:
        raw = str(value).strip()
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return int(parsed.timestamp() * 1000)


def sheet_by_name(workbook, fragment: str):
    needle = fragment.casefold()
    for name in workbook.sheetnames:
        if needle in name.casefold():
            return workbook[name]
    raise RuntimeError(f"Sheet not found: {fragment}")


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def positive_number(value: object, *, allow_zero: bool = False) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    parsed = float(value)
    if not math.isfinite(parsed) or parsed < 0 or (parsed == 0 and not allow_zero):
        return None
    return parsed


def _strict_json_decoder() -> json.JSONDecoder:
    def reject_constant(value: str):
        raise ValueError(f"invalid JSON constant {value}")

    def reject_duplicates(pairs):
        result = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"duplicate JSON key {key!r}")
            result[key] = value
        return result

    return json.JSONDecoder(
        parse_constant=reject_constant,
        object_pairs_hook=reject_duplicates,
    )


def _strict_json(body: bytes) -> dict:
    value = _strict_json_decoder().decode(body.decode("utf-8"))
    if type(value) is not dict:
        raise ValueError("YouTube next response is not a JSON object")
    return value


def _aware_iso_timestamp(value: object, field_name: str) -> float:
    raw = clean_text(value)
    if not raw:
        raise RuntimeError(f"YouTube WatchPage has no exact {field_name}")
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError as exc:
        raise RuntimeError(
            f"YouTube WatchPage has a malformed {field_name}"
        ) from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise RuntimeError(
            f"YouTube WatchPage {field_name} is not timezone-aware"
        )
    timestamp = parsed.timestamp()
    if not math.isfinite(timestamp) or timestamp <= 0:
        raise RuntimeError(f"YouTube WatchPage has an invalid {field_name}")
    return timestamp


class YouTubeWatchPageLiveProof:
    """Prove public official live state from YouTube's unauthenticated watch page."""

    def __init__(self, *, retries: int = 2, timeout_seconds: int = 20) -> None:
        self.retries = max(0, int(retries))
        self.timeout_seconds = max(1, int(timeout_seconds))

    @staticmethod
    def _validate_final_url(final_url: str, video_id: str) -> None:
        try:
            parsed = urllib.parse.urlparse(final_url)
            query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
            port = parsed.port
        except ValueError as exc:
            raise RuntimeError("YouTube WatchPage returned a malformed URL") from exc
        if (
            parsed.scheme != "https"
            or (parsed.hostname or "").casefold() != "www.youtube.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.path != "/watch"
            or parsed.params
            or parsed.fragment
            or query != [("v", video_id), ("hl", "en"), ("gl", "US")]
        ):
            raise RuntimeError("YouTube WatchPage redirected away from the exact public URL")

    @staticmethod
    def _player_response(body: bytes) -> dict:
        text = body.decode("utf-8")
        assignments = list(WATCH_PLAYER_ASSIGNMENT.finditer(text))
        if len(assignments) != 1:
            raise RuntimeError("YouTube WatchPage has ambiguous player response evidence")
        try:
            value, end = _strict_json_decoder().raw_decode(text, assignments[0].end())
        except ValueError as exc:
            raise RuntimeError("YouTube WatchPage player response is malformed") from exc
        if type(value) is not dict:
            raise RuntimeError("YouTube WatchPage player response is not an object")
        trailing = text[end:]
        if not re.match(r"^\s*;", trailing):
            raise RuntimeError("YouTube WatchPage player response is not terminated")
        return value

    @staticmethod
    def _project(payload: dict, video_id: str) -> dict:
        playability = payload.get("playabilityStatus")
        details = payload.get("videoDetails")
        microformat = payload.get("microformat")
        if (
            type(playability) is not dict
            or type(details) is not dict
            or type(microformat) is not dict
            or type(microformat.get("playerMicroformatRenderer")) is not dict
        ):
            raise RuntimeError(f"YouTube WatchPage evidence is incomplete for {video_id}")
        micro = microformat["playerMicroformatRenderer"]
        live = micro.get("liveBroadcastDetails")
        if type(live) is not dict:
            raise RuntimeError(f"YouTube WatchPage has no broadcast facts for {video_id}")

        title = details.get("title")
        micro_title = micro.get("title")
        canonical_url = f"https://www.youtube.com/watch?v={video_id}"
        if (
            details.get("videoId") != video_id
            or micro.get("externalVideoId") != video_id
            or details.get("channelId") != OFFICIAL_CHANNEL_ID
            or micro.get("externalChannelId") != OFFICIAL_CHANNEL_ID
            or details.get("author") != "Lofi Girl"
            or micro.get("ownerChannelName") != "Lofi Girl"
            or not isinstance(title, str)
            or not title.strip()
            or type(micro_title) is not dict
            or micro_title.get("simpleText") != title
            or micro.get("canonicalUrl") != canonical_url
            or details.get("isPrivate") is not False
            or details.get("isCrawlable") is not True
            or micro.get("isUnlisted") is not False
            or details.get("isLiveContent") is not True
        ):
            raise RuntimeError(f"YouTube WatchPage public identity mismatch for {video_id}")

        started = _aware_iso_timestamp(live.get("startTimestamp"), "startTimestamp")
        is_live_now = live.get("isLiveNow")
        status = playability.get("status")
        if is_live_now is True:
            if (
                status != "OK"
                or details.get("isLive") is not True
                or clean_text(live.get("endTimestamp"))
            ):
                raise RuntimeError(
                    f"YouTube WatchPage does not prove an active public live for {video_id}"
                )
            return {
                "id": video_id,
                "channel_id": OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "title": title,
                "is_live": True,
                "live_status": "is_live",
                "release_timestamp": started,
                "detail_source": "youtube_watch_page",
            }

        if is_live_now is not False or details.get("isLive") is True:
            raise RuntimeError(f"YouTube WatchPage live state is ambiguous for {video_id}")
        ended = _aware_iso_timestamp(live.get("endTimestamp"), "endTimestamp")
        if ended < started:
            raise RuntimeError(f"YouTube WatchPage end precedes start for {video_id}")
        if not (
            status == "OK"
            or (
                status == "UNPLAYABLE"
                and playability.get("reason") == ENDED_UNPLAYABLE_REASON
            )
        ):
            raise RuntimeError(
                f"YouTube WatchPage does not prove a public ended live for {video_id}"
            )
        return {
            "id": video_id,
            "channel_id": OFFICIAL_CHANNEL_ID,
            "availability": "public",
            "title": title,
            "is_live": False,
            "live_status": "was_live",
            "release_timestamp": started,
            "end_timestamp": ended,
            "detail_source": "youtube_watch_page",
        }

    def _fetch_once(self, video_id: str) -> tuple[dict, str]:
        url = WATCH_ENDPOINT + "?" + urllib.parse.urlencode(
            [("v", video_id), ("hl", "en"), ("gl", "US")]
        )
        request = urllib.request.Request(
            url,
            headers={
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
                "Cache-Control": "no-cache",
                "User-Agent": (
                    "Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/140.0.0.0 Safari/537.36"
                ),
            },
        )
        with urllib.request.urlopen(request, timeout=self.timeout_seconds) as response:
            final_url = str(response.geturl())
            body = response.read(WATCH_MAX_HTML_BYTES + 1)
        if len(body) > WATCH_MAX_HTML_BYTES:
            raise RuntimeError("YouTube WatchPage exceeded the bounded HTML size")
        self._validate_final_url(final_url, video_id)
        return self._player_response(body), final_url

    def extract_info(self, video_id: str) -> dict:
        if not VIDEO_ID.fullmatch(video_id):
            raise RuntimeError(f"Invalid YouTube video ID: {video_id!r}")
        for attempt in range(self.retries + 1):
            try:
                payload, _ = self._fetch_once(video_id)
                return self._project(payload, video_id)
            except urllib.error.HTTPError as exc:
                if exc.code not in WATCH_TRANSIENT_HTTP_CODES or attempt >= self.retries:
                    raise RuntimeError(
                        f"YouTube WatchPage HTTP {exc.code} for {video_id}"
                    ) from exc
            except (urllib.error.URLError, TimeoutError, OSError) as exc:
                if attempt >= self.retries:
                    raise RuntimeError(
                        f"YouTube WatchPage network failure for {video_id}: "
                        f"{type(exc).__name__}"
                    ) from exc
            except UnicodeDecodeError as exc:
                raise RuntimeError(
                    f"YouTube WatchPage is not strict UTF-8 for {video_id}"
                ) from exc
            time.sleep(2 ** attempt)
        raise RuntimeError("YouTube WatchPage retry loop exhausted")


class YouTubeNextLiveFallback:
    """Read a currently-live public counter when yt-dlp hits YouTube's bot gate.

    The public ``next`` response has no exact start time. This adapter therefore
    accepts a separate factual start timestamp and uses the human-readable date
    only as a consistency check. Relative dates are allowed solely when the exact
    timestamp was just proved by :class:`YouTubeWatchPageLiveProof`.
    """

    def __init__(self, *, retries: int = 2, timeout_seconds: int = 20) -> None:
        self.retries = max(0, int(retries))
        self.timeout_seconds = max(1, int(timeout_seconds))

    @staticmethod
    def _mapping_at(value: object, *keys: str) -> dict:
        current = value
        for key in keys:
            if type(current) is not dict:
                return {}
            current = current.get(key)
        return current if type(current) is dict else {}

    @staticmethod
    def _runs_text(value: object) -> str:
        if type(value) is not dict or type(value.get("runs")) is not list:
            return ""
        runs = value["runs"]
        if not runs or any(type(run) is not dict or not isinstance(run.get("text"), str) for run in runs):
            return ""
        return "".join(run["text"] for run in runs)

    @staticmethod
    def _started_date(value: object) -> date | None:
        match = re.fullmatch(
            r"Started streaming on ([A-Z][a-z]{2}) (\d{1,2}), (\d{4})",
            clean_text(value),
        )
        if not match:
            return None
        months = {
            name: index
            for index, name in enumerate(
                ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"),
                1,
            )
        }
        try:
            return date(int(match.group(3)), months[match.group(1)], int(match.group(2)))
        except (KeyError, ValueError):
            return None

    @staticmethod
    def _validate_final_url(final_url: str) -> None:
        try:
            parsed = urllib.parse.urlparse(final_url)
            query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
            port = parsed.port
        except ValueError as exc:
            raise RuntimeError("YouTube next returned a malformed URL") from exc
        if (
            parsed.scheme != "https"
            or (parsed.hostname or "").casefold() != "www.youtube.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.path != "/youtubei/v1/next"
            or parsed.fragment
            or query != [("prettyPrint", "false")]
        ):
            raise RuntimeError("YouTube next redirected away from its public endpoint")

    def _fetch_once(self, video_id: str) -> tuple[dict, str]:
        request = urllib.request.Request(
            NEXT_ENDPOINT,
            data=json.dumps(
                {
                    "context": {
                        "client": {
                            "clientName": "WEB",
                            "clientVersion": NEXT_CLIENT_VERSION,
                            "hl": "en",
                            "gl": "US",
                        }
                    },
                    "videoId": video_id,
                },
                separators=(",", ":"),
            ).encode("utf-8"),
            headers={
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
                "Content-Type": "application/json",
                "User-Agent": (
                    "Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/140.0.0.0 Safari/537.36"
                ),
                "X-YouTube-Client-Name": "1",
                "X-YouTube-Client-Version": NEXT_CLIENT_VERSION,
            },
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=self.timeout_seconds) as response:
            final_url = str(response.geturl())
            body = response.read(NEXT_MAX_JSON_BYTES + 1)
        if len(body) > NEXT_MAX_JSON_BYTES:
            raise RuntimeError("YouTube next response exceeded the bounded JSON size")
        self._validate_final_url(final_url)
        return _strict_json(body), final_url

    def _project(
        self,
        payload: dict,
        video_id: str,
        started_ms: object,
        *,
        start_source: str = "previous_verified_asset",
        expected_title: str | None = None,
    ) -> dict:
        current = self._mapping_at(payload, "currentVideoEndpoint")
        watch = self._mapping_at(current, "watchEndpoint")
        web = self._mapping_at(current, "commandMetadata", "webCommandMetadata")
        if watch.get("videoId") != video_id or web.get("webPageType") != "WEB_PAGE_TYPE_WATCH":
            raise RuntimeError(f"YouTube next identity mismatch for {video_id}")
        parsed_watch = urllib.parse.urlparse(clean_text(web.get("url")))
        watch_query = urllib.parse.parse_qs(parsed_watch.query, keep_blank_values=True)
        if (
            parsed_watch.scheme
            or parsed_watch.netloc
            or parsed_watch.path != "/watch"
            or parsed_watch.params
            or parsed_watch.fragment
            or watch_query.get("v") != [video_id]
            or any(key not in {"v", "pp"} for key in watch_query)
            or len(watch_query.get("pp", [])) > 1
            or any(not value for value in watch_query.get("pp", []))
        ):
            raise RuntimeError(f"YouTube next watch URL mismatch for {video_id}")

        contents = self._mapping_at(
            payload, "contents", "twoColumnWatchNextResults", "results", "results"
        ).get("contents")
        if type(contents) is not list:
            raise RuntimeError(f"YouTube next has no primary content for {video_id}")
        primary = [
            item.get("videoPrimaryInfoRenderer")
            for item in contents
            if type(item) is dict and type(item.get("videoPrimaryInfoRenderer")) is dict
        ]
        secondary = [
            item.get("videoSecondaryInfoRenderer")
            for item in contents
            if type(item) is dict and type(item.get("videoSecondaryInfoRenderer")) is dict
        ]
        if len(primary) != 1 or len(secondary) != 1:
            raise RuntimeError(f"YouTube next has ambiguous primary/owner data for {video_id}")

        title = self._runs_text(primary[0].get("title"))
        view = self._mapping_at(primary[0], "viewCount", "videoViewCountRenderer")
        original = view.get("originalViewCount")
        runs = (view.get("viewCount") or {}).get("runs")
        if (
            not title
            or (expected_title is not None and title != expected_title)
            or view.get("isLive") is not True
            or not isinstance(original, str)
            or re.fullmatch(r"[0-9]+", original, flags=re.ASCII) is None
            or type(runs) is not list
            or len(runs) != 2
            or type(runs[0]) is not dict
            or type(runs[1]) is not dict
            or runs[1].get("text") != " watching now"
            or not isinstance(runs[0].get("text"), str)
            or not re.fullmatch(r"[0-9]{1,3}(?:,[0-9]{3})*|[0-9]+", runs[0]["text"], flags=re.ASCII)
            or int(runs[0]["text"].replace(",", "")) != int(original)
        ):
            raise RuntimeError(f"YouTube next has no exact live counter for {video_id}")

        owner = self._mapping_at(secondary[0], "owner", "videoOwnerRenderer")
        owner_browse = self._mapping_at(owner, "navigationEndpoint", "browseEndpoint").get("browseId")
        owner_runs = (owner.get("title") or {}).get("runs")
        if type(owner_runs) is not list or len(owner_runs) != 1 or type(owner_runs[0]) is not dict:
            raise RuntimeError(f"YouTube next has no exact owner for {video_id}")
        title_browse = self._mapping_at(owner_runs[0], "navigationEndpoint", "browseEndpoint").get("browseId")
        if (
            owner_browse != OFFICIAL_CHANNEL_ID
            or title_browse != OFFICIAL_CHANNEL_ID
            or owner_runs[0].get("text") != "Lofi Girl"
        ):
            raise RuntimeError(f"YouTube next owner mismatch for {video_id}")

        started = positive_number(started_ms)
        if started is None:
            raise RuntimeError(
                f"YouTube next has no exact start time for new official stream {video_id}"
            )
        raw_date_text = clean_text((primary[0].get("dateText") or {}).get("simpleText"))
        date_text = self._started_date(raw_date_text)
        factual_date = datetime.fromtimestamp(started / 1000, tz=timezone.utc).date()
        relative_date = re.fullmatch(
            r"Started streaming (?:[1-9][0-9]*|one|an?) "
            r"(?:seconds?|minutes?|hours?|days?|weeks?|months?|years?) ago",
            raw_date_text,
            flags=re.IGNORECASE,
        )
        if date_text is not None:
            if abs((date_text - factual_date).days) > 1:
                raise RuntimeError(f"YouTube next start date mismatch for {video_id}")
        elif not (relative_date and start_source == "youtube_watch_page"):
            raise RuntimeError(f"YouTube next start date mismatch for {video_id}")
        return {
            "id": video_id,
            "channel_id": OFFICIAL_CHANNEL_ID,
            "availability": "public",
            "title": title,
            "is_live": True,
            "live_status": "is_live",
            "concurrent_view_count": int(original),
            "release_timestamp": int(started / 1000),
            "detail_source": (
                "youtube_next_watch_page_start"
                if start_source == "youtube_watch_page"
                else "youtube_next_previous_start"
            ),
        }

    def extract_info(
        self,
        video_id: str,
        *,
        started_ms: object,
        start_source: str = "previous_verified_asset",
        expected_title: str | None = None,
    ) -> dict:
        if not VIDEO_ID.fullmatch(video_id):
            raise RuntimeError(f"Invalid YouTube video ID: {video_id!r}")
        for attempt in range(self.retries + 1):
            try:
                payload, _ = self._fetch_once(video_id)
                return self._project(
                    payload,
                    video_id,
                    started_ms,
                    start_source=start_source,
                    expected_title=expected_title,
                )
            except urllib.error.HTTPError as exc:
                if exc.code not in NEXT_TRANSIENT_HTTP_CODES or attempt >= self.retries:
                    raise RuntimeError(f"YouTube next HTTP {exc.code} for {video_id}") from exc
            except (urllib.error.URLError, TimeoutError, OSError) as exc:
                if attempt >= self.retries:
                    raise RuntimeError(
                        f"YouTube next network failure for {video_id}: {type(exc).__name__}"
                    ) from exc
            except (UnicodeDecodeError, ValueError) as exc:
                if attempt >= self.retries:
                    raise RuntimeError(
                        f"YouTube next returned malformed JSON for {video_id}"
                    ) from exc
            except RuntimeError:
                if attempt >= self.retries:
                    raise
            time.sleep(2 ** attempt)
        raise RuntimeError("YouTube next retry loop exhausted")


def is_youtube_antibot_error(error: BaseException) -> bool:
    message = str(error).casefold().replace("’", "'")
    return (
        "sign in to confirm you're not a bot" in message
        or "sign in to confirm you are not a bot" in message
    )


def is_youtube_ended_recording_error(
    error: BaseException, video_id: str
) -> bool:
    """Recognize only yt-dlp's exact public-ended recording failure."""
    message = str(error).strip()
    return message in {
        ENDED_UNPLAYABLE_REASON,
        f"[youtube] {video_id}: {ENDED_UNPLAYABLE_REASON}",
        f"ERROR: [youtube] {video_id}: {ENDED_UNPLAYABLE_REASON}",
    }


def official_live_row(info: dict, observed_ms: int) -> tuple[dict, tuple[int, int]]:
    """Project one verified active official stream without inventing counters."""
    video_id = clean_text(info.get("id"))
    if not VIDEO_ID.fullmatch(video_id):
        raise RuntimeError("Official stream response has no exact video ID")
    if clean_text(info.get("channel_id")) != OFFICIAL_CHANNEL_ID:
        raise RuntimeError(f"Official stream {video_id} belongs to an unexpected channel")
    if clean_text(info.get("availability")) != "public":
        raise RuntimeError(f"Official stream {video_id} is not verifiably public")
    if info.get("is_live") is not True or clean_text(info.get("live_status")) != "is_live":
        raise RuntimeError(f"Official stream {video_id} is not verifiably live")
    viewers = positive_number(info.get("concurrent_view_count"), allow_zero=True)
    if viewers is None or not viewers.is_integer():
        raise RuntimeError(f"Official stream {video_id} has no factual concurrent viewer count")
    title = TITLE_SCAN_SUFFIX.sub("", clean_text(info.get("title")))
    if not title:
        raise RuntimeError(f"Official stream {video_id} has no title")
    started_seconds = positive_number(
        info.get("release_timestamp") or info.get("timestamp")
    )
    if started_seconds is None:
        raise RuntimeError(f"Official stream {video_id} has no factual start timestamp")
    row = {
        "vid": video_id,
        "channel": "Lofi Girl",
        "channelId": OFFICIAL_CHANNEL_ID,
        "chUrl": f"https://www.youtube.com/channel/{OFFICIAL_CHANNEL_ID}",
        "title": title,
        "url": f"https://www.youtube.com/watch?v={video_id}",
        "started": int(started_seconds * 1000),
        "disc": "official Lofi Girl /streams",
        "audiences": ["youtube"],
        "liveStatus": "is_live",
        "source": "Official Lofi Girl streams scan",
    }
    return row, (int(observed_ms), int(viewers))


def discover_official_live_streams(
    observed_ms: int | None = None,
    *,
    flat_reader=None,
    detail_reader=None,
    watch_reader=None,
    next_reader=None,
    previous_active: dict[str, dict] | None = None,
) -> dict:
    """Discover every currently-live radio exposed by the official streams tab."""
    observed_ms = int(observed_ms or utc_now_ms())
    if flat_reader is None or detail_reader is None:
        import yt_dlp
    if flat_reader is None:
        flat_reader = yt_dlp.YoutubeDL({
            "quiet": True,
            "no_warnings": True,
            "skip_download": True,
            "extract_flat": True,
            "playlistend": OFFICIAL_LISTING_LIMIT,
            "socket_timeout": 30,
            "retries": 2,
            "extractor_retries": 2,
            "ignoreerrors": False,
            "cachedir": False,
        })
    if detail_reader is None:
        detail_reader = yt_dlp.YoutubeDL({
            "quiet": True,
            "no_warnings": True,
            "skip_download": True,
            "socket_timeout": 30,
            "retries": 2,
            "extractor_retries": 2,
            "ignoreerrors": False,
            "cachedir": False,
        })
    streams_listing = flat_reader.extract_info(OFFICIAL_STREAMS_URL, download=False) or {}
    uploads_listing = flat_reader.extract_info(OFFICIAL_UPLOADS_URL, download=False) or {}

    def listing_ids(listing: dict, *statuses: str) -> list[str]:
        allowed = set(statuses)
        return list(dict.fromkeys(
            clean_text(entry.get("id"))
            for entry in (listing.get("entries") or [])
            if entry
            and VIDEO_ID.fullmatch(clean_text(entry.get("id")))
            and clean_text(entry.get("live_status")) in allowed
        ))

    # A missing flat status is hydrated because extractor drift must not hide a
    # just-started radio. An explicit upcoming status is a different factual
    # state: count it, but never route it into live detail/WatchPage projection.
    streams_flat_active_ids = listing_ids(streams_listing, "is_live")
    streams_unknown_ids = listing_ids(streams_listing, "")
    streams_upcoming_ids = listing_ids(streams_listing, "is_upcoming")
    uploads_active_ids = listing_ids(uploads_listing, "is_live")
    uploads_upcoming_ids = listing_ids(uploads_listing, "is_upcoming")
    flat_active_set = set(streams_flat_active_ids) | set(uploads_active_ids)
    upcoming_set = set(streams_upcoming_ids) | set(uploads_upcoming_ids)
    conflicting_flat_ids = flat_active_set & upcoming_set
    if conflicting_flat_ids:
        raise RuntimeError(
            "Official Lofi Girl listings disagree on live/upcoming state for "
            + ", ".join(sorted(conflicting_flat_ids))
        )
    streams_unknown_ids = [
        video_id for video_id in streams_unknown_ids if video_id not in upcoming_set
    ]
    streams_candidate_ids = list(dict.fromkeys(
        streams_flat_active_ids + streams_unknown_ids
    ))
    candidate_ids = list(dict.fromkeys(
        streams_flat_active_ids + streams_unknown_ids + uploads_active_ids
    ))
    previous_active = previous_active or {}
    if not candidate_ids and not previous_active:
        raise RuntimeError("Official Lofi Girl /streams returned no active livestreams")

    rows: list[dict] = []
    points: dict[str, list[tuple[int, int]]] = {}
    detail_cache: dict[str, dict] = {}
    watch_page_fallbacks = 0
    next_fallbacks = 0
    ended_ids: set[str] = set()

    def detail(video_id: str) -> dict:
        nonlocal watch_reader, next_reader, watch_page_fallbacks, next_fallbacks
        if video_id not in detail_cache:
            try:
                detail_cache[video_id] = detail_reader.extract_info(
                    f"https://www.youtube.com/watch?v={video_id}", download=False
                ) or {}
            except Exception as exc:
                if not (
                    is_youtube_antibot_error(exc)
                    or is_youtube_ended_recording_error(exc, video_id)
                ):
                    raise
                if video_id not in candidate_ids and video_id not in previous_active:
                    raise RuntimeError(
                        f"Unexpected official live fallback target {video_id}"
                    ) from exc
                if watch_reader is None:
                    watch_reader = YouTubeWatchPageLiveProof()
                watch_info = watch_reader.extract_info(video_id) or {}
                watch_page_fallbacks += 1
                if (
                    clean_text(watch_info.get("id")) != video_id
                    or clean_text(watch_info.get("channel_id")) != OFFICIAL_CHANNEL_ID
                    or clean_text(watch_info.get("availability")) != "public"
                ):
                    raise RuntimeError(
                        f"YouTube WatchPage proof identity mismatch for {video_id}"
                    ) from exc
                if (
                    watch_info.get("is_live") is False
                    and clean_text(watch_info.get("live_status"))
                    in {"was_live", "post_live"}
                ):
                    detail_cache[video_id] = watch_info
                    return detail_cache[video_id]
                if (
                    watch_info.get("is_live") is not True
                    or clean_text(watch_info.get("live_status")) != "is_live"
                ):
                    raise RuntimeError(
                        f"YouTube WatchPage live state is ambiguous for {video_id}"
                    ) from exc
                title = clean_text(watch_info.get("title"))
                started_seconds = positive_number(watch_info.get("release_timestamp"))
                if not title or started_seconds is None:
                    raise RuntimeError(
                        f"YouTube WatchPage proof is incomplete for {video_id}"
                    ) from exc
                if next_reader is None:
                    next_reader = YouTubeNextLiveFallback()
                started_ms = int(started_seconds * 1000)
                next_info = next_reader.extract_info(
                    video_id,
                    started_ms=started_ms,
                    start_source="youtube_watch_page",
                    expected_title=title,
                )
                next_started = positive_number(next_info.get("release_timestamp"))
                if (
                    clean_text(next_info.get("id")) != video_id
                    or clean_text(next_info.get("channel_id")) != OFFICIAL_CHANNEL_ID
                    or clean_text(next_info.get("availability")) != "public"
                    or clean_text(next_info.get("title")) != title
                    or next_info.get("is_live") is not True
                    or clean_text(next_info.get("live_status")) != "is_live"
                    or next_started is None
                    or int(next_started * 1000) != started_ms
                ):
                    raise RuntimeError(
                        f"YouTube next diverges from WatchPage proof for {video_id}"
                    ) from exc
                viewers = positive_number(
                    next_info.get("concurrent_view_count"), allow_zero=True
                )
                if viewers is None or not viewers.is_integer():
                    raise RuntimeError(
                        f"YouTube next has no exact concurrent count for {video_id}"
                    ) from exc
                detail_cache[video_id] = dict(watch_info)
                detail_cache[video_id]["concurrent_view_count"] = int(viewers)
                detail_cache[video_id]["detail_source"] = (
                    "youtube_watch_page+youtube_next"
                )
                next_fallbacks += 1
        return detail_cache[video_id]

    active_ids: list[str] = []
    confirmed_listing_ended = 0
    for video_id in candidate_ids:
        info = detail(video_id)
        if clean_text(info.get("id")) != video_id:
            raise RuntimeError(f"Official stream detail identity mismatch for {video_id}")
        if (
            clean_text(info.get("channel_id")) == OFFICIAL_CHANNEL_ID
            and clean_text(info.get("availability")) == "public"
            and info.get("is_live") is False
            and clean_text(info.get("live_status")) in {"was_live", "post_live"}
        ):
            confirmed_listing_ended += 1
            ended_ids.add(video_id)
            continue
        if (
            clean_text(info.get("channel_id")) == OFFICIAL_CHANNEL_ID
            and clean_text(info.get("availability")) == "public"
            and info.get("is_live") is False
            and clean_text(info.get("live_status")) == "is_upcoming"
        ):
            confirmed_listing_ended += 1
            continue
        row, point = official_live_row(info, observed_ms)
        rows.append(row)
        points[video_id] = [point]
        active_ids.append(video_id)

    recovered = 0
    confirmed_ended = 0
    legacy_rejected = 0
    listing_set = set(active_ids)
    missing_previous = sorted(set(previous_active) - listing_set)
    for video_id in missing_previous:
        info = detail(video_id)
        trusted = previous_active[video_id].get("trusted") is True
        exact_identity = clean_text(info.get("id")) == video_id
        exact_channel = clean_text(info.get("channel_id")) == OFFICIAL_CHANNEL_ID
        if not exact_identity or not exact_channel:
            if trusted:
                raise RuntimeError(
                    f"Previously verified official stream {video_id} has conflicting identity"
                )
            legacy_rejected += 1
            continue
        if (
            clean_text(info.get("availability")) == "public"
            and info.get("is_live") is True
            and clean_text(info.get("live_status")) == "is_live"
        ):
            row, point = official_live_row(info, observed_ms)
            rows.append(row)
            points[video_id] = [point]
            recovered += 1
            continue
        if (
            clean_text(info.get("availability")) == "public"
            and info.get("is_live") is False
            and clean_text(info.get("live_status")) in {"was_live", "post_live"}
        ):
            confirmed_ended += 1
            ended_ids.add(video_id)
            continue
        raise RuntimeError(
            f"Previously active official stream {video_id} has ambiguous current status"
        )

    expected_ids = set(active_ids) | {
        video_id for video_id in missing_previous if video_id in points
    }
    if not rows:
        raise RuntimeError("Official Lofi Girl streams scan verified no active livestreams")
    if len(rows) != len(expected_ids) or set(points) != expected_ids:
        raise RuntimeError("Official Lofi Girl streams scan is incomplete")
    return {
        "rows": rows,
        "points": points,
        "endedIds": sorted(ended_ids),
        "metrics": {
            "expected": len(expected_ids),
            "verified": len(rows),
            "observedT": observed_ms,
            "listingActive": len(active_ids),
            "streamsTabCandidates": len(streams_candidate_ids),
            "streamsTabActive": len(streams_flat_active_ids),
            "streamsTabUnknown": len(streams_unknown_ids),
            "listingUpcoming": len(upcoming_set),
            "uploadsPlaylistActive": len(uploads_active_ids),
            "listingConfirmedEnded": confirmed_listing_ended,
            "previousActive": len(previous_active),
            "missingFromListing": len(missing_previous),
            "recoveredStillLive": recovered,
            "confirmedEnded": confirmed_ended,
            "legacyRejected": legacy_rejected,
            "watchPageFallbacks": watch_page_fallbacks,
            "nextFallbacks": next_fallbacks,
        },
    }


def parse_live_catalogue(workbook) -> list[dict]:
    rows: list[dict] = []
    sheet = sheet_by_name(workbook, "Live Streams")
    for values in sheet.iter_rows(min_row=2, values_only=True):
        values = tuple(values) + (None,) * max(0, 7 - len(values))
        video_id = clean_text(values[0])
        if not VIDEO_ID.match(video_id):
            continue
        row = {
            "vid": video_id,
            "channel": clean_text(values[1]),
            "title": TITLE_SCAN_SUFFIX.sub("", clean_text(values[2])),
            "url": clean_text(values[3])
            or f"https://www.youtube.com/watch?v={video_id}",
            "started": to_ms(values[4]),
            "disc": clean_text(values[5]),
        }
        audience = clean_text(values[6])
        if re.search(r"\bkids?\b", audience, re.IGNORECASE):
            row["audiences"] = ["kids"]
        rows.append(row)
    return rows


def parse_history_points(workbook, fragment: str) -> dict[str, list[tuple[int, int]]]:
    points: dict[str, list[tuple[int, int]]] = defaultdict(list)
    sheet = sheet_by_name(workbook, fragment)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if len(values) < 3:
            continue
        video_id = clean_text(values[0])
        timestamp = to_ms(values[1])
        try:
            viewers = int(float(values[2]))
        except (TypeError, ValueError):
            continue
        if VIDEO_ID.match(video_id) and timestamp is not None:
            points[video_id].append((timestamp, viewers))
    return points


def load_existing_live_metadata(snapshot_path: Path | None) -> dict[str, dict]:
    if not snapshot_path or not snapshot_path.exists():
        return {}
    raw = snapshot_path.read_text(encoding="utf-8")
    raw = re.sub(r"^window\.LOFI_DATA=", "", raw).rstrip(";\n")
    payload = json.loads(raw)
    return {
        str(row.get("vid") or ""): row
        for row in ((payload.get("d") or {}).get("lives") or [])
        if VIDEO_ID.match(str(row.get("vid") or ""))
    }


def load_previous_active_official_lives(path: Path | None) -> dict[str, dict]:
    """Load the prior active official cohort so a partial listing cannot shrink it."""
    if not path or not path.exists():
        return {}
    raw = path.read_text(encoding="utf-8")
    prefix = "window.LOFI_LIVE_DATA="
    if not raw.startswith(prefix):
        raise RuntimeError("Existing livestream bootstrap has an unexpected format")
    payload = json.loads(raw[len(prefix):].rstrip(";\n"))
    data = payload.get("d") or {}
    summaries = data.get("liveSummary") or {}
    metrics = payload.get("metrics") or {}
    proof_present = (
        positive_number(metrics.get("officialExpected"), allow_zero=True) is not None
        and metrics.get("officialVerified") is not None
        and int(metrics.get("officialExpected"))
        == int(metrics.get("officialVerified"))
    )
    result: dict[str, dict] = {}
    for row in data.get("lives") or []:
        video_id = clean_text(row.get("vid"))
        if not VIDEO_ID.fullmatch(video_id):
            continue
        trusted = (
            proof_present
            and clean_text(row.get("channelId")) == OFFICIAL_CHANNEL_ID
            and clean_text(row.get("source")) == "Official Lofi Girl streams scan"
            and clean_text(row.get("liveStatus")) == "is_live"
        )
        legacy = (
            not proof_present
            and clean_text(row.get("channel")) == "Lofi Girl"
            and (summaries.get(video_id) or {}).get("active") is True
        )
        if trusted or legacy:
            result[video_id] = {
                "trusted": trusted,
                "started": int(row.get("started"))
                if positive_number(row.get("started")) is not None
                else None,
            }
    return result


def build_payload(
    workbook,
    snapshot_path: Path | None = None,
    official_snapshot: dict | None = None,
) -> dict:
    lives = parse_live_catalogue(workbook)
    sheet_live_ids = {row["vid"] for row in lives}
    live_by_id = {row["vid"]: row for row in lives}
    official_snapshot = official_snapshot or {"rows": [], "points": {}, "metrics": {}}
    official_rows = official_snapshot.get("rows") or []
    ended_ids = {
        clean_text(video_id) for video_id in (official_snapshot.get("endedIds") or [])
    }
    if any(not VIDEO_ID.fullmatch(video_id) for video_id in ended_ids):
        raise RuntimeError("Official livestream snapshot has an invalid ended ID")
    official_ids: set[str] = set()
    for source in official_rows:
        video_id = clean_text(source.get("vid"))
        if not VIDEO_ID.fullmatch(video_id) or video_id in official_ids:
            raise RuntimeError("Official livestream snapshot has invalid or duplicate IDs")
        official_ids.add(video_id)
        current = live_by_id.get(video_id)
        if current is None:
            current = dict(source)
            lives.append(current)
            live_by_id[video_id] = current
        else:
            discovery = current.get("disc")
            current.update(source)
            if discovery:
                current["disc"] = discovery
    if official_ids & ended_ids:
        raise RuntimeError("Official livestream cannot be both active and ended")
    live_ids = {row["vid"] for row in lives}
    existing = load_existing_live_metadata(snapshot_path)
    for row in lives:
        previous = existing.get(row["vid"], {})
        for key in ("madeForKids", "audiences", "channelId", "chUrl", "subs"):
            if key in previous and key not in row:
                row[key] = previous[key]

    points: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for fragment in ("Live History", "Live Hourly"):
        for video_id, series in parse_history_points(workbook, fragment).items():
            if video_id in live_ids:
                points[video_id].extend(series)
    sheet_history_times = {
        video_id: {timestamp for timestamp, _ in series}
        for video_id, series in points.items()
        if series
    }
    sheet_source_latest = max(
        (timestamp for series in points.values() for timestamp, _ in series),
        default=0,
    )
    official_points = official_snapshot.get("points") or {}
    if set(official_points) != official_ids:
        raise RuntimeError("Official livestream rows and counter evidence do not match")
    for video_id, series in official_points.items():
        for point in series:
            if (
                not isinstance(point, (list, tuple))
                or len(point) != 2
                or positive_number(point[0]) is None
                or positive_number(point[1], allow_zero=True) is None
            ):
                raise RuntimeError(f"Official livestream {video_id} has invalid counter evidence")
            points[video_id].append((int(point[0]), int(point[1])))

    source_latest = max(
        (timestamp for series in points.values() for timestamp, _ in series),
        default=0,
    )
    cutoff_24h = source_latest - DAY_MS
    summary: dict[str, dict] = {}
    for video_id in sorted(live_ids):
        series = points.get(video_id) or []
        if not series:
            continue
        series.sort(key=lambda point: point[0])
        latest_t, now = series[-1]
        # A single collector observation is a current point, not a factual peak.
        # Wait for at least two distinct Sheet timestamps before exposing peaks.
        has_history_coverage = len(sheet_history_times.get(video_id, set())) >= 2
        peak_all = max((viewers for _, viewers in series), default=None) if has_history_coverage else None
        recent = (
            [viewers for timestamp, viewers in series if timestamp >= cutoff_24h]
            if has_history_coverage
            else []
        )
        summary[video_id] = {
            "now": now,
            "latestT": latest_t,
            "peak24": max(recent) if recent else None,
            "peakAll": peak_all,
            "active": (
                video_id not in ended_ids
                and (
                    (video_id in official_ids and now >= 0)
                    or (video_id not in official_ids and now > 0)
                )
                and source_latest - latest_t <= 3 * 3_600_000
            ),
        }

    if not lives or not summary or not source_latest:
        raise RuntimeError("Refusing to publish an empty livestream bootstrap")
    active = sum(1 for item in summary.values() if item.get("active") is True)
    official_metrics = official_snapshot.get("metrics") or {}
    official_expected = int(official_metrics.get("expected") or len(official_ids))
    official_verified = int(official_metrics.get("verified") or len(official_ids))
    official_observed = int(official_metrics.get("observedT") or 0)
    if official_expected != len(official_ids) or official_verified != len(official_ids):
        raise RuntimeError("Official livestream discovery proof is inconsistent")
    if official_ids and positive_number(official_observed) is None:
        raise RuntimeError("Official livestream discovery has no factual observation time")
    return {
        "t": source_latest,
        "d": {"lives": lives, "liveSummary": summary},
        "metrics": {
            "tracked": len(lives),
            "summarized": len(summary),
            "active": active,
            "sourceLatestT": source_latest,
            "sheetSourceLatestT": sheet_source_latest,
            "officialExpected": official_expected,
            "officialVerified": official_verified,
            "officialAdded": len(official_ids - sheet_live_ids),
            "officialUpdated": len(official_ids & sheet_live_ids),
            "officialObservedT": official_observed,
            "officialListingActive": int(official_metrics.get("listingActive") or 0),
            "officialStreamsTabCandidates": int(official_metrics.get("streamsTabCandidates") or 0),
            "officialStreamsTabActive": int(official_metrics.get("streamsTabActive") or 0),
            "officialStreamsTabUnknown": int(official_metrics.get("streamsTabUnknown") or 0),
            "officialListingUpcoming": int(official_metrics.get("listingUpcoming") or 0),
            "officialUploadsPlaylistActive": int(official_metrics.get("uploadsPlaylistActive") or 0),
            "officialListingConfirmedEnded": int(official_metrics.get("listingConfirmedEnded") or 0),
            "officialPreviousActive": int(official_metrics.get("previousActive") or 0),
            "officialMissingFromListing": int(official_metrics.get("missingFromListing") or 0),
            "officialRecoveredStillLive": int(official_metrics.get("recoveredStillLive") or 0),
            "officialConfirmedEnded": int(official_metrics.get("confirmedEnded") or 0),
            "officialLegacyRejected": int(official_metrics.get("legacyRejected") or 0),
            "officialWatchPageFallbacks": int(official_metrics.get("watchPageFallbacks") or 0),
            "officialNextFallbacks": int(official_metrics.get("nextFallbacks") or 0),
        },
    }


def read_workbook(source: str | None):
    if source:
        return openpyxl.load_workbook(source, read_only=True, data_only=True)
    request = urllib.request.Request(XLSX_URL, headers={"User-Agent": "LofiRadar/1.0"})
    with urllib.request.urlopen(request, timeout=90) as response:
        content = response.read()
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def write_payload(path: Path, payload: dict) -> None:
    content = (
        "window.LOFI_LIVE_DATA="
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def verify_published_asset(
    local_path: Path,
    base_url: str,
    *,
    timeout_seconds: float = 900,
    interval_seconds: float = 15,
) -> None:
    """Wait until Pages serves byte-for-byte the locally committed live asset."""
    if not local_path.is_file():
        raise RuntimeError(f"Local livestream asset is missing: {local_path}")
    parsed_base = urllib.parse.urlparse(base_url)
    if (
        parsed_base.scheme != "https"
        or not parsed_base.hostname
        or parsed_base.username is not None
        or parsed_base.password is not None
        or parsed_base.port is not None
        or parsed_base.query
        or parsed_base.fragment
    ):
        raise RuntimeError("Pages verification requires a plain HTTPS base URL")
    expected = local_path.read_bytes()
    if not expected:
        raise RuntimeError("Local livestream asset is empty")
    asset_url = urllib.parse.urljoin(
        base_url.rstrip("/") + "/",
        urllib.parse.quote(local_path.name),
    )
    parsed_asset = urllib.parse.urlparse(asset_url)
    deadline = time.monotonic() + max(1.0, float(timeout_seconds))
    interval = max(0.1, float(interval_seconds))
    last_error = "the public asset did not match"
    attempt = 0
    while True:
        attempt += 1
        query = urllib.parse.urlencode({"live_verify": f"{utc_now_ms()}-{attempt}"})
        request_url = asset_url + "?" + query
        request = urllib.request.Request(
            request_url,
            headers={"Cache-Control": "no-cache", "User-Agent": "LofiRadar/1.0"},
        )
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                final = urllib.parse.urlparse(str(response.geturl()))
                body = response.read(len(expected) + 1)
            if (
                final.scheme != parsed_asset.scheme
                or final.hostname != parsed_asset.hostname
                or final.port != parsed_asset.port
                or final.path != parsed_asset.path
                or final.params
                or final.fragment
                or final.query != query
            ):
                raise RuntimeError("Pages verification redirected away from the exact asset")
            if body == expected:
                return
            last_error = (
                f"public asset differs ({len(body)} bytes served, "
                f"{len(expected)} bytes expected)"
            )
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, OSError) as exc:
            last_error = f"{type(exc).__name__}: {exc}"
        if time.monotonic() >= deadline:
            raise RuntimeError(
                f"Timed out waiting for exact Pages livestream asset: {last_error}"
            )
        time.sleep(min(interval, max(0.0, deadline - time.monotonic())))


def ensure_sheet_freshness(
    payload: dict, max_age_hours: float, *, now_ms: int | None = None
) -> None:
    """Keep a fresh official observation from masking stale Sheet history."""
    if max_age_hours <= 0:
        return
    sheet_latest = int((payload.get("metrics") or {}).get("sheetSourceLatestT") or 0)
    if sheet_latest <= 0:
        raise RuntimeError("Livestream Sheet has no factual history timestamp")
    age_ms = int(now_ms or utc_now_ms()) - sheet_latest
    if age_ms > max_age_hours * 3_600_000:
        raise RuntimeError(
            f"Livestream Sheet is {age_ms / 3_600_000:.1f}h old; "
            "keeping the previous bootstrap"
        )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", help="Use a local workbook instead of the public Sheet")
    parser.add_argument("--snapshot", type=Path, default=Path("Lofi_Radar_data.js"))
    parser.add_argument("--output", type=Path, default=Path("Lofi_Radar_live_data.js"))
    parser.add_argument(
        "--previous-asset",
        type=Path,
        help="Prior live asset used for factual cohort continuity",
    )
    parser.add_argument("--max-age-hours", type=float, default=0)
    parser.add_argument("--verify-base-url")
    parser.add_argument("--verify-timeout", type=float, default=900)
    parser.add_argument("--verify-interval", type=float, default=15)
    args = parser.parse_args()

    if args.verify_base_url:
        verify_published_asset(
            args.output,
            args.verify_base_url,
            timeout_seconds=args.verify_timeout,
            interval_seconds=args.verify_interval,
        )
        print(f"Verified exact Pages livestream asset: {args.output.name}")
        return 0

    previous_active = load_previous_active_official_lives(
        args.previous_asset or args.output
    )
    official_snapshot = discover_official_live_streams(previous_active=previous_active)
    workbook = read_workbook(args.xlsx)
    try:
        payload = build_payload(workbook, args.snapshot, official_snapshot)
    finally:
        workbook.close()
    ensure_sheet_freshness(payload, args.max_age_hours)
    write_payload(args.output, payload)
    print(json.dumps(payload["metrics"], sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

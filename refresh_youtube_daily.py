"""Daily YouTube Radar refresh: discovery, current metrics and history.

The workflow runs this collector in deterministic shards, uploads one JSON
artifact per shard, then calls the same script with ``--merge-dir``.  Only the
merge phase writes the public snapshot, so partial scans never look fresh.
"""
from __future__ import annotations

import argparse
import atexit
import concurrent.futures
import hashlib
import http.client
import io
import json
import math
import os
import re
import shutil
import socket
import subprocess
import sys
import threading
import time
import urllib.parse
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from generate_youtube_recommendation_pool import POOL_PREFIX, write_recommendation_pool

ROOT = Path(__file__).resolve().parent
DEFAULT_SNAPSHOT = ROOT / "Lofi_Radar_data.js"
DEFAULT_AVATARS = ROOT / "Lofi_Radar_new_channel_avatars.js"
DEFAULT_HISTORY_DIR = ROOT / "video_history"
DEFAULT_RECOMMENDATION_POOL = ROOT / "Lofi_Radar_recommendation_pool.js"
RADAR_TIMEZONE_NAME = "Europe/Paris"
RADAR_TIMEZONE = ZoneInfo(RADAR_TIMEZONE_NAME)
DAILY_VIEW_HISTORY_START_MS = int(datetime(2026, 7, 20, tzinfo=RADAR_TIMEZONE).timestamp() * 1000)
SHEET_EXPORT = (
    "https://docs.google.com/spreadsheets/d/"
    "1XE_M9pQWn8w2Qu83vV_tv9sEFDFQ13fTePseG6mh1vI/export?format=xlsx"
)

MIN_SECONDS = 20 * 60
MIN_ALL_VIEWS = 1_000_000
MIN_TREND_VIEWS = 500_000
MAX_TREND_AGE_MONTHS = 12
MIN_NEWS_VIEWS = 100_000
MIN_NEWS_VPM = 10_000
MAX_NEWS_AGE_MONTHS = 3
MAX_NEWS_ROWS = 1_000
MIN_KIDS_VIEWS = 100_000
MIN_KIDS_VPM = 5_000
SCAN_SCOPES = ("all", "standard", "kids")
METADATA_SOURCE_YTDLP = "youtube_yt_dlp"
METADATA_SOURCE_SHEET = "dashboard_sheet"
METADATA_SOURCE_PRESERVED = "preserved_existing"
METADATA_SOURCE_API = "youtube_data_api"
METADATA_SOURCE_PRIORITY = {
    "": 0,
    METADATA_SOURCE_YTDLP: 1,
    METADATA_SOURCE_SHEET: 1,
    METADATA_SOURCE_PRESERVED: 2,
    METADATA_SOURCE_API: 3,
}
CARD_BUCKETS = ("all", "trends", "news", "ours", "kids")
CARD_METADATA_FIELDS = (
    "title", "url", "durH", "durationSource", "pub", "channel",
    "chUrl", "channelId", "subs",
)
SEARCH_RESULTS = int(os.environ.get("RADAR_SEARCH_RESULTS", "10"))
KIDS_SEARCH_RESULTS = int(os.environ.get("RADAR_KIDS_SEARCH_RESULTS", "100"))
KIDS_BOOTSTRAP_SEARCH_RESULTS = int(os.environ.get("RADAR_KIDS_BOOTSTRAP_RESULTS", "100"))
MAX_KIDS_SEARCH_CALLS = 80
KIDS_MIN_RESULTS_EXAMINED = 2_000
KIDS_SEARCH_LANES = ("viewCount", "relevance", "date")
KIDS_YTDLP_SEARCH_PARAMS = {
    # YouTube search protobuf filters: long videos with the requested order.
    "viewCount": "CAMSAhgC",
    "relevance": "EgIYAg%3D%3D",
    "date": "CAISBAgCEAE%3D",
}
KIDS_DOM_PAGE_LOAD_TIMEOUT_MS = int(os.environ.get("RADAR_KIDS_DOM_PAGE_TIMEOUT_MS", "15000"))
KIDS_DOM_SCRIPT_TIMEOUT_MS = int(os.environ.get("RADAR_KIDS_DOM_SCRIPT_TIMEOUT_MS", "5000"))
KIDS_DOM_HTTP_TIMEOUT_SECONDS = int(os.environ.get("RADAR_KIDS_DOM_HTTP_TIMEOUT_SECONDS", "20"))
KIDS_DOM_MARKER_WAIT_SECONDS = float(os.environ.get("RADAR_KIDS_DOM_MARKER_WAIT_SECONDS", "8"))
KIDS_CANARY_RETRIES = int(os.environ.get("RADAR_KIDS_CANARY_RETRIES", "2"))
KIDS_CANARY_RETRY_DELAY_SECONDS = float(
    os.environ.get("RADAR_KIDS_CANARY_RETRY_DELAY_SECONDS", "0.75")
)
KIDS_INNERTUBE_CLIENT_VERSION = (
    os.environ.get("RADAR_KIDS_INNERTUBE_CLIENT_VERSION", "21.02.35").strip()
    or "21.02.35"
)
KIDS_INNERTUBE_HTTP_RETRIES = int(
    os.environ.get("RADAR_KIDS_INNERTUBE_HTTP_RETRIES", "2")
)
KIDS_INNERTUBE_HTTP_TIMEOUT_SECONDS = int(
    os.environ.get("RADAR_KIDS_INNERTUBE_HTTP_TIMEOUT_SECONDS", "20")
)
KIDS_INNERTUBE_RETRY_DELAY_SECONDS = float(
    os.environ.get("RADAR_KIDS_INNERTUBE_RETRY_DELAY_SECONDS", "0.75")
)
KIDS_INNERTUBE_MAX_JSON_BYTES = int(
    os.environ.get("RADAR_KIDS_INNERTUBE_MAX_JSON_BYTES", str(4 * 1024 * 1024))
)
KIDS_NEXT_CLIENT_VERSION = "2.20260114.08.00"
KIDS_NEXT_HTTP_RETRIES = int(
    os.environ.get("RADAR_KIDS_NEXT_HTTP_RETRIES", "2")
)
KIDS_NEXT_HTTP_TIMEOUT_SECONDS = int(
    os.environ.get("RADAR_KIDS_NEXT_HTTP_TIMEOUT_SECONDS", "20")
)
KIDS_NEXT_RETRY_DELAY_SECONDS = float(
    os.environ.get("RADAR_KIDS_NEXT_RETRY_DELAY_SECONDS", "0.75")
)
KIDS_NEXT_MAX_JSON_BYTES = int(
    os.environ.get("RADAR_KIDS_NEXT_MAX_JSON_BYTES", str(2 * 1024 * 1024))
)
KIDS_WATCH_HTTP_RETRIES = int(
    os.environ.get("RADAR_KIDS_WATCH_HTTP_RETRIES", "2")
)
KIDS_WATCH_HTTP_TIMEOUT_SECONDS = int(
    os.environ.get("RADAR_KIDS_WATCH_HTTP_TIMEOUT_SECONDS", "30")
)
KIDS_WATCH_RETRY_DELAY_SECONDS = float(
    os.environ.get("RADAR_KIDS_WATCH_RETRY_DELAY_SECONDS", "0.75")
)
KIDS_WATCH_MAX_HTML_BYTES = int(
    os.environ.get("RADAR_KIDS_WATCH_MAX_HTML_BYTES", str(8 * 1024 * 1024))
)
TRACK_WORKERS = int(os.environ.get("RADAR_TRACK_WORKERS", "12"))
SEARCH_WORKERS = int(os.environ.get("RADAR_SEARCH_WORKERS", "4"))
MIN_TRACK_RATIO = 0.90
MIN_PUBLISH_TRACK_RATIO = 0.99
HISTORY_RETENTION_DAYS = 400
OWN_CHANNEL_HANDLES = ("@LofiGirl",)
OWN_UPLOADS_PER_CHANNEL = 50
THREAD = threading.local()
YOUTUBE_API_RETRIES = int(os.environ.get("RADAR_YOUTUBE_API_RETRIES", "1"))

# Genre words such as "hip hop" are intentionally not rejected: this is an
# instrumental long-form radar.  We reject explicit vocal/performance signals.
VOCAL = re.compile(
    r"\b(?:lyrics?|lyric\s+video|official\s+(?:music\s+)?video|music\s+video|"
    r"vocals?|vocal\s+(?:mix|edit|version)|singer|singing|sung|rap(?:ping)?|"
    r"feat(?:uring)?\.?|ft\.?|acap+ella|a\s+cappella|live\s+performance|concert|"
    r"sing[ -]?along|karaoke|story\s*time|storytelling|bedtime\s+story|spoken|"
    r"voice[ -]?over|voices?|choir|choral|humming|mantra|narrat(?:ion|ed|or)|"
    r"podcast|guided|affirmations?|chanting|asmr)\b",
    re.I,
)
KIDS_AMBIGUOUS = re.compile(r"\b(?:songs?|nursery\s+rhymes?|lullab(?:y|ies))\b", re.I)
KIDS_STRONG_INSTRUMENTAL = re.compile(
    r"\b(?:instrumental|no\s+(?:lyrics?|vocals?)|without\s+vocals?|music\s+box|"
    r"piano|classical|ambient|soundscape|lofi|lo[ -]?fi|jazz|bossa|guitar|"
    r"chill\s+house|drum\s+(?:and|&)\s+bass|dnb|synthwave|background\s+music|"
    r"sleep\s+music|study\s+music|focus\s+music|relaxing\s+music|"
    r"calming\s+music|soothing\s+music)\b",
    re.I,
)
KIDS_EXPLICIT_INSTRUMENTAL = re.compile(
    r"\b(?:instrumental|no\s+(?:lyrics?|vocals?|voices?)|without\s+(?:lyrics?|vocals?|voices?)|music\s+box)\b",
    re.I,
)
KIDS_CONTEXTUAL_MUSICAL_PROOF = re.compile(
    r"\b(?:piano|classical|mozart|brahms|music\s+box|ambient|soundscape|lofi|lo[ -]?fi|"
    r"jazz|bossa|guitar|chill\s+house|drum\s+(?:and|&)\s+bass|dnb|synthwave)\b",
    re.I,
)
KIDS_LULLABY_PROOF = re.compile(
    r"\b(?:piano|classical|mozart|brahms|music\s+box|instrumental|"
    r"no\s+(?:lyrics?|vocals?|voices?)|without\s+(?:lyrics?|vocals?|voices?))\b",
    re.I,
)
KIDS_HARD_AMBIGUOUS = re.compile(r"\b(?:songs?|nursery\s+rhymes?)\b", re.I)
KIDS_CONFIRMED_VOCAL_VIDEO_IDS = frozenset({
    "eNSCeIa5_5g",
    "Mi0XBUz562Y",
})
NEGATED_VOCAL = re.compile(
    r"\b(?:no|without)\s+(?:lyrics?|vocals?|voices?)\b",
    re.I,
)
VIDEO_ID = re.compile(r"^[\w-]{11}$")
CHANNEL_ID = re.compile(r"^UC[\w-]{22}$")
ISO_DURATION = re.compile(
    r"^P(?:(?P<days>\d+)D)?(?:T(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?)?$"
)
WATCH_LENGTH_SECONDS = re.compile(r'"lengthSeconds"\s*:\s*"(?P<seconds>\d+)"')
DEFERRED_GENRE = re.compile(r"\bphonk\b", re.I)

KIDS_QUERY_EXCLUSIONS = (
    "-singalong -karaoke -storytime -storytelling "
    "-spoken -podcast -guided -affirmations -chanting -asmr"
)
KIDS_QUERY_SPECS = (
    ("baby sleep music instrumental", "Baby sleep", "Relaxation / meditation"),
    ("newborn sleep music instrumental", "Baby sleep", "Relaxation / meditation"),
    ("infant sleep music instrumental", "Baby sleep", "Relaxation / meditation"),
    ("toddler sleep music instrumental", "Baby sleep", "Relaxation / meditation"),
    ("kids bedtime music instrumental", "Baby sleep", "Relaxation / meditation"),
    ("nap time music for kids instrumental", "Baby sleep", "Relaxation / meditation"),
    ("instrumental lullabies for babies", "Baby sleep", "Relaxation / meditation"),
    ("baby music box no vocals", "Baby sleep", "Relaxation / meditation"),
    ("piano lullabies for babies instrumental", "Piano", "Relaxation / meditation"),
    ("calming music for babies instrumental", "Baby sleep", "Relaxation / meditation"),
    ("soothing baby music no lyrics", "Baby sleep", "Relaxation / meditation"),
    ("baby sensory music instrumental", "Kids background", "Relaxation / meditation"),
    ("relaxing music for toddlers instrumental", "Baby sleep", "Relaxation / meditation"),
    ("calm down music for kids instrumental", "Kids background", "Relaxation / meditation"),
    ("kids study music instrumental", "Kids focus", "Study / focus / work"),
    ("homework music for kids no lyrics", "Kids focus", "Study / focus / work"),
    ("focus music for children instrumental", "Kids focus", "Study / focus / work"),
    ("classroom background music instrumental", "Kids focus", "Study / focus / work"),
    ("preschool quiet time music instrumental", "Kids focus", "Study / focus / work"),
    ("playtime music for kids instrumental", "Kids background", "Relaxation / meditation"),
    ("daycare background music instrumental", "Kids background", "Relaxation / meditation"),
    ("Montessori music for children instrumental", "Kids background", "Study / focus / work"),
    ("reading music for kids instrumental", "Kids focus", "Study / focus / work"),
    ("drawing music for kids instrumental", "Kids focus", "Study / focus / work"),
    ("kids yoga music instrumental", "Kids background", "Relaxation / meditation"),
    ("baby lofi sleep music instrumental", "Lofi / chillhop", "Relaxation / meditation"),
    ("kids lofi study beats instrumental", "Lofi / chillhop", "Study / focus / work"),
    ("jazz for babies instrumental", "Jazz", "Relaxation / meditation"),
    ("classical music for babies instrumental", "Classical", "Relaxation / meditation"),
    ("piano music for babies instrumental", "Piano", "Relaxation / meditation"),
    ("ambient music for babies sleep instrumental", "Ambient", "Relaxation / meditation"),
    ("baby sleep music 8 hours instrumental", "Baby sleep", "Relaxation / meditation"),
    ("baby sleep music all night no vocals", "Baby sleep", "Relaxation / meditation"),
    ("toddler bedtime music 3 hours instrumental", "Baby sleep", "Relaxation / meditation"),
    ("kids study music 1 hour instrumental", "Kids focus", "Study / focus / work"),
    ("bossa nova for babies instrumental", "Jazz", "Relaxation / meditation"),
    ("guitar music for babies instrumental", "Guitar", "Relaxation / meditation"),
    ("chill house for kids instrumental", "Chill house", "Relaxation / meditation"),
    ("drum and bass for kids instrumental", "Drum & Bass", "Gaming / night drive"),
    ("synthwave for kids instrumental", "Synthwave", "Gaming / night drive"),
)
KIDS_DOM_POSITIVE_CANARIES = ("L1Y-GbKA0PM", "qXcMNBQnQMM")
KIDS_DOM_NEGATIVE_CANARY = "XVFUtEh9zrY"

_KIDS_DOM_VALIDATOR = None
_KIDS_DOM_VALIDATOR_LOCK = threading.Lock()


def is_deferred_row(row: dict) -> bool:
    """Phonk is paused in the YouTube roadmap and must not re-enter the radar."""
    fields = ("genre", "title", "kw", "disc", "niche", "concept", "scene", "style", "channel")
    return bool(DEFERRED_GENRE.search(" ".join(str(row.get(field) or "") for field in fields)))


def prune_deferred_rows(data: dict) -> set[str]:
    """Remove paused genres from public buckets and return their video IDs."""
    dropped: set[str] = set()
    for bucket in ("all", "trends", "news", "ours", "kids"):
        visible = []
        for row in list(data.get(bucket) or []):
            if is_deferred_row(row):
                video_id = str(row.get("vid") or "")
                if VIDEO_ID.match(video_id):
                    dropped.add(video_id)
                continue
            visible.append(row)
        data[bucket] = visible
    for bucket in ("recos", "roadmap", "lives"):
        data[bucket] = [
            row for row in (data.get(bucket) or [])
            if not is_deferred_row(row) and str(row.get("vid") or "") not in dropped
        ]
    for bucket in ("hist", "liveHist", "liveHourly"):
        history = data.get(bucket)
        if isinstance(history, dict):
            for video_id in dropped:
                history.pop(video_id, None)
    return dropped


def prune_news_below_view_floor(data: dict) -> int:
    """Keep daily discoveries only once they reach the public view floor."""
    rows = list(data.get("news") or [])
    data["news"] = [
        row for row in rows if int(row.get("views") or 0) >= MIN_NEWS_VIEWS
    ]
    return len(rows) - len(data["news"])


def utc_now_ms() -> int:
    return int(datetime.now(timezone.utc).timestamp() * 1000)


def atomic_write_text(path: Path, content: str) -> None:
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(content, encoding="utf-8")
    tmp.replace(path)


def youtube_api_json(url: str, *, timeout: int = 30) -> dict:
    """Read JSON with one bounded retry for transient API failures."""
    for attempt in range(YOUTUBE_API_RETRIES + 1):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as response:
                return json.load(response)
        except urllib.error.HTTPError as exc:
            try:
                error_payload = json.loads(exc.read().decode("utf-8", "ignore"))
                reasons = {
                    str(item.get("reason") or "")
                    for item in ((error_payload.get("error") or {}).get("errors") or [])
                }
            except (ValueError, AttributeError):
                reasons = set()
            retryable = (
                exc.code == 429
                or exc.code >= 500
                or "rateLimitExceeded" in reasons
            ) and "quotaExceeded" not in reasons
            if not retryable or attempt >= YOUTUBE_API_RETRIES:
                raise
            retry_after = str((exc.headers or {}).get("Retry-After") or "")
            try:
                delay = max(float(retry_after), 0.5)
            except ValueError:
                delay = 1.0 + attempt * 0.75
            time.sleep(delay)
        except (urllib.error.URLError, TimeoutError):
            if attempt >= YOUTUBE_API_RETRIES:
                raise
            time.sleep(1.0 + attempt * 0.75)
    raise RuntimeError("YouTube API retry loop exhausted")


def parse_snapshot_text(raw: str) -> dict:
    return json.loads(re.sub(r"^window\.LOFI_DATA=", "", raw).rstrip(";\n"))


def read_snapshot(path: Path) -> dict:
    return parse_snapshot_text(path.read_text(encoding="utf-8"))


def write_snapshot(path: Path, payload: dict) -> None:
    atomic_write_text(
        path,
        "window.LOFI_DATA="
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";",
    )


def split_keywords(value: object) -> list[str]:
    return [part.strip() for part in re.split(r"\s*;\s*|[\r\n]+", str(value or "")) if part.strip()]


def parse_nonnegative_count(value: object) -> int | None:
    """Accept exact public counters without coercing missing or invalid values to zero."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value >= 0 else None
    if isinstance(value, float):
        return int(value) if math.isfinite(value) and value >= 0 and value.is_integer() else None
    text = str(value).strip()
    return int(text) if re.fullmatch(r"\d+", text) else None


def stable_shard(value: str, shards: int) -> int:
    digest = hashlib.sha256(value.encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") % shards


def to_ms(info: dict) -> int | None:
    stamp = info.get("timestamp") or info.get("release_timestamp")
    if isinstance(stamp, (int, float)):
        return int(stamp * 1000)
    date = str(info.get("upload_date") or "")
    try:
        return int(datetime.strptime(date, "%Y%m%d").replace(tzinfo=timezone.utc).timestamp() * 1000)
    except ValueError:
        return None


def parse_iso_duration(value: str | None) -> float | None:
    match = ISO_DURATION.match(value or "")
    if not match:
        return None
    parts = {key: int(number or 0) for key, number in match.groupdict().items()}
    return float(
        parts["days"] * 86400
        + parts["hours"] * 3600
        + parts["minutes"] * 60
        + parts["seconds"]
    )


def parse_watch_duration(value: str) -> float | None:
    """Read YouTube's public lengthSeconds value without guessing a duration."""
    match = WATCH_LENGTH_SECONDS.search(value or "")
    if not match:
        return None
    seconds = int(match.group("seconds"))
    return float(seconds) if seconds > 0 else None


def fetch_public_duration(video_id: str) -> float | None:
    request = urllib.request.Request(
        f"https://www.youtube.com/watch?v={video_id}",
        headers={"User-Agent": "Mozilla/5.0", "Accept-Language": "en-US,en;q=0.8"},
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        return parse_watch_duration(response.read().decode("utf-8", "ignore"))


def owned_genre_from_title(title: object) -> str:
    """Classify only explicit genre words in an official upload title."""
    text = re.sub(r"[-_]", " ", str(title or "").casefold())
    if re.search(r"\b(?:halloween)\b", text) and re.search(r"\b(?:lofi|lo fi|chillhop)\b", text):
        return "Halloween Lofi"
    if re.search(r"\b(?:christmas|xmas)\b", text) and re.search(r"\b(?:lofi|lo fi|chillhop)\b", text):
        return "Christmas Lofi"
    rules = (
        (r"\b(?:lofi|lo fi|chillhop)\b", "Lofi / chillhop"),
        (r"\b(?:synthwave|retrowave|outrun)\b", "Synthwave"),
        (r"\b(?:drum and bass|drum bass|dnb|liquid jungle)\b", "Drum & Bass"),
        (r"\b(?:chill house|lofi house|lo fi house|deep house|ambient house|downtempo house)\b", "Chill house"),
        (r"\b(?:jazz|jazzhop|bossa)\b", "Jazz"),
        (r"\b(?:classical|baroque)\b", "Classical"),
        (r"\b(?:guitar|acoustic|fingerstyle)\b", "Guitar"),
        (r"\b(?:piano)\b", "Piano"),
        (r"\b(?:ambient|soundscape)\b", "Ambient"),
    )
    return next((label for pattern, label in rules if re.search(pattern, text)), "")


def kids_genre_from_metadata(row: dict) -> str:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("title", "_scanDescription", "_scanTags")
    )
    explicit = owned_genre_from_title(text)
    if explicit:
        return explicit
    if re.search(r"\bmusic\s+box\b", text, re.I):
        return "Piano"
    return "Other / multi-genre"


def add_owned_metadata(row: dict) -> None:
    if not row.get("genre"):
        genre = owned_genre_from_title(row.get("title"))
        if genre:
            row["genre"] = genre
            row["genreSource"] = "title_explicit"


def age_months(published_ms: int | None, now_ms: int) -> float | None:
    if not published_ms:
        return None
    return max((now_ms - published_ms) / 2_629_746_000, 0.1)


def cluster_for(title: str, fallback: str = "") -> str:
    text = title.lower()
    if any(word in text for word in ("baby", "sleep", "bedtime", "lullaby", "nap time", "all night")):
        return "Sleep / night"
    if any(word in text for word in ("work", "focus", "study", "coding", "office")):
        return "Study / focus / work"
    if any(word in text for word in ("drive", "car", "night", "drift", "gym", "gaming")):
        return "Gaming / night drive"
    return fallback or "Relaxation / meditation"


def is_instrumental(info: dict) -> bool:
    duration = info.get("duration")
    if not isinstance(duration, (int, float)) or duration < MIN_SECONDS:
        return False
    text = " ".join(str(info.get(key) or "") for key in ("title", "description", "channel", "uploader"))
    return not has_vocal_signal(text)


def has_vocal_signal(text: str) -> bool:
    return bool(VOCAL.search(NEGATED_VOCAL.sub("", text or "")))


def kids_instrumental_evidence(row: dict) -> str:
    """Return the auditable proof used by the final fail-closed Kids gate."""
    if str(row.get("vid") or "") in KIDS_CONFIRMED_VOCAL_VIDEO_IDS:
        return ""
    duration_hours = row.get("durH")
    if not isinstance(duration_hours, (int, float)) or duration_hours * 3600 < MIN_SECONDS:
        return ""
    content_text = " ".join(
        str(row.get(key) or "")
        for key in ("title", "_scanDescription", "_scanTags")
    )
    negative_text = content_text + " " + str(row.get("channel") or "")
    if has_vocal_signal(negative_text):
        return ""
    strong = bool(KIDS_STRONG_INSTRUMENTAL.search(content_text))
    explicit = bool(KIDS_EXPLICIT_INSTRUMENTAL.search(content_text))
    # ``songs`` and ``nursery rhymes`` remain fail-closed because those labels
    # overwhelmingly denote vocal programmes even on Made-for-Kids videos.
    if KIDS_HARD_AMBIGUOUS.search(content_text):
        return ""
    if explicit and strong:
        return "metadata_explicit"
    # Lullabies can be purely instrumental, but need a musical proof stronger
    # than the word itself (piano/classical/composer/music-box or explicit).
    if KIDS_AMBIGUOUS.search(content_text) and not KIDS_LULLABY_PROOF.search(content_text):
        return ""
    if strong and KIDS_CONTEXTUAL_MUSICAL_PROOF.search(content_text):
        return "made_for_kids_contextual_metadata"
    return ""


def is_kids_instrumental(row: dict) -> bool:
    """Fail closed for Kids: long-form, no vocal signals and clear instrumental metadata."""
    return bool(kids_instrumental_evidence(row))


def is_kids_marker_href(href: object) -> bool:
    """Recognize only YouTube's two rendered Family Options destinations."""
    try:
        parsed = urllib.parse.urlparse(str(href or ""))
    except ValueError:
        return False
    host = (parsed.hostname or "").casefold()
    if parsed.scheme != "https":
        return False
    if host == "ytkids.app.goo.gl":
        return True
    return (
        host in {"youtube.com", "www.youtube.com"}
        and parsed.path == "/myfamily/"
        and parsed.fragment == "mf-compare"
    )


class KidsDomCanaryError(RuntimeError):
    """No public YouTube source proved a trustworthy Kids signal."""


class KidsDomProbeError(RuntimeError):
    """A public YouTube source could not produce a trustworthy yes/no answer."""


class KidsPlayerIndeterminateError(KidsDomProbeError):
    """A retryable public player response was not conclusive."""


class YouTubeInnertubePlayerClient:
    """Validate Kids restrictions in YouTube's public Android player response."""

    _ENDPOINT = "https://www.youtube.com/youtubei/v1/player?prettyPrint=false"
    _TRANSIENT_HTTP_CODES = frozenset({403, 408, 425, 429, 500, 502, 503, 504})
    _KIDS_RESTRICTION_TEXTS = frozenset({
        "Miniplayer is off for videos made for kids",
        "Miniplayer is off for videos made for kids. Tap play to resume",
    })
    _PAUSED_MODE = "PLAYBACK_MODE_PAUSED_ONLY"
    _ALLOW_MODE = "PLAYBACK_MODE_ALLOW"

    def __init__(
        self,
        *,
        retries: int = KIDS_INNERTUBE_HTTP_RETRIES,
        retry_delay_seconds: float = KIDS_INNERTUBE_RETRY_DELAY_SECONDS,
        timeout_seconds: int = KIDS_INNERTUBE_HTTP_TIMEOUT_SECONDS,
        client_version: str = KIDS_INNERTUBE_CLIENT_VERSION,
    ) -> None:
        self.retries = max(0, int(retries))
        self.retry_delay_seconds = max(0.0, float(retry_delay_seconds))
        self.timeout_seconds = max(1, int(timeout_seconds))
        self.client_version = str(client_version).strip()
        if not re.fullmatch(r"\d+(?:\.\d+)+", self.client_version):
            raise ValueError(
                f"Invalid Innertube Android client version: {client_version!r}"
            )

    def _request(self, video_id: str) -> urllib.request.Request:
        payload = {
            "context": {
                "client": {
                    "clientName": "ANDROID",
                    "clientVersion": self.client_version,
                    "androidSdkVersion": 35,
                    "hl": "en",
                    "gl": "US",
                },
            },
            "videoId": video_id,
            "contentCheckOk": True,
            "racyCheckOk": True,
        }
        return urllib.request.Request(
            self._ENDPOINT,
            data=json.dumps(payload, separators=(",", ":")).encode("utf-8"),
            headers={
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
                "Content-Type": "application/json",
                "User-Agent": (
                    f"com.google.android.youtube/{self.client_version} "
                    "(Linux; U; Android 15) gzip"
                ),
                "X-YouTube-Client-Name": "3",
                "X-YouTube-Client-Version": self.client_version,
            },
            method="POST",
        )

    @staticmethod
    def _strict_json(raw: bytes) -> dict:
        def object_without_duplicates(pairs: list[tuple[str, object]]) -> dict:
            result: dict = {}
            for key, value in pairs:
                if key in result:
                    raise ValueError(f"duplicate JSON key: {key}")
                result[key] = value
            return result

        def reject_constant(value: str) -> None:
            raise ValueError(f"invalid JSON constant: {value}")

        try:
            document = json.loads(
                raw.decode("utf-8"),
                object_pairs_hook=object_without_duplicates,
                parse_constant=reject_constant,
            )
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            raise KidsPlayerIndeterminateError(
                f"YouTube player returned malformed JSON: {exc}"
            ) from exc
        if type(document) is not dict:
            raise KidsPlayerIndeterminateError(
                "YouTube player returned a non-object JSON response"
            )
        return document

    @classmethod
    def _validate_final_url(cls, final_url: str) -> None:
        try:
            parsed = urllib.parse.urlparse(final_url)
            port = parsed.port
        except ValueError as exc:
            raise KidsPlayerIndeterminateError(
                "YouTube player returned a malformed final URL"
            ) from exc
        query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
        if (
            parsed.scheme != "https"
            or (parsed.hostname or "").casefold() != "www.youtube.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.path != "/youtubei/v1/player"
            or parsed.fragment
            or query != [("prettyPrint", "false")]
        ):
            raise KidsPlayerIndeterminateError(
                "YouTube player redirected away from the expected endpoint"
            )

    def _fetch_once(self, video_id: str) -> tuple[dict, str]:
        request = self._request(video_id)
        with urllib.request.urlopen(
            request, timeout=self.timeout_seconds
        ) as response:
            final_url = (
                str(response.geturl())
                if callable(getattr(response, "geturl", None))
                else request.full_url
            )
            body = response.read(KIDS_INNERTUBE_MAX_JSON_BYTES + 1)
        if len(body) > KIDS_INNERTUBE_MAX_JSON_BYTES:
            raise KidsPlayerIndeterminateError(
                "YouTube player response exceeded the bounded JSON size"
            )
        self._validate_final_url(final_url)
        return self._strict_json(body), final_url

    @staticmethod
    def _mapping_at(value: object, *keys: str) -> dict:
        current = value
        for key in keys:
            if type(current) is not dict:
                return {}
            current = current.get(key)
        return current if type(current) is dict else {}

    @classmethod
    def _has_support_answer(cls, value: object) -> bool:
        if not isinstance(value, str):
            return False
        try:
            parsed = urllib.parse.urlparse(value)
            port = parsed.port
        except ValueError:
            return False
        if (
            parsed.scheme not in {"", "https"}
            or (not parsed.scheme and not value.startswith("//"))
            or (parsed.hostname or "").casefold() != "support.google.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.fragment
        ):
            return False
        if parsed.path == "/youtube/answer/9632097":
            return True
        answers = urllib.parse.parse_qs(
            parsed.query, keep_blank_values=True
        ).get("answer")
        return (
            parsed.path == "/youtube/bin/answer.py"
            and answers == ["9632097"]
        )

    @classmethod
    def _classify(cls, payload: dict, final_url: str, video_id: str) -> bool:
        cls._validate_final_url(final_url)
        return cls._classify_player_payload(payload, video_id)

    @classmethod
    def _classify_player_payload(cls, payload: dict, video_id: str) -> bool:
        """Classify only the official miniplayer restriction triple.

        Android and the public web watch bootstrap expose the same restriction,
        but use different renderer paths and text shapes.  Keep those two exact
        shapes explicit so unrelated page copy can never become Kids evidence.
        """
        details = payload.get("videoDetails")
        playability = payload.get("playabilityStatus")
        if type(details) is not dict or type(playability) is not dict:
            raise KidsPlayerIndeterminateError(
                "YouTube player response is missing top-level player objects"
            )
        if details.get("videoId") != video_id:
            raise KidsPlayerIndeterminateError(
                f"YouTube player videoDetails mismatch for {video_id}: "
                f"{details.get('videoId')!r}"
            )
        if playability.get("status") != "OK":
            raise KidsPlayerIndeterminateError(
                f"YouTube player status is not OK for {video_id}: "
                f"{playability.get('status')!r}"
            )
        renderer = cls._mapping_at(
            playability, "miniplayer", "miniplayerRenderer"
        )
        if not renderer:
            raise KidsPlayerIndeterminateError(
                f"YouTube player has no miniplayerRenderer for {video_id}"
            )
        mode = renderer.get("playbackMode")
        notification_candidates = [
            cls._mapping_at(
                renderer,
                "minimizedEndpoint",
                "addToToastAction",
                "item",
                "notificationActionRenderer",
            ),
            cls._mapping_at(
                renderer,
                "minimizedEndpoint",
                "openPopupAction",
                "popup",
                "notificationActionRenderer",
            ),
        ]
        notifications = [value for value in notification_candidates if value]
        if len(notifications) > 1:
            raise KidsPlayerIndeterminateError(
                f"YouTube player has ambiguous Kids notifications for {video_id}"
            )
        notification = notifications[0] if notifications else {}
        response_text = cls._mapping_at(notification, "responseText")
        runs = response_text.get("runs")
        runs_signal = (
            isinstance(runs, list)
            and any(
                type(run) is dict
                and run.get("text") in cls._KIDS_RESTRICTION_TEXTS
                for run in runs
            )
        )
        simple_text = response_text.get("simpleText")
        text_signal = (
            runs_signal
            or (
                isinstance(simple_text, str)
                and simple_text in cls._KIDS_RESTRICTION_TEXTS
            )
        )
        button = cls._mapping_at(
            notification, "actionButton", "buttonRenderer"
        )
        support_urls = []
        for endpoint_name in ("navigationEndpoint", "command"):
            value = cls._mapping_at(
                button, endpoint_name, "urlEndpoint"
            ).get("url")
            if isinstance(value, str):
                support_urls.append(value)
        support_signal = (
            len(support_urls) == 1
            and cls._has_support_answer(support_urls[0])
        )
        signals = (
            mode == cls._PAUSED_MODE,
            text_signal,
            support_signal,
        )
        signal_count = sum(signals)
        if mode == cls._PAUSED_MODE and signal_count == len(signals):
            return True
        if mode == cls._ALLOW_MODE and signal_count == 0:
            return False
        raise KidsPlayerIndeterminateError(
            f"YouTube player Kids restrictions are inconclusive for {video_id}: "
            f"mode={mode!r}, signals={signal_count}/{len(signals)}"
        )

    def has_kids_player_signals(self, video_id: str) -> bool:
        if not VIDEO_ID.fullmatch(video_id):
            raise KidsDomProbeError(f"Invalid YouTube video ID: {video_id!r}")
        for attempt in range(self.retries + 1):
            try:
                payload, final_url = self._fetch_once(video_id)
                return self._classify(payload, final_url, video_id)
            except urllib.error.HTTPError as exc:
                if (
                    exc.code not in self._TRANSIENT_HTTP_CODES
                    or attempt >= self.retries
                ):
                    raise KidsDomProbeError(
                        f"YouTube player HTTP {exc.code} for {video_id}"
                    ) from exc
            except (
                urllib.error.URLError,
                TimeoutError,
                OSError,
                http.client.IncompleteRead,
            ) as exc:
                if attempt >= self.retries:
                    raise KidsDomProbeError(
                        f"YouTube player network failure for {video_id}: "
                        f"{type(exc).__name__}: {exc}"
                    ) from exc
            except KidsPlayerIndeterminateError:
                if attempt >= self.retries:
                    raise
            if self.retry_delay_seconds:
                time.sleep(self.retry_delay_seconds * (2 ** attempt))
        raise KidsDomProbeError("YouTube player retry loop exhausted")

    def close(self) -> None:
        pass


class YouTubeWatchNextClient:
    """Classify Kids from YouTube's complete public WEB ``next`` response."""

    _ENDPOINT = "https://www.youtube.com/youtubei/v1/next?prettyPrint=false"
    _TRANSIENT_HTTP_CODES = YouTubeInnertubePlayerClient._TRANSIENT_HTTP_CODES
    _KIDS_NOTIFICATION_TEXT = "This action is turned off for content made for kids"

    def __init__(
        self,
        *,
        retries: int = KIDS_NEXT_HTTP_RETRIES,
        retry_delay_seconds: float = KIDS_NEXT_RETRY_DELAY_SECONDS,
        timeout_seconds: int = KIDS_NEXT_HTTP_TIMEOUT_SECONDS,
    ) -> None:
        self.retries = max(0, int(retries))
        self.retry_delay_seconds = max(0.0, float(retry_delay_seconds))
        self.timeout_seconds = max(1, int(timeout_seconds))

    def _request(self, video_id: str) -> urllib.request.Request:
        payload = {
            "context": {
                "client": {
                    "clientName": "WEB",
                    "clientVersion": KIDS_NEXT_CLIENT_VERSION,
                    "hl": "en",
                    "gl": "US",
                },
            },
            "videoId": video_id,
        }
        return urllib.request.Request(
            self._ENDPOINT,
            data=json.dumps(payload, separators=(",", ":")).encode("utf-8"),
            headers={
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
                "Content-Type": "application/json",
                "User-Agent": (
                    "Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/140.0.0.0 Safari/537.36"
                ),
                "X-YouTube-Client-Name": "1",
                "X-YouTube-Client-Version": KIDS_NEXT_CLIENT_VERSION,
            },
            method="POST",
        )

    @classmethod
    def _validate_final_url(cls, final_url: str) -> None:
        try:
            parsed = urllib.parse.urlparse(final_url)
            port = parsed.port
        except ValueError as exc:
            raise KidsPlayerIndeterminateError(
                "YouTube next returned a malformed final URL"
            ) from exc
        query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
        if (
            parsed.scheme != "https"
            or (parsed.hostname or "").casefold() != "www.youtube.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.path != "/youtubei/v1/next"
            or parsed.fragment
            or query != [("prettyPrint", "false")]
        ):
            raise KidsPlayerIndeterminateError(
                "YouTube next redirected away from the expected endpoint"
            )

    def _fetch_once(self, video_id: str) -> tuple[dict, str]:
        request = self._request(video_id)
        with urllib.request.urlopen(
            request, timeout=self.timeout_seconds
        ) as response:
            final_url = (
                str(response.geturl())
                if callable(getattr(response, "geturl", None))
                else request.full_url
            )
            body = response.read(KIDS_NEXT_MAX_JSON_BYTES + 1)
        if len(body) > KIDS_NEXT_MAX_JSON_BYTES:
            raise KidsPlayerIndeterminateError(
                "YouTube next response exceeded the bounded JSON size"
            )
        self._validate_final_url(final_url)
        return YouTubeInnertubePlayerClient._strict_json(body), final_url

    @staticmethod
    def _mapping_at(value: object, *keys: str) -> dict:
        return YouTubeInnertubePlayerClient._mapping_at(value, *keys)

    @staticmethod
    def _list_at(value: object, *keys: str) -> list:
        current = value
        for key in keys:
            if type(current) is not dict:
                return []
            current = current.get(key)
        return current if type(current) is list else []

    @classmethod
    def _complete_watch_content(
        cls, payload: dict, video_id: str
    ) -> tuple[list, dict]:
        current = cls._mapping_at(payload, "currentVideoEndpoint")
        web = cls._mapping_at(
            current, "commandMetadata", "webCommandMetadata"
        )
        watch = cls._mapping_at(current, "watchEndpoint")
        web_url = web.get("url")
        try:
            web_parsed = urllib.parse.urlparse(web_url)
            web_query = urllib.parse.parse_qsl(
                web_parsed.query, keep_blank_values=True
            )
        except (TypeError, ValueError) as exc:
            raise KidsPlayerIndeterminateError(
                f"YouTube next has malformed currentVideoEndpoint URL for {video_id}"
            ) from exc
        query_keys = [key for key, _ in web_query]
        web_video_values = [value for key, value in web_query if key == "v"]
        web_url_ok = (
            web_parsed.scheme == ""
            and web_parsed.netloc == ""
            and web_parsed.path == "/watch"
            and web_parsed.params == ""
            and web_parsed.fragment == ""
            and web_video_values == [video_id]
            and query_keys.count("v") == 1
            and query_keys.count("pp") <= 1
            and all(key in {"v", "pp"} for key in query_keys)
            and all(value for key, value in web_query if key == "pp")
        )
        if (
            watch.get("videoId") != video_id
            or not web_url_ok
            or web.get("webPageType") != "WEB_PAGE_TYPE_WATCH"
        ):
            raise KidsPlayerIndeterminateError(
                f"YouTube next currentVideoEndpoint mismatch for {video_id}"
            )

        two_column = cls._mapping_at(
            payload, "contents", "twoColumnWatchNextResults"
        )
        primary = cls._list_at(
            two_column, "results", "results", "contents"
        )
        primary_results = cls._mapping_at(two_column, "results", "results")
        secondary_results = cls._mapping_at(
            two_column,
            "secondaryResults",
            "secondaryResults",
        )
        secondary = secondary_results.get("results")
        secondary_renderer_names = {
            "compactVideoRenderer",
            "compactPlaylistRenderer",
            "compactRadioRenderer",
            "continuationItemRenderer",
            "lockupViewModel",
            "reelShelfRenderer",
        }
        secondary_is_complete = (
            type(secondary) is list
            and bool(secondary)
            and isinstance(secondary_results.get("trackingParams"), str)
            and bool(secondary_results.get("trackingParams"))
            and secondary_results.get("targetId") == "watch-next-feed"
            and all(
                type(item) is dict
                and len(item) == 1
                and next(iter(item)) in secondary_renderer_names
                and type(next(iter(item.values()))) is dict
                and bool(next(iter(item.values())))
                for item in secondary
            )
        )
        if (
            not primary
            or not isinstance(primary_results.get("trackingParams"), str)
            or not primary_results.get("trackingParams")
            or not secondary_is_complete
        ):
            raise KidsPlayerIndeterminateError(
                f"YouTube next lacks complete primary/secondary content for {video_id}"
            )
        primary_info = [
            item.get("videoPrimaryInfoRenderer")
            for item in primary
            if type(item) is dict
            and type(item.get("videoPrimaryInfoRenderer")) is dict
        ]
        secondary_info = [
            item.get("videoSecondaryInfoRenderer")
            for item in primary
            if type(item) is dict
            and type(item.get("videoSecondaryInfoRenderer")) is dict
        ]
        if len(primary_info) != 1 or len(secondary_info) != 1:
            raise KidsPlayerIndeterminateError(
                f"YouTube next has {len(primary_info)} videoPrimaryInfoRenderer and "
                f"{len(secondary_info)} videoSecondaryInfoRenderer objects for {video_id}"
            )
        owner = cls._mapping_at(
            secondary_info[0], "owner", "videoOwnerRenderer"
        )
        if not primary_info[0] or not owner:
            raise KidsPlayerIndeterminateError(
                f"YouTube next has incomplete primary/secondary info for {video_id}"
            )
        return primary, secondary_info[0]

    @classmethod
    def _notification_renderers(cls, secondary_info: dict) -> list[dict]:
        notifications: list[tuple[tuple[object, ...], dict]] = []

        def visit(value: object, path: tuple[object, ...] = ()) -> None:
            if type(value) is dict:
                for key, child in value.items():
                    child_path = path + (key,)
                    if key == "notificationActionRenderer" and type(child) is dict:
                        notifications.append((child_path, child))
                    visit(child, child_path)
            elif type(value) is list:
                for index, child in enumerate(value):
                    visit(child, path + (index,))

        visit(secondary_info)
        misplaced = [
            path
            for path, _ in notifications
            if not cls._is_allowed_notification_path(path)
        ]
        if misplaced:
            raise KidsPlayerIndeterminateError(
                "YouTube next has notificationActionRenderer outside approved "
                f"subscription paths: {misplaced[0]!r}"
            )
        return [renderer for _, renderer in notifications]

    @classmethod
    def _notification_without_tracking(cls, value: object) -> object:
        """Remove only opaque request-specific tokens before deduplication."""
        if type(value) is dict:
            return {
                key: cls._notification_without_tracking(child)
                for key, child in value.items()
                if key not in {"trackingParams", "clickTrackingParams"}
            }
        if type(value) is list:
            return [cls._notification_without_tracking(child) for child in value]
        return value

    @staticmethod
    def _path_matches(
        path: tuple[object, ...], pattern: tuple[object, ...]
    ) -> bool:
        return len(path) == len(pattern) and all(
            isinstance(actual, int) if expected is int else actual == expected
            for actual, expected in zip(path, pattern)
        )

    @classmethod
    def _is_allowed_notification_path(cls, path: tuple[object, ...]) -> bool:
        legacy = (
            "subscribeButton", "subscribeButtonRenderer",
            "notificationPreferenceButton",
            "subscriptionNotificationToggleButtonRenderer", "command",
            "commandExecutorCommand", "commands", int,
            "openPopupAction", "popup", "menuPopupRenderer", "items", int,
            "menuServiceItemRenderer", "command", "signalServiceEndpoint",
            "actions", int, "openPopupAction", "popup",
            "notificationActionRenderer",
        )
        modern_tail = (
            "panelLoadingStrategy", "inlineContent", "dialogViewModel",
            "customContent", "listViewModel", "listItems", int,
            "listItemViewModel", "trailingButtons", "buttons", int,
            "subscribeButtonViewModel", "onShowSubscriptionOptions",
            "innertubeCommand", "showSheetCommand", "panelLoadingStrategy",
            "inlineContent", "sheetViewModel", "content", "listViewModel",
            "listItems", int, "listItemViewModel", "rendererContext",
            "commandContext", "onTap", "innertubeCommand",
            "signalServiceEndpoint", "actions", int, "openPopupAction",
            "popup", "notificationActionRenderer",
        )
        modern_prefixes = (
            (
                "owner", "videoOwnerRenderer", "navigationEndpoint",
                "showDialogCommand",
            ),
            (
                "owner", "videoOwnerRenderer", "attributedTitle",
                "commandRuns", int, "onTap", "innertubeCommand",
                "showDialogCommand",
            ),
            (
                "subscribeButton", "subscribeButtonRenderer",
                "onSubscribeEndpoints", int, "showDialogCommand",
            ),
            (
                "subscribeButton", "subscribeButtonRenderer",
                "onUnsubscribeEndpoints", int, "showDialogCommand",
            ),
        )
        return cls._path_matches(path, legacy) or any(
            cls._path_matches(path, prefix + modern_tail)
            for prefix in modern_prefixes
        )

    @classmethod
    def _notification_signals(cls, secondary_info: dict) -> tuple[bool, bool]:
        notifications = cls._notification_renderers(secondary_info)
        canonical = {
            json.dumps(
                cls._notification_without_tracking(value),
                sort_keys=True,
                separators=(",", ":"),
            )
            for value in notifications
        }
        if len(canonical) > 1:
            raise KidsPlayerIndeterminateError(
                "YouTube next has contradictory Kids notification renderers"
            )
        if not canonical:
            return False, False
        notification = json.loads(next(iter(canonical)))
        text = cls._mapping_at(notification, "responseText").get("simpleText")
        text_signal = text == cls._KIDS_NOTIFICATION_TEXT
        command = cls._mapping_at(
            notification, "actionButton", "buttonRenderer", "command"
        )
        endpoint_url = cls._mapping_at(command, "urlEndpoint").get("url")
        metadata_url = cls._mapping_at(
            command, "commandMetadata", "webCommandMetadata"
        ).get("url")
        support_signal = (
            isinstance(endpoint_url, str)
            and endpoint_url == metadata_url
            and YouTubeInnertubePlayerClient._has_support_answer(endpoint_url)
        )
        return text_signal, support_signal

    @classmethod
    def _carousel_marker_signal(cls, primary: list) -> bool:
        carousels = []
        for item in primary:
            section = cls._mapping_at(item, "itemSectionRenderer")
            section_contents = section.get("contents")
            if type(section_contents) is not list:
                continue
            for content in section_contents:
                carousel = cls._mapping_at(
                    content, "videoMetadataCarouselViewModel"
                )
                if carousel:
                    carousels.append(carousel)

        marker_urls = []
        for carousel in carousels:
            items = carousel.get("carouselItems")
            if type(items) is not list:
                continue
            for item in items:
                text_item = cls._mapping_at(
                    item,
                    "carouselItemViewModel",
                    "carouselItem",
                    "ctaCarouselItemViewModel",
                    "textCarousel",
                    "textCarouselItemViewModel",
                )
                if not text_item:
                    continue
                on_tap = cls._mapping_at(
                    text_item, "onTap", "innertubeCommand"
                )
                button_tap = cls._mapping_at(
                    text_item,
                    "button",
                    "buttonViewModel",
                    "onTap",
                    "innertubeCommand",
                )
                endpoint_url = cls._mapping_at(
                    on_tap, "urlEndpoint"
                ).get("url")
                button_url = cls._mapping_at(
                    button_tap, "urlEndpoint"
                ).get("url")
                if (
                    isinstance(endpoint_url, str)
                    and endpoint_url == button_url
                ):
                    if cls._is_official_carousel_marker(endpoint_url):
                        marker_urls.append(endpoint_url)
        if len(carousels) > 1 or len(marker_urls) > 1:
            raise KidsPlayerIndeterminateError(
                "YouTube next has ambiguous metadata carousels or Kids markers"
            )
        return len(marker_urls) == 1

    @staticmethod
    def _is_official_carousel_marker(value: object) -> bool:
        try:
            parsed = urllib.parse.urlparse(str(value or ""))
            port = parsed.port
        except ValueError:
            return False
        host = (parsed.hostname or "").casefold()
        if (
            parsed.scheme != "https"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
        ):
            return False
        if host == "ytkids.app.goo.gl":
            return parsed.path == "/nou5" and not (
                parsed.params or parsed.query or parsed.fragment
            )
        return (
            host == "www.youtube.com"
            and parsed.path == "/myfamily/"
            and not parsed.params
            and not parsed.query
            and parsed.fragment == "mf-compare"
        )

    @staticmethod
    def _all_strings(value: object):
        if type(value) is dict:
            for child in value.values():
                yield from YouTubeWatchNextClient._all_strings(child)
        elif type(value) is list:
            for child in value:
                yield from YouTubeWatchNextClient._all_strings(child)
        elif isinstance(value, str):
            yield value

    @classmethod
    def _classify(cls, payload: dict, final_url: str, video_id: str) -> bool:
        cls._validate_final_url(final_url)
        primary, secondary_info = cls._complete_watch_content(payload, video_id)
        text_signal, support_signal = cls._notification_signals(secondary_info)
        marker_signal = cls._carousel_marker_signal(primary)
        signals = (text_signal, support_signal, marker_signal)
        signal_count = sum(signals)
        if signal_count == len(signals):
            return True

        strings = tuple(cls._all_strings(payload))
        misplaced = (
            (not text_signal and cls._KIDS_NOTIFICATION_TEXT in strings)
            or (
                not support_signal
                and any(
                    YouTubeInnertubePlayerClient._has_support_answer(value)
                    for value in strings
                )
            )
            or (
                not marker_signal
                and any(cls._is_official_carousel_marker(value) for value in strings)
            )
        )
        if signal_count == 0 and not misplaced:
            return False
        raise KidsPlayerIndeterminateError(
            f"YouTube next Kids restrictions are inconclusive for {video_id}: "
            f"signals={signal_count}/{len(signals)}, misplaced={misplaced}"
        )

    def has_kids_player_signals(self, video_id: str) -> bool:
        if not VIDEO_ID.fullmatch(video_id):
            raise KidsDomProbeError(f"Invalid YouTube video ID: {video_id!r}")
        for attempt in range(self.retries + 1):
            try:
                payload, final_url = self._fetch_once(video_id)
                return self._classify(payload, final_url, video_id)
            except urllib.error.HTTPError as exc:
                if (
                    exc.code not in self._TRANSIENT_HTTP_CODES
                    or attempt >= self.retries
                ):
                    raise KidsDomProbeError(
                        f"YouTube next HTTP {exc.code} for {video_id}"
                    ) from exc
            except (
                urllib.error.URLError,
                TimeoutError,
                OSError,
                http.client.IncompleteRead,
            ) as exc:
                if attempt >= self.retries:
                    raise KidsDomProbeError(
                        f"YouTube next network failure for {video_id}: "
                        f"{type(exc).__name__}: {exc}"
                    ) from exc
            except KidsPlayerIndeterminateError:
                if attempt >= self.retries:
                    raise
            if self.retry_delay_seconds:
                time.sleep(self.retry_delay_seconds * (2 ** attempt))
        raise KidsDomProbeError("YouTube next retry loop exhausted")

    def close(self) -> None:
        pass


class YouTubeWatchPagePlayerClient:
    """Read the public watch-page player bootstrap without cookies or API keys."""

    _ENDPOINT = "https://www.youtube.com/watch"
    _PLAYER_MARKER = re.compile(
        rb"(?:^|[>;])\s*var\s+ytInitialPlayerResponse\s*=\s*",
        re.MULTILINE,
    )
    _TRANSIENT_HTTP_CODES = YouTubeInnertubePlayerClient._TRANSIENT_HTTP_CODES

    def __init__(
        self,
        *,
        retries: int = KIDS_WATCH_HTTP_RETRIES,
        retry_delay_seconds: float = KIDS_WATCH_RETRY_DELAY_SECONDS,
        timeout_seconds: int = KIDS_WATCH_HTTP_TIMEOUT_SECONDS,
    ) -> None:
        self.retries = max(0, int(retries))
        self.retry_delay_seconds = max(0.0, float(retry_delay_seconds))
        self.timeout_seconds = max(1, int(timeout_seconds))

    @classmethod
    def _watch_url(cls, video_id: str) -> str:
        return cls._ENDPOINT + "?" + urllib.parse.urlencode({
            "v": video_id,
            "hl": "en",
            "gl": "US",
        })

    @classmethod
    def _validate_final_url(cls, final_url: str, video_id: str) -> None:
        try:
            parsed = urllib.parse.urlparse(final_url)
            port = parsed.port
        except ValueError as exc:
            raise KidsPlayerIndeterminateError(
                "YouTube watch page returned a malformed final URL"
            ) from exc
        query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
        if (
            parsed.scheme != "https"
            or (parsed.hostname or "").casefold() != "www.youtube.com"
            or parsed.username is not None
            or parsed.password is not None
            or port is not None
            or parsed.path != "/watch"
            or parsed.fragment
            or query != [("v", video_id), ("hl", "en"), ("gl", "US")]
        ):
            raise KidsPlayerIndeterminateError(
                "YouTube watch page redirected away from the expected video"
            )

    @classmethod
    def _extract_player_payload(cls, body: bytes) -> dict:
        markers = list(cls._PLAYER_MARKER.finditer(body))
        if len(markers) != 1:
            raise KidsPlayerIndeterminateError(
                "YouTube watch page must contain exactly one player bootstrap"
            )
        start = markers[0].end()
        if start >= len(body) or body[start] != ord("{"):
            raise KidsPlayerIndeterminateError(
                "YouTube watch page player bootstrap is not a JSON object"
            )

        depth = 0
        in_string = False
        escaped = False
        end = -1
        for index in range(start, len(body)):
            value = body[index]
            if in_string:
                if escaped:
                    escaped = False
                elif value == ord("\\"):
                    escaped = True
                elif value == ord('"'):
                    in_string = False
                continue
            if value == ord('"'):
                in_string = True
            elif value == ord("{"):
                depth += 1
            elif value == ord("}"):
                depth -= 1
                if depth == 0:
                    end = index + 1
                    break
                if depth < 0:
                    break
        if end < 0 or in_string or depth != 0:
            raise KidsPlayerIndeterminateError(
                "YouTube watch page player bootstrap is truncated"
            )
        suffix = end
        while suffix < len(body) and body[suffix] in b" \t\r\n":
            suffix += 1
        if suffix >= len(body) or body[suffix] != ord(";"):
            raise KidsPlayerIndeterminateError(
                "YouTube watch page player bootstrap has no statement boundary"
            )
        return YouTubeInnertubePlayerClient._strict_json(body[start:end])

    def _request(self, video_id: str) -> urllib.request.Request:
        return urllib.request.Request(
            self._watch_url(video_id),
            headers={
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
                "User-Agent": (
                    "Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/140.0.0.0 Safari/537.36"
                ),
            },
            method="GET",
        )

    def _fetch_once(self, video_id: str) -> tuple[dict, str]:
        request = self._request(video_id)
        with urllib.request.urlopen(
            request, timeout=self.timeout_seconds
        ) as response:
            final_url = (
                str(response.geturl())
                if callable(getattr(response, "geturl", None))
                else request.full_url
            )
            body = response.read(KIDS_WATCH_MAX_HTML_BYTES + 1)
        if len(body) > KIDS_WATCH_MAX_HTML_BYTES:
            raise KidsPlayerIndeterminateError(
                "YouTube watch page exceeded the bounded HTML size"
            )
        self._validate_final_url(final_url, video_id)
        return self._extract_player_payload(body), final_url

    def has_kids_player_signals(self, video_id: str) -> bool:
        if not VIDEO_ID.fullmatch(video_id):
            raise KidsDomProbeError(f"Invalid YouTube video ID: {video_id!r}")
        for attempt in range(self.retries + 1):
            try:
                payload, _ = self._fetch_once(video_id)
                return YouTubeInnertubePlayerClient._classify_player_payload(
                    payload, video_id
                )
            except urllib.error.HTTPError as exc:
                if (
                    exc.code not in self._TRANSIENT_HTTP_CODES
                    or attempt >= self.retries
                ):
                    raise KidsDomProbeError(
                        f"YouTube watch page HTTP {exc.code} for {video_id}"
                    ) from exc
            except (
                urllib.error.URLError,
                TimeoutError,
                OSError,
                http.client.IncompleteRead,
            ) as exc:
                if attempt >= self.retries:
                    raise KidsDomProbeError(
                        f"YouTube watch page network failure for {video_id}: "
                        f"{type(exc).__name__}: {exc}"
                    ) from exc
            except KidsPlayerIndeterminateError:
                if attempt >= self.retries:
                    raise
            if self.retry_delay_seconds:
                time.sleep(self.retry_delay_seconds * (2 ** attempt))
        raise KidsDomProbeError("YouTube watch-page retry loop exhausted")

    def close(self) -> None:
        pass


class YouTubePublicPlayerClient:
    """Select one canary-proven public player representation for the whole shard."""

    def __init__(
        self,
        innertube_client: object | None = None,
        watch_client: object | None = None,
        next_client: object | None = None,
    ) -> None:
        self._backends = (
            ("innertube_android", innertube_client or YouTubeInnertubePlayerClient()),
            ("watch_next", next_client or YouTubeWatchNextClient()),
            ("watch_page", watch_client or YouTubeWatchPagePlayerClient()),
        )
        self._selected_backend: tuple[str, object] | None = None

    @staticmethod
    def _probe_backend(name: str, client: object, video_id: str) -> bool:
        probe = getattr(client, "has_kids_player_signals", None)
        if not callable(probe):
            raise KidsDomProbeError(f"{name} has no public player probe")
        answer = probe(video_id)
        if type(answer) is not bool:
            raise KidsDomProbeError(
                f"{name} returned a non-boolean Kids answer"
            )
        return answer

    def select_backend(
        self,
        positive_ids: tuple[str, ...],
        negative_id: str,
    ) -> str:
        """Lock one backend only if it passes the complete canary batch alone."""
        if self._selected_backend is not None:
            return self._selected_backend[0]
        errors = []
        for name, client in self._backends:
            try:
                positives = {
                    video_id: self._probe_backend(name, client, video_id)
                    for video_id in positive_ids
                }
                negative = self._probe_backend(name, client, negative_id)
                if all(positives.values()) and not negative:
                    self._selected_backend = (name, client)
                    print(
                        f"INFO Kids public player backend selected: {name}",
                        file=sys.stderr,
                        flush=True,
                    )
                    return name
                errors.append(
                    f"{name}=positive={positives}, negative={negative}"
                )
            except Exception as exc:
                errors.append(f"{name}={type(exc).__name__}: {exc}")
        raise KidsDomProbeError(
            "No single public YouTube player source passed every canary: "
            + "; ".join(errors)
        )

    def has_kids_player_signals(self, video_id: str) -> bool:
        if self._selected_backend is None:
            raise KidsDomProbeError(
                "Public YouTube player source was not selected by canaries"
            )
        name, client = self._selected_backend
        # Never fail over per candidate: the selected backend's canary proof is
        # the trust boundary for this entire shard.
        return self._probe_backend(name, client, video_id)

    def close(self) -> None:
        for _, client in self._backends:
            close = getattr(client, "close", None)
            if callable(close):
                close()


class ChromeWebDriverClient:
    """Minimal standard-library W3C WebDriver client; never reads page source."""

    _MARKER_SCRIPT = r"""
const expectedVideoId = arguments[0];
const visible = element => {
  if (!element || !element.isConnected || !element.getClientRects().length) return false;
  const style = window.getComputedStyle(element);
  return style.display !== 'none' && style.visibility !== 'hidden' &&
         style.visibility !== 'collapse' && Number(style.opacity || 1) > 0 &&
         element.getAttribute('aria-hidden') !== 'true';
};
const locationUrl = new URL(document.location.href);
const blockers = Array.from(document.querySelectorAll(
  'ytd-consent-bump-v2-lightbox, form[action*="consent"], iframe[src*="recaptcha"], #captcha'
));
if (locationUrl.hostname === 'consent.youtube.com' || blockers.some(visible)) {
  return 'blocked';
}
if (locationUrl.pathname !== '/watch' || locationUrl.searchParams.get('v') !== expectedVideoId) {
  return 'loading';
}
const watch = document.querySelector('ytd-watch-flexy');
if (!watch || watch.getAttribute('video-id') !== expectedVideoId) return 'loading';
const marker = href => {
  try {
    const url = new URL(href, document.baseURI);
    const host = url.hostname.toLowerCase();
    return url.protocol === 'https:' && (host === 'ytkids.app.goo.gl' ||
      ((host === 'youtube.com' || host === 'www.youtube.com') &&
       url.pathname === '/myfamily/' && url.hash === '#mf-compare'));
  } catch (_) {
    return false;
  }
};
return Array.from(document.querySelectorAll('yt-video-metadata-carousel-view-model'))
  .some(card => visible(card) && Array.from(card.querySelectorAll('a[href]'))
    .some(link => marker(link.href))) ? 'marker' : 'ready';
"""

    _CONSENT_SCRIPT = r"""
const visible = element => {
  if (!element || !element.isConnected || !element.getClientRects().length) return false;
  const style = window.getComputedStyle(element);
  return style.display !== 'none' && style.visibility !== 'hidden' &&
         style.visibility !== 'collapse' && Number(style.opacity || 1) > 0 &&
         element.getAttribute('aria-hidden') !== 'true';
};
const direct = document.querySelector(
  'ytd-consent-bump-v2-lightbox #reject-button button, '
  + 'ytd-consent-bump-v2-lightbox button[aria-label*="Reject" i]'
);
if (visible(direct)) {
  direct.click();
  return 'clicked';
}
const reject = Array.from(document.querySelectorAll('button, [role="button"], input[type="submit"]'))
  .filter(visible)
  .find(element => /^(?:reject all|reject|decline|tout refuser|alle ablehnen|rechazar todo|rifiuta tutto)\b/i
    .test(String(element.innerText || element.value || element.getAttribute('aria-label') || '').trim()));
if (!reject) return 'unhandled';
reject.click();
return 'clicked';
"""

    def __init__(self) -> None:
        self.process: subprocess.Popen | None = None
        self.base_url = ""
        self.session_id = ""
        self._start()

    @staticmethod
    def _driver_path() -> str:
        configured = os.environ.get("CHROMEWEBDRIVER", "").strip()
        if configured:
            resolved = shutil.which(configured)
            if resolved:
                return resolved
            path = Path(configured).expanduser()
            if path.is_dir():
                for name in ("chromedriver", "chromedriver.exe"):
                    child = path / name
                    if child.is_file():
                        return str(child)
            if path.is_file():
                return str(path)
            raise RuntimeError(f"CHROMEWEBDRIVER does not exist: {configured}")
        resolved = shutil.which("chromedriver") or shutil.which("chromedriver.exe")
        if not resolved:
            raise RuntimeError("ChromeDriver not found via CHROMEWEBDRIVER or PATH")
        return resolved

    def _request(self, method: str, path: str, payload: dict | None = None) -> object:
        body = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            self.base_url + path,
            data=body,
            method=method,
            headers={"Content-Type": "application/json; charset=utf-8"},
        )
        try:
            with urllib.request.urlopen(
                request, timeout=KIDS_DOM_HTTP_TIMEOUT_SECONDS
            ) as response:
                document = json.load(response)
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", "replace")[:500]
            raise RuntimeError(f"WebDriver HTTP {exc.code}: {detail}") from exc
        if not isinstance(document, dict) or "value" not in document:
            raise RuntimeError("Malformed WebDriver response")
        value = document["value"]
        if isinstance(value, dict) and value.get("error"):
            raise RuntimeError(
                f"WebDriver {value.get('error')}: {value.get('message') or 'unknown error'}"
            )
        return value

    def _start(self) -> None:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as reservation:
            reservation.bind(("127.0.0.1", 0))
            port = reservation.getsockname()[1]
        self.base_url = f"http://127.0.0.1:{port}"
        self.process = subprocess.Popen(
            [self._driver_path(), f"--port={port}", "--allowed-origins=*"],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            if self.process.poll() is not None:
                raise RuntimeError(f"ChromeDriver exited with code {self.process.returncode}")
            try:
                self._request("GET", "/status")
                break
            except Exception:
                time.sleep(0.1)
        else:
            self.close()
            raise RuntimeError("ChromeDriver did not become ready within 10 seconds")
        value = self._request("POST", "/session", {
            "capabilities": {
                "alwaysMatch": {
                    "browserName": "chrome",
                    "pageLoadStrategy": "eager",
                    "goog:chromeOptions": {
                        "args": [
                            "--headless=new",
                            "--disable-gpu",
                            "--disable-dev-shm-usage",
                            "--mute-audio",
                            "--no-sandbox",
                            "--autoplay-policy=user-gesture-required",
                            "--lang=en-US",
                            "--window-size=1280,900",
                        ]
                    },
                }
            }
        })
        if not isinstance(value, dict) or not value.get("sessionId"):
            self.close()
            raise RuntimeError("ChromeDriver did not return a session ID")
        self.session_id = str(value["sessionId"])
        self._request("POST", f"/session/{self.session_id}/timeouts", {
            "implicit": 0,
            "pageLoad": KIDS_DOM_PAGE_LOAD_TIMEOUT_MS,
            "script": KIDS_DOM_SCRIPT_TIMEOUT_MS,
        })

    def has_family_options_marker(self, video_id: str) -> bool:
        if not VIDEO_ID.fullmatch(video_id):
            return False
        base = f"/session/{self.session_id}"
        self._request("POST", base + "/url", {
            "url": f"https://www.youtube.com/watch?v={video_id}&hl=en&gl=US"
        })
        started = time.monotonic()
        deadline = started + min(max(KIDS_DOM_MARKER_WAIT_SECONDS, 5), 8)
        last_state = "loading"
        consent_started = 0.0
        while time.monotonic() < deadline:
            last_state = self._request("POST", base + "/execute/sync", {
                "script": self._MARKER_SCRIPT,
                "args": [video_id],
            })
            if last_state == "marker":
                return True
            if last_state == "blocked":
                if not consent_started:
                    consent_result = self._request("POST", base + "/execute/sync", {
                        "script": self._CONSENT_SCRIPT,
                        "args": [],
                    })
                    if consent_result != "clicked":
                        raise KidsDomProbeError(
                            "YouTube watch page is blocked by consent or captcha"
                        )
                    consent_started = time.monotonic()
                    started = consent_started
                    deadline = consent_started + min(max(KIDS_DOM_MARKER_WAIT_SECONDS, 5), 8)
                elif time.monotonic() - consent_started >= 5:
                    raise KidsDomProbeError(
                        "YouTube consent did not resolve the expected watch page"
                    )
                time.sleep(0.25)
                continue
            if last_state not in {"loading", "ready"}:
                raise KidsDomProbeError(f"Unexpected rendered watch-page state: {last_state!r}")
            if last_state == "ready" and time.monotonic() - started >= 5:
                return False
            time.sleep(0.25)
        if last_state == "ready":
            return False
        raise KidsDomProbeError("YouTube watch page did not load the expected video in time")

    def close(self) -> None:
        if self.session_id:
            try:
                self._request("DELETE", f"/session/{self.session_id}")
            except Exception:
                pass
            self.session_id = ""
        if self.process and self.process.poll() is None:
            self.process.terminate()
            try:
                self.process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                self.process.kill()
                self.process.wait(timeout=2)
        self.process = None


class KidsDomValidator:
    """Serialize one canary-proven public client for every Kids query in a shard."""

    def __init__(
        self,
        client: object | None = None,
        *,
        canary_retries: int = KIDS_CANARY_RETRIES,
        canary_retry_delay_seconds: float = KIDS_CANARY_RETRY_DELAY_SECONDS,
    ) -> None:
        self._client = client
        self._lock = threading.Lock()
        self._canaries_checked = False
        self._canary_error = ""
        self._canary_retries = max(0, int(canary_retries))
        self._canary_retry_delay_seconds = max(
            0.0, float(canary_retry_delay_seconds)
        )

    def _get_client(self) -> object:
        if self._client is None:
            self._client = YouTubePublicPlayerClient()
        return self._client

    def _probe(self, video_id: str) -> bool:
        client = self._get_client()
        probe = getattr(client, "has_kids_player_signals", None)
        if not callable(probe):
            probe = getattr(client, "has_kids_watch_page_signals", None)
        if not callable(probe):
            probe = client.has_family_options_marker
        return probe(video_id) is True

    def _check_canaries(self) -> None:
        if self._canaries_checked:
            if self._canary_error:
                raise KidsDomCanaryError(self._canary_error)
            return
        self._canaries_checked = True
        last_error: Exception | None = None
        for attempt in range(self._canary_retries + 1):
            try:
                client = self._get_client()
                select_backend = getattr(client, "select_backend", None)
                if callable(select_backend):
                    select_backend(
                        KIDS_DOM_POSITIVE_CANARIES,
                        KIDS_DOM_NEGATIVE_CANARY,
                    )
                    return
                else:
                    positives = {
                        video_id: self._probe(video_id)
                        for video_id in KIDS_DOM_POSITIVE_CANARIES
                    }
                    negative = self._probe(KIDS_DOM_NEGATIVE_CANARY)
                    if all(positives.values()) and not negative:
                        return
                    last_error = RuntimeError(
                        f"positive={positives}, negative={negative}"
                    )
            except Exception as exc:
                last_error = exc
            if (
                attempt < self._canary_retries
                and self._canary_retry_delay_seconds
            ):
                time.sleep(
                    self._canary_retry_delay_seconds * (2 ** attempt)
                )
        assert last_error is not None
        self._canary_error = (
            f"Kids DOM canaries failed closed: "
            f"{type(last_error).__name__}: {last_error}"
        )
        raise KidsDomCanaryError(self._canary_error) from last_error

    def ensure_canaries(self) -> None:
        """Validate the rendered Kids signal even when a search has no candidates."""
        with self._lock:
            self._check_canaries()

    def is_made_for_kids(self, video_id: str) -> bool:
        with self._lock:
            self._check_canaries()
            try:
                return self._probe(video_id)
            except KidsDomProbeError:
                raise
            except Exception as exc:
                raise KidsDomProbeError(
                    f"Kids DOM verification failed for {video_id}: {type(exc).__name__}: {exc}"
                ) from exc

    def close(self) -> None:
        with self._lock:
            if self._client is not None:
                self._client.close()
                self._client = None


def kids_dom_validator() -> KidsDomValidator:
    global _KIDS_DOM_VALIDATOR
    with _KIDS_DOM_VALIDATOR_LOCK:
        if _KIDS_DOM_VALIDATOR is None:
            _KIDS_DOM_VALIDATOR = KidsDomValidator()
            atexit.register(_KIDS_DOM_VALIDATOR.close)
        return _KIDS_DOM_VALIDATOR


def ydl():
    if not hasattr(THREAD, "ydl"):
        import yt_dlp

        THREAD.ydl = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "playlistend": SEARCH_RESULTS,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "ignore_no_formats_error": True,
                "cachedir": False,
                "geo_bypass_country": "FR",
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return THREAD.ydl


def search_ydl():
    if not hasattr(THREAD, "search_ydl"):
        import yt_dlp

        THREAD.search_ydl = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "extract_flat": True,
                "playlistend": SEARCH_RESULTS,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "cachedir": False,
                "geo_bypass_country": "FR",
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return THREAD.search_ydl


def kids_search_ydl(result_limit: int):
    """Flat yt-dlp search reader with an explicit Kids top-N cap."""
    import yt_dlp

    readers = getattr(THREAD, "kids_search_ydl", None)
    if readers is None:
        readers = {}
        THREAD.kids_search_ydl = readers
    if result_limit not in readers:
        readers[result_limit] = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "extract_flat": True,
                "playlistend": result_limit,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "cachedir": False,
                "geo_bypass_country": "US",
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return readers[result_limit]


def owned_ydl():
    """Use a dedicated playlist reader so the official channel is not capped at 10 uploads."""
    if not hasattr(THREAD, "owned_ydl"):
        import yt_dlp

        THREAD.owned_ydl = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "extract_flat": True,
                "playlistend": OWN_UPLOADS_PER_CHANNEL,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "cachedir": False,
                "geo_bypass_country": "FR",
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return THREAD.owned_ydl


def info_to_row(info: dict, now_ms: int, *, genre: str = "", cluster: str = "", query: str = "") -> dict | None:
    video_id = str(info.get("id") or "")
    views = info.get("view_count")
    duration = info.get("duration")
    if not VIDEO_ID.match(video_id) or not isinstance(views, (int, float)):
        return None
    published = to_ms(info)
    age = age_months(published, now_ms)
    row = {
        "title": str(info.get("title") or "").strip(),
        "vid": video_id,
        "url": info.get("webpage_url") or f"https://www.youtube.com/watch?v={video_id}",
        "views": int(views),
        "channel": info.get("channel") or info.get("uploader") or "Unknown channel",
        "chUrl": info.get("channel_url") or info.get("uploader_url") or "",
        "metadataSource": METADATA_SOURCE_YTDLP,
    }
    channel_id = str(info.get("channel_id") or "")
    if CHANNEL_ID.match(channel_id):
        row["channelId"] = channel_id
    followers = info.get("channel_follower_count")
    if isinstance(followers, (int, float)) and followers > 0:
        row["subs"] = int(followers)
    comments = parse_nonnegative_count(info.get("comment_count"))
    if comments is not None:
        row["comments"] = comments
    if isinstance(duration, (int, float)):
        row["durH"] = float(duration) / 3600
    if published:
        row["pub"] = published
        row["pubSource"] = METADATA_SOURCE_YTDLP
    if age is not None:
        row["ageM"] = age
        row["vpm"] = int(views) / age
    if genre:
        row["genre"] = genre
        row["cluster"] = cluster_for(row["title"], cluster)
    if query:
        row["kw"] = query
        row["kwCount"] = 1
        row["pattern"] = "Daily keyword scan"
        row["added"] = now_ms
    return row


def fetch_one_video(video_id: str, now_ms: int) -> dict | None:
    info = ydl().extract_info(f"https://www.youtube.com/watch?v={video_id}", download=False)
    return info_to_row(info or {}, now_ms)


def public_oembed_confirms_unavailable(video_id: str) -> bool:
    """Return true only when YouTube's public, cookie-free oEmbed endpoint says 404.

    yt-dlp deliberately uses ``ignoreerrors`` for the large daily refresh, so a
    deleted/private video and a transient extractor failure both surface as a
    missing row.  The former can be quarantined immediately, while the latter
    must keep the existing two-scan fail-closed policy.  A public oEmbed 404 is
    a narrow, authoritative signal that the video is no longer publicly
    addressable; every other response or network error remains inconclusive.
    """
    if not VIDEO_ID.match(str(video_id or "")):
        return False
    watch_url = f"https://www.youtube.com/watch?v={video_id}"
    endpoint = "https://www.youtube.com/oembed?" + urllib.parse.urlencode({
        "url": watch_url,
        "format": "json",
    })
    request = urllib.request.Request(
        endpoint,
        headers={
            "Accept": "application/json",
            "Accept-Language": "en",
            "User-Agent": "Mozilla/5.0 (compatible; LofiRadar/1.0)",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=10) as response:
            response.read(1)
        return False
    except urllib.error.HTTPError as exc:
        return exc.code == 404
    except (urllib.error.URLError, TimeoutError, OSError):
        return False


def fetch_search(spec: dict, now_ms: int) -> tuple[list[dict], int, int]:
    # Current yt-dlp releases no longer expose ``ytsearchdate``.  Use the
    # actual YouTube "This month" search filter; known catalogue videos are
    # refreshed separately by ID, while this pass discovers new uploads.
    search_url = (
        "https://www.youtube.com/results?search_query="
        + urllib.parse.quote_plus(spec["query"])
        + "&sp=EgIIBA%3D%3D"
    )
    info = search_ydl().extract_info(search_url, download=False) or {}
    entries = [item for item in (info.get("entries") or []) if item]
    if not entries:
        raise RuntimeError("YouTube returned no raw search results")
    rows: list[dict] = []
    enriched = 0
    for rank, item in enumerate(entries, start=1):
        video_id = str(item.get("id") or "")
        if not VIDEO_ID.match(video_id):
            continue
        try:
            full = ydl().extract_info(
                f"https://www.youtube.com/watch?v={video_id}", download=False
            ) or {}
        except Exception:
            continue
        if not full:
            continue
        enriched += 1
        if not is_instrumental(full):
            continue
        row = info_to_row(
            full,
            now_ms,
            genre=spec["genre"],
            cluster=spec["cluster"],
            query=spec["query"],
        )
        if row:
            row["rank"] = rank
            row["audiences"] = ["youtube"]
            rows.append(row)
    if enriched / len(entries) < 0.50:
        raise RuntimeError(f"Only {enriched}/{len(entries)} search results could be enriched")
    return rows, len(entries), enriched


def fetch_api_rows(
    video_ids: list[str],
    now_ms: int,
    key: str,
    *,
    include_scan_text: bool = False,
) -> dict[str, dict]:
    """Fast official metrics path when a YOUTUBE_API_KEY secret is present."""
    out: dict[str, dict] = {}
    for start in range(0, len(video_ids), 50):
        batch = video_ids[start : start + 50]
        query = urllib.parse.urlencode({
            "part": "snippet,contentDetails,statistics,status",
            "id": ",".join(batch),
            "key": key,
        })
        payload = youtube_api_json(
            "https://www.googleapis.com/youtube/v3/videos?" + query,
            timeout=30,
        )
        for item in payload.get("items") or []:
            snippet = item.get("snippet") or {}
            statistics = item.get("statistics") or {}
            duration = parse_iso_duration((item.get("contentDetails") or {}).get("duration"))
            try:
                published = int(
                    datetime.fromisoformat(snippet["publishedAt"].replace("Z", "+00:00")).timestamp()
                    * 1000
                )
                views = int(statistics["viewCount"])
            except (KeyError, TypeError, ValueError):
                continue
            age = age_months(published, now_ms)
            video_id = item.get("id")
            row = {
                "title": snippet.get("title") or "",
                "vid": video_id,
                "url": f"https://www.youtube.com/watch?v={video_id}",
                "views": views,
                "pub": published,
                "ageM": age,
                "vpm": views / age if age else None,
                "channel": snippet.get("channelTitle") or "Unknown channel",
                "chUrl": f"https://www.youtube.com/channel/{snippet.get('channelId', '')}",
                "channelId": snippet.get("channelId") or "",
                "metadataSource": METADATA_SOURCE_API,
                "pubSource": METADATA_SOURCE_API,
            }
            comments = parse_nonnegative_count(statistics.get("commentCount"))
            if comments is not None:
                row["comments"] = comments
            made_for_kids = (item.get("status") or {}).get("madeForKids")
            if isinstance(made_for_kids, bool):
                row["madeForKids"] = made_for_kids
                row["madeForKidsSource"] = "youtube_data_api_status"
            if duration is not None:
                row["durH"] = duration / 3600
            if include_scan_text:
                row["_scanDescription"] = snippet.get("description") or ""
                row["_scanTags"] = " ".join(snippet.get("tags") or [])
                row["_liveBroadcastContent"] = snippet.get("liveBroadcastContent") or "none"
            out[video_id] = row
    return out


def empty_kids_funnel() -> dict[str, int]:
    return {
        "raw": 0,
        "unique": 0,
        "enriched": 0,
        "rejected_missing_metadata": 0,
        "rejected_prefilter": 0,
        "rejected_enrichment_unavailable": 0,
        "rejected_live": 0,
        "rejected_duration": 0,
        "rejected_views": 0,
        "rejected_vpm": 0,
        "rejected_vocal_or_instrumental": 0,
        "rejected_made_for_kids": 0,
        "kept": 0,
        "duration_fallback_from_search": 0,
        "duration_missing": 0,
        "duration_below_minimum": 0,
        "lane_calls_expected": 0,
        "lane_calls_completed": 0,
    }


def merge_kids_funnel(target: dict[str, int], source: dict | None) -> None:
    for key, value in (source or {}).items():
        if isinstance(value, (int, float)):
            target[key] = int(target.get(key, 0)) + int(value)


def validate_kids_funnel(funnel: dict[str, int], raw: int, kept: int) -> None:
    rejected = sum(
        int(value)
        for key, value in funnel.items()
        if key.startswith("rejected_")
    )
    if int(funnel.get("raw", 0)) != int(raw):
        raise RuntimeError("Kids funnel rejected: raw results do not match examined results")
    if int(funnel.get("kept", 0)) != int(kept):
        raise RuntimeError("Kids funnel rejected: kept results do not match candidate results")
    if int(funnel.get("unique", 0)) != rejected + int(funnel.get("kept", 0)):
        raise RuntimeError("Kids funnel rejected: unique results are not fully accounted for")
    if int(funnel.get("lane_calls_completed", 0)) != int(
        funnel.get("lane_calls_expected", 0)
    ):
        raise RuntimeError("Kids funnel rejected: search lane coverage is incomplete")


def fetch_kids_search(spec: dict, now_ms: int, key: str) -> tuple[list[dict], int, int, dict]:
    """Search two bounded lanes and validate every deduplicated row officially."""
    if not key:
        raise RuntimeError("Kids discovery requires YOUTUBE_API_KEY")
    video_ids: list[str] = []
    ranks: dict[str, int] = {}
    discovery_lanes: dict[str, set[str]] = defaultdict(set)
    result_limit = int(spec.get("searchResults") or KIDS_SEARCH_RESULTS)
    lanes = tuple(spec.get("searchLanes") or ("viewCount",))
    funnel = empty_kids_funnel()
    funnel["lane_calls_expected"] = len(lanes)
    lane_limit = max(1, result_limit // len(lanes))
    for lane in lanes:
        page_token = ""
        lane_seen = 0
        pages = 0
        max_pages = max(1, (lane_limit + 49) // 50)
        while lane_seen < lane_limit and pages < max_pages:
            pages += 1
            params: dict[str, object] = {
                "part": "snippet",
                "q": spec["query"] + " " + KIDS_QUERY_EXCLUSIONS,
                "type": "video",
                "order": lane,
                "videoDuration": "long",
                "safeSearch": "strict",
                "relevanceLanguage": "en",
                "regionCode": "US",
                "maxResults": min(50, lane_limit - lane_seen),
                "key": key,
            }
            if page_token:
                params["pageToken"] = page_token
            payload = youtube_api_payload("search", params)
            items = payload.get("items") or []
            if items:
                funnel["lane_calls_completed"] += 1
            funnel["raw"] += len(items)
            for item in items:
                video_id = str((item.get("id") or {}).get("videoId") or "")
                if not VIDEO_ID.match(video_id):
                    continue
                lane_seen += 1
                discovery_lanes[video_id].add(lane)
                if video_id not in ranks:
                    ranks[video_id] = len(video_ids) + 1
                    video_ids.append(video_id)
                if lane_seen >= lane_limit:
                    break
            page_token = str(payload.get("nextPageToken") or "")
            if not page_token or not items:
                break
    if not video_ids:
        raise RuntimeError("YouTube Data API returned no Kids search results")
    funnel["unique"] = len(video_ids)

    official = fetch_api_rows(
        video_ids,
        now_ms,
        key,
        include_scan_text=True,
    )
    rows: list[dict] = []
    for video_id in video_ids:
        row = official.get(video_id)
        if not row:
            funnel["rejected_missing_metadata"] += 1
            continue
        if row.get("_liveBroadcastContent") != "none":
            funnel["rejected_live"] += 1
            continue
        duration = row.get("durH")
        if not isinstance(duration, (int, float)) or duration * 3600 < MIN_SECONDS:
            funnel["rejected_duration"] += 1
            continue
        if int(row.get("views") or 0) < MIN_KIDS_VIEWS:
            funnel["rejected_views"] += 1
            continue
        if not isinstance(row.get("vpm"), (int, float)) or float(row["vpm"]) < MIN_KIDS_VPM:
            funnel["rejected_vpm"] += 1
            continue
        evidence = kids_instrumental_evidence(row)
        if not evidence:
            funnel["rejected_vocal_or_instrumental"] += 1
            continue
        if row.get("madeForKids") is not True:
            funnel["rejected_made_for_kids"] += 1
            continue
        row = dict(row)
        row["genre"] = kids_genre_from_metadata(row)
        row["cluster"] = cluster_for(row.get("title") or "", spec["cluster"])
        row["kw"] = spec["query"]
        row["kwCount"] = 1
        row["pattern"] = "Daily Kids keyword scan"
        row["added"] = now_ms
        row["rank"] = ranks[video_id]
        row["audiences"] = ["kids"]
        row["instrumentalVerified"] = True
        row["instrumentalEvidenceSource"] = evidence
        row["discoveryLanes"] = sorted(discovery_lanes[video_id])
        row["liveStatus"] = "none"
        row.pop("_scanDescription", None)
        row.pop("_scanTags", None)
        row.pop("_liveBroadcastContent", None)
        rows.append(row)
    funnel["enriched"] = len(official)
    funnel["kept"] = len(rows)
    return rows, funnel["raw"], len(official), funnel


def is_kids_flat_candidate(info: dict) -> bool:
    """Cheap structural gate; unknown flat metadata is enriched, never accepted."""
    video_id = str(info.get("id") or "")
    title = str(info.get("title") or "").strip()
    duration = info.get("duration")
    views = info.get("view_count")
    live_status = str(info.get("live_status") or "").casefold()
    if (
        not VIDEO_ID.fullmatch(video_id)
        or not title
        or info.get("is_live") is True
        or live_status in {"is_live", "is_upcoming", "post_live", "was_live"}
    ):
        return False
    if isinstance(duration, (int, float)) and duration < MIN_SECONDS:
        return False
    if isinstance(views, (int, float)) and views < MIN_KIDS_VIEWS:
        return False
    return not has_vocal_signal(title)


def fetch_kids_search_ydl(spec: dict, now_ms: int) -> tuple[list[dict], int, int, dict]:
    """No-key Kids fallback: flat search, strict enrichment, then rendered DOM truth."""
    result_limit = int(spec.get("searchResults") or KIDS_SEARCH_RESULTS)
    lanes = tuple(spec.get("searchLanes") or ("viewCount",))
    lane_limit = max(1, result_limit // len(lanes))
    entries_by_id: dict[str, dict] = {}
    discovery_lanes: dict[str, set[str]] = defaultdict(set)
    lane_ranks: dict[str, list[int]] = defaultdict(list)
    raw_received = 0
    lane_calls_completed = 0
    for lane in lanes:
        search = (
            "https://www.youtube.com/results?search_query="
            + urllib.parse.quote_plus(spec["query"] + " " + KIDS_QUERY_EXCLUSIONS)
            + "&sp=" + KIDS_YTDLP_SEARCH_PARAMS[lane]
        )
        info = kids_search_ydl(lane_limit).extract_info(search, download=False) or {}
        lane_entries = [item for item in (info.get("entries") or []) if item][:lane_limit]
        if lane_entries:
            lane_calls_completed += 1
        raw_received += len(lane_entries)
        for rank, item in enumerate(lane_entries, start=1):
            video_id = str(item.get("id") or "")
            if not VIDEO_ID.fullmatch(video_id):
                continue
            discovery_lanes[video_id].add(lane)
            lane_ranks[video_id].append(rank)
            entries_by_id.setdefault(video_id, item)
    entries = list(entries_by_id.values())
    if not entries:
        raise RuntimeError("yt-dlp returned no raw Kids search results")
    prefiltered = [item for item in entries if is_kids_flat_candidate(item)]
    funnel = empty_kids_funnel()
    funnel["lane_calls_expected"] = len(lanes)
    funnel["lane_calls_completed"] = lane_calls_completed
    funnel["raw"] = raw_received
    funnel["unique"] = len(entries)
    funnel["rejected_prefilter"] = len(entries) - len(prefiltered)
    ranks = {str(item.get("id")): rank for rank, item in enumerate(entries, start=1)}
    validator = kids_dom_validator()
    validator.ensure_canaries()
    rows: list[dict] = []
    enriched = 0
    enrichment_attempts = 0
    enrichment_failures = 0
    for item in prefiltered:
        video_id = str(item.get("id") or "")
        enrichment_attempts += 1
        try:
            full = ydl().extract_info(
                f"https://www.youtube.com/watch?v={video_id}", download=False
            ) or {}
        except Exception as exc:
            enrichment_failures += 1
            funnel["rejected_enrichment_unavailable"] += 1
            print(
                f"WARN Kids enrichment {video_id}: {type(exc).__name__}: {exc}",
                file=sys.stderr,
            )
            continue
        if not full:
            enrichment_failures += 1
            funnel["rejected_enrichment_unavailable"] += 1
            continue
        enriched += 1
        if full.get("is_live") is True or str(full.get("live_status") or "").casefold() in {
            "is_live", "is_upcoming", "post_live", "was_live",
        }:
            funnel["rejected_live"] += 1
            continue
        row = info_to_row(full, now_ms)
        if not row:
            funnel["rejected_missing_metadata"] += 1
            continue
        if row.get("vid") != video_id:
            raise RuntimeError(
                f"Kids enrichment returned a different video ID for {video_id}"
            )
        full_duration = row.get("durH")
        flat_duration = item.get("duration")
        full_duration_valid = (
            isinstance(full_duration, (int, float))
            and not isinstance(full_duration, bool)
            and math.isfinite(float(full_duration))
            and float(full_duration) > 0
        )
        if (
            not full_duration_valid
            and isinstance(flat_duration, (int, float))
            and not isinstance(flat_duration, bool)
            and math.isfinite(float(flat_duration))
            and float(flat_duration) > 0
        ):
            # On hosted runners YouTube can omit player-duration fields from
            # the detailed response while the same video's search renderer
            # still provides its factual duration. The flat row was bound to
            # this exact video ID before enrichment, so retaining that value
            # restores information instead of guessing it.
            row["durH"] = float(flat_duration) / 3600
            row["durationSource"] = "youtube_search_result"
            funnel["duration_fallback_from_search"] += 1
        row["_scanDescription"] = str(full.get("description") or "")
        tags = full.get("tags") or []
        row["_scanTags"] = " ".join(str(value) for value in tags) if isinstance(tags, list) else str(tags)
        evidence = kids_instrumental_evidence(row)
        duration = row.get("durH")
        if (
            not isinstance(duration, (int, float))
            or isinstance(duration, bool)
            or not math.isfinite(float(duration))
        ):
            funnel["duration_missing"] += 1
            funnel["rejected_duration"] += 1
            continue
        if float(duration) * 3600 < MIN_SECONDS:
            funnel["duration_below_minimum"] += 1
            funnel["rejected_duration"] += 1
            continue
        if int(row.get("views") or 0) < MIN_KIDS_VIEWS:
            funnel["rejected_views"] += 1
            continue
        if not isinstance(row.get("vpm"), (int, float)) or float(row["vpm"]) < MIN_KIDS_VPM:
            funnel["rejected_vpm"] += 1
            continue
        if not evidence:
            funnel["rejected_vocal_or_instrumental"] += 1
            continue
        if not validator.is_made_for_kids(video_id):
            funnel["rejected_made_for_kids"] += 1
            continue
        row["madeForKids"] = True
        row["madeForKidsSource"] = "youtube_public_player_restrictions"
        row["genre"] = kids_genre_from_metadata(row)
        row["cluster"] = cluster_for(row.get("title") or "", spec["cluster"])
        row["kw"] = spec["query"]
        row["kwCount"] = 1
        row["pattern"] = "Daily Kids keyword scan"
        row["added"] = now_ms
        row["rank"] = min(lane_ranks[video_id])
        row["audiences"] = ["kids"]
        row["instrumentalVerified"] = True
        row["instrumentalEvidenceSource"] = evidence
        row["discoveryLanes"] = sorted(discovery_lanes[video_id])
        row["liveStatus"] = "none"
        row.pop("_scanDescription", None)
        row.pop("_scanTags", None)
        rows.append(row)
    if enrichment_attempts and enriched / enrichment_attempts < 0.80:
        raise RuntimeError(
            f"Only {enriched}/{enrichment_attempts} prefiltered Kids candidates "
            "could be enriched"
        )
    funnel["enriched"] = enriched
    funnel["kept"] = len(rows)
    return rows, raw_received, enriched, funnel


def youtube_api_payload(path: str, params: dict[str, object]) -> dict:
    """Load one YouTube Data API response without exposing the API key in logs."""
    query = urllib.parse.urlencode(params)
    return youtube_api_json(
        "https://www.googleapis.com/youtube/v3/" + path + "?" + query,
        timeout=30,
    )


def fetch_owned_upload_ids(api_key: str) -> list[str]:
    """Return recent public uploads from the official Lofi Girl channel(s)."""
    ids: list[str] = []
    for handle in OWN_CHANNEL_HANDLES:
        channels = youtube_api_payload(
            "channels",
            {"part": "contentDetails", "forHandle": handle, "key": api_key},
        ).get("items") or []
        if not channels:
            raise RuntimeError(f"Official channel lookup returned no channel for {handle}")
        uploads = ((channels[0].get("contentDetails") or {}).get("relatedPlaylists") or {}).get("uploads")
        if not uploads:
            raise RuntimeError(f"Official channel {handle} has no uploads playlist")
        page_token = ""
        while len(ids) < OWN_UPLOADS_PER_CHANNEL:
            params: dict[str, object] = {
                "part": "contentDetails",
                "playlistId": uploads,
                "maxResults": min(50, OWN_UPLOADS_PER_CHANNEL - len(ids)),
                "key": api_key,
            }
            if page_token:
                params["pageToken"] = page_token
            payload = youtube_api_payload("playlistItems", params)
            ids.extend(
                str((item.get("contentDetails") or {}).get("videoId") or "")
                for item in payload.get("items") or []
            )
            page_token = str(payload.get("nextPageToken") or "")
            if not page_token:
                break
    return list(dict.fromkeys(video_id for video_id in ids if VIDEO_ID.match(video_id)))


def fetch_owned_api_rows(now_ms: int, api_key: str) -> dict[str, dict]:
    """Refresh recent official uploads so Analyse cannot miss a new release."""
    rows = fetch_api_rows(fetch_owned_upload_ids(api_key), now_ms, api_key)
    if not rows:
        raise RuntimeError("Official Lofi Girl upload scan returned no usable videos")
    for row in rows.values():
        add_owned_metadata(row)
        row["source"] = "Official Lofi Girl daily scan"
    return rows


def fetch_owned_ydl_rows(now_ms: int) -> dict[str, dict]:
    """Fallback when the repository has no YouTube API secret configured."""
    info = owned_ydl().extract_info("https://www.youtube.com/@LofiGirl/videos", download=False) or {}
    ids = [
        str(item.get("id") or "")
        for item in info.get("entries") or []
        if item and VIDEO_ID.match(str(item.get("id") or ""))
    ]
    if not ids:
        raise RuntimeError("Official Lofi Girl channel returned no recent uploads")
    rows: dict[str, dict] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=TRACK_WORKERS) as pool:
        future_to_id = {pool.submit(fetch_one_video, video_id, now_ms): video_id for video_id in ids}
        for future in concurrent.futures.as_completed(future_to_id):
            row = future.result()
            if row:
                add_owned_metadata(row)
                row["source"] = "Official Lofi Girl daily scan"
                rows[row["vid"]] = row
    if not rows:
        raise RuntimeError("Official Lofi Girl upload scan returned no usable videos")
    # yt-dlp occasionally returns a valid counter without a duration. Keep the
    # newest upload complete by reading YouTube's public lengthSeconds field.
    latest = next((rows[video_id] for video_id in ids if video_id in rows), None)
    if latest and not isinstance(latest.get("durH"), (int, float)):
        try:
            duration = fetch_public_duration(latest["vid"])
            if duration is not None:
                latest["durH"] = duration / 3600
                latest["durationSource"] = "youtube_public_length"
        except Exception as exc:
            print(
                f"WARN official duration {latest['vid']}: {type(exc).__name__}: {exc}",
                file=sys.stderr,
            )
    return rows


def kids_search_lanes(day: str | None = None) -> list[tuple[str, ...]]:
    """Use exactly two calls/query while rotating relevance and recent daily."""
    query_day = day or history_day_key(utc_now_ms())
    day_parity = datetime.fromisoformat(query_day).date().toordinal() % 2
    return [
        ("viewCount", "date" if (index + day_parity) % 2 else "relevance")
        for index in range(len(KIDS_QUERY_SPECS))
    ]


def query_specs(
    payload: dict,
    *,
    include_kids: bool = True,
    kids_day: str | None = None,
) -> list[dict]:
    votes: dict[str, dict[str, Counter]] = defaultdict(lambda: {"genre": Counter(), "cluster": Counter()})
    for bucket in ("all", "trends", "news"):
        for row in payload.get("d", {}).get(bucket, []):
            if is_deferred_row(row):
                continue
            for query in split_keywords(row.get("kw")):
                votes[query]["genre"][str(row.get("genre") or "Other")] += 1
                votes[query]["cluster"][str(row.get("cluster") or "Relaxation / meditation")] += 1
    regular = [
        {
            "query": query,
            "genre": data["genre"].most_common(1)[0][0],
            "cluster": data["cluster"].most_common(1)[0][0],
            "audience": "youtube",
        }
        for query, data in sorted(votes.items())
    ]
    if not include_kids:
        return regular
    kids_result_limit = max(KIDS_SEARCH_RESULTS, KIDS_BOOTSTRAP_SEARCH_RESULTS)
    kids_lanes = kids_search_lanes(kids_day)
    search_calls = sum(
        len(lanes) * max(1, math.ceil((kids_result_limit / len(lanes)) / 50))
        for lanes in kids_lanes
    )
    if search_calls > MAX_KIDS_SEARCH_CALLS:
        raise RuntimeError(
            f"Kids search budget exceeded: {search_calls}>{MAX_KIDS_SEARCH_CALLS} calls"
        )
    kids = [
        {
            "query": query,
            "genre": genre,
            "cluster": cluster,
            "audience": "kids",
            "searchResults": kids_result_limit,
        }
        for (query, genre, cluster), lanes in zip(KIDS_QUERY_SPECS, kids_lanes)
    ]
    for spec, lanes in zip(kids, kids_lanes):
        spec["searchLanes"] = list(lanes)
    return regular + kids


SHEET_VIDEO_TAB_FRAGMENTS = ("All Videos", "Trends", "News", "Our Videos")
SHEET_VIDEO_URL = re.compile(
    r"(?:watch\?v=|/vi/|youtu\.be/|/embed/|/shorts/)([\w-]{11})",
    re.I,
)


def sheet_cell_video_id(cell: object) -> str | None:
    """Read a YouTube ID from a raw ID, hyperlink target or HYPERLINK formula."""
    hyperlink = getattr(cell, "hyperlink", None)
    candidates = (
        getattr(cell, "value", None),
        getattr(hyperlink, "target", None) if hyperlink else None,
    )
    for candidate in candidates:
        text = str(candidate or "").strip()
        if VIDEO_ID.fullmatch(text):
            return text
        match = SHEET_VIDEO_URL.search(text)
        if match:
            return match.group(1)
    return None


def sheet_number(value: object) -> float | None:
    """Mirror the dashboard Sheet parser for factual numeric metadata."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
    else:
        try:
            parsed = float(re.sub(r"[,\s]", "", str(value)))
        except (TypeError, ValueError):
            return None
    return parsed if math.isfinite(parsed) else None


def sheet_date_ms(value: object, epoch: datetime) -> int | None:
    """Convert an XLSX date exactly enough to match the browser's ``toMs``."""
    if isinstance(value, bool) or value is None:
        return None
    parsed: datetime | None = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
    elif isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return None
        parsed = epoch + timedelta(days=float(value))
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    stamp = int(round(parsed.timestamp() * 1000))
    return stamp if stamp > 0 else None


def canonical_ours_metadata_rows(
    ours_ids: list[str] | set[str],
    rows_by_id: dict[str, dict],
) -> list[dict]:
    """Return one explicit Sheet visibility record for every canonical ID."""
    rows: list[dict] = []
    for video_id in sorted(ours_ids):
        source = rows_by_id.get(video_id)
        if not isinstance(source, dict):
            raise RuntimeError(
                f"Canonical Our Videos metadata is missing for {video_id}"
            )
        pub = source.get("pub")
        duration = source.get("durH")
        if pub is not None and (
            isinstance(pub, bool)
            or not isinstance(pub, (int, float))
            or not math.isfinite(float(pub))
            or float(pub) <= 0
        ):
            raise RuntimeError(
                f"Canonical Our Videos publication metadata is invalid for {video_id}"
            )
        if duration is not None and (
            isinstance(duration, bool)
            or not isinstance(duration, (int, float))
            or not math.isfinite(float(duration))
            or float(duration) < 0
        ):
            raise RuntimeError(
                f"Canonical Our Videos duration metadata is invalid for {video_id}"
            )
        rows.append({
            "vid": video_id,
            "pub": int(pub) if pub is not None else None,
            "durH": float(duration) if duration is not None else None,
        })
    return rows


def canonical_ours_metadata_digest(rows: list[dict]) -> str:
    rendered = json.dumps(
        rows,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(rendered.encode("utf-8")).hexdigest()


def sheet_video_catalog() -> dict[str, object]:
    """Load visible video IDs and the canonical Analyse cohort in one request."""
    try:
        from openpyxl import load_workbook

        with urllib.request.urlopen(SHEET_EXPORT, timeout=45) as response:
            payload = response.read()
        workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=False
        )
        values_workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True
        )
        titles = [
            name
            for name in workbook.sheetnames
            if any(fragment in name for fragment in SHEET_VIDEO_TAB_FRAGMENTS)
        ]
        if not any("Our Videos" in name for name in titles):
            raise RuntimeError("Our Videos tab is missing from the radar Sheet")
        ids: set[str] = set()
        ours_ids: set[str] = set()
        ours_rows: dict[str, dict] = {}
        for title in titles:
            is_ours = "Our Videos" in title
            max_col = 6 if is_ours else 2
            value_rows = (
                values_workbook[title].iter_rows(min_row=1, max_col=max_col)
                if is_ours
                else None
            )
            for row in workbook[title].iter_rows(min_row=1, max_col=max_col):
                value_row = next(value_rows) if value_rows is not None else None
                cells = row[:1] if is_ours else row
                for cell in cells:
                    video_id = sheet_cell_video_id(cell)
                    if video_id:
                        ids.add(video_id)
                        if is_ours:
                            ours_ids.add(video_id)
                            ours_rows[video_id] = {
                                "vid": video_id,
                                "pub": sheet_date_ms(
                                    value_row[2].value,
                                    values_workbook.epoch,
                                ),
                                "durH": sheet_number(value_row[5].value),
                            }
                        break
        if not ids:
            raise RuntimeError("No dashboard video ID found in the radar Sheet")
        if not ours_ids:
            raise RuntimeError("Our Videos tab contains no valid video ID")
        if set(ours_rows) != ours_ids:
            raise RuntimeError("Our Videos metadata coverage is incomplete")
        return {"all": ids, "ours": ours_ids, "ours_rows": ours_rows}
    except Exception as exc:
        raise RuntimeError(
            f"Could not load the canonical dashboard video list: {type(exc).__name__}: {exc}"
        ) from exc


def sheet_video_ids() -> set[str]:
    """Load every video ID that the live dashboard can display from the Sheet."""
    return sheet_video_catalog()["all"]


def payload_bucket_ids(payload: dict, buckets: tuple[str, ...]) -> set[str]:
    return {
        str(row.get("vid"))
        for bucket in buckets
        for row in payload.get("d", {}).get(bucket, [])
        if not is_deferred_row(row)
        if VIDEO_ID.match(str(row.get("vid") or ""))
    }


def tracked_ids(
    payload: dict,
    scan_scope: str = "all",
    canonical_sheet_ids: set[str] | None = None,
) -> list[str]:
    if scan_scope not in SCAN_SCOPES:
        raise ValueError(f"Unknown scan scope: {scan_scope}")
    if scan_scope == "kids":
        return sorted(payload_bucket_ids(payload, ("kids",)))
    unavailable = {
        str(video_id)
        for video_id in ((payload.get("videoMetrics") or {}).get("unavailable_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    standard_ids = payload_bucket_ids(payload, ("all", "trends", "news", "ours"))
    kids_ids = payload_bucket_ids(payload, ("kids",))
    # The standard job owns daily counters for every existing public card,
    # including the already-classified Kids cohort. Kids *discovery* remains a
    # separate scope; adding these IDs here only performs the ordinary metrics
    # lookup and prevents their cards/history from freezing between Kids scans.
    ids = set(standard_ids) | kids_ids
    ids.update(canonical_sheet_ids if canonical_sheet_ids is not None else sheet_video_ids())
    return sorted(ids - unavailable)


def write_tracked_manifest(
    snapshot: Path,
    output: Path,
    scan_scope: str = "all",
) -> dict:
    """Resolve the canonical tracked set once for every parallel scan shard."""
    payload = read_snapshot(snapshot)
    sheet_catalog = (
        {"all": set(), "ours": set(), "ours_rows": {}}
        if scan_scope == "kids"
        else sheet_video_catalog()
    )
    ids = tracked_ids(payload, scan_scope, sheet_catalog["all"])
    if not ids and scan_scope != "kids":
        raise RuntimeError("Canonical tracked-video manifest is empty")
    quarantine_ids = {
        str(video_id)
        for video_id in ((payload.get("videoMetrics") or {}).get("unavailable_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    if scan_scope == "kids":
        quarantine_ids.clear()
    canonical_ours_ids = sorted(sheet_catalog["ours"])
    canonical_ours_digest = hashlib.sha256(
        "\n".join(canonical_ours_ids).encode("utf-8")
    ).hexdigest()
    canonical_ours_metadata = canonical_ours_metadata_rows(
        canonical_ours_ids,
        sheet_catalog["ours_rows"],
    )
    canonical_ours_metadata_digest_value = canonical_ours_metadata_digest(
        canonical_ours_metadata
    )
    manifest = {
        "version": 2,
        "scan_scope": scan_scope,
        "generated_ms": utc_now_ms(),
        "snapshot_metrics_ms": int(payload.get("videoMetricsT") or 0),
        "ids": ids,
        # Keep the full canonical Sheet cohort even when an ID is quarantined:
        # publication must never silently turn 83/83 into a false-green 82/82.
        "ours_ids": canonical_ours_ids,
        "ours_total": len(canonical_ours_ids),
        "ours_digest": canonical_ours_digest,
        # Publication date and duration are part of the browser's visibility
        # predicate. Carry them through the one shared manifest so a public
        # counter row cannot lose Sheet metadata in a no-API shard.
        "ours_metadata": canonical_ours_metadata,
        "ours_metadata_total": len(canonical_ours_metadata),
        "ours_metadata_digest": canonical_ours_metadata_digest_value,
        "quarantine_ids": sorted(quarantine_ids),
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    atomic_write_text(output, json.dumps(manifest, ensure_ascii=False, separators=(",", ":")))
    print(json.dumps({"tracked_manifest": len(ids), "output": str(output)}))
    return manifest


def read_tracked_manifest(path: Path, expected_scope: str | None = None) -> list[str]:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    ids = [str(video_id) for video_id in (manifest.get("ids") or []) if VIDEO_ID.match(str(video_id or ""))]
    manifest_scope = str(manifest.get("scan_scope") or "all")
    if int(manifest.get("version") or 0) not in (1, 2) or (
        not ids and manifest_scope != "kids"
    ):
        raise RuntimeError(f"Invalid or empty tracked-video manifest: {path}")
    if expected_scope and manifest_scope != expected_scope:
        raise RuntimeError(
            f"Tracked-video manifest scope mismatch: expected {expected_scope}, got {manifest_scope}"
        )
    if len(ids) != len(set(ids)):
        raise RuntimeError(f"Duplicate IDs in tracked-video manifest: {path}")
    return sorted(ids)


def read_ours_manifest(
    path: Path,
    expected_scope: str | None = None,
) -> tuple[list[str], bool, int, str, dict[str, dict], int, str]:
    """Return canonical Our Videos IDs and whether the manifest proves coverage."""
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest_scope = str(manifest.get("scan_scope") or "all")
    if expected_scope and manifest_scope != expected_scope:
        raise RuntimeError(
            f"Tracked-video manifest scope mismatch: expected {expected_scope}, got {manifest_scope}"
        )
    if manifest_scope == "kids":
        return [], int(manifest.get("version") or 0) >= 2, 0, "", {}, 0, ""
    if int(manifest.get("version") or 0) < 2:
        return [], False, 0, "", {}, 0, ""
    ours_ids = [
        str(video_id)
        for video_id in (manifest.get("ours_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    ]
    if not ours_ids or len(ours_ids) != len(set(ours_ids)):
        raise RuntimeError(f"Invalid canonical Our Videos cohort in tracked manifest: {path}")
    ours_ids = sorted(ours_ids)
    ours_total = int(manifest.get("ours_total") or 0)
    ours_digest = str(manifest.get("ours_digest") or "")
    expected_digest = hashlib.sha256("\n".join(ours_ids).encode("utf-8")).hexdigest()
    if ours_total != len(ours_ids) or ours_digest != expected_digest:
        raise RuntimeError(f"Invalid canonical Our Videos proof in tracked manifest: {path}")
    raw_metadata = manifest.get("ours_metadata") or []
    if not isinstance(raw_metadata, list):
        raise RuntimeError(
            f"Invalid canonical Our Videos metadata proof in tracked manifest: {path}"
        )
    metadata_rows: list[dict] = []
    metadata_by_id: dict[str, dict] = {}
    for source in raw_metadata:
        if not isinstance(source, dict):
            raise RuntimeError(
                f"Invalid canonical Our Videos metadata proof in tracked manifest: {path}"
            )
        video_id = str(source.get("vid") or "")
        if not VIDEO_ID.fullmatch(video_id) or video_id in metadata_by_id:
            raise RuntimeError(
                f"Invalid canonical Our Videos metadata proof in tracked manifest: {path}"
            )
        normalized = canonical_ours_metadata_rows([video_id], {video_id: source})[0]
        metadata_rows.append(normalized)
        metadata_by_id[video_id] = normalized
    metadata_rows.sort(key=lambda row: row["vid"])
    metadata_total = int(manifest.get("ours_metadata_total") or 0)
    metadata_digest = str(manifest.get("ours_metadata_digest") or "")
    expected_metadata_digest = canonical_ours_metadata_digest(metadata_rows)
    if (
        set(metadata_by_id) != set(ours_ids)
        or metadata_total != len(ours_ids)
        or len(metadata_rows) != metadata_total
        or metadata_digest != expected_metadata_digest
    ):
        raise RuntimeError(
            f"Invalid canonical Our Videos metadata proof in tracked manifest: {path}"
        )
    return (
        ours_ids,
        True,
        ours_total,
        ours_digest,
        metadata_by_id,
        metadata_total,
        metadata_digest,
    )


def read_quarantine_manifest(path: Path) -> list[str]:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    return sorted({
        str(video_id)
        for video_id in (manifest.get("quarantine_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    })


def merge_keyword_rows(rows: list[dict]) -> list[dict]:
    by_id: dict[str, dict] = {}
    keywords: dict[str, set[str]] = defaultdict(set)
    ranks: dict[str, list[int]] = defaultdict(list)
    audiences: dict[str, set[str]] = defaultdict(set)
    kids_status: dict[str, list[bool]] = defaultdict(list)
    kids_sources: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        video_id = row["vid"]
        old = by_id.get(video_id)
        if not old or int(row.get("views") or 0) >= int(old.get("views") or 0):
            by_id[video_id] = dict(row)
        keywords[video_id].update(split_keywords(row.get("kw")))
        audiences[video_id].update(
            str(value).lower()
            for value in (row.get("audiences") or [])
            if str(value).lower() in {"youtube", "kids"}
        )
        if isinstance(row.get("madeForKids"), bool):
            kids_status[video_id].append(row["madeForKids"])
        if row.get("madeForKids") is True and row.get("madeForKidsSource"):
            kids_sources[video_id].add(str(row["madeForKidsSource"]))
        if isinstance(row.get("rank"), (int, float)):
            ranks[video_id].append(int(row["rank"]))
    for video_id, row in by_id.items():
        if keywords[video_id]:
            row["kw"] = "; ".join(sorted(keywords[video_id], key=str.lower))
            row["kwCount"] = len(keywords[video_id])
        if ranks[video_id]:
            row["rank"] = min(ranks[video_id])
        if audiences[video_id]:
            row["audiences"] = sorted(audiences[video_id], key=("youtube", "kids").index)
        if True in kids_status[video_id]:
            row["madeForKids"] = True
            if kids_sources[video_id]:
                row["madeForKidsSource"] = sorted(kids_sources[video_id])[0]
        elif False in kids_status[video_id]:
            row["madeForKids"] = False
    return list(by_id.values())


def preserve_audience_classification(winner: dict, other: dict | None) -> dict:
    """Keep official Kids truth and discovery provenance when one fresher row wins."""
    if not other:
        return winner
    merged = dict(winner)
    statuses = [
        value
        for value in (winner.get("madeForKids"), other.get("madeForKids"))
        if isinstance(value, bool)
    ]
    if True in statuses:
        merged["madeForKids"] = True
        for source_row in (winner, other):
            if source_row.get("madeForKids") is True and source_row.get("madeForKidsSource"):
                merged["madeForKidsSource"] = source_row["madeForKidsSource"]
                break
    elif False in statuses:
        merged["madeForKids"] = False
    audiences = {
        str(value).lower()
        for row in (winner, other)
        for value in (row.get("audiences") or [])
        if str(value).lower() in {"youtube", "kids"}
    }
    if audiences:
        merged["audiences"] = sorted(audiences, key=("youtube", "kids").index)
    # Counter selection and metadata selection are intentionally independent.
    # A yt-dlp row can carry the newest counter while the official API row has
    # the precise publication timestamp and canonical snippet metadata.
    for key in CARD_METADATA_FIELDS:
        if other.get(key) in (None, ""):
            continue
        if metadata_priority(other, key) > metadata_priority(merged, key):
            merged[key] = other[key]
            if key == "pub" and other.get("pubSource"):
                merged["pubSource"] = other["pubSource"]
    if metadata_priority(other) > metadata_priority(merged):
        merged["metadataSource"] = other.get("metadataSource")
    return merged


def fetch_discovery_spec(spec: dict, now_ms: int, api_key: str) -> tuple:
    if spec.get("audience") == "kids":
        if api_key:
            return fetch_kids_search(spec, now_ms, api_key)
        return fetch_kids_search_ydl(spec, now_ms)
    return fetch_search(spec, now_ms)


def run_shard(
    snapshot: Path,
    output: Path,
    shard: int,
    shards: int,
    tracked_manifest: Path | None = None,
    scan_scope: str = "all",
) -> dict:
    if scan_scope not in SCAN_SCOPES:
        raise ValueError(f"Unknown scan scope: {scan_scope}")
    payload = read_snapshot(snapshot)
    now_ms = utc_now_ms()
    all_tracked_ids = (
        read_tracked_manifest(tracked_manifest, scan_scope)
        if tracked_manifest
        else tracked_ids(payload, scan_scope)
    )
    if tracked_manifest:
        (
            all_canonical_ours_ids,
            canonical_ours_manifest,
            canonical_ours_total,
            canonical_ours_digest,
            all_canonical_ours_metadata,
            canonical_ours_metadata_total,
            canonical_ours_metadata_digest_value,
        ) = read_ours_manifest(
            tracked_manifest, scan_scope
        )
    elif scan_scope == "kids":
        all_canonical_ours_ids, canonical_ours_manifest = [], True
        canonical_ours_total, canonical_ours_digest = 0, ""
        all_canonical_ours_metadata = {}
        canonical_ours_metadata_total = 0
        canonical_ours_metadata_digest_value = ""
    else:
        # Collector mode without a shared manifest remains supported locally,
        # but still resolves the canonical Analyse cohort fail-closed.
        sheet_catalog = sheet_video_catalog()
        all_canonical_ours_ids = sorted(sheet_catalog["ours"])
        canonical_ours_manifest = True
        canonical_ours_total = len(all_canonical_ours_ids)
        canonical_ours_digest = hashlib.sha256(
            "\n".join(all_canonical_ours_ids).encode("utf-8")
        ).hexdigest()
        metadata_rows = canonical_ours_metadata_rows(
            all_canonical_ours_ids,
            sheet_catalog["ours_rows"],
        )
        all_canonical_ours_metadata = {
            row["vid"]: row for row in metadata_rows
        }
        canonical_ours_metadata_total = len(metadata_rows)
        canonical_ours_metadata_digest_value = canonical_ours_metadata_digest(
            metadata_rows
        )
    all_quarantine_ids = (
        []
        if scan_scope == "kids"
        else (
            read_quarantine_manifest(tracked_manifest)
            if tracked_manifest
            else [
                str(video_id)
                for video_id in ((payload.get("videoMetrics") or {}).get("unavailable_ids") or [])
                if VIDEO_ID.match(str(video_id or ""))
            ]
        )
    )
    ids = [video_id for video_id in all_tracked_ids if stable_shard(video_id, shards) == shard]
    canonical_ours_ids = [
        video_id
        for video_id in all_canonical_ours_ids
        if stable_shard(video_id, shards) == shard
    ]
    canonical_ours_metadata = [
        all_canonical_ours_metadata[video_id]
        for video_id in canonical_ours_ids
    ]
    quarantine_ids = [
        video_id for video_id in all_quarantine_ids if stable_shard(video_id, shards) == shard
    ]
    lookup_ids = sorted(set(ids) | set(quarantine_ids))
    fresh: dict[str, dict] = {}
    track_failed = 0
    api_key = os.environ.get("YOUTUBE_API_KEY", "").strip()
    all_specs = query_specs(payload, include_kids=scan_scope != "standard")
    specs = [
        spec
        for spec in all_specs
        if (
            scan_scope == "all"
            or (scan_scope == "kids" and spec.get("audience") == "kids")
            or (scan_scope == "standard" and spec.get("audience") != "kids")
        )
        and stable_shard(spec["query"], shards) == shard
    ]
    owned_fresh: dict[str, dict] = {}
    owned_ok = True
    live_audiences: dict[str, dict] = {}
    # Official-upload discovery belongs only to the standard radar.
    if shard == 0 and scan_scope != "kids":
        try:
            owned_fresh = fetch_owned_api_rows(now_ms, api_key) if api_key else fetch_owned_ydl_rows(now_ms)
            fresh.update(owned_fresh)
        except Exception as exc:
            owned_ok = False
            print(f"WARN official-upload discovery: {type(exc).__name__}: {exc}", file=sys.stderr)
        if api_key:
            live_ids = sorted({
                str(row.get("vid") or "")
                for row in payload.get("d", {}).get("lives", [])
                if VIDEO_ID.match(str(row.get("vid") or ""))
            })
            if live_ids:
                try:
                    live_rows = fetch_api_rows(live_ids, now_ms, api_key)
                    live_audiences = {
                        video_id: {
                            "madeForKids": row["madeForKids"],
                            "audiences": ["kids"] if row["madeForKids"] else ["youtube"],
                        }
                        for video_id, row in live_rows.items()
                        if isinstance(row.get("madeForKids"), bool)
                    }
                except Exception as exc:
                    print(
                        f"WARN livestream audience refresh: {type(exc).__name__}: {exc}",
                        file=sys.stderr,
                    )
    if api_key:
        try:
            fresh.update(fetch_api_rows(lookup_ids, now_ms, api_key))
        except Exception as exc:
            raise RuntimeError(f"YouTube Data API metrics failed closed: {type(exc).__name__}: {exc}") from exc
    else:
        with concurrent.futures.ThreadPoolExecutor(max_workers=TRACK_WORKERS) as pool:
            future_to_id = {
                pool.submit(fetch_one_video, video_id, now_ms): video_id for video_id in lookup_ids
            }
            for future in concurrent.futures.as_completed(future_to_id):
                video_id = future_to_id[future]
                try:
                    row = future.result()
                    if row:
                        fresh[video_id] = row
                    else:
                        track_failed += 1
                except Exception as exc:
                    track_failed += 1
                    print(f"WARN tracked {video_id}: {type(exc).__name__}: {exc}", file=sys.stderr)

    tracked_fresh_ids = sorted(set(ids) & set(fresh))
    tracked_failed_ids = sorted(set(ids) - set(tracked_fresh_ids))
    if api_key:
        tracked_unavailable_ids = tracked_failed_ids
    else:
        tracked_unavailable_ids = sorted(
            video_id
            for video_id in tracked_failed_ids
            if public_oembed_confirms_unavailable(video_id)
        )
    tracked_recovered_ids = sorted(set(quarantine_ids) & set(fresh))
    if tracked_unavailable_ids:
        print(
            f"INFO shard {shard}: {len(tracked_unavailable_ids)} tracked IDs "
            "confirmed publicly unavailable by YouTube oEmbed: "
            + ", ".join(tracked_unavailable_ids[:24]),
            file=sys.stderr,
        )
    if tracked_failed_ids:
        print(
            f"WARN shard {shard}: {len(tracked_failed_ids)} tracked IDs missing: "
            + ", ".join(tracked_failed_ids[:24]),
            file=sys.stderr,
        )

    candidates: list[dict] = []
    query_failed = 0
    query_raw = 0
    query_enriched = 0
    kids_queries_total = sum(1 for spec in specs if spec.get("audience") == "kids")
    kids_query_failed = 0
    kids_results_examined = 0
    kids_candidates_kept = 0
    kids_funnel = empty_kids_funnel()
    kids_funnel_queries = 0
    # A canary failure is not an ordinary partial keyword miss: without a
    # trustworthy rendered signal, no Kids classification may be published.
    if kids_queries_total and not api_key:
        kids_dom_validator().ensure_canaries()
    with concurrent.futures.ThreadPoolExecutor(max_workers=SEARCH_WORKERS) as pool:
        future_to_spec = {
            pool.submit(fetch_discovery_spec, spec, now_ms, api_key): spec
            for spec in specs
        }
        for future in concurrent.futures.as_completed(future_to_spec):
            spec = future_to_spec[future]
            try:
                result = future.result()
                if len(result) == 4:
                    rows, raw_count, enriched_count, funnel = result
                else:
                    rows, raw_count, enriched_count = result
                    funnel = None
                if spec.get("audience") == "kids":
                    merge_kids_funnel(kids_funnel, funnel)
                    if isinstance(funnel, dict):
                        kids_funnel_queries += 1
                query_raw += raw_count
                query_enriched += enriched_count
                if spec.get("audience") == "kids":
                    kids_results_examined += raw_count
                    kids_candidates_kept += len(rows)
                candidates.extend(rows)
                for row in rows:
                    previous = fresh.get(row["vid"])
                    if not previous or int(row.get("views") or 0) >= int(previous.get("views") or 0):
                        fresh[row["vid"]] = preserve_audience_classification(row, previous)
                    else:
                        fresh[row["vid"]] = preserve_audience_classification(previous, row)
            except Exception as exc:
                query_failed += 1
                if spec.get("audience") == "kids":
                    kids_query_failed += 1
                print(f"WARN query {spec['query']}: {type(exc).__name__}: {exc}", file=sys.stderr)

    track_ok = len(tracked_fresh_ids)
    query_ok = len(specs) - query_failed
    if ids and track_ok / len(ids) < MIN_TRACK_RATIO:
        raise RuntimeError(f"Shard {shard}: only {track_ok}/{len(ids)} tracked videos refreshed")
    if query_ok < len(specs):
        print(
            f"WARN shard {shard}: discovery is partial at {query_ok}/{len(specs)} keyword searches",
            file=sys.stderr,
        )
    if (
        kids_queries_total
        and kids_query_failed == 0
        and kids_funnel_queries == kids_queries_total
    ):
        validate_kids_funnel(kids_funnel, kids_results_examined, kids_candidates_kept)

    artifact = {
        "version": 2,
        "scan_scope": scan_scope,
        "generated_at": datetime.fromtimestamp(now_ms / 1000, timezone.utc).isoformat(),
        "generated_ms": now_ms,
        "shard": shard,
        "shards": shards,
        "tracked_total": len(ids),
        "tracked_ok": track_ok,
        "queries_total": len(specs),
        "queries_ok": query_ok,
        "queries_raw": query_raw,
        "queries_enriched": query_enriched,
        "kids_queries_total": kids_queries_total,
        "kids_queries_ok": kids_queries_total - kids_query_failed,
        "kids_results_examined": kids_results_examined,
        "kids_candidates_kept": kids_candidates_kept,
        "kids_funnel": kids_funnel,
        "kids_funnel_queries": kids_funnel_queries,
        "owned_ok": owned_ok,
        "tracked_ids": ids,
        "canonical_ours_manifest": canonical_ours_manifest,
        "canonical_ours_ids": canonical_ours_ids,
        "canonical_ours_total": canonical_ours_total,
        "canonical_ours_digest": canonical_ours_digest,
        "canonical_ours_metadata": canonical_ours_metadata,
        "canonical_ours_metadata_total": canonical_ours_metadata_total,
        "canonical_ours_metadata_digest": canonical_ours_metadata_digest_value,
        "tracked_fresh_ids": tracked_fresh_ids,
        "tracked_failed_ids": tracked_failed_ids,
        "tracked_unavailable_ids": tracked_unavailable_ids,
        "tracked_recovered_ids": tracked_recovered_ids,
        "owned_fresh": list(owned_fresh.values()),
        "live_audiences": live_audiences,
        "fresh": list(fresh.values()),
        "candidates": merge_keyword_rows(candidates),
    }
    atomic_write_text(output, json.dumps(artifact, ensure_ascii=False, separators=(",", ":")))
    print(json.dumps({key: artifact[key] for key in ("shard", "tracked_total", "tracked_ok", "queries_total", "queries_ok")}))
    return artifact


def metadata_priority(row: dict, field: str | None = None) -> int:
    source = str(row.get("metadataSource") or "")
    if field == "pub":
        source = str(row.get("pubSource") or source)
    elif field == "durH":
        source = str(row.get("durationSource") or source)
    if not source and row.get("madeForKidsSource") == "youtube_data_api_status":
        source = METADATA_SOURCE_API
    return METADATA_SOURCE_PRIORITY.get(source, 0)


def should_replace_metadata(existing: dict, fresh: dict, key: str) -> bool:
    if existing.get(key) in (None, ""):
        return True
    existing_priority = metadata_priority(existing, key)
    fresh_priority = metadata_priority(fresh, key)
    # Legacy card metadata often came from the official API before provenance
    # fields existed. Never let a lower-precision yt-dlp discovery row silently
    # replace it; the next API refresh can still update it authoritatively.
    if existing_priority == 0 and fresh_priority == 1:
        if key == "pub":
            existing["pubSource"] = METADATA_SOURCE_PRESERVED
        else:
            existing["metadataSource"] = METADATA_SOURCE_PRESERVED
        return False
    return fresh_priority >= existing_priority


def update_row(existing: dict, fresh: dict, now_ms: int) -> None:
    if isinstance(fresh.get("views"), (int, float)):
        existing["views"] = int(fresh["views"])
    comments = parse_nonnegative_count(fresh.get("comments"))
    if comments is not None:
        existing["comments"] = comments
    for key in CARD_METADATA_FIELDS:
        if fresh.get(key) not in (None, "") and should_replace_metadata(existing, fresh, key):
            existing[key] = fresh[key]
            if key == "pub" and fresh.get("pubSource"):
                existing["pubSource"] = fresh["pubSource"]
            elif key == "durH":
                duration_source = fresh.get("durationSource") or fresh.get("metadataSource")
                if duration_source:
                    existing["durationSource"] = duration_source
    if fresh.get("metadataSource") and metadata_priority(fresh) >= metadata_priority(existing):
        existing["metadataSource"] = fresh["metadataSource"]
    for key in (
        "madeForKidsSource", "instrumentalVerified", "instrumentalEvidenceSource",
        "discoveryLanes", "liveStatus",
    ):
        if fresh.get(key) not in (None, ""):
            existing[key] = fresh[key]
    if isinstance(fresh.get("madeForKids"), bool):
        existing["madeForKids"] = fresh["madeForKids"]
    if isinstance(fresh.get("audiences"), list):
        existing["audiences"] = list(dict.fromkeys(
            str(value).lower()
            for value in fresh["audiences"]
            if str(value).lower() in {"youtube", "kids"}
        ))
    if not str(existing.get("genre") or "").strip() and fresh.get("genre"):
        existing["genre"] = fresh["genre"]
        if fresh.get("genreSource"):
            existing["genreSource"] = fresh["genreSource"]
    published = existing.get("pub")
    views = existing.get("views")
    age = age_months(int(published), now_ms) if isinstance(published, (int, float)) else None
    if age is not None:
        existing["ageM"] = age
        if isinstance(views, (int, float)):
            existing["vpm"] = views / age


def merge_owned_metadata(existing: dict, discovered: dict) -> None:
    """Use official-upload discovery to add/fill cards, never as counter truth."""
    for key in CARD_METADATA_FIELDS + (
        "metadataSource", "pubSource", "source", "genre", "genreSource",
    ):
        if existing.get(key) in (None, "") and discovered.get(key) not in (None, ""):
            existing[key] = discovered[key]


def merge_sheet_ours_metadata(existing: dict, sheet_row: dict) -> None:
    """Fill the same Our Videos fields that the browser retains from Sheet."""
    published = sheet_row.get("pub")
    if existing.get("pub") in (None, "") and isinstance(published, (int, float)):
        existing["pub"] = int(published)
        existing["pubSource"] = METADATA_SOURCE_SHEET
    duration = sheet_row.get("durH")
    if existing.get("durH") in (None, "") and isinstance(duration, (int, float)):
        existing["durH"] = float(duration)
        existing["durationSource"] = METADATA_SOURCE_SHEET


def merge_discovery_fields(existing: dict, discovered: dict) -> None:
    keywords = set(split_keywords(existing.get("kw")))
    keywords.update(split_keywords(discovered.get("kw")))
    if keywords:
        existing["kw"] = "; ".join(sorted(keywords, key=str.lower))
        existing["kwCount"] = len(keywords)
    ranks = [
        int(value)
        for value in (existing.get("rank"), discovered.get("rank"))
        if isinstance(value, (int, float)) and value > 0
    ]
    if ranks:
        existing["rank"] = min(ranks)
    audiences = {
        str(value).lower()
        for row in (existing, discovered)
        for value in (row.get("audiences") or [])
        if str(value).lower() in {"youtube", "kids"}
    }
    if audiences:
        existing["audiences"] = sorted(audiences, key=("youtube", "kids").index)
    if discovered.get("madeForKids") is True or existing.get("madeForKids") is True:
        existing["madeForKids"] = True
        if discovered.get("madeForKidsSource"):
            existing["madeForKidsSource"] = discovered["madeForKidsSource"]
    elif isinstance(discovered.get("madeForKids"), bool):
        existing["madeForKids"] = discovered["madeForKids"]


def history_day_key(timestamp_ms: int) -> str:
    """Return the dashboard's business day for a UTC timestamp."""
    return datetime.fromtimestamp(timestamp_ms / 1000, timezone.utc).astimezone(RADAR_TIMEZONE).date().isoformat()


def normalize_daily_points(points: list, now_ms: int) -> list[list[int]]:
    by_day: dict[object, list[int]] = {}
    for point in points or []:
        if isinstance(point, list) and len(point) >= 2:
            try:
                parsed = [int(point[0]), int(point[1])]
                day = history_day_key(parsed[0])
                if day not in by_day or parsed[0] >= by_day[day][0]:
                    by_day[day] = parsed
            except (TypeError, ValueError):
                pass
    cutoff = max(now_ms - HISTORY_RETENTION_DAYS * 86400000, DAILY_VIEW_HISTORY_START_MS)
    return sorted(
        (point for point in by_day.values() if point[0] >= cutoff),
        key=lambda point: point[0],
    )


def append_daily_point(points: list, now_ms: int, views: int) -> list[list[int]]:
    return normalize_daily_points(list(points or []) + [[now_ms, int(views)]], now_ms)


def history_shard_name(video_id: str) -> str:
    return f"{ord(video_id[0]):02x}.json"


def update_history_shards(
    history_dir: Path,
    desired_ids: set[str],
    fresh: dict[str, dict],
    legacy: dict,
    now_ms: int,
    scope_only: bool = False,
) -> tuple[int, int]:
    """Update selected histories without changing points owned by another scope."""
    history_dir.mkdir(parents=True, exist_ok=True)
    names = {history_shard_name(video_id) for video_id in desired_ids}
    names.update(history_shard_name(video_id) for video_id in legacy if VIDEO_ID.match(video_id))
    if not scope_only:
        # The standard publication verifier expects every existing shard header
        # to carry the current standard refresh timestamp.
        names.update(path.name for path in history_dir.glob("*.json"))
    total_ids = 0
    written = 0
    for name in sorted(names):
        path = history_dir / name
        current = {}
        if path.exists():
            try:
                current = (json.loads(path.read_text(encoding="utf-8")).get("d") or {})
            except (OSError, ValueError, AttributeError):
                current = {}
        updated: dict[str, list[list[int]]] = {}
        candidate_ids = list(current)
        candidate_ids.extend(sorted(
            video_id
            for video_id in desired_ids
            if history_shard_name(video_id) == name and video_id not in current
        ))
        for video_id in candidate_ids:
            points = list(current.get(video_id) or [])
            if video_id in desired_ids:
                points += list(legacy.get(video_id) or [])
                row = fresh.get(video_id)
                if row and isinstance(row.get("views"), (int, float)):
                    points.append([now_ms, int(row["views"])])
                clean = normalize_daily_points(points, now_ms)
            else:
                # Preserve another scope byte-for-value at the series level.
                clean = points
            if clean:
                updated[video_id] = clean
        total_ids += len(updated)
        rendered = json.dumps(
            {"version": 1, "updated": now_ms, "d": updated},
            ensure_ascii=False,
            separators=(",", ":"),
        )
        previous = path.read_text(encoding="utf-8") if path.exists() else None
        if previous != rendered:
            atomic_write_text(path, rendered)
            written += 1
    return total_ids, written


def validate_history_refresh(
    history_dir: Path,
    expected_views: dict[str, int],
    now_ms: int,
) -> int:
    """Fail the publication if a refreshed counter did not reach daily history."""
    expected_day = history_day_key(now_ms)
    missing: list[str] = []
    for video_id, views in sorted(expected_views.items()):
        path = history_dir / history_shard_name(video_id)
        try:
            points = (json.loads(path.read_text(encoding="utf-8")).get("d") or {}).get(video_id) or []
        except (OSError, ValueError, AttributeError):
            points = []
        day_points = [
            point
            for point in points
            if isinstance(point, list)
            and len(point) >= 2
            and history_day_key(int(point[0])) == expected_day
        ]
        if not day_points or int(day_points[-1][1]) != int(views):
            missing.append(video_id)
    if missing:
        sample = ", ".join(missing[:8])
        raise RuntimeError(
            f"History refresh rejected for {len(missing)}/{len(expected_views)} videos "
            f"on {expected_day} ({sample})"
        )
    return len(expected_views)


def validate_card_refresh(
    data: dict,
    fresh: dict[str, dict],
    refreshed_ids: set[str],
    now_ms: int,
) -> tuple[int, int]:
    """Fail closed when a refreshed history counter did not reach every card."""
    expected = 0
    failures: list[str] = []
    for bucket in CARD_BUCKETS:
        for index, row in enumerate(data.get(bucket) or []):
            video_id = str(row.get("vid") or "")
            if video_id not in refreshed_ids:
                continue
            expected += 1
            current = fresh.get(video_id) or {}
            fresh_views = current.get("views")
            published = row.get("pub")
            fresh_published = current.get("pub")
            label = f"{bucket}[{index}]/{video_id}"
            if not isinstance(fresh_views, (int, float)) or int(row.get("views") or -1) != int(fresh_views):
                failures.append(label + ":views")
                continue
            if not isinstance(published, (int, float)) or not isinstance(fresh_published, (int, float)):
                failures.append(label + ":pub")
                continue
            if (
                metadata_priority(current, "pub") >= metadata_priority(row, "pub")
                and int(published) != int(fresh_published)
            ):
                failures.append(label + ":pub")
                continue
            expected_age = age_months(int(published), now_ms)
            age = row.get("ageM")
            vpm = row.get("vpm")
            expected_vpm = float(fresh_views) / expected_age if expected_age else None
            if (
                expected_age is None
                or not isinstance(age, (int, float))
                or abs(float(age) - expected_age) > max(1e-9, expected_age * 1e-9)
            ):
                failures.append(label + ":ageM")
                continue
            if (
                expected_vpm is None
                or not isinstance(vpm, (int, float))
                or abs(float(vpm) - expected_vpm) > max(1e-6, expected_vpm * 1e-9)
            ):
                failures.append(label + ":vpm")
    if not expected:
        raise RuntimeError("Card refresh rejected: no refreshed card row was validated")
    if failures:
        sample = ", ".join(failures[:8])
        raise RuntimeError(
            f"Card refresh rejected for {len(failures)}/{expected} rows ({sample})"
        )
    return expected, expected


def is_analysis_card(row: dict) -> bool:
    """Mirror the Analyse UI's factual visibility predicate exactly."""
    published = row.get("pub")
    if not isinstance(published, (int, float)) or published <= 0:
        return False
    duration = row.get("durH")
    # anaResolvedDurationHours() treats zero/invalid direct durations as
    # unknown; an unknown duration remains visible. Only a known positive
    # short-form duration is excluded.
    if not isinstance(duration, (int, float)) or not math.isfinite(float(duration)):
        return True
    return float(duration) <= 0 or float(duration) >= 0.15


def write_avatar_overlay(payload: dict, path: Path) -> int:
    channels: dict[str, str] = {}
    aliases: dict[str, str] = {}
    for bucket in ("all", "trends", "news", "kids"):
        for row in payload.get("d", {}).get(bucket, []):
            channel_url = str(row.get("chUrl") or "")
            id_match = re.search(r"/channel/(UC[\w-]+)", channel_url)
            handle_match = re.search(r"youtube\.com/@([^/?#]+)", channel_url, re.I)
            channel_id = str(row.get("channelId") or (id_match.group(1) if id_match else ""))
            handle = "@" + handle_match.group(1) if handle_match else ""
            if CHANNEL_ID.match(channel_id):
                channels[channel_id] = f"https://unavatar.io/youtube/{channel_id}?fallback=false"
                if handle:
                    aliases[handle] = channel_id
            elif handle:
                channels[handle] = f"https://unavatar.io/youtube/{handle}?fallback=false"
    rendered = (
        "/* Channel logos refreshed automatically alongside the daily YouTube radar. */\n"
        "window.YT_CHANNEL_AVATARS=window.YT_CHANNEL_AVATARS||{channels:{},videos:{}};\n"
        "window.YT_CHANNEL_AVATARS.channels=window.YT_CHANNEL_AVATARS.channels||{};\n"
        "window.YT_CHANNEL_AVATARS.aliases=window.YT_CHANNEL_AVATARS.aliases||{};\n"
        "(()=>{const atlas=window.YT_CHANNEL_AVATARS,fresh="
        + json.dumps(channels, ensure_ascii=False, separators=(",", ":"))
        + ";Object.keys(fresh).forEach(key=>{if(!atlas.channels[key])atlas.channels[key]=fresh[key];});"
        "Object.assign(atlas.aliases,"
        + json.dumps(aliases, ensure_ascii=False, separators=(",", ":"))
        + ");})();\n"
    )
    atomic_write_text(path, rendered)
    return len(channels)


KIDS_VERIFICATION_SOURCES = {
    "youtube_data_api_status",
    "youtube_innertube_android_player_restrictions",
    "youtube_public_player_restrictions",
}


def load_scoped_artifacts(
    merge_dir: Path,
    expected_shards: int,
    scan_scope: str,
) -> list[dict]:
    files = sorted(merge_dir.rglob("youtube-shard-*.json"))
    artifacts = [json.loads(path.read_text(encoding="utf-8")) for path in files]
    if len(artifacts) != expected_shards:
        raise RuntimeError(
            f"Expected exactly {expected_shards} shard artifacts, got {len(artifacts)}"
        )
    shard_values = [int(artifact.get("shard", -1)) for artifact in artifacts]
    if len(set(shard_values)) != len(shard_values):
        raise RuntimeError(f"Duplicate shard artifacts: {shard_values}")
    if set(shard_values) != set(range(expected_shards)):
        raise RuntimeError(
            f"Expected shards 0..{expected_shards - 1}, got {sorted(shard_values)}"
        )
    if any(int(artifact.get("shards", -1)) != expected_shards for artifact in artifacts):
        raise RuntimeError("Shard-count mismatch in artifacts")
    scopes = {str(artifact.get("scan_scope") or "all") for artifact in artifacts}
    if scopes != {scan_scope}:
        raise RuntimeError(
            f"Artifact scope mismatch: expected {scan_scope}, got {sorted(scopes)}"
        )
    return artifacts


def is_verified_kids_candidate(row: dict) -> bool:
    audiences = {str(value).lower() for value in (row.get("audiences") or [])}
    duration = row.get("durH")
    views = row.get("views")
    vpm = row.get("vpm")
    return (
        str(row.get("vid") or "") not in KIDS_CONFIRMED_VOCAL_VIDEO_IDS
        and audiences == {"kids"}
        and row.get("madeForKids") is True
        and row.get("madeForKidsSource") in KIDS_VERIFICATION_SOURCES
        and row.get("instrumentalVerified") is True
        and row.get("liveStatus") == "none"
        and isinstance(duration, (int, float))
        and duration * 3600 >= MIN_SECONDS
        and isinstance(views, (int, float))
        and int(views) >= MIN_KIDS_VIEWS
        and isinstance(vpm, (int, float))
        and float(vpm) >= MIN_KIDS_VPM
        and not is_deferred_row(row)
    )


def merge_kids_artifacts(
    snapshot: Path,
    avatars: Path,
    merge_dir: Path,
    expected_shards: int,
    history_dir: Path | None,
    require_kids: bool,
) -> dict:
    artifacts = load_scoped_artifacts(merge_dir, expected_shards, "kids")
    payload = read_snapshot(snapshot)
    data = payload.setdefault("d", {})
    previous_kids = list(data.get("kids") or [])
    bootstrap = not bool(previous_kids)

    tracked_total = sum(int(a.get("tracked_total", 0)) for a in artifacts)
    tracked_ok = sum(int(a.get("tracked_ok", 0)) for a in artifacts)
    queries_total = sum(int(a.get("queries_total", 0)) for a in artifacts)
    queries_ok = sum(int(a.get("queries_ok", 0)) for a in artifacts)
    queries_raw = sum(int(a.get("queries_raw", 0)) for a in artifacts)
    queries_enriched = sum(int(a.get("queries_enriched", 0)) for a in artifacts)
    kids_queries_total = sum(int(a.get("kids_queries_total", 0)) for a in artifacts)
    kids_queries_ok = sum(int(a.get("kids_queries_ok", 0)) for a in artifacts)
    kids_results_examined = sum(int(a.get("kids_results_examined", 0)) for a in artifacts)
    kids_candidates_kept = sum(int(a.get("kids_candidates_kept", 0)) for a in artifacts)
    kids_funnel = empty_kids_funnel()
    for artifact in artifacts:
        merge_kids_funnel(kids_funnel, artifact.get("kids_funnel"))
    funnel_contract = all(
        int(artifact.get("version") or 0) >= 2 for artifact in artifacts
    )
    if require_kids and not funnel_contract:
        raise RuntimeError("Kids merge rejected: artifacts have no verified funnel contract")
    if require_kids and funnel_contract:
        validate_kids_funnel(kids_funnel, kids_results_examined, kids_candidates_kept)

    if require_kids and (
        kids_queries_total != len(KIDS_QUERY_SPECS)
        or kids_queries_ok != len(KIDS_QUERY_SPECS)
    ):
        raise RuntimeError(
            f"Merge rejected: expected {len(KIDS_QUERY_SPECS)}/{len(KIDS_QUERY_SPECS)} "
            f"Kids queries, got {kids_queries_ok}/{kids_queries_total}"
        )
    if queries_total != kids_queries_total or queries_ok != kids_queries_ok:
        raise RuntimeError("Kids merge rejected: artifact contains non-Kids discovery queries")
    verified_candidates = merge_keyword_rows([
        row
        for artifact in artifacts
        for row in (artifact.get("candidates") or [])
        if is_verified_kids_candidate(row)
    ])
    if require_kids and (
        kids_results_examined < KIDS_MIN_RESULTS_EXAMINED
        or kids_candidates_kept <= 0
        or not verified_candidates
    ):
        raise RuntimeError(
            "Kids merge rejected: discovery coverage or verified candidates are insufficient"
        )
    expected_lane_calls = len(KIDS_QUERY_SPECS) * 2
    if require_kids and funnel_contract and (
        int(kids_funnel.get("lane_calls_expected", 0)) != expected_lane_calls
        or int(kids_funnel.get("lane_calls_completed", 0)) != expected_lane_calls
    ):
        raise RuntimeError(
            "Kids merge rejected: expected 80/80 search lanes, got "
            f"{int(kids_funnel.get('lane_calls_completed', 0))}/"
            f"{int(kids_funnel.get('lane_calls_expected', 0))}"
        )
    if tracked_total:
        if tracked_ok / tracked_total < MIN_PUBLISH_TRACK_RATIO:
            raise RuntimeError(
                f"Kids merge rejected: {tracked_ok}/{tracked_total} tracked videos refreshed"
            )
    elif previous_kids:
        raise RuntimeError("Kids merge rejected: existing cohort produced no tracked IDs")

    tracked_fresh_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_fresh_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    tracked_failed_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_failed_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    if len(tracked_fresh_ids) != tracked_ok:
        raise RuntimeError(
            f"Kids merge rejected: {tracked_ok} refreshed counters but "
            f"{len(tracked_fresh_ids)} traceable refreshed IDs"
        )
    if (
        len(tracked_failed_ids) != tracked_total - tracked_ok
        or tracked_failed_ids & tracked_fresh_ids
    ):
        raise RuntimeError("Kids merge rejected: inconsistent tracked failure IDs")

    now_ms = max(int(a.get("generated_ms", 0)) for a in artifacts) or utc_now_ms()
    fresh: dict[str, dict] = {}
    candidates: list[dict] = []
    for artifact in artifacts:
        candidates.extend(artifact.get("candidates") or [])
        for row in artifact.get("fresh") or []:
            video_id = str(row.get("vid") or "")
            if not VIDEO_ID.match(video_id):
                continue
            previous = fresh.get(video_id)
            if not previous or int(row.get("views") or 0) >= int(previous.get("views") or 0):
                fresh[video_id] = preserve_audience_classification(row, previous)
            else:
                fresh[video_id] = preserve_audience_classification(previous, row)

    kids_rows = list(previous_kids)
    by_kids = {
        str(row.get("vid")): row
        for row in kids_rows
        if VIDEO_ID.match(str(row.get("vid") or ""))
    }
    for row in kids_rows:
        current = fresh.get(str(row.get("vid") or ""))
        if current:
            update_row(row, current, now_ms)

    inserted_kids = 0
    for row in merge_keyword_rows(candidates):
        if not is_verified_kids_candidate(row):
            continue
        current = by_kids.get(row["vid"])
        if current:
            update_row(current, row, now_ms)
            merge_discovery_fields(current, row)
            current["instrumentalVerified"] = True
            current["liveStatus"] = "none"
        else:
            added = dict(row)
            kids_rows.append(added)
            by_kids[added["vid"]] = added
            inserted_kids += 1

    kids_rows = [
        row
        for row in kids_rows
        if is_verified_kids_candidate(row)
    ]
    kids_rows.sort(key=lambda row: row.get("vpm") or 0, reverse=True)
    if require_kids and bootstrap and not kids_rows:
        raise RuntimeError("Merge rejected: initial Kids scan returned no verified candidates")
    data["kids"] = kids_rows

    desired_ids = {
        str(row.get("vid"))
        for row in kids_rows
        if VIDEO_ID.match(str(row.get("vid") or ""))
    }
    resolved_history_dir = history_dir or snapshot.parent / "video_history"
    history_ids, history_files = update_history_shards(
        resolved_history_dir,
        desired_ids,
        fresh,
        {},
        now_ms,
        scope_only=True,
    )
    expected_history_views = {
        video_id: int(fresh[video_id]["views"])
        for video_id in desired_ids & set(fresh)
        if isinstance(fresh[video_id].get("views"), (int, float))
    }
    history_updated = validate_history_refresh(
        resolved_history_dir,
        expected_history_views,
        now_ms,
    )
    day = history_day_key(now_ms)
    payload["kidsMetricsT"] = now_ms
    payload["kidsMetrics"] = {
        "day": day,
        "day_timezone": RADAR_TIMEZONE_NAME,
        "tracked": len(kids_rows),
        "updated": len(expected_history_views),
        "queries": kids_queries_total,
        "queries_ok": kids_queries_ok,
        "search_results": queries_raw,
        "search_results_enriched": queries_enriched,
        "results_examined": kids_results_examined,
        "results_examined_minimum": KIDS_MIN_RESULTS_EXAMINED,
        "candidates_kept": kids_candidates_kept,
        "verified_unique": len(verified_candidates),
        "duplicate_occurrences": max(0, kids_candidates_kept - len(verified_candidates)),
        "funnel": kids_funnel,
        "search_lanes_expected": int(kids_funnel.get("lane_calls_expected", 0)),
        "search_lanes_completed": int(kids_funnel.get("lane_calls_completed", 0)),
        "added": inserted_kids,
        "history_updated": history_updated,
        "history_day": day,
        "partial": (
            kids_queries_ok < kids_queries_total
            or (tracked_total > 0 and tracked_ok < tracked_total)
        ),
        "missing_ids": sorted(tracked_failed_ids),
    }
    avatar_count = write_avatar_overlay(payload, avatars)
    write_snapshot(snapshot, payload)
    summary = {
        "scan_scope": "kids",
        "tracked": len(kids_rows),
        "updated": len(expected_history_views),
        "kids_added": inserted_kids,
        "kids_queries": kids_queries_total,
        "kids_queries_ok": kids_queries_ok,
        "kids_results_examined": kids_results_examined,
        "kids_candidates_kept": kids_candidates_kept,
        "kids_verified_unique": len(verified_candidates),
        "kids_funnel": kids_funnel,
        "kids_search_lanes_expected": int(kids_funnel.get("lane_calls_expected", 0)),
        "kids_search_lanes_completed": int(kids_funnel.get("lane_calls_completed", 0)),
        "history_ids": history_ids,
        "history_files": history_files,
        "history_updated": history_updated,
        "history_day": day,
        "avatars": avatar_count,
        "timestamp": datetime.fromtimestamp(now_ms / 1000, timezone.utc).isoformat(),
    }
    print(json.dumps(summary, ensure_ascii=False))
    return summary



def merge_artifacts(
    snapshot: Path,
    avatars: Path,
    merge_dir: Path,
    expected_shards: int,
    history_dir: Path | None = None,
    recommendation_pool: Path | None = None,
    generate_recommendations: bool = True,
    require_kids: bool = False,
    scan_scope: str = "all",
) -> dict:
    if scan_scope not in SCAN_SCOPES:
        raise ValueError(f"Unknown scan scope: {scan_scope}")
    if scan_scope == "kids":
        if generate_recommendations:
            raise RuntimeError("Kids-only merge cannot generate recommendation data")
        return merge_kids_artifacts(
            snapshot,
            avatars,
            merge_dir,
            expected_shards,
            history_dir,
            require_kids,
        )
    if scan_scope == "standard" and require_kids:
        raise RuntimeError("--require-kids is incompatible with --scan-scope standard")
    artifacts = load_scoped_artifacts(merge_dir, expected_shards, scan_scope)

    tracked_total = sum(int(a.get("tracked_total", 0)) for a in artifacts)
    tracked_ok = sum(int(a.get("tracked_ok", 0)) for a in artifacts)
    queries_total = sum(int(a.get("queries_total", 0)) for a in artifacts)
    queries_ok = sum(int(a.get("queries_ok", 0)) for a in artifacts)
    queries_raw = sum(int(a.get("queries_raw", 0)) for a in artifacts)
    queries_enriched = sum(int(a.get("queries_enriched", 0)) for a in artifacts)
    kids_queries_total = sum(int(a.get("kids_queries_total", 0)) for a in artifacts)
    kids_queries_ok = sum(int(a.get("kids_queries_ok", 0)) for a in artifacts)
    kids_results_examined = sum(int(a.get("kids_results_examined", 0)) for a in artifacts)
    kids_candidates_kept = sum(int(a.get("kids_candidates_kept", 0)) for a in artifacts)
    payload = read_snapshot(snapshot)
    if require_kids:
        expected_kids_queries = len(KIDS_QUERY_SPECS)
        if (
            kids_queries_total != expected_kids_queries
            or kids_queries_ok != expected_kids_queries
        ):
            raise RuntimeError(
                f"Merge rejected: expected {expected_kids_queries}/{expected_kids_queries} "
                f"Kids queries, got {kids_queries_ok}/{kids_queries_total}"
            )
    if not tracked_total or tracked_ok / tracked_total < MIN_PUBLISH_TRACK_RATIO:
        raise RuntimeError(f"Merge rejected: {tracked_ok}/{tracked_total} tracked videos refreshed")
    owned_ok = all(bool(artifact.get("owned_ok", True)) for artifact in artifacts)
    tracked_fresh_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_fresh_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    tracked_failed_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_failed_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    if len(tracked_fresh_ids) != tracked_ok:
        raise RuntimeError(
            f"Merge rejected: {tracked_ok} refreshed counters but "
            f"{len(tracked_fresh_ids)} traceable refreshed IDs"
        )
    expected_failed = tracked_total - tracked_ok
    if len(tracked_failed_ids) != expected_failed or tracked_failed_ids & tracked_fresh_ids:
        raise RuntimeError(
            f"Merge rejected: {expected_failed} missing counters but "
            f"{len(tracked_failed_ids)} traceable missing IDs"
        )
    canonical_ours_declared = any(
        artifact.get("canonical_ours_manifest") is True for artifact in artifacts
    )
    if canonical_ours_declared and not all(
        artifact.get("canonical_ours_manifest") is True for artifact in artifacts
    ):
        raise RuntimeError("Merge rejected: canonical Our Videos manifest coverage is incomplete")
    canonical_ours_totals = {
        int(artifact.get("canonical_ours_total") or 0) for artifact in artifacts
    }
    canonical_ours_digests = {
        str(artifact.get("canonical_ours_digest") or "") for artifact in artifacts
    }
    if canonical_ours_declared and (
        len(canonical_ours_totals) != 1
        or len(canonical_ours_digests) != 1
        or not next(iter(canonical_ours_digests), "")
    ):
        raise RuntimeError("Merge rejected: canonical Our Videos proof differs across shards")
    canonical_ours_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("canonical_ours_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    if canonical_ours_declared and scan_scope != "kids" and not canonical_ours_ids:
        raise RuntimeError("Merge rejected: canonical Our Videos cohort is empty")
    if canonical_ours_declared and scan_scope != "kids":
        canonical_ours_total = next(iter(canonical_ours_totals))
        canonical_ours_digest = next(iter(canonical_ours_digests))
        actual_digest = hashlib.sha256(
            "\n".join(sorted(canonical_ours_ids)).encode("utf-8")
        ).hexdigest()
        if (
            len(canonical_ours_ids) != canonical_ours_total
            or actual_digest != canonical_ours_digest
        ):
            raise RuntimeError(
                "Merge rejected: canonical Our Videos cohort is truncated or inconsistent"
            )
    canonical_ours_metadata: dict[str, dict] = {}
    if canonical_ours_declared and scan_scope != "kids":
        metadata_totals = {
            int(artifact.get("canonical_ours_metadata_total") or 0)
            for artifact in artifacts
        }
        metadata_digests = {
            str(artifact.get("canonical_ours_metadata_digest") or "")
            for artifact in artifacts
        }
        if (
            len(metadata_totals) != 1
            or len(metadata_digests) != 1
            or not next(iter(metadata_digests), "")
        ):
            raise RuntimeError(
                "Merge rejected: canonical Our Videos metadata proof differs across shards"
            )
        metadata_rows: list[dict] = []
        for artifact in artifacts:
            shard_ids = {
                str(video_id)
                for video_id in (artifact.get("canonical_ours_ids") or [])
                if VIDEO_ID.fullmatch(str(video_id or ""))
            }
            shard_metadata_ids: set[str] = set()
            for source in artifact.get("canonical_ours_metadata") or []:
                if not isinstance(source, dict):
                    raise RuntimeError(
                        "Merge rejected: canonical Our Videos metadata is invalid"
                    )
                video_id = str(source.get("vid") or "")
                if (
                    not VIDEO_ID.fullmatch(video_id)
                    or video_id in shard_metadata_ids
                    or video_id in canonical_ours_metadata
                ):
                    raise RuntimeError(
                        "Merge rejected: canonical Our Videos metadata contains duplicates"
                    )
                normalized = canonical_ours_metadata_rows(
                    [video_id], {video_id: source}
                )[0]
                shard_metadata_ids.add(video_id)
                canonical_ours_metadata[video_id] = normalized
                metadata_rows.append(normalized)
            if shard_metadata_ids != shard_ids:
                raise RuntimeError(
                    "Merge rejected: canonical Our Videos metadata shard is truncated"
                )
        metadata_rows.sort(key=lambda row: row["vid"])
        metadata_total = next(iter(metadata_totals))
        metadata_digest = next(iter(metadata_digests))
        if (
            set(canonical_ours_metadata) != canonical_ours_ids
            or len(metadata_rows) != metadata_total
            or metadata_total != canonical_ours_total
            or canonical_ours_metadata_digest(metadata_rows) != metadata_digest
        ):
            raise RuntimeError(
                "Merge rejected: canonical Our Videos metadata is truncated or inconsistent"
            )

    previous_unavailable_ids = {
        str(video_id)
        for video_id in ((payload.get("videoMetrics") or {}).get("unavailable_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    previous_missing_ids = {
        str(video_id)
        for video_id in ((payload.get("videoMetrics") or {}).get("missing_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    newly_unavailable_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_unavailable_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    # The official Data API can authoritatively quarantine an omission in one
    # pass.  The public yt-dlp fallback is deliberately more conservative: an
    # ID must be absent from two consecutive complete scans before it leaves
    # the active denominator.  Quarantined IDs remain recovery probes.
    newly_unavailable_ids.update(tracked_failed_ids & previous_missing_ids)
    recovered_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_recovered_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    confirmed_recovered_ids = recovered_ids & previous_unavailable_ids
    unavailable_ids = sorted(
        (previous_unavailable_ids | newly_unavailable_ids) - confirmed_recovered_ids
    )
    unavailable_set = set(unavailable_ids)
    newly_quarantined_active_ids = tracked_failed_ids & unavailable_set
    active_tracked_total = (
        tracked_total
        - len(newly_quarantined_active_ids)
        + len(confirmed_recovered_ids)
    )
    active_updated_total = len(tracked_fresh_ids | confirmed_recovered_ids)
    if active_tracked_total < active_updated_total:
        raise RuntimeError(
            f"Merge rejected: {active_updated_total} refreshed videos exceed "
            f"the {active_tracked_total} active tracked videos"
        )
    missing_ids = sorted(tracked_failed_ids - unavailable_set)
    data = payload.setdefault("d", {})
    preserved_kids = list(data.get("kids") or []) if scan_scope == "standard" else None
    bootstrap_kids = not bool(data.get("kids"))
    if preserved_kids is not None:
        data["kids"] = []
    prune_deferred_rows(data)
    legacy_history = data.pop("hist", {})
    now_ms = max(int(a.get("generated_ms", 0)) for a in artifacts) or utc_now_ms()
    fresh: dict[str, dict] = {}
    owned_fresh: dict[str, dict] = {}
    candidates: list[dict] = []
    for artifact in artifacts:
        candidates.extend(artifact.get("candidates") or [])
        for row in artifact.get("fresh") or []:
            video_id = row.get("vid")
            if not VIDEO_ID.match(str(video_id or "")):
                continue
            previous = fresh.get(video_id)
            if not previous or int(row.get("views") or 0) >= int(previous.get("views") or 0):
                fresh[video_id] = preserve_audience_classification(row, previous)
            else:
                fresh[video_id] = preserve_audience_classification(previous, row)
        for row in artifact.get("owned_fresh") or []:
            video_id = row.get("vid")
            if VIDEO_ID.match(str(video_id or "")):
                owned_fresh[video_id] = row

    for bucket in ("all", "trends", "news", "ours", "kids"):
        for row in data.setdefault(bucket, []):
            current = fresh.get(row.get("vid"))
            if current:
                update_row(row, current, now_ms)
    if preserved_kids is not None:
        for row in preserved_kids:
            current = fresh.get(str(row.get("vid") or ""))
            if current:
                update_row(row, current, now_ms)
    data["kids"] = [
        row
        for row in data["kids"]
        if row.get("madeForKids") is True
        and int(row.get("views") or 0) >= MIN_KIDS_VIEWS
        and float(row.get("vpm") or 0) >= MIN_KIDS_VPM
    ]
    removed_low_view_news = prune_news_below_view_floor(data)

    live_audiences: dict[str, dict] = {}
    for artifact in artifacts:
        live_audiences.update(artifact.get("live_audiences") or {})
    for row in data.setdefault("lives", []):
        audience = live_audiences.get(str(row.get("vid") or ""))
        if audience:
            row.update(audience)

    by_ours = {row.get("vid"): row for row in data.setdefault("ours", [])}
    inserted_ours = 0
    canonical_ours_rows: list[dict] = []
    for video_id in sorted(canonical_ours_ids):
        current = by_ours.get(video_id)
        if current is None:
            current = {"vid": video_id}
            data["ours"].append(current)
            by_ours[video_id] = current
            inserted_ours += 1
        # Match Object.assign(Sheet row, snapshot row) in the client: Sheet
        # publication/duration metadata exists before the fresh counter row is
        # applied, and remains when a no-API collector omitted those fields.
        merge_sheet_ours_metadata(current, canonical_ours_metadata[video_id])
        authoritative = fresh.get(video_id)
        if authoritative:
            update_row(current, authoritative, now_ms)
        canonical_ours_rows.append(current)
    sheet_ours_expected = len(canonical_ours_ids)
    sheet_ours_updated = 0
    if canonical_ours_declared:
        _, sheet_ours_updated = validate_card_refresh(
            {"ours": canonical_ours_rows},
            fresh,
            set(canonical_ours_ids),
            now_ms,
        )
    for row in owned_fresh.values():
        current = by_ours.get(row["vid"])
        if current:
            merge_owned_metadata(current, row)
        else:
            added = {"vid": row["vid"]}
            merge_owned_metadata(added, row)
            authoritative = fresh.get(row["vid"])
            if authoritative:
                update_row(added, authoritative, now_ms)
            data["ours"].append(added)
            by_ours[added["vid"]] = added
            inserted_ours += 1
    data["ours"].sort(key=lambda row: row.get("pub") or 0, reverse=True)

    by_all = {row.get("vid"): row for row in data["all"]}
    by_trends = {row.get("vid"): row for row in data["trends"]}
    by_news = {row.get("vid"): row for row in data["news"]}
    by_kids = {row.get("vid"): row for row in data["kids"]}
    inserted_all = 0
    inserted_trends = 0
    inserted_news = 0
    inserted_kids = 0
    for row in merge_keyword_rows(candidates):
        if is_deferred_row(row):
            continue
        views = int(row.get("views") or 0)
        age = row.get("ageM")
        for current in (
            by_all.get(row["vid"]),
            by_trends.get(row["vid"]),
            by_news.get(row["vid"]),
            by_kids.get(row["vid"]),
        ):
            if current:
                merge_discovery_fields(current, row)
        audiences = {
            str(value).lower()
            for value in (row.get("audiences") or ["youtube"])
            if str(value).lower() in {"youtube", "kids"}
        }
        is_youtube = "youtube" in audiences
        is_kids = "kids" in audiences and row.get("madeForKids") is True
        if (
            is_kids
            and views >= MIN_KIDS_VIEWS
            and float(row.get("vpm") or 0) >= MIN_KIDS_VPM
            and row["vid"] not in by_kids
        ):
            added = dict(row)
            data["kids"].append(added)
            by_kids[row["vid"]] = added
            inserted_kids += 1
        if is_youtube and views >= MIN_ALL_VIEWS and row["vid"] not in by_all:
            data["all"].append(row)
            by_all[row["vid"]] = row
            inserted_all += 1
        if (
            is_youtube
            and views >= MIN_TREND_VIEWS
            and isinstance(age, (int, float))
            and age <= MAX_TREND_AGE_MONTHS
            and row["vid"] not in by_trends
        ):
            data["trends"].append(dict(row))
            by_trends[row["vid"]] = row
            inserted_trends += 1
        if (
            is_youtube
            and views >= MIN_NEWS_VIEWS
            and isinstance(age, (int, float))
            and age <= MAX_NEWS_AGE_MONTHS
            and (row.get("vpm") or 0) >= MIN_NEWS_VPM
            and row["vid"] not in by_news
        ):
            discovered = dict(row)
            discovered["disc"] = row.get("kw") or ""
            discovered["days"] = 1
            discovered["why"] = (
                "Discovered by the daily recent-video scan; "
                f"{views:,} views at {age:.1f} months "
                f"({int(row.get('vpm') or 0):,} views/month)."
            )
            data["news"].append(discovered)
            by_news[row["vid"]] = discovered
            inserted_news += 1

    if require_kids and bootstrap_kids and not data["kids"]:
        raise RuntimeError(
            "Merge rejected: initial Kids scan returned no verified candidates"
        )

    for row in data["news"]:
        if isinstance(row.get("added"), (int, float)):
            row["days"] = max(1, int((now_ms - row["added"]) / 86400000) + 1)

    data["trends"] = [
        row
        for row in data["trends"]
        if int(row.get("views") or 0) >= MIN_TREND_VIEWS
        and isinstance(row.get("ageM"), (int, float))
        and row["ageM"] <= MAX_TREND_AGE_MONTHS
    ]
    for bucket in ("all", "trends", "news", "kids"):
        data[bucket].sort(key=lambda row: row.get("vpm") or 0, reverse=True)
    if len(data["news"]) > MAX_NEWS_ROWS:
        data["news"] = sorted(
            data["news"], key=lambda row: row.get("added") or 0, reverse=True
        )[:MAX_NEWS_ROWS]
    prune_deferred_rows(data)
    if preserved_kids is not None:
        # Preserve the dedicated cohort/membership and Kids scan metadata while
        # publishing its newly refreshed standard counters.
        data["kids"] = preserved_kids

    # Prove exactly the cohort the browser can display. This must happen only
    # after every public prune, and unavailable rows are excluded just like
    # removeUnavailableVideoRows() in the client.
    analysis_rows = [
        row
        for row in data["ours"]
        if str(row.get("vid") or "") not in unavailable_set
        and is_analysis_card(row)
    ]
    analysis_ids = {str(row.get("vid") or "") for row in analysis_rows}
    if len(analysis_ids) != len(analysis_rows):
        raise RuntimeError("Merge rejected: duplicate visible Analyse video IDs")
    analysis_rows_expected = len(analysis_rows)
    analysis_rows_updated = 0
    if canonical_ours_declared:
        _, analysis_rows_updated = validate_card_refresh(
            {"ours": analysis_rows},
            fresh,
            analysis_ids,
            now_ms,
        )

    desired_ids = {
        str(video_id)
        for artifact in artifacts
        for video_id in (artifact.get("tracked_ids") or [])
        if VIDEO_ID.match(str(video_id or ""))
    }
    desired_ids.update(
        str(row.get("vid"))
        for bucket in ("all", "trends", "news", "ours", "kids")
        for row in data[bucket]
        if VIDEO_ID.match(str(row.get("vid") or ""))
    )
    desired_ids.update(confirmed_recovered_ids)
    resolved_history_dir = history_dir or snapshot.parent / "video_history"
    history_ids, history_files = update_history_shards(
        resolved_history_dir,
        desired_ids,
        fresh,
        legacy_history,
        now_ms,
    )
    refreshed_history_ids = tracked_fresh_ids | confirmed_recovered_ids
    expected_history_views = {
        video_id: int(fresh[video_id]["views"])
        for video_id in refreshed_history_ids
        if video_id in fresh and isinstance(fresh[video_id].get("views"), (int, float))
    }
    if len(expected_history_views) != active_updated_total:
        raise RuntimeError(
            f"Merge rejected: {active_updated_total} refreshed videos but "
            f"{len(expected_history_views)} usable history values"
        )
    history_updated = validate_history_refresh(resolved_history_dir, expected_history_views, now_ms)
    card_rows_expected, card_rows_updated = validate_card_refresh(
        data,
        fresh,
        refreshed_history_ids,
        now_ms,
    )
    history_day = history_day_key(now_ms)

    if preserved_kids is not None and data.get("kids"):
        kids_ids = {
            str(row.get("vid"))
            for row in data["kids"]
            if VIDEO_ID.match(str(row.get("vid") or ""))
        }
        active_kids_ids = kids_ids - unavailable_set
        refreshed_kids_ids = active_kids_ids & refreshed_history_ids
        kids_metrics = dict(payload.get("kidsMetrics") or {})
        # The standard run refreshes only the existing Kids cohort's factual
        # counters. Keep discovery/query statistics from the most recent real
        # Kids scan so this daily maintenance cannot masquerade as discovery.
        kids_metrics.update({
            "day": history_day,
            "day_timezone": RADAR_TIMEZONE_NAME,
            "tracked": len(active_kids_ids),
            "updated": len(refreshed_kids_ids),
            "history_updated": len(refreshed_kids_ids),
            "history_day": history_day,
            "partial": len(refreshed_kids_ids) < len(active_kids_ids),
            "missing_ids": sorted(active_kids_ids - refreshed_kids_ids),
        })
        payload["kidsMetricsT"] = now_ms
        payload["kidsMetrics"] = kids_metrics

    payload["t"] = now_ms
    payload["videoMetricsT"] = now_ms
    payload["videoMetrics"] = {
        "tracked": active_tracked_total,
        "updated": active_updated_total,
        "keywords": queries_total,
        "keywords_ok": queries_ok,
        "discovery_partial": queries_ok < queries_total or not owned_ok,
        "owned_discovery_ok": owned_ok,
        "search_results": queries_raw,
        "search_results_enriched": queries_enriched,
        "kids_queries": kids_queries_total,
        "kids_queries_ok": kids_queries_ok,
        "kids_results_examined": kids_results_examined,
        "kids_candidates_kept": kids_candidates_kept,
        "history_updated": history_updated,
        "card_rows_expected": card_rows_expected,
        "card_rows_updated": card_rows_updated,
        "sheet_ours_expected": sheet_ours_expected,
        "sheet_ours_updated": sheet_ours_updated,
        "analysis_rows_expected": analysis_rows_expected,
        "analysis_rows_updated": analysis_rows_updated,
        "history_day": history_day,
        "day_timezone": RADAR_TIMEZONE_NAME,
        "partial": active_updated_total < active_tracked_total,
        "unavailable_ids": unavailable_ids,
        "missing_ids": missing_ids,
    }
    payload["videoHistory"] = {
        "layout": "video_history/{first_char_hex}.json",
        "retention_days": HISTORY_RETENTION_DAYS,
        "updated": now_ms,
        "day": history_day,
        "day_timezone": RADAR_TIMEZONE_NAME,
    }
    avatar_count = write_avatar_overlay(payload, avatars)
    write_snapshot(snapshot, payload)
    pool_payload = None
    if generate_recommendations:
        pool_data = dict(data)
        pool_data["videoMetricsT"] = now_ms
        pool_payload = write_recommendation_pool(
            pool_data,
            recommendation_pool or snapshot.parent / DEFAULT_RECOMMENDATION_POOL.name,
            generated_ms=now_ms,
        )
    summary = {
        "tracked": active_tracked_total,
        "updated": active_updated_total,
        "keywords": queries_total,
        "keywords_ok": queries_ok,
        "discovery_partial": queries_ok < queries_total or not owned_ok,
        "history_ids": history_ids,
        "history_files": history_files,
        "history_updated": history_updated,
        "card_rows_expected": card_rows_expected,
        "card_rows_updated": card_rows_updated,
        "sheet_ours_expected": sheet_ours_expected,
        "sheet_ours_updated": sheet_ours_updated,
        "analysis_rows_expected": analysis_rows_expected,
        "analysis_rows_updated": analysis_rows_updated,
        "history_day": history_day,
        "unavailable": len(unavailable_ids),
        "missing": len(missing_ids),
        "all_added": inserted_all,
        "trends_added": inserted_trends,
        "news_added": inserted_news,
        "kids_added": inserted_kids,
        "kids_queries": kids_queries_total,
        "kids_queries_ok": kids_queries_ok,
        "kids_results_examined": kids_results_examined,
        "kids_candidates_kept": kids_candidates_kept,
        "news_removed_below_view_floor": removed_low_view_news,
        "ours_added": inserted_ours,
        "avatars": avatar_count,
        "recommendations_generated": len(pool_payload["items"]) if pool_payload else None,
        "timestamp": datetime.fromtimestamp(now_ms / 1000, timezone.utc).isoformat(),
    }
    print(json.dumps(summary, ensure_ascii=False))
    return summary


def snapshot_freshness(snapshot: Path, now_ms: int | None = None) -> dict:
    """Return a machine-readable daily health decision for the watchdog cron."""
    payload = read_snapshot(snapshot)
    metrics = payload.get("videoMetrics") or {}
    stamp = int(payload.get("videoMetricsT") or 0)
    now_ms = int(now_ms or utc_now_ms())
    tracked = int(metrics.get("tracked") or 0)
    updated = int(metrics.get("updated") or 0)
    keywords = int(metrics.get("keywords") or 0)
    keywords_ok = int(metrics.get("keywords_ok") or 0)
    history_updated = int(metrics.get("history_updated") or 0)
    card_rows_expected = int(metrics.get("card_rows_expected") or 0)
    card_rows_updated = int(metrics.get("card_rows_updated") or 0)
    sheet_ours_expected = int(metrics.get("sheet_ours_expected") or 0)
    sheet_ours_updated = int(metrics.get("sheet_ours_updated") or 0)
    analysis_rows_expected = int(metrics.get("analysis_rows_expected") or 0)
    analysis_rows_updated = int(metrics.get("analysis_rows_updated") or 0)
    same_day = bool(stamp) and history_day_key(stamp) == history_day_key(now_ms)
    fresh = (
        same_day
        and tracked > 0
        and updated == tracked
        and history_updated == updated
        and card_rows_expected > 0
        and card_rows_updated == card_rows_expected
        and sheet_ours_expected > 0
        and sheet_ours_updated == sheet_ours_expected
        and analysis_rows_expected > 0
        and analysis_rows_updated == analysis_rows_expected
        and not bool(metrics.get("partial"))
        and metrics.get("history_day") == history_day_key(stamp)
        and metrics.get("day_timezone") == RADAR_TIMEZONE_NAME
    )
    return {
        "fresh": fresh,
        "today": history_day_key(now_ms),
        "snapshot_day": history_day_key(stamp) if stamp else None,
        "tracked": tracked,
        "updated": updated,
        "keywords": keywords,
        "keywords_ok": keywords_ok,
        "history_updated": history_updated,
        "card_rows_expected": card_rows_expected,
        "card_rows_updated": card_rows_updated,
        "sheet_ours_expected": sheet_ours_expected,
        "sheet_ours_updated": sheet_ours_updated,
        "analysis_rows_expected": analysis_rows_expected,
        "analysis_rows_updated": analysis_rows_updated,
    }


def verify_publication(
    base_url: str,
    snapshot: Path,
    history_dir: Path,
    timeout_seconds: int = 900,
    interval_seconds: int = 15,
    recommendation_pool: Path | None = None,
    require_recommendation_pool: bool = True,
) -> dict:
    """Wait until Pages serves the factual snapshot/history and optional ideas."""
    local = read_snapshot(snapshot)
    expected = int(local.get("videoMetricsT") or 0)
    if not expected:
        raise RuntimeError("Local snapshot has no validated videoMetricsT")
    shards = sorted(history_dir.glob("*.json"))
    if not shards:
        raise RuntimeError("Local history directory has no shards")
    local_history: dict[str, dict] = {}
    for shard in shards:
        document = json.loads(shard.read_text(encoding="utf-8"))
        if not isinstance(document, dict) or not isinstance(document.get("d"), dict):
            raise RuntimeError(f"Local history shard {shard.name} is malformed")
        local_history[shard.name] = document
    local_metrics = local.get("videoMetrics") or {}
    expected_history_updated = int(local_metrics.get("history_updated") or 0)
    expected_card_rows = int(local_metrics.get("card_rows_expected") or 0)
    updated_card_rows = int(local_metrics.get("card_rows_updated") or 0)
    expected_sheet_ours = int(local_metrics.get("sheet_ours_expected") or 0)
    updated_sheet_ours = int(local_metrics.get("sheet_ours_updated") or 0)
    expected_analysis_rows = int(local_metrics.get("analysis_rows_expected") or 0)
    updated_analysis_rows = int(local_metrics.get("analysis_rows_updated") or 0)
    if expected_card_rows <= 0 or updated_card_rows != expected_card_rows:
        raise RuntimeError(
            f"Local factual cards are incomplete at {updated_card_rows}/{expected_card_rows}"
        )
    if expected_sheet_ours <= 0 or updated_sheet_ours != expected_sheet_ours:
        raise RuntimeError(
            f"Local Our Videos cards are incomplete at {updated_sheet_ours}/{expected_sheet_ours}"
        )
    if expected_analysis_rows <= 0 or updated_analysis_rows != expected_analysis_rows:
        raise RuntimeError(
            f"Local Analyse cards are incomplete at {updated_analysis_rows}/{expected_analysis_rows}"
        )
    local_latest_count = sum(
        1
        for document in local_history.values()
        for points in document["d"].values()
        if isinstance(points, list) and points and isinstance(points[-1], list) and len(points[-1]) >= 2
    )
    if expected_history_updated <= 0 or local_latest_count < expected_history_updated:
        raise RuntimeError(
            f"Local factual history is incomplete at {local_latest_count}/{expected_history_updated}"
        )
    local_pool_path = recommendation_pool or snapshot.parent / DEFAULT_RECOMMENDATION_POOL.name
    if require_recommendation_pool and not local_pool_path.exists():
        raise RuntimeError("Local renewable recommendation pool is missing")
    root = base_url.rstrip("/") + "/"
    deadline = time.monotonic() + max(timeout_seconds, 1)
    last_error = "not attempted"
    while time.monotonic() < deadline:
        cache_buster = urllib.parse.urlencode({"expected": expected, "attempt": int(time.time())})
        try:
            with urllib.request.urlopen(root + "Lofi_Radar_data.js?" + cache_buster, timeout=30) as response:
                remote = parse_snapshot_text(response.read().decode("utf-8"))
            remote_stamp = int(remote.get("videoMetricsT") or 0)
            if remote_stamp < expected:
                last_error = f"served snapshot={remote_stamp}, expected={expected}"
                time.sleep(max(interval_seconds, 1))
                continue
            remote_metrics = remote.get("videoMetrics") or {}
            remote_card_expected = int(remote_metrics.get("card_rows_expected") or 0)
            remote_card_updated = int(remote_metrics.get("card_rows_updated") or 0)
            remote_sheet_ours_expected = int(remote_metrics.get("sheet_ours_expected") or 0)
            remote_sheet_ours_updated = int(remote_metrics.get("sheet_ours_updated") or 0)
            remote_analysis_expected = int(remote_metrics.get("analysis_rows_expected") or 0)
            remote_analysis_updated = int(remote_metrics.get("analysis_rows_updated") or 0)
            if (
                remote_card_expected != expected_card_rows
                or remote_card_updated != expected_card_rows
            ):
                last_error = (
                    f"served cards={remote_card_updated}/{remote_card_expected}, "
                    f"expected={expected_card_rows}/{expected_card_rows}"
                )
                time.sleep(max(interval_seconds, 1))
                continue
            if (
                remote_analysis_expected != expected_analysis_rows
                or remote_analysis_updated != expected_analysis_rows
            ):
                last_error = (
                    f"served Analyse={remote_analysis_updated}/{remote_analysis_expected}, "
                    f"expected={expected_analysis_rows}/{expected_analysis_rows}"
                )
                time.sleep(max(interval_seconds, 1))
                continue
            if (
                remote_sheet_ours_expected != expected_sheet_ours
                or remote_sheet_ours_updated != expected_sheet_ours
            ):
                last_error = (
                    f"served Our Videos={remote_sheet_ours_updated}/{remote_sheet_ours_expected}, "
                    f"expected={expected_sheet_ours}/{expected_sheet_ours}"
                )
                time.sleep(max(interval_seconds, 1))
                continue
            pool_count = None
            if require_recommendation_pool:
                with urllib.request.urlopen(
                    root + DEFAULT_RECOMMENDATION_POOL.name + "?" + cache_buster,
                    timeout=30,
                ) as response:
                    pool_raw = response.read().decode("utf-8").strip()
                if not pool_raw.startswith(POOL_PREFIX):
                    raise RuntimeError("served recommendation pool is malformed")
                remote_pool = json.loads(pool_raw[len(POOL_PREFIX):].rstrip(";\n "))
                pool_stamp = int(remote_pool.get("sourceT") or 0)
                pool_count = len(remote_pool.get("items") or [])
                if pool_stamp < expected or pool_count <= 1_000:
                    last_error = f"served recommendation pool={pool_count}@{pool_stamp}, expected >1000@{expected}"
                    time.sleep(max(interval_seconds, 1))
                    continue
            stale_shards: list[str] = []
            history_stamps: list[int] = []
            verified_points = 0
            for shard in shards:
                with urllib.request.urlopen(
                    root + "video_history/" + shard.name + "?" + cache_buster,
                    timeout=30,
                ) as response:
                    remote_history = json.load(response) or {}
                history_stamp = int(remote_history.get("updated") or 0)
                history_stamps.append(history_stamp)
                if history_stamp < expected:
                    stale_shards.append(shard.name)
                    continue
                remote_rows = remote_history.get("d") if isinstance(remote_history, dict) else None
                if not isinstance(remote_rows, dict):
                    stale_shards.append(shard.name)
                    continue
                for video_id, local_points in local_history[shard.name]["d"].items():
                    if not isinstance(local_points, list) or not local_points:
                        continue
                    local_point = local_points[-1]
                    remote_points = remote_rows.get(video_id)
                    if not isinstance(local_point, list) or len(local_point) < 2 or not isinstance(remote_points, list) or not remote_points:
                        stale_shards.append(shard.name)
                        break
                    remote_point = remote_points[-1]
                    if (
                        not isinstance(remote_point, list)
                        or len(remote_point) < 2
                        or int(remote_point[0]) < int(local_point[0])
                        or (
                            int(remote_point[0]) == int(local_point[0])
                            and int(remote_point[1]) != int(local_point[1])
                        )
                    ):
                        stale_shards.append(shard.name)
                        break
                    verified_points += 1
            if not stale_shards:
                result = {
                    "published": True,
                    "expected": expected,
                    "snapshot": remote_stamp,
                    "history_min": min(history_stamps),
                    "history_shards": len(history_stamps),
                    "history_points": verified_points,
                    "card_rows": remote_card_updated,
                    "recommendations": pool_count,
                }
                print(json.dumps(result))
                return result
            last_error = f"{len(stale_shards)}/{len(shards)} stale history shards: {', '.join(stale_shards[:8])}"
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
        time.sleep(max(interval_seconds, 1))
    raise RuntimeError(f"GitHub Pages did not publish the validated YouTube refresh: {last_error}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--snapshot", type=Path, default=DEFAULT_SNAPSHOT)
    parser.add_argument("--avatars", type=Path, default=DEFAULT_AVATARS)
    parser.add_argument("--history-dir", type=Path, default=DEFAULT_HISTORY_DIR)
    parser.add_argument("--recommendation-pool", type=Path, default=DEFAULT_RECOMMENDATION_POOL)
    parser.add_argument("--skip-recommendation-pool", action="store_true")
    parser.add_argument("--require-kids", action="store_true")
    parser.add_argument("--scan-scope", choices=SCAN_SCOPES, default="all")
    parser.add_argument("--shard", type=int)
    parser.add_argument("--shards", type=int, default=10)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--merge-dir", type=Path)
    parser.add_argument("--tracked-manifest", type=Path)
    parser.add_argument("--write-tracked-manifest", type=Path)
    parser.add_argument("--check-fresh-today", action="store_true")
    parser.add_argument("--verify-base-url")
    parser.add_argument("--verify-core-only", action="store_true")
    parser.add_argument("--verify-timeout", type=int, default=900)
    parser.add_argument("--verify-interval", type=int, default=15)
    args = parser.parse_args()
    if args.write_tracked_manifest:
        write_tracked_manifest(args.snapshot, args.write_tracked_manifest, args.scan_scope)
        return
    if args.check_fresh_today:
        health = snapshot_freshness(args.snapshot)
        print(json.dumps(health))
        raise SystemExit(0 if health["fresh"] else 1)
    if args.verify_base_url:
        verify_publication(
            args.verify_base_url,
            args.snapshot,
            args.history_dir,
            args.verify_timeout,
            args.verify_interval,
            args.recommendation_pool,
            not args.verify_core_only,
        )
        return
    if args.merge_dir:
        merge_artifacts(
            args.snapshot,
            args.avatars,
            args.merge_dir,
            args.shards,
            args.history_dir,
            args.recommendation_pool,
            not args.skip_recommendation_pool,
            args.require_kids,
            args.scan_scope,
        )
        return
    if args.shard is None or args.output is None:
        parser.error("collector mode requires --shard and --output")
    if args.shard < 0 or args.shard >= args.shards:
        parser.error("--shard must be in [0, --shards)")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    run_shard(
        args.snapshot,
        args.output,
        args.shard,
        args.shards,
        args.tracked_manifest,
        args.scan_scope,
    )


if __name__ == "__main__":
    main()

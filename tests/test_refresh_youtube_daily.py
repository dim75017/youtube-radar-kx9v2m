import io
import json
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

import refresh_youtube_daily as radar


class InnertubeResponse(io.BytesIO):
    def __init__(
        self,
        payload,
        final_url=(
            "https://www.youtube.com/youtubei/v1/player"
            "?prettyPrint=false"
        ),
    ):
        if isinstance(payload, bytes):
            body = payload
        elif isinstance(payload, str):
            body = payload.encode("utf-8")
        else:
            body = json.dumps(payload, separators=(",", ":")).encode("utf-8")
        super().__init__(body)
        self.final_url = final_url

    def geturl(self):
        return self.final_url

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()


class DailyHistoryTests(unittest.TestCase):
    def test_official_api_hydration_preserves_made_for_kids_tristate(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        captured = {}
        items = []
        comment_counts = {
            "abcdefghijk": "42",
            "zyxwvutsrqp": "0",
            "mnopqrstuvw": "-1",
        }
        for video_id, status in (
            ("abcdefghijk", {"madeForKids": True}),
            ("zyxwvutsrqp", {"madeForKids": False}),
            ("mnopqrstuvw", {}),
        ):
            items.append({
                "id": video_id,
                "snippet": {
                    "title": "Long instrumental mix",
                    "publishedAt": "2026-01-01T00:00:00Z",
                    "channelTitle": "Channel",
                    "channelId": "UC1234567890123456789012",
                },
                "contentDetails": {"duration": "PT1H"},
                "statistics": {
                    "viewCount": "1000000",
                    "commentCount": comment_counts[video_id],
                },
                "status": status,
            })

        class Response(io.BytesIO):
            def __enter__(self):
                return self

            def __exit__(self, *args):
                self.close()

        def fake_open(url, timeout=30):
            captured["url"] = url
            return Response(json.dumps({"items": items}).encode())

        with patch.object(radar.urllib.request, "urlopen", side_effect=fake_open):
            rows = radar.fetch_api_rows(
                ["abcdefghijk", "zyxwvutsrqp", "mnopqrstuvw"],
                now,
                "secret",
            )
        query = radar.urllib.parse.parse_qs(radar.urllib.parse.urlparse(captured["url"]).query)
        self.assertIn("status", query["part"][0].split(","))
        self.assertIs(rows["abcdefghijk"]["madeForKids"], True)
        self.assertIs(rows["zyxwvutsrqp"]["madeForKids"], False)
        self.assertNotIn("madeForKids", rows["mnopqrstuvw"])
        self.assertEqual(rows["abcdefghijk"]["comments"], 42)
        self.assertEqual(rows["zyxwvutsrqp"]["comments"], 0)
        self.assertNotIn("comments", rows["mnopqrstuvw"])
        self.assertEqual(rows["abcdefghijk"]["metadataSource"], radar.METADATA_SOURCE_API)
        self.assertEqual(rows["abcdefghijk"]["pubSource"], radar.METADATA_SOURCE_API)

    def test_kids_search_requires_official_true_and_instrumental_long_form(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        ids = ["abcdefghijk", "zyxwvutsrqp", "mnopqrstuvw", "qrstuvwxyz0"]
        search_payload = {
            "items": [{"id": {"videoId": video_id}} for video_id in ids],
        }
        official = {
            "abcdefghijk": {
                "vid": "abcdefghijk", "title": "Baby sleep music instrumental · 3 hours",
                "views": 2_000_000, "durH": 3, "channel": "Calm Baby",
                "madeForKids": True, "_scanDescription": "instrumental sleep music",
            },
            "zyxwvutsrqp": {
                "vid": "zyxwvutsrqp", "title": "Baby sleep music instrumental",
                "views": 3_000_000, "durH": 2, "channel": "General Music",
                "madeForKids": False, "_scanDescription": "instrumental",
            },
            "mnopqrstuvw": {
                "vid": "mnopqrstuvw", "title": "Kids sing along instrumental songs",
                "views": 4_000_000, "durH": 2, "channel": "Singing Kids",
                "madeForKids": True, "_scanDescription": "sing along with vocals",
            },
            "qrstuvwxyz0": {
                "vid": "qrstuvwxyz0", "title": "Baby sleep music instrumental live",
                "views": 5_000_000, "durH": 8, "channel": "Calm Baby",
                "madeForKids": True, "_scanDescription": "instrumental",
                "_liveBroadcastContent": "upcoming",
            },
        }
        for row in official.values():
            row.setdefault("_liveBroadcastContent", "none")
            row.setdefault("ageM", 1)
            row.setdefault("vpm", row["views"])
        spec = {
            "query": "baby sleep music instrumental",
            "genre": "Baby sleep",
            "cluster": "Relaxation / meditation",
            "audience": "kids",
        }
        with patch.object(radar, "youtube_api_payload", return_value=search_payload), patch.object(
            radar, "fetch_api_rows", return_value=official
        ):
            rows, raw, enriched, funnel = radar.fetch_kids_search(spec, now, "secret")
        self.assertEqual((raw, enriched), (4, 4))
        self.assertEqual([row["vid"] for row in rows], ["abcdefghijk"])
        self.assertIs(rows[0]["madeForKids"], True)
        self.assertEqual(rows[0]["audiences"], ["kids"])
        self.assertNotIn("_scanDescription", rows[0])
        self.assertEqual(funnel["kept"], 1)
        with self.assertRaisesRegex(RuntimeError, "requires YOUTUBE_API_KEY"):
            radar.fetch_kids_search(spec, now, "")

    def test_kids_queries_are_builtin_with_top_100_daily_and_safe_lane_budget(self):
        specs = [s for s in radar.query_specs({"d": {}}, include_kids=True, kids_day="2026-08-13") if s["audience"] == "kids"]
        queries = [s["query"].lower() for s in specs]
        self.assertEqual(radar.KIDS_BOOTSTRAP_SEARCH_RESULTS, 100)
        self.assertEqual(radar.KIDS_SEARCH_RESULTS, 100)
        self.assertTrue(all(s["searchResults"] == 100 for s in specs))
        calls = sum(len(s["searchLanes"]) for s in specs)
        self.assertLessEqual(calls, radar.MAX_KIDS_SEARCH_CALLS)
        daily = [
            s
            for s in radar.query_specs({"d": {"kids": [{"vid": "abcdefghijk"}]}}, include_kids=True, kids_day="2026-08-13")
            if s["audience"] == "kids"
        ]
        self.assertTrue(all(s["searchResults"] == 100 for s in daily))
        empty_but_bootstrapped = [
            s
            for s in radar.query_specs({
                "d": {"kids": []},
                "videoMetrics": {"kids_queries": len(radar.KIDS_QUERY_SPECS)},
            }, include_kids=True, kids_day="2026-08-13")
            if s["audience"] == "kids"
        ]
        self.assertEqual(len(empty_but_bootstrapped), 40)
        self.assertTrue(all(s["searchResults"] == 100 for s in empty_but_bootstrapped))
        self.assertEqual(sum("viewCount" in s["searchLanes"] for s in specs), 40)
        self.assertEqual(sum("relevance" in s["searchLanes"] for s in specs), 20)
        self.assertEqual(sum("date" in s["searchLanes"] for s in specs), 20)
        self.assertEqual(len(queries), len(set(queries)))
        for fragment in (
            "baby sleep", "toddler", "kids lofi", "ambient music for babies",
            "piano music for babies", "classical music for babies", "jazz for babies",
            "bossa nova for babies", "chill house for kids", "drum and bass for kids",
        ):
            self.assertTrue(any(fragment in query for query in queries), fragment)
        self.assertFalse(any("phonk" in query for query in queries))
        self.assertTrue(all(
            any(signal in query for signal in ("instrumental", "no vocals", "no lyrics"))
            for query in queries
        ))

    def test_kids_lane_plan_rotates_secondary_lane_next_day(self):
        first = radar.kids_search_lanes("2026-08-13")
        second = radar.kids_search_lanes("2026-08-14")
        self.assertEqual(len(first), len(radar.KIDS_QUERY_SPECS))
        self.assertTrue(all(lanes[0] == "viewCount" for lanes in first))
        self.assertEqual(
            [lanes[1] for lanes in first],
            ["date" if lanes[1] == "relevance" else "relevance" for lanes in second],
        )

    def test_dom_marker_accepts_only_exact_family_options_destinations(self):
        for href in (
            "https://www.youtube.com/myfamily/#mf-compare",
            "https://youtube.com/myfamily/#mf-compare",
            "https://ytkids.app.goo.gl/abc123",
        ):
            self.assertTrue(radar.is_kids_marker_href(href), href)
        for href in (
            "https://www.youtube.com/myfamily/",
            "https://www.youtube.com/myfamily/#other",
            "https://evil.example/?next=https://youtube.com/myfamily/#mf-compare",
            "https://not-ytkids.app.goo.gl/abc123",
            "http://ytkids.app.goo.gl/abc123",
            "javascript:void(0)",
        ):
            self.assertFalse(radar.is_kids_marker_href(href), href)

    @staticmethod
    def _kids_player_payload(
        video_id,
        *,
        mode="PLAYBACK_MODE_PAUSED_ONLY",
        status="OK",
        include_text=True,
        support_url=(
            "//support.google.com/youtube/bin/answer.py"
            "?answer=9632097&nohelpkit=1&hl=en"
        ),
    ):
        renderer = {"playbackMode": mode}
        if include_text or support_url is not None:
            notification = {}
            if include_text:
                notification["responseText"] = {
                    "runs": [{
                        "text": (
                            "Miniplayer is off for videos made for kids. "
                            "Tap play to resume"
                        ),
                    }],
                }
            if support_url is not None:
                notification["actionButton"] = {
                    "buttonRenderer": {
                        "navigationEndpoint": {
                            "urlEndpoint": {"url": support_url},
                        },
                    },
                }
            renderer["minimizedEndpoint"] = {
                "addToToastAction": {
                    "item": {"notificationActionRenderer": notification},
                },
            }
        return {
            "videoDetails": {"videoId": video_id},
            "playabilityStatus": {
                "status": status,
                "miniplayer": {"miniplayerRenderer": renderer},
            },
        }

    @staticmethod
    def _watch_player_payload(
        video_id,
        *,
        mode="PLAYBACK_MODE_PAUSED_ONLY",
        status="OK",
        include_text=True,
        support_url=(
            "//support.google.com/youtube/bin/answer.py"
            "?answer=9632097&nohelpkit=1&hl=en"
        ),
    ):
        renderer = {"playbackMode": mode}
        if include_text or support_url is not None:
            notification = {}
            if include_text:
                notification["responseText"] = {
                    "simpleText": (
                        "Miniplayer is off for videos made for kids. "
                        "Tap play to resume"
                    ),
                }
            if support_url is not None:
                notification["actionButton"] = {
                    "buttonRenderer": {
                        "command": {
                            "urlEndpoint": {"url": support_url},
                        },
                    },
                }
            renderer["minimizedEndpoint"] = {
                "openPopupAction": {
                    "popup": {"notificationActionRenderer": notification},
                    "popupType": "TOAST",
                },
            }
        return {
            "videoDetails": {"videoId": video_id},
            "playabilityStatus": {
                "status": status,
                "miniplayer": {"miniplayerRenderer": renderer},
            },
        }

    @staticmethod
    def _watch_html(payload):
        return (
            b'<!doctype html><html><head></head><body><script nonce="test">'
            b"var ytInitialPlayerResponse = "
            + json.dumps(payload, separators=(",", ":")).encode("utf-8")
            + b";var nextBootstrap = true;</script></body></html>"
        )

    @staticmethod
    def _next_payload(
        video_id,
        *,
        include_text=True,
        include_support=True,
        include_marker=True,
        support_url=(
            "//support.google.com/youtube/bin/answer.py"
            "?answer=9632097&nohelpkit=1&hl=en"
        ),
        marker_url="https://ytkids.app.goo.gl/nou5",
    ):
        notification = {}
        if include_text:
            notification["responseText"] = {
                "simpleText": "This action is turned off for content made for kids",
            }
        if include_support:
            notification["actionButton"] = {
                "buttonRenderer": {
                    "command": {
                        "commandMetadata": {
                            "webCommandMetadata": {"url": support_url},
                        },
                        "urlEndpoint": {"url": support_url},
                    },
                },
            }
        secondary_info = {
            "owner": {
                "videoOwnerRenderer": {
                    "title": {"simpleText": "Owner"},
                },
            },
        }
        if include_text or include_support:
            secondary_info["subscribeButton"] = {
                "subscribeButtonRenderer": {
                    "notificationPreferenceButton": {
                        "subscriptionNotificationToggleButtonRenderer": {
                            "command": {
                                "commandExecutorCommand": {
                                    "commands": [{
                                        "openPopupAction": {
                                            "popup": {
                                                "menuPopupRenderer": {
                                                    "items": [{
                                                        "menuServiceItemRenderer": {
                                                            "command": {
                                                                "signalServiceEndpoint": {
                                                                    "actions": [{
                                                                        "openPopupAction": {
                                                                            "popup": {
                                                                                "notificationActionRenderer": notification,
                                                                            },
                                                                        },
                                                                    }],
                                                                },
                                                            },
                                                        },
                                                    }],
                                                },
                                            },
                                        },
                                    }],
                                },
                            },
                        },
                    },
                },
            }
        carousel = {
            "carouselTitles": [{"carouselTitleViewModel": {"title": "Info"}}],
            "carouselItems": [],
        }
        if include_marker:
            command = {"urlEndpoint": {"url": marker_url}}
            carousel["carouselItems"] = [{
                "carouselItemViewModel": {
                    "carouselItem": {
                        "ctaCarouselItemViewModel": {
                            "textCarousel": {
                                "textCarouselItemViewModel": {
                                    "onTap": {"innertubeCommand": command},
                                    "button": {
                                        "buttonViewModel": {
                                            "onTap": {"innertubeCommand": command},
                                        },
                                    },
                                },
                            },
                        },
                    },
                },
            }]
        return {
            "currentVideoEndpoint": {
                "commandMetadata": {
                    "webCommandMetadata": {
                        "url": f"/watch?v={video_id}",
                        "webPageType": "WEB_PAGE_TYPE_WATCH",
                    },
                },
                "watchEndpoint": {"videoId": video_id},
            },
            "contents": {
                "twoColumnWatchNextResults": {
                    "results": {
                        "results": {
                            "contents": [
                                {
                                    "videoPrimaryInfoRenderer": {
                                        "title": {"simpleText": "Video title"},
                                    },
                                },
                                {"videoSecondaryInfoRenderer": secondary_info},
                                {
                                    "itemSectionRenderer": {
                                        "contents": [{
                                            "videoMetadataCarouselViewModel": carousel,
                                        }],
                                    },
                                },
                            ],
                            "trackingParams": "primary-tracking",
                        },
                    },
                    "secondaryResults": {
                        "secondaryResults": {
                            "results": [{"compactVideoRenderer": {"videoId": "abcdefghijk"}}],
                            "trackingParams": "secondary-tracking",
                            "targetId": "watch-next-feed",
                        },
                    },
                },
            },
        }

    def test_innertube_player_uses_exact_android_post_and_three_kids_signals(self):
        video_id = "Pk7UDVYh2bs"
        payload = self._kids_player_payload(video_id)
        captured = {}

        def fake_open(request, timeout):
            captured["request"] = request
            captured["timeout"] = timeout
            return InnertubeResponse(payload)

        client = radar.YouTubeInnertubePlayerClient(retries=0)
        with patch.object(radar.urllib.request, "urlopen", side_effect=fake_open):
            self.assertTrue(client.has_kids_player_signals(video_id))

        request = captured["request"]
        parsed = radar.urllib.parse.urlparse(request.full_url)
        self.assertEqual(request.get_method(), "POST")
        self.assertEqual(parsed.scheme, "https")
        self.assertEqual(parsed.hostname, "www.youtube.com")
        self.assertEqual(parsed.path, "/youtubei/v1/player")
        self.assertEqual(
            radar.urllib.parse.parse_qs(parsed.query),
            {"prettyPrint": ["false"]},
        )
        self.assertNotIn("key", radar.urllib.parse.parse_qs(parsed.query))
        body = json.loads(request.data.decode("utf-8"))
        self.assertEqual(body["videoId"], video_id)
        self.assertEqual(body["context"]["client"], {
            "clientName": "ANDROID",
            "clientVersion": radar.KIDS_INNERTUBE_CLIENT_VERSION,
            "androidSdkVersion": 35,
            "hl": "en",
            "gl": "US",
        })
        self.assertIs(body["contentCheckOk"], True)
        self.assertIs(body["racyCheckOk"], True)
        headers = {
            key.casefold(): value for key, value in request.header_items()
        }
        self.assertEqual(headers["x-youtube-client-name"], "3")
        self.assertEqual(
            headers["x-youtube-client-version"],
            radar.KIDS_INNERTUBE_CLIENT_VERSION,
        )
        self.assertEqual(headers["accept-language"], "en-US,en;q=0.9")
        self.assertNotIn("cookie", headers)
        self.assertIsInstance(
            radar.KidsDomValidator()._get_client(),
            radar.YouTubePublicPlayerClient,
        )

    def test_innertube_player_accepts_only_coherent_positive_or_negative(self):
        kids_id = "Pk7UDVYh2bs"
        normal_id = "XVFUtEh9zrY"
        client = radar.YouTubeInnertubePlayerClient(retries=2)
        negative = self._kids_player_payload(
            normal_id,
            mode="PLAYBACK_MODE_ALLOW",
            include_text=False,
            support_url=None,
        )
        negative["unrelatedRecommendation"] = {
            "playbackMode": "PLAYBACK_MODE_PAUSED_ONLY",
            "text": "Miniplayer is off for videos made for kids. Tap play to resume",
            "url": "https://support.google.com/youtube/bin/answer.py?answer=9632097",
        }
        with patch.object(
            radar.urllib.request,
            "urlopen",
            side_effect=lambda *args, **kwargs: InnertubeResponse(negative),
        ) as opened, patch.object(radar.time, "sleep") as slept:
            self.assertFalse(client.has_kids_player_signals(normal_id))
        self.assertEqual(opened.call_count, 1)
        slept.assert_not_called()

        invalid_payloads = {
            "wrong video": self._kids_player_payload(kids_id),
            "non-OK": self._kids_player_payload(kids_id, status="LOGIN_REQUIRED"),
            "missing renderer": {
                "videoDetails": {"videoId": kids_id},
                "playabilityStatus": {"status": "OK"},
            },
            "unknown mode": self._kids_player_payload(kids_id, mode="UNKNOWN"),
        }
        invalid_payloads["wrong video"]["videoDetails"]["videoId"] = normal_id
        for label, invalid in invalid_payloads.items():
            strict_client = radar.YouTubeInnertubePlayerClient(retries=0)
            with self.subTest(case=label), patch.object(
                radar.urllib.request,
                "urlopen",
                side_effect=lambda *args, payload=invalid, **kwargs: InnertubeResponse(payload),
            ):
                with self.assertRaises(radar.KidsDomProbeError):
                    strict_client.has_kids_player_signals(kids_id)

    def test_innertube_player_rejects_partial_or_misplaced_signals(self):
        video_id = "Pk7UDVYh2bs"
        partials = {
            2: self._kids_player_payload(video_id, support_url=None),
            1: self._kids_player_payload(
                video_id, include_text=False, support_url=None
            ),
        }
        for signal_count, payload in partials.items():
            client = radar.YouTubeInnertubePlayerClient(
                retries=1, retry_delay_seconds=0.25
            )
            with self.subTest(signals=signal_count), patch.object(
                radar.urllib.request,
                "urlopen",
                side_effect=lambda *args, body=payload, **kwargs: InnertubeResponse(body),
            ) as opened, patch.object(radar.time, "sleep") as slept:
                with self.assertRaisesRegex(
                    radar.KidsDomProbeError, f"signals={signal_count}/3"
                ):
                    client.has_kids_player_signals(video_id)
            self.assertEqual(opened.call_count, 2)
            slept.assert_called_once_with(0.25)

        allow_with_marker = self._kids_player_payload(
            video_id, mode="PLAYBACK_MODE_ALLOW"
        )
        with patch.object(
            radar.urllib.request,
            "urlopen",
            return_value=InnertubeResponse(allow_with_marker),
        ):
            with self.assertRaises(radar.KidsDomProbeError):
                radar.YouTubeInnertubePlayerClient(retries=0).has_kids_player_signals(video_id)

    def test_innertube_support_url_is_strict(self):
        valid = (
            "//support.google.com/youtube/bin/answer.py"
            "?answer=9632097&nohelpkit=1&hl=en"
        )
        self.assertTrue(
            radar.YouTubeInnertubePlayerClient._has_support_answer(valid)
        )
        self.assertTrue(radar.YouTubeInnertubePlayerClient._has_support_answer(
            "https://support.google.com/youtube/answer/9632097"
        ))
        for value in (
            "http://support.google.com/youtube/bin/answer.py?answer=9632097",
            "https://evil.example/youtube/bin/answer.py?answer=9632097",
            "https://support.google.com.evil.example/youtube/bin/answer.py?answer=9632097",
            "https://support.google.com:443/youtube/bin/answer.py?answer=9632097",
            "https://support.google.com/youtube/bin/answer.py?answer=96320970",
            "https://support.google.com/youtube/bin/answer.py?answer=9632097&answer=9632097",
            "https://evil.example/?next=https://support.google.com/youtube/answer/9632097",
            "//support.google.com/other?answer=9632097",
        ):
            with self.subTest(url=value):
                self.assertFalse(
                    radar.YouTubeInnertubePlayerClient._has_support_answer(value)
                )

    def test_innertube_player_rejects_redirected_endpoint(self):
        video_id = "Pk7UDVYh2bs"
        payload = self._kids_player_payload(video_id)
        invalid_urls = (
            "http://www.youtube.com/youtubei/v1/player?prettyPrint=false",
            "https://youtube.com/youtubei/v1/player?prettyPrint=false",
            "https://www.youtube.com:443/youtubei/v1/player?prettyPrint=false",
            "https://www.youtube.com/youtubei/v1/next?prettyPrint=false",
            "https://www.youtube.com/youtubei/v1/player?prettyPrint=false&key=public",
        )
        for final_url in invalid_urls:
            with self.subTest(url=final_url), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(payload, final_url),
            ):
                with self.assertRaisesRegex(
                    radar.KidsDomProbeError, "expected endpoint"
                ):
                    radar.YouTubeInnertubePlayerClient(retries=0).has_kids_player_signals(video_id)

    def test_innertube_player_retries_transient_and_malformed_responses(self):
        video_id = "Pk7UDVYh2bs"
        good = self._kids_player_payload(video_id)
        endpoint = radar.YouTubeInnertubePlayerClient._ENDPOINT
        failures = (
            radar.urllib.error.HTTPError(
                endpoint, 503, "busy", {}, io.BytesIO(b"busy")
            ),
            radar.urllib.error.HTTPError(
                endpoint, 403, "blocked", {}, io.BytesIO(b"blocked")
            ),
            radar.http.client.IncompleteRead(b"partial", 100),
            b"{not-json",
            {"videoDetails": {"videoId": video_id}},
        )
        for failure in failures:
            responses = iter((failure, good))

            def fake_open(*args, **kwargs):
                value = next(responses)
                if isinstance(value, Exception):
                    raise value
                return InnertubeResponse(value)

            client = radar.YouTubeInnertubePlayerClient(
                retries=1, retry_delay_seconds=0.25
            )
            with self.subTest(failure=type(failure).__name__), patch.object(
                radar.urllib.request, "urlopen", side_effect=fake_open
            ) as opened, patch.object(radar.time, "sleep") as slept:
                self.assertTrue(client.has_kids_player_signals(video_id))
            self.assertEqual(opened.call_count, 2)
            slept.assert_called_once_with(0.25)

    def test_innertube_player_fails_closed_after_malformed_json_exhaustion(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeInnertubePlayerClient(
            retries=1, retry_delay_seconds=0
        )
        with patch.object(
            radar.urllib.request,
            "urlopen",
            side_effect=lambda *args, **kwargs: InnertubeResponse(b"{bad-json"),
        ) as opened:
            with self.assertRaisesRegex(radar.KidsDomProbeError, "malformed JSON"):
                client.has_kids_player_signals(video_id)
        self.assertEqual(opened.call_count, 2)

    def test_innertube_player_rejects_noncanonical_json_and_oversize(self):
        video_id = "Pk7UDVYh2bs"
        malformed = (
            b'{"videoDetails":{},"videoDetails":{}}',
            b'{"value":NaN}',
            b'[]',
        )
        for raw in malformed:
            with self.subTest(raw=raw), patch.object(
                radar.urllib.request,
                "urlopen",
                side_effect=lambda *args, body=raw, **kwargs: InnertubeResponse(body),
            ):
                with self.assertRaises(radar.KidsDomProbeError):
                    radar.YouTubeInnertubePlayerClient(retries=0).has_kids_player_signals(video_id)

        with patch.object(radar, "KIDS_INNERTUBE_MAX_JSON_BYTES", 8), patch.object(
            radar.urllib.request,
            "urlopen",
            return_value=InnertubeResponse(b"123456789"),
        ):
            with self.assertRaisesRegex(radar.KidsDomProbeError, "bounded JSON size"):
                radar.YouTubeInnertubePlayerClient(retries=0).has_kids_player_signals(video_id)

        with self.assertRaisesRegex(ValueError, "Invalid Innertube Android"):
            radar.YouTubeInnertubePlayerClient(client_version="bad\nheader")

    def test_watch_next_uses_exact_public_web_post_without_credentials(self):
        video_id = "Pk7UDVYh2bs"
        payload = self._next_payload(video_id)
        captured = {}

        def fake_open(request, timeout):
            captured["request"] = request
            captured["timeout"] = timeout
            return InnertubeResponse(
                payload, radar.YouTubeWatchNextClient._ENDPOINT
            )

        client = radar.YouTubeWatchNextClient(retries=0)
        with patch.object(radar.urllib.request, "urlopen", side_effect=fake_open):
            self.assertTrue(client.has_kids_player_signals(video_id))

        request = captured["request"]
        self.assertEqual(request.get_method(), "POST")
        self.assertEqual(request.full_url, radar.YouTubeWatchNextClient._ENDPOINT)
        body = json.loads(request.data.decode("utf-8"))
        self.assertEqual(body, {
            "context": {
                "client": {
                    "clientName": "WEB",
                    "clientVersion": radar.KIDS_NEXT_CLIENT_VERSION,
                    "hl": "en",
                    "gl": "US",
                },
            },
            "videoId": video_id,
        })
        headers = {
            key.casefold(): value for key, value in request.header_items()
        }
        self.assertEqual(headers["x-youtube-client-name"], "1")
        self.assertEqual(
            headers["x-youtube-client-version"],
            "2.20260114.08.00",
        )
        self.assertNotIn("cookie", headers)
        self.assertNotIn("authorization", headers)
        self.assertNotIn(
            "key",
            radar.urllib.parse.parse_qs(
                radar.urllib.parse.urlparse(request.full_url).query
            ),
        )

    def test_watch_next_accepts_only_complete_positive_or_complete_negative(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        for marker_url in (
            "https://ytkids.app.goo.gl/nou5",
            "https://www.youtube.com/myfamily/#mf-compare",
        ):
            with self.subTest(marker=marker_url):
                self.assertTrue(client._classify(
                    self._next_payload(video_id, marker_url=marker_url),
                    client._ENDPOINT,
                    video_id,
                ))
        self.assertFalse(client._classify(
            self._next_payload(
                video_id,
                include_text=False,
                include_support=False,
                include_marker=False,
            ),
            client._ENDPOINT,
            video_id,
        ))

        for signals in (
            (True, False, False),
            (False, True, False),
            (False, False, True),
            (True, True, False),
            (True, False, True),
            (False, True, True),
        ):
            with self.subTest(signals=signals), self.assertRaisesRegex(
                radar.KidsDomProbeError,
                f"signals={sum(signals)}/3",
            ):
                client._classify(
                    self._next_payload(
                        video_id,
                        include_text=signals[0],
                        include_support=signals[1],
                        include_marker=signals[2],
                    ),
                    client._ENDPOINT,
                    video_id,
                )

    def test_watch_next_accepts_exact_relative_watch_url_with_optional_pp(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        for web_url in (
            f"/watch?v={video_id}",
            f"/watch?v={video_id}&pp=opaque_context",
            f"/watch?pp=opaque_context&v={video_id}",
        ):
            payload = self._next_payload(video_id)
            payload["currentVideoEndpoint"]["commandMetadata"][
                "webCommandMetadata"
            ]["url"] = web_url
            with self.subTest(url=web_url):
                self.assertTrue(client._classify(
                    payload, client._ENDPOINT, video_id
                ))

        invalid_urls = (
            f"https://www.youtube.com/watch?v={video_id}",
            f"/watch?v={video_id}&pp=",
            f"/watch?v={video_id}&pp=a&pp=b",
            f"/watch?v={video_id}&feature=x",
            f"/watch?v={video_id}&v={video_id}",
            f"/watch?v=XVFUtEh9zrY&pp=opaque",
            f"/watch;unexpected?v={video_id}",
            f"/watch?v={video_id}#fragment",
        )
        for web_url in invalid_urls:
            payload = self._next_payload(video_id)
            payload["currentVideoEndpoint"]["commandMetadata"][
                "webCommandMetadata"
            ]["url"] = web_url
            with self.subTest(url=web_url), self.assertRaisesRegex(
                radar.KidsDomProbeError, "currentVideoEndpoint mismatch"
            ):
                client._classify(payload, client._ENDPOINT, video_id)

    def test_watch_next_dedupes_only_request_specific_notification_tracking(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        payload = self._next_payload(video_id)
        secondary = payload["contents"]["twoColumnWatchNextResults"]["results"][
            "results"
        ]["contents"][1]["videoSecondaryInfoRenderer"]
        notification = client._notification_renderers(secondary)[0]
        legacy_subscribe = json.loads(json.dumps(secondary["subscribeButton"]))
        secondary.pop("subscribeButton")
        secondary["owner"] = {
            "videoOwnerRenderer": {
                "title": {"simpleText": "Owner"},
                "navigationEndpoint": {
                    "showDialogCommand": {
                        "panelLoadingStrategy": {
                            "inlineContent": {
                                "dialogViewModel": {
                                    "customContent": {
                                        "listViewModel": {
                                            "listItems": [{
                                                "listItemViewModel": {
                                                    "trailingButtons": {
                                                        "buttons": [{
                                                            "subscribeButtonViewModel": {
                                                                "onShowSubscriptionOptions": {
                                                                    "innertubeCommand": {
                                                                        "showSheetCommand": {
                                                                            "panelLoadingStrategy": {
                                                                                "inlineContent": {
                                                                                    "sheetViewModel": {
                                                                                        "content": {
                                                                                            "listViewModel": {
                                                                                                "listItems": [{
                                                                                                    "listItemViewModel": {
                                                                                                        "rendererContext": {
                                                                                                            "commandContext": {
                                                                                                                "onTap": {
                                                                                                                    "innertubeCommand": {
                                                                                                                        "signalServiceEndpoint": {
                                                                                                                            "actions": [{
                                                                                                                                "openPopupAction": {
                                                                                                                                    "popup": {
                                                                                                                                        "notificationActionRenderer": notification,
                                                                                                                                    },
                                                                                                                                },
                                                                                                                            }],
                                                                                                                        },
                                                                                                                    },
                                                                                                                },
                                                                                                            },
                                                                                                        },
                                                                                                    },
                                                                                                }],
                                                                                            },
                                                                                        },
                                                                                    },
                                                                                },
                                                                            },
                                                                        },
                                                                    },
                                                                },
                                                            },
                                                        }],
                                                    },
                                                },
                                            }],
                                        },
                                    },
                                },
                            },
                        },
                    },
                },
            },
        }
        secondary["subscribeButton"] = legacy_subscribe
        notifications = client._notification_renderers(secondary)
        self.assertEqual(len(notifications), 2)
        for index, value in enumerate(notifications):
            value["trackingParams"] = f"renderer-tracking-{index}"
            button = value["actionButton"]["buttonRenderer"]
            button["trackingParams"] = f"button-tracking-{index}"
            button["text"] = {"simpleText": "Learn more"}
            button["command"]["clickTrackingParams"] = (
                f"click-tracking-{index}"
            )
        self.assertTrue(client._classify(payload, client._ENDPOINT, video_id))

        contradiction = json.loads(json.dumps(payload))
        conflicting_secondary = contradiction["contents"][
            "twoColumnWatchNextResults"
        ]["results"]["results"]["contents"][1]["videoSecondaryInfoRenderer"]
        conflicting = client._notification_renderers(conflicting_secondary)[1]
        conflicting["responseText"]["simpleText"] = "Different notification"
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "contradictory"
        ):
            client._classify(contradiction, client._ENDPOINT, video_id)

        url_contradiction = json.loads(json.dumps(payload))
        url_secondary = url_contradiction["contents"][
            "twoColumnWatchNextResults"
        ]["results"]["results"]["contents"][1]["videoSecondaryInfoRenderer"]
        url_renderer = client._notification_renderers(url_secondary)[1]
        url_command = url_renderer["actionButton"]["buttonRenderer"][
            "command"
        ]
        different_official_url = (
            "https://support.google.com/youtube/answer/9632097"
        )
        url_command["urlEndpoint"]["url"] = different_official_url
        url_command["commandMetadata"]["webCommandMetadata"][
            "url"
        ] = different_official_url
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "contradictory"
        ):
            client._classify(
                url_contradiction, client._ENDPOINT, video_id
            )

        non_tracking_difference = json.loads(json.dumps(payload))
        different_secondary = non_tracking_difference["contents"][
            "twoColumnWatchNextResults"
        ]["results"]["results"]["contents"][1]["videoSecondaryInfoRenderer"]
        different = client._notification_renderers(different_secondary)[1]
        different["actionButton"]["buttonRenderer"]["text"][
            "simpleText"
        ] = "Unexpected label"
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "contradictory"
        ):
            client._classify(
                non_tracking_difference, client._ENDPOINT, video_id
            )

    def test_watch_next_requires_single_primary_renderer_and_unambiguous_marker(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        missing_primary = self._next_payload(video_id)
        contents = missing_primary["contents"]["twoColumnWatchNextResults"][
            "results"
        ]["results"]["contents"]
        contents.pop(0)
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "videoPrimaryInfoRenderer"
        ):
            client._classify(missing_primary, client._ENDPOINT, video_id)

        duplicate_primary = self._next_payload(video_id)
        duplicate_primary["contents"]["twoColumnWatchNextResults"]["results"][
            "results"
        ]["contents"].append({"videoPrimaryInfoRenderer": {}})
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "videoPrimaryInfoRenderer"
        ):
            client._classify(duplicate_primary, client._ENDPOINT, video_id)

        invalid_markers = (
            "https://ytkids.app.goo.gl/other",
            "https://ytkids.app.goo.gl/nou5?x=1",
            "https://ytkids.app.goo.gl/nou5#fragment",
            "https://youtube.com/myfamily/#mf-compare",
            "https://www.youtube.com/myfamily/#other",
            "https://www.youtube.com/myfamily/?x=1#mf-compare",
        )
        for marker_url in invalid_markers:
            payload = self._next_payload(video_id, marker_url=marker_url)
            with self.subTest(marker=marker_url), self.assertRaises(
                radar.KidsDomProbeError
            ):
                client._classify(payload, client._ENDPOINT, video_id)

        duplicate_carousel = self._next_payload(video_id)
        primary = duplicate_carousel["contents"]["twoColumnWatchNextResults"][
            "results"
        ]["results"]["contents"]
        primary.append(json.loads(json.dumps(primary[2])))
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "ambiguous metadata carousels"
        ):
            client._classify(duplicate_carousel, client._ENDPOINT, video_id)

    def test_watch_next_rejects_misplaced_kids_signals_and_incomplete_structure(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        misplaced = self._next_payload(
            video_id,
            include_text=False,
            include_support=False,
            include_marker=False,
        )
        misplaced["unrelatedRenderer"] = {
            "simpleText": "This action is turned off for content made for kids",
        }
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "misplaced=True"
        ):
            client._classify(misplaced, client._ENDPOINT, video_id)

        misplaced_renderer = self._next_payload(video_id)
        secondary_info = misplaced_renderer["contents"][
            "twoColumnWatchNextResults"
        ]["results"]["results"]["contents"][1]["videoSecondaryInfoRenderer"]
        exact_notification = client._notification_renderers(secondary_info)[0]
        secondary_info["arbitraryKey"] = {
            "notificationActionRenderer": json.loads(
                json.dumps(exact_notification)
            ),
        }
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "outside approved subscription paths"
        ):
            client._classify(
                misplaced_renderer, client._ENDPOINT, video_id
            )

        invalid = []
        wrong_endpoint = self._next_payload(
            video_id,
            include_text=False,
            include_support=False,
            include_marker=False,
        )
        wrong_endpoint["currentVideoEndpoint"]["watchEndpoint"]["videoId"] = (
            "XVFUtEh9zrY"
        )
        invalid.append(wrong_endpoint)
        missing_primary = self._next_payload(video_id)
        missing_primary["contents"]["twoColumnWatchNextResults"]["results"] = {}
        invalid.append(missing_primary)
        missing_secondary = self._next_payload(video_id)
        missing_secondary["contents"]["twoColumnWatchNextResults"][
            "secondaryResults"
        ] = {}
        invalid.append(missing_secondary)
        empty_secondary_feed = self._next_payload(
            video_id,
            include_text=False,
            include_support=False,
            include_marker=False,
        )
        empty_secondary_feed["contents"]["twoColumnWatchNextResults"][
            "secondaryResults"
        ]["secondaryResults"]["results"] = [{}]
        invalid.append(empty_secondary_feed)
        empty_secondary_info = self._next_payload(
            video_id,
            include_text=False,
            include_support=False,
            include_marker=False,
        )
        empty_secondary_info["contents"]["twoColumnWatchNextResults"][
            "results"
        ]["results"]["contents"][1]["videoSecondaryInfoRenderer"] = {}
        invalid.append(empty_secondary_info)
        duplicate_info = self._next_payload(video_id)
        primary = duplicate_info["contents"]["twoColumnWatchNextResults"][
            "results"
        ]["results"]["contents"]
        primary.append({"videoSecondaryInfoRenderer": {}})
        invalid.append(duplicate_info)
        for payload in invalid:
            with self.subTest(keys=list(payload)), self.assertRaises(
                radar.KidsDomProbeError
            ):
                client._classify(payload, client._ENDPOINT, video_id)

    def test_watch_next_rejects_malformed_json_redirect_and_oversize(self):
        video_id = "Pk7UDVYh2bs"
        client = radar.YouTubeWatchNextClient(retries=0)
        malformed = (
            b'{"currentVideoEndpoint":{},"currentVideoEndpoint":{}}',
            b'{"value":NaN}',
            b"[]",
            b"{bad-json",
        )
        for body in malformed:
            with self.subTest(body=body), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(
                    body, radar.YouTubeWatchNextClient._ENDPOINT
                ),
            ):
                with self.assertRaises(radar.KidsDomProbeError):
                    client.has_kids_player_signals(video_id)

        invalid_urls = (
            "http://www.youtube.com/youtubei/v1/next?prettyPrint=false",
            "https://youtube.com/youtubei/v1/next?prettyPrint=false",
            "https://www.youtube.com/youtubei/v1/player?prettyPrint=false",
            "https://www.youtube.com/youtubei/v1/next?prettyPrint=false&key=x",
        )
        for final_url in invalid_urls:
            with self.subTest(url=final_url), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(
                    self._next_payload(video_id), final_url
                ),
            ):
                with self.assertRaisesRegex(
                    radar.KidsDomProbeError, "expected endpoint"
                ):
                    client.has_kids_player_signals(video_id)

        with patch.object(radar, "KIDS_NEXT_MAX_JSON_BYTES", 8), patch.object(
            radar.urllib.request,
            "urlopen",
            return_value=InnertubeResponse(
                b"123456789", radar.YouTubeWatchNextClient._ENDPOINT
            ),
        ):
            with self.assertRaisesRegex(radar.KidsDomProbeError, "bounded JSON"):
                client.has_kids_player_signals(video_id)

    def test_public_player_prefers_complete_next_backend_before_watch(self):
        class Backend:
            def __init__(self, answers=None, error=None):
                self.answers = answers or {}
                self.error = error
                self.calls = []

            def has_kids_player_signals(self, video_id):
                self.calls.append(video_id)
                if self.error:
                    raise self.error
                return self.answers[video_id]

            def close(self):
                pass

        first_positive, second_positive = radar.KIDS_DOM_POSITIVE_CANARIES
        answers = {
            first_positive: True,
            second_positive: True,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
            "abcdefghijk": True,
        }
        android = Backend(error=radar.KidsDomProbeError("android unavailable"))
        next_backend = Backend(answers=answers)
        watch = Backend(answers=answers)
        public = radar.YouTubePublicPlayerClient(
            android, watch_client=watch, next_client=next_backend
        )
        with patch("builtins.print") as printed:
            selected = public.select_backend(
                radar.KIDS_DOM_POSITIVE_CANARIES,
                radar.KIDS_DOM_NEGATIVE_CANARY,
            )
        self.assertEqual(selected, "watch_next")
        self.assertTrue(public.has_kids_player_signals("abcdefghijk"))
        self.assertEqual(android.calls, [first_positive])
        self.assertEqual(next_backend.calls, [
            first_positive,
            second_positive,
            radar.KIDS_DOM_NEGATIVE_CANARY,
            "abcdefghijk",
        ])
        self.assertEqual(watch.calls, [])
        printed.assert_called_once()

    def test_watch_page_player_uses_exact_public_get_and_web_restriction_shape(self):
        video_id = "Pk7UDVYh2bs"
        payload = self._watch_player_payload(video_id)
        final_url = radar.YouTubeWatchPagePlayerClient._watch_url(video_id)
        captured = {}

        def fake_open(request, timeout):
            captured["request"] = request
            captured["timeout"] = timeout
            return InnertubeResponse(self._watch_html(payload), final_url)

        client = radar.YouTubeWatchPagePlayerClient(retries=0)
        with patch.object(radar.urllib.request, "urlopen", side_effect=fake_open):
            self.assertTrue(client.has_kids_player_signals(video_id))

        request = captured["request"]
        self.assertEqual(request.get_method(), "GET")
        parsed = radar.urllib.parse.urlparse(request.full_url)
        self.assertEqual(parsed.scheme, "https")
        self.assertEqual(parsed.hostname, "www.youtube.com")
        self.assertEqual(parsed.path, "/watch")
        self.assertEqual(
            radar.urllib.parse.parse_qs(parsed.query),
            {"v": [video_id], "hl": ["en"], "gl": ["US"]},
        )
        headers = {
            key.casefold(): value for key, value in request.header_items()
        }
        self.assertNotIn("cookie", headers)
        self.assertNotIn("authorization", headers)
        self.assertNotIn("key", radar.urllib.parse.parse_qs(parsed.query))

    def test_watch_page_player_accepts_only_coherent_positive_or_negative(self):
        kids_id = "Pk7UDVYh2bs"
        normal_id = "XVFUtEh9zrY"
        cases = (
            (kids_id, self._watch_player_payload(kids_id), True),
            (
                normal_id,
                self._watch_player_payload(
                    normal_id,
                    mode="PLAYBACK_MODE_ALLOW",
                    include_text=False,
                    support_url=None,
                ),
                False,
            ),
        )
        for video_id, payload, expected in cases:
            with self.subTest(video_id=video_id), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(
                    self._watch_html(payload),
                    radar.YouTubeWatchPagePlayerClient._watch_url(video_id),
                ),
            ):
                self.assertIs(
                    radar.YouTubeWatchPagePlayerClient(
                        retries=0
                    ).has_kids_player_signals(video_id),
                    expected,
                )

        partial = self._watch_player_payload(kids_id, support_url=None)
        with patch.object(
            radar.urllib.request,
            "urlopen",
            return_value=InnertubeResponse(
                self._watch_html(partial),
                radar.YouTubeWatchPagePlayerClient._watch_url(kids_id),
            ),
        ):
            with self.assertRaisesRegex(radar.KidsDomProbeError, "signals=2/3"):
                radar.YouTubeWatchPagePlayerClient(
                    retries=0
                ).has_kids_player_signals(kids_id)

    def test_watch_page_player_rejects_ambiguous_or_malformed_bootstrap(self):
        video_id = "Pk7UDVYh2bs"
        payload = self._watch_player_payload(video_id)
        good = self._watch_html(payload)
        duplicate = good.replace(
            b"</body>",
            b"<script>var ytInitialPlayerResponse = {};</script></body>",
        )
        malformed = (
            b"<script>var ytInitialPlayerResponse = "
            b'{"videoDetails":{},"videoDetails":{}};'
            b"</script>"
        )
        invalid_documents = (
            b"<html>no player bootstrap</html>",
            duplicate,
            malformed,
            good.replace(b";var nextBootstrap", b" var nextBootstrap"),
        )
        for body in invalid_documents:
            with self.subTest(body=body[:40]), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(
                    body,
                    radar.YouTubeWatchPagePlayerClient._watch_url(video_id),
                ),
            ):
                with self.assertRaises(radar.KidsDomProbeError):
                    radar.YouTubeWatchPagePlayerClient(
                        retries=0
                    ).has_kids_player_signals(video_id)

    def test_watch_page_player_rejects_redirect_and_oversize(self):
        video_id = "Pk7UDVYh2bs"
        body = self._watch_html(self._watch_player_payload(video_id))
        invalid_urls = (
            f"https://www.youtube.com/watch?v=XVFUtEh9zrY&hl=en&gl=US",
            f"https://youtube.com/watch?v={video_id}&hl=en&gl=US",
            f"http://www.youtube.com/watch?v={video_id}&hl=en&gl=US",
            f"https://www.youtube.com/embed/{video_id}?hl=en&gl=US",
        )
        for final_url in invalid_urls:
            with self.subTest(url=final_url), patch.object(
                radar.urllib.request,
                "urlopen",
                return_value=InnertubeResponse(body, final_url),
            ):
                with self.assertRaisesRegex(
                    radar.KidsDomProbeError, "expected video"
                ):
                    radar.YouTubeWatchPagePlayerClient(
                        retries=0
                    ).has_kids_player_signals(video_id)

        with patch.object(radar, "KIDS_WATCH_MAX_HTML_BYTES", 8), patch.object(
            radar.urllib.request,
            "urlopen",
            return_value=InnertubeResponse(
                b"123456789",
                radar.YouTubeWatchPagePlayerClient._watch_url(video_id),
            ),
        ):
            with self.assertRaisesRegex(radar.KidsDomProbeError, "bounded HTML"):
                radar.YouTubeWatchPagePlayerClient(
                    retries=0
                ).has_kids_player_signals(video_id)

    def test_public_player_selects_one_complete_backend_for_canaries_and_candidates(self):
        class Backend:
            def __init__(self, answers=None, error=None):
                self.answers = answers or {}
                self.error = error
                self.calls = []
                self.closed = False

            def has_kids_player_signals(self, video_id):
                self.calls.append(video_id)
                if self.error:
                    raise self.error
                return self.answers[video_id]

            def close(self):
                self.closed = True

        first_positive, second_positive = radar.KIDS_DOM_POSITIVE_CANARIES
        primary = Backend(error=radar.KidsPlayerIndeterminateError("missing objects"))
        next_backend = Backend(
            error=radar.KidsPlayerIndeterminateError("next missing objects")
        )
        watch = Backend(answers={
            first_positive: True,
            second_positive: True,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
            "abcdefghijk": True,
        })
        public = radar.YouTubePublicPlayerClient(
            primary, watch_client=watch, next_client=next_backend
        )
        validator = radar.KidsDomValidator(public, canary_retries=0)
        validator.ensure_canaries()
        self.assertTrue(validator.is_made_for_kids("abcdefghijk"))
        self.assertEqual(primary.calls, [first_positive])
        self.assertEqual(next_backend.calls, [first_positive])
        self.assertEqual(watch.calls, [
            first_positive,
            second_positive,
            radar.KIDS_DOM_NEGATIVE_CANARY,
            "abcdefghijk",
        ])
        validator.close()
        self.assertTrue(primary.closed)
        self.assertTrue(watch.closed)

    def test_public_player_fails_closed_when_both_sources_are_indeterminate(self):
        class BrokenBackend:
            def __init__(self, label):
                self.label = label

            def has_kids_player_signals(self, video_id):
                raise radar.KidsDomProbeError(self.label)

            def close(self):
                pass

        client = radar.YouTubePublicPlayerClient(
            BrokenBackend("android unavailable"),
            watch_client=BrokenBackend("watch unavailable"),
            next_client=BrokenBackend("next unavailable"),
        )
        with self.assertRaisesRegex(
            radar.KidsDomProbeError,
            "No single public YouTube player source passed every canary",
        ):
            client.select_backend(
                radar.KIDS_DOM_POSITIVE_CANARIES,
                radar.KIDS_DOM_NEGATIVE_CANARY,
            )

    def test_public_player_never_combines_partial_canary_backends(self):
        first_positive, second_positive = radar.KIDS_DOM_POSITIVE_CANARIES

        class PartialBackend:
            def __init__(self, answers):
                self.answers = answers
                self.calls = []

            def has_kids_player_signals(self, video_id):
                self.calls.append(video_id)
                return self.answers[video_id]

            def close(self):
                pass

        android = PartialBackend({
            first_positive: True,
            second_positive: False,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
        })
        watch = PartialBackend({
            first_positive: False,
            second_positive: True,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
        })
        validator = radar.KidsDomValidator(
            radar.YouTubePublicPlayerClient(
                android,
                watch_client=watch,
                next_client=PartialBackend({
                    first_positive: False,
                    second_positive: False,
                    radar.KIDS_DOM_NEGATIVE_CANARY: False,
                }),
            ),
            canary_retries=0,
        )
        with self.assertRaisesRegex(
            radar.KidsDomCanaryError,
            "No single public YouTube player source passed every canary",
        ):
            validator.ensure_canaries()
        expected = [
            first_positive,
            second_positive,
            radar.KIDS_DOM_NEGATIVE_CANARY,
        ]
        self.assertEqual(android.calls, expected)
        self.assertEqual(watch.calls, expected)

    def test_public_player_never_fails_over_a_candidate_after_selection(self):
        first_positive, second_positive = radar.KIDS_DOM_POSITIVE_CANARIES
        candidate = "abcdefghijk"

        class Backend:
            def __init__(self, answers, candidate_error=None):
                self.answers = answers
                self.candidate_error = candidate_error
                self.calls = []

            def has_kids_player_signals(self, video_id):
                self.calls.append(video_id)
                if video_id == candidate and self.candidate_error:
                    raise self.candidate_error
                return self.answers.get(video_id, False)

            def close(self):
                pass

        complete = {
            first_positive: True,
            second_positive: True,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
        }
        selected = Backend(
            complete,
            radar.KidsDomProbeError("selected source drifted"),
        )
        unused = Backend({**complete, candidate: True})
        validator = radar.KidsDomValidator(
            radar.YouTubePublicPlayerClient(
                selected, watch_client=unused, next_client=unused
            ),
            canary_retries=0,
        )
        validator.ensure_canaries()
        with self.assertRaisesRegex(
            radar.KidsDomProbeError, "selected source drifted"
        ):
            validator.is_made_for_kids(candidate)
        self.assertNotIn(candidate, unused.calls)

    def test_dom_probe_rejects_consent_then_continues(self):
        client = object.__new__(radar.ChromeWebDriverClient)
        client.session_id = "session"
        marker_states = iter(("blocked", "marker"))
        requests = []

        def fake_request(method, path, payload=None):
            requests.append((method, path, payload))
            if path.endswith("/url"):
                return None
            if payload["script"] == radar.ChromeWebDriverClient._MARKER_SCRIPT:
                return next(marker_states)
            if payload["script"] == radar.ChromeWebDriverClient._CONSENT_SCRIPT:
                return "clicked"
            raise AssertionError(payload)

        client._request = fake_request
        with patch.object(
            radar.time, "monotonic", side_effect=(0.0, 0.0, 0.1, 0.2)
        ), patch.object(radar.time, "sleep"):
            self.assertTrue(client.has_family_options_marker("abcdefghijk"))
        self.assertIn("&gl=US", requests[0][2]["url"])
        self.assertTrue(any(
            payload and payload.get("script") == radar.ChromeWebDriverClient._CONSENT_SCRIPT
            for _, _, payload in requests
        ))

    def test_dom_validator_runs_canaries_once_and_fails_closed(self):
        class FakeClient:
            def __init__(self, answers):
                self.answers = answers
                self.calls = []

            def has_family_options_marker(self, video_id):
                self.calls.append(video_id)
                return self.answers.get(video_id, False)

            def close(self):
                pass

        answers = {
            radar.KIDS_DOM_POSITIVE_CANARIES[0]: True,
            radar.KIDS_DOM_POSITIVE_CANARIES[1]: True,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
            "abcdefghijk": True,
            "zyxwvutsrqp": False,
        }
        client = FakeClient(answers)
        validator = radar.KidsDomValidator(client)
        self.assertTrue(validator.is_made_for_kids("abcdefghijk"))
        self.assertFalse(validator.is_made_for_kids("zyxwvutsrqp"))
        for canary in (*radar.KIDS_DOM_POSITIVE_CANARIES, radar.KIDS_DOM_NEGATIVE_CANARY):
            self.assertEqual(client.calls.count(canary), 1)

        failing = FakeClient({
            radar.KIDS_DOM_POSITIVE_CANARIES[0]: True,
            radar.KIDS_DOM_POSITIVE_CANARIES[1]: False,
            radar.KIDS_DOM_NEGATIVE_CANARY: False,
        })
        failed_validator = radar.KidsDomValidator(
            failing,
            canary_retries=1,
            canary_retry_delay_seconds=0,
        )
        with self.assertRaisesRegex(radar.KidsDomCanaryError, "failed closed"):
            failed_validator.is_made_for_kids("abcdefghijk")
        for canary in (*radar.KIDS_DOM_POSITIVE_CANARIES, radar.KIDS_DOM_NEGATIVE_CANARY):
            self.assertEqual(failing.calls.count(canary), 2)
        calls_after_failure = list(failing.calls)
        with self.assertRaises(radar.KidsDomCanaryError):
            failed_validator.is_made_for_kids("abcdefghijk")
        self.assertEqual(failing.calls, calls_after_failure)

    def test_dom_validator_retries_complete_canary_batch_then_succeeds(self):
        first_positive, second_positive = radar.KIDS_DOM_POSITIVE_CANARIES
        batches = (
            {
                first_positive: True,
                second_positive: False,
                radar.KIDS_DOM_NEGATIVE_CANARY: False,
            },
            {
                first_positive: True,
                second_positive: True,
                radar.KIDS_DOM_NEGATIVE_CANARY: False,
            },
        )

        class BatchClient:
            def __init__(self):
                self.calls = []

            def has_family_options_marker(self, video_id):
                batch_index = min(
                    len(self.calls) // 3,
                    len(batches) - 1,
                )
                self.calls.append(video_id)
                return batches[batch_index][video_id]

            def close(self):
                pass

        client = BatchClient()
        validator = radar.KidsDomValidator(
            client,
            canary_retries=1,
            canary_retry_delay_seconds=0.25,
        )
        with patch.object(radar.time, "sleep") as slept:
            validator.ensure_canaries()
        expected_batch = [
            first_positive,
            second_positive,
            radar.KIDS_DOM_NEGATIVE_CANARY,
        ]
        self.assertEqual(client.calls, expected_batch * 2)
        slept.assert_called_once_with(0.25)
        validator.ensure_canaries()
        self.assertEqual(client.calls, expected_batch * 2)

    def test_no_key_shard_still_runs_all_40_kids_queries(self):
        class Validator:
            def __init__(self):
                self.canary_checks = 0

            def ensure_canaries(self):
                self.canary_checks += 1

        validator = Validator()
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            output = root / "youtube-shard-0.json"
            manifest = root / "tracked.json"
            radar.write_snapshot(snapshot, {
                "d": {
                    "all": [{"vid": "abcdefghijk", "title": "Focus music"}],
                    "trends": [], "news": [], "ours": [], "kids": [], "lives": [],
                }
            })
            manifest.write_text(json.dumps({
                "version": 1, "ids": ["abcdefghijk"], "quarantine_ids": [],
            }), encoding="utf-8")
            with patch.dict(radar.os.environ, {"YOUTUBE_API_KEY": ""}), patch.object(
                radar, "fetch_owned_ydl_rows", return_value={}
            ), patch.object(
                radar, "fetch_one_video",
                return_value={"vid": "abcdefghijk", "views": 200_000},
            ), patch.object(
                radar, "fetch_discovery_spec", return_value=([], 1, 1)
            ) as discovery, patch.object(
                radar, "kids_dom_validator", return_value=validator
            ):
                artifact = radar.run_shard(snapshot, output, 0, 1, manifest)
        self.assertEqual(len(radar.KIDS_QUERY_SPECS), 40)
        self.assertEqual(artifact["queries_total"], 40)
        self.assertEqual(artifact["kids_queries_total"], 40)
        self.assertEqual(artifact["kids_queries_ok"], 40)
        self.assertEqual(validator.canary_checks, 1)
        self.assertEqual(discovery.call_count, 40)
        self.assertTrue(all(call.args[2] == "" for call in discovery.call_args_list))

    def test_no_key_fallback_prefilters_then_sets_dom_kids_provenance(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        good = {
            "id": "abcdefghijk", "title": "SLEEP MUSIC FOR KIDS - Peaceful Background Music",
            "duration": 3 * 3600, "view_count": 2_000_000,
            "is_live": False,
        }
        too_short = dict(good, id="zyxwvutsrqp", duration=19 * 60)

        class Reader:
            def __init__(self, result):
                self.result = result
                self.calls = []

            def extract_info(self, target, download=False):
                self.calls.append(target)
                return self.result

        flat_reader = Reader({"entries": [good, too_short]})
        full = dict(
            good,
            upload_date="20260801",
            description="Instrumental sleep music without vocals",
            tags=["instrumental", "baby sleep"],
            channel="Calm Baby",
            channel_url="https://www.youtube.com/@CalmBaby",
        )
        full_reader = Reader(full)

        class Validator:
            def __init__(self):
                self.calls = []
                self.canary_checks = 0

            def ensure_canaries(self):
                self.canary_checks += 1

            def is_made_for_kids(self, video_id):
                self.calls.append(video_id)
                return True

        validator = Validator()
        spec = {
            "query": "baby sleep music instrumental",
            "genre": "Baby sleep",
            "cluster": "Relaxation / meditation",
            "audience": "kids",
            "searchResults": 100,
            "searchLanes": ["viewCount", "relevance"],
        }
        with patch.object(radar, "kids_search_ydl", return_value=flat_reader), patch.object(
            radar, "ydl", return_value=full_reader
        ), patch.object(radar, "kids_dom_validator", return_value=validator):
            rows, raw, enriched, funnel = radar.fetch_kids_search_ydl(spec, now)
        self.assertEqual((raw, enriched), (4, 1))
        self.assertIn("sp=CAMSAhgC", flat_reader.calls[0])
        self.assertIn("sp=EgIYAg%3D%3D", flat_reader.calls[1])
        self.assertEqual(len(flat_reader.calls), 2)
        self.assertEqual(len(full_reader.calls), 1)
        self.assertEqual(validator.canary_checks, 1)
        self.assertEqual(validator.calls, ["abcdefghijk"])
        self.assertEqual([row["vid"] for row in rows], ["abcdefghijk"])
        self.assertIs(rows[0]["madeForKids"], True)
        self.assertEqual(
            rows[0]["madeForKidsSource"],
            "youtube_public_player_restrictions",
        )
        self.assertEqual(rows[0]["audiences"], ["kids"])
        self.assertEqual(rows[0]["discoveryLanes"], ["relevance", "viewCount"])
        self.assertEqual(funnel["raw"], 4)
        self.assertEqual(funnel["unique"], 2)
        self.assertEqual(funnel["kept"], 1)
        self.assertEqual(funnel["lane_calls_expected"], 2)
        self.assertEqual(funnel["lane_calls_completed"], 2)
        rejected = sum(
            value for key, value in funnel.items() if key.startswith("rejected_")
        )
        self.assertEqual(funnel["unique"], rejected + funnel["kept"])
        radar.validate_kids_funnel(funnel, raw, len(rows))

        rejecting = Validator()
        rejecting.is_made_for_kids = lambda video_id: False
        unused_full_reader = Reader(full)
        with patch.object(radar, "kids_search_ydl", return_value=flat_reader), patch.object(
            radar, "ydl", return_value=unused_full_reader
        ), patch.object(radar, "kids_dom_validator", return_value=rejecting):
            rejected_rows, _, rejected_enriched, rejected_funnel = radar.fetch_kids_search_ydl(spec, now)
        self.assertEqual(rejected_rows, [])
        self.assertEqual(rejected_enriched, 1)
        self.assertEqual(len(unused_full_reader.calls), 1)
        self.assertEqual(rejected_funnel["rejected_made_for_kids"], 1)

        class BrokenValidator:
            def ensure_canaries(self):
                pass

            def is_made_for_kids(self, video_id):
                raise radar.KidsDomProbeError("webdriver lost")

        with patch.object(radar, "kids_search_ydl", return_value=flat_reader), patch.object(
            radar, "ydl", return_value=unused_full_reader
        ), patch.object(radar, "kids_dom_validator", return_value=BrokenValidator()):
            with self.assertRaisesRegex(radar.KidsDomProbeError, "webdriver lost"):
                radar.fetch_kids_search_ydl(spec, now)

        class BrokenReader:
            def extract_info(self, target, download=False):
                raise RuntimeError("yt-dlp extraction failed")

        with patch.object(radar, "kids_search_ydl", return_value=flat_reader), patch.object(
            radar, "ydl", return_value=BrokenReader()
        ), patch.object(radar, "kids_dom_validator", return_value=validator):
            with self.assertRaisesRegex(RuntimeError, "could be enriched"):
                radar.fetch_kids_search_ydl(spec, now)

    def test_no_key_fallback_runs_canaries_when_prefilter_is_empty(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)

        class Reader:
            def __init__(self, result):
                self.result = result
                self.calls = []

            def extract_info(self, target, download=False):
                self.calls.append(target)
                return self.result

        flat_reader = Reader({"entries": [{
            "id": "abcdefghijk",
            "title": "Baby sleep music with vocals",
            "duration": None,
            "view_count": 2_000_000,
        }]})
        full_reader = Reader({})

        class Validator:
            def __init__(self):
                self.canary_checks = 0
                self.calls = []

            def ensure_canaries(self):
                self.canary_checks += 1

            def is_made_for_kids(self, video_id):
                self.calls.append(video_id)
                return True

        validator = Validator()
        spec = {
            "query": "baby sleep music instrumental",
            "genre": "Baby sleep",
            "cluster": "Relaxation / meditation",
            "audience": "kids",
            "searchResults": 100,
        }
        with patch.object(radar, "kids_search_ydl", return_value=flat_reader), patch.object(
            radar, "ydl", return_value=full_reader
        ), patch.object(radar, "kids_dom_validator", return_value=validator):
            rows, raw, enriched, funnel = radar.fetch_kids_search_ydl(spec, now)
        self.assertEqual((rows, raw, enriched), ([], 1, 0))
        self.assertEqual(funnel["rejected_prefilter"], 1)
        self.assertEqual(validator.canary_checks, 1)
        self.assertEqual(validator.calls, [])
        self.assertEqual(full_reader.calls, [])

    def test_no_key_fallback_preserves_exact_flat_duration_when_cloud_detail_omits_it(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        flat = {
            "id": "abcdefghijk",
            "title": "Baby sleep music instrumental no vocals",
            "duration": 3 * 3600,
            "view_count": 2_000_000,
        }
        full = {
            "id": "abcdefghijk",
            "title": flat["title"],
            "duration": None,
            "view_count": 2_000_000,
            "upload_date": "20260801",
            "description": "Long instrumental piano music without vocals",
            "tags": ["instrumental", "baby sleep"],
            "channel": "Calm Baby",
        }

        class Reader:
            def __init__(self, result):
                self.result = result

            def extract_info(self, target, download=False):
                return self.result

        class Validator:
            def ensure_canaries(self):
                pass

            def is_made_for_kids(self, video_id):
                return video_id == "abcdefghijk"

        spec = {
            "query": "baby sleep music instrumental",
            "genre": "Baby sleep",
            "cluster": "Relaxation / meditation",
            "audience": "kids",
            "searchResults": 100,
            "searchLanes": ["viewCount", "relevance"],
        }
        with patch.object(
            radar, "kids_search_ydl", return_value=Reader({"entries": [flat]})
        ), patch.object(radar, "ydl", return_value=Reader(full)), patch.object(
            radar, "kids_dom_validator", return_value=Validator()
        ):
            rows, raw, enriched, funnel = radar.fetch_kids_search_ydl(spec, now)

        self.assertEqual((raw, enriched), (2, 1))
        self.assertEqual([row["vid"] for row in rows], ["abcdefghijk"])
        self.assertEqual(rows[0]["durH"], 3.0)
        self.assertEqual(rows[0]["durationSource"], "youtube_search_result")
        self.assertEqual(funnel["duration_fallback_from_search"], 1)
        self.assertEqual(funnel["duration_missing"], 0)
        self.assertEqual(funnel["duration_below_minimum"], 0)
        self.assertEqual(funnel["rejected_duration"], 0)
        radar.validate_kids_funnel(funnel, raw, len(rows))

        for invalid_detail_duration in (float("nan"), float("inf"), 0, -1):
            detail = dict(full, duration=invalid_detail_duration)
            with self.subTest(detail_duration=invalid_detail_duration), patch.object(
                radar, "kids_search_ydl", return_value=Reader({"entries": [flat]})
            ), patch.object(radar, "ydl", return_value=Reader(detail)), patch.object(
                radar, "kids_dom_validator", return_value=Validator()
            ):
                fallback_rows, _, _, fallback_funnel = radar.fetch_kids_search_ydl(
                    spec, now
                )
            self.assertEqual(fallback_rows[0]["durH"], 3.0)
            self.assertEqual(
                fallback_rows[0]["durationSource"], "youtube_search_result"
            )
            self.assertEqual(fallback_funnel["duration_fallback_from_search"], 1)

        for invalid_flat_duration in (True, float("nan"), float("inf"), 0, -1):
            invalid_flat = dict(flat, duration=invalid_flat_duration)
            with self.subTest(flat_duration=invalid_flat_duration), patch.object(
                radar,
                "kids_search_ydl",
                return_value=Reader({"entries": [invalid_flat]}),
            ), patch.object(radar, "ydl", return_value=Reader(full)), patch.object(
                radar, "kids_dom_validator", return_value=Validator()
            ):
                rejected_rows, _, _, rejected_funnel = radar.fetch_kids_search_ydl(
                    spec, now
                )
            self.assertEqual(rejected_rows, [])
            self.assertEqual(rejected_funnel["duration_fallback_from_search"], 0)
            self.assertEqual(
                rejected_funnel["duration_missing"]
                + rejected_funnel["rejected_prefilter"],
                1,
            )

        mismatched = dict(full, id="zyxwvutsrqp")
        with patch.object(
            radar, "kids_search_ydl", return_value=Reader({"entries": [flat]})
        ), patch.object(radar, "ydl", return_value=Reader(mismatched)), patch.object(
            radar, "kids_dom_validator", return_value=Validator()
        ):
            with self.assertRaisesRegex(RuntimeError, "different video ID"):
                radar.fetch_kids_search_ydl(spec, now)

    def test_no_key_lane_plan_is_exactly_eighty_distinct_searches_max(self):
        specs = [
            spec for spec in radar.query_specs(
                {"d": {"kids": []}}, include_kids=True, kids_day="2026-08-13"
            )
            if spec["audience"] == "kids"
        ]
        urls = []
        for spec in specs:
            for lane in spec["searchLanes"]:
                urls.append(
                    "https://www.youtube.com/results?search_query="
                    + radar.urllib.parse.quote_plus(
                        spec["query"] + " " + radar.KIDS_QUERY_EXCLUSIONS
                    )
                    + "&sp=" + radar.KIDS_YTDLP_SEARCH_PARAMS[lane]
                )
        self.assertEqual(len(urls), radar.MAX_KIDS_SEARCH_CALLS)
        self.assertEqual(len(set(urls)), radar.MAX_KIDS_SEARCH_CALLS)
        self.assertEqual(sum("sp=CAMSAhgC" in url for url in urls), 40)
        self.assertEqual(sum("sp=EgIYAg%3D%3D" in url for url in urls), 20)
        self.assertEqual(sum("sp=CAISBAgCEAE%3D" in url for url in urls), 20)

    def test_instrumental_filter_rejects_kids_vocal_and_spoken_signals(self):
        base = {"duration": 3 * 3600, "title": "Baby sleep music instrumental"}
        self.assertTrue(radar.is_instrumental(base))
        for signal in (
            "lyrics", "vocals", "sing along", "children singing", "bedtime story",
            "spoken word", "voice-over", "guided affirmations", "choir", "humming",
            "children voices", "mantra",
        ):
            row = dict(base, description=signal)
            self.assertFalse(radar.is_instrumental(row), signal)
        self.assertFalse(radar.is_instrumental(dict(base, duration=19 * 60)))
        self.assertTrue(radar.is_instrumental(
            dict(base, title="Baby sleep music · no vocals")
        ))

        self.assertTrue(radar.is_instrumental(
            dict(base, title="Baby sleep music without lyrics")
        ))
        self.assertTrue(radar.is_instrumental(
            dict(base, title="Nursery rhyme piano instrumental")
        ))

    def test_contextual_kids_instrumental_proof_requires_distinct_musical_signal(self):
        base = {
            "durH": 3,
            "channel": "Calm Baby",
            "_scanDescription": "Peaceful ambient soundscape for newborn sleep",
        }
        for video_id, title in (
            ("ctxsleep001", "Baby sleep music for deep rest"),
            ("ctxsleep002", "Calming music for babies and toddlers"),
        ):
            row = dict(base, vid=video_id, title=title)
            self.assertEqual(
                radar.kids_instrumental_evidence(row),
                "made_for_kids_contextual_metadata",
            )
            self.assertTrue(radar.is_kids_instrumental(row))

        self.assertFalse(radar.is_kids_instrumental({
            "vid": "abcdefghijk",
            "durH": 3,
            "title": "Baby sleep music for deep rest",
            "_scanDescription": "Calming bedtime music for newborn sleep",
            "channel": "Calm Baby",
        }))

        mozart = dict(
            base,
            vid="abcdefghijk",
            title="Sleep instantly with Mozart and Brahms lullabies",
            _scanDescription="Long classical piano music for baby sleep",
        )
        self.assertTrue(radar.is_kids_instrumental(mozart))
        for title in (
            "Nursery rhymes for baby sleep",
            "Kids songs for bedtime",
            "Gentle lullabies for newborns",
        ):
            self.assertFalse(
                radar.is_kids_instrumental(dict(base, vid="zyxwvutsrqp", title=title)),
                title,
            )

    def test_kids_filter_rejects_confirmed_vocal_false_positives(self):
        base = {
            "durH": 3,
            "title": "Baby sleep music instrumental",
            "_scanDescription": "instrumental music box, no vocals",
        }
        verified = dict(
            base,
            vid="abcdefghijk",
            audiences=["kids"],
            madeForKids=True,
            madeForKidsSource="youtube_innertube_android_player_restrictions",
            instrumentalVerified=True,
            liveStatus="none",
            views=200_000,
            vpm=20_000,
        )
        self.assertTrue(radar.is_kids_instrumental(verified))
        self.assertTrue(radar.is_verified_kids_candidate(verified))
        for video_id in ("eNSCeIa5_5g", "Mi0XBUz562Y"):
            denied = dict(verified, vid=video_id)
            self.assertFalse(radar.is_kids_instrumental(denied), video_id)
            self.assertFalse(radar.is_verified_kids_candidate(denied), video_id)

    def test_merge_keyword_rows_preserves_kids_truth_in_both_orders(self):
        youtube = {
            "vid": "abcdefghijk", "views": 2_000_000, "kw": "focus music",
            "audiences": ["youtube"],
        }
        kids = {
            "vid": "abcdefghijk", "views": 2_000_000,
            "kw": "baby sleep music instrumental", "audiences": ["kids"],
            "madeForKids": True,
        }
        for rows in ([youtube, kids], [kids, youtube]):
            merged = radar.merge_keyword_rows(rows)[0]
            self.assertIs(merged["madeForKids"], True)
            self.assertEqual(merged["audiences"], ["youtube", "kids"])
            self.assertEqual(merged["kwCount"], 2)

    def test_update_row_changes_kids_status_only_when_official_value_exists(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        existing = {"vid": "abcdefghijk", "madeForKids": True}
        radar.update_row(existing, {"vid": "abcdefghijk", "views": 2}, now)
        self.assertIs(existing["madeForKids"], True)
        radar.update_row(existing, {"vid": "abcdefghijk", "madeForKids": False}, now)
        self.assertIs(existing["madeForKids"], False)

    def test_update_row_persists_only_nonnegative_comment_counts(self):
        now = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
        existing = {"vid": "abcdefghijk", "comments": 12}
        radar.update_row(existing, {"vid": "abcdefghijk", "views": 2}, now)
        self.assertEqual(existing["comments"], 12)
        radar.update_row(existing, {"vid": "abcdefghijk", "comments": 0}, now)
        self.assertEqual(existing["comments"], 0)
        radar.update_row(existing, {"vid": "abcdefghijk", "comments": -1}, now)
        self.assertEqual(existing["comments"], 0)

    def test_avatar_overlay_includes_kids_channels(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "avatars.js"
            count = radar.write_avatar_overlay({
                "d": {
                    "all": [], "trends": [], "news": [],
                    "kids": [{
                        "chUrl": "https://www.youtube.com/channel/UC1234567890123456789012",
                        "channelId": "UC1234567890123456789012",
                    }],
                }
            }, output)
            rendered = output.read_text(encoding="utf-8")
        self.assertEqual(count, 1)
        self.assertIn("UC1234567890123456789012", rendered)

    def test_kids_candidate_stays_in_separate_public_bucket(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            radar.write_snapshot(snapshot, {
                "d": {
                    "all": [{"vid": "abcdefghijk", "views": 1_000_000, "pub": 1700000000000}],
                    "trends": [], "news": [], "ours": [], "recos": [], "roadmap": [], "lives": [],
                }
            })
            generated = int(datetime(2026, 8, 10, 8, tzinfo=timezone.utc).timestamp() * 1000)
            artifact = {
                "version": 1, "generated_ms": generated, "shard": 0, "shards": 1,
                "tracked_total": 1, "tracked_ok": 1,
                "tracked_ids": ["abcdefghijk"], "tracked_fresh_ids": ["abcdefghijk"],
                "tracked_failed_ids": [], "tracked_unavailable_ids": [], "tracked_recovered_ids": [],
                "queries_total": 1, "queries_ok": 1, "queries_raw": 3, "queries_enriched": 3,
                "fresh": [{"vid": "abcdefghijk", "views": 1_000_001, "pub": 1700000000000}],
                "owned_fresh": [], "live_audiences": {},
                "candidates": [{
                    "vid": "zyxwvutsrqp", "title": "Baby sleep music instrumental",
                    "views": 150_000, "pub": generated - 10 * 86400000, "ageM": 1,
                    "vpm": 150_000, "genre": "Baby sleep", "cluster": "Relaxation / meditation",
                    "kw": "baby sleep music instrumental", "audiences": ["kids"],
                    "madeForKids": True, "durH": 2, "channel": "Calm Baby",
                }],
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "expected .* Kids queries"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    require_kids=True,
                )
            artifact["kids_queries_total"] = len(radar.KIDS_QUERY_SPECS)
            artifact["kids_queries_ok"] = len(radar.KIDS_QUERY_SPECS) - 2
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            with self.assertRaisesRegex(RuntimeError, "40/40.*38/40"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    require_kids=True,
                )
            artifact["kids_queries_ok"] = len(radar.KIDS_QUERY_SPECS)
            verified_candidates = artifact["candidates"]
            artifact["candidates"] = []
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            with self.assertRaisesRegex(RuntimeError, "no verified candidates"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    require_kids=True,
                )
            artifact["candidates"] = verified_candidates
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            summary = radar.merge_artifacts(
                snapshot, avatars, shards, 1,
                generate_recommendations=False,
                require_kids=True,
            )
            artifact["generated_ms"] = generated + 3600000
            artifact["fresh"][0]["views"] = 1_000_002
            artifact["candidates"] = []
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            daily_summary = radar.merge_artifacts(
                snapshot, avatars, shards, 1,
                generate_recommendations=False,
                require_kids=True,
            )
            data = radar.read_snapshot(snapshot)["d"]
        self.assertEqual([row["vid"] for row in data["kids"]], ["zyxwvutsrqp"])
        self.assertNotIn("zyxwvutsrqp", {row["vid"] for row in data["all"]})
        self.assertEqual(summary["kids_added"], 1)
        self.assertEqual(daily_summary["kids_added"], 0)
        self.assertEqual(
            [row["vid"] for row in data["kids"]],
            ["zyxwvutsrqp"],
        )

    def test_rerun_replaces_same_utc_day(self):
        day = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        later = day + 4 * 3600000
        points = radar.append_daily_point([[day, 100]], later, 125)
        self.assertEqual(points, [[later, 125]])

    def test_merge_updates_existing_video_and_adds_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            payload = {
                "t": 1,
                "d": {
                    "all": [{"vid": "abcdefghijk", "title": "Old", "views": 100, "pub": 1700000000000, "kw": "focus music"}],
                    "trends": [],
                    "news": [],
                    "recos": [],
                    "roadmap": [],
                },
            }
            radar.write_snapshot(snapshot, payload)
            generated = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
            artifact = {
                "version": 1,
                "generated_ms": generated,
                "shard": 0,
                "shards": 1,
                "tracked_total": 1,
                "tracked_ok": 1,
                "tracked_ids": ["abcdefghijk"],
                "tracked_fresh_ids": ["abcdefghijk"],
                "queries_total": 1,
                "queries_ok": 1,
                "queries_raw": 10,
                "queries_enriched": 10,
                "fresh": [{
                    "vid": "abcdefghijk",
                    "title": "Fresh",
                    "views": 150,
                    "pub": 1700000000000,
                    "chUrl": "https://www.youtube.com/@FocusChannel",
                    "channelId": "UC1234567890123456789012",
                }],
                "candidates": [],
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            summary = radar.merge_artifacts(snapshot, avatars, shards, 1)
            merged = radar.read_snapshot(snapshot)
            history = json.loads((root / "video_history" / "61.json").read_text(encoding="utf-8"))
            self.assertEqual(merged["d"]["all"][0]["views"], 150)
            self.assertEqual(merged["d"]["all"][0]["title"], "Fresh")
            self.assertEqual(merged["d"]["all"][0]["channelId"], "UC1234567890123456789012")
            self.assertNotIn("hist", merged["d"])
            self.assertEqual(history["d"]["abcdefghijk"], [[generated, 150]])
            self.assertEqual(merged["videoMetricsT"], generated)
            self.assertEqual(merged["videoMetrics"]["search_results"], 10)
            self.assertEqual(merged["videoMetrics"]["search_results_enriched"], 10)
            self.assertEqual(summary["updated"], 1)

    def test_discovery_failure_never_blocks_factual_daily_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            radar.write_snapshot(snapshot, {
                "t": 1,
                "d": {
                    "all": [{"vid": "abcdefghijk", "title": "Tracked", "views": 100, "pub": 1700000000000}],
                    "trends": [], "news": [], "recos": [], "roadmap": [],
                },
            })
            manifest = root / "tracked.json"
            ours_digest = radar.hashlib.sha256(b"abcdefghijk").hexdigest()
            ours_metadata = [{
                "vid": "abcdefghijk",
                "pub": 1_700_000_000_000,
                "durH": 1.0,
            }]
            manifest.write_text(
                json.dumps({
                    "version": 2,
                    "scan_scope": "all",
                    "ids": ["abcdefghijk"],
                    "ours_ids": ["abcdefghijk"],
                    "ours_total": 1,
                    "ours_digest": ours_digest,
                    "ours_metadata": ours_metadata,
                    "ours_metadata_total": 1,
                    "ours_metadata_digest": radar.canonical_ours_metadata_digest(
                        ours_metadata
                    ),
                    "quarantine_ids": [],
                }),
                encoding="utf-8",
            )
            generated = int(datetime(2026, 8, 4, 8, tzinfo=timezone.utc).timestamp() * 1000)
            with patch.object(radar, "utc_now_ms", return_value=generated), patch.object(
                radar, "fetch_owned_ydl_rows", side_effect=RuntimeError("owned discovery unavailable")
            ), patch.object(
                radar, "query_specs", return_value=[{"query": "focus music"}]
            ), patch.object(
                radar, "fetch_search", side_effect=RuntimeError("search unavailable")
            ), patch.object(
                radar,
                "fetch_one_video",
                return_value={"vid": "abcdefghijk", "title": "Tracked", "views": 150, "pub": 1700000000000},
            ), patch.dict(radar.os.environ, {"YOUTUBE_API_KEY": ""}):
                artifact = radar.run_shard(
                    snapshot,
                    shards / "youtube-shard-0.json",
                    0,
                    1,
                    tracked_manifest=manifest,
                )

            self.assertEqual(artifact["tracked_ok"], 1)
            self.assertEqual(artifact["queries_ok"], 0)
            self.assertFalse(artifact["owned_ok"])
            radar.merge_artifacts(
                snapshot,
                avatars,
                shards,
                1,
                generate_recommendations=False,
            )
            merged = radar.read_snapshot(snapshot)
            metrics = merged["videoMetrics"]
            self.assertFalse(metrics["partial"])
            self.assertTrue(metrics["discovery_partial"])
            self.assertFalse(metrics["owned_discovery_ok"])
            self.assertEqual(metrics["history_updated"], 1)
            self.assertTrue(radar.snapshot_freshness(snapshot, generated)["fresh"])

    def test_daily_history_keeps_latest_point_per_paris_day(self):
        morning = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        evening = morning + 10 * 3600000
        next_day = morning + 24 * 3600000
        points = radar.normalize_daily_points(
            [[morning, 100], [evening, 125], [next_day, 140]], next_day
        )
        self.assertEqual(points, [[evening, 125], [next_day, 140]])

    def test_scans_on_both_sides_of_paris_midnight_are_not_deduplicated(self):
        july_27_paris = int(datetime(2026, 7, 27, 16, 7, tzinfo=timezone.utc).timestamp() * 1000)
        july_28_paris = int(datetime(2026, 7, 27, 23, 49, tzinfo=timezone.utc).timestamp() * 1000)
        points = radar.normalize_daily_points(
            [[july_27_paris, 100], [july_28_paris, 125]],
            july_28_paris,
        )
        self.assertEqual(points, [[july_27_paris, 100], [july_28_paris, 125]])
        self.assertEqual(radar.history_day_key(july_27_paris), "2026-07-27")
        self.assertEqual(radar.history_day_key(july_28_paris), "2026-07-28")

    def test_daily_history_drops_points_before_20_july_2026(self):
        before = int(datetime(2026, 7, 19, 20, tzinfo=timezone.utc).timestamp() * 1000)
        start = int(datetime(2026, 7, 20, 20, tzinfo=timezone.utc).timestamp() * 1000)
        next_day = start + 86400000
        points = radar.normalize_daily_points([[before, 100], [start, 120], [next_day, 150]], next_day)
        self.assertEqual(points, [[start, 120], [next_day, 150]])

    def test_history_is_not_erased_when_a_source_temporarily_omits_an_id(self):
        with tempfile.TemporaryDirectory() as tmp:
            history_dir = Path(tmp)
            stamp = int(datetime(2026, 7, 27, 16, tzinfo=timezone.utc).timestamp() * 1000)
            path = history_dir / "61.json"
            path.write_text(
                json.dumps({"version": 1, "updated": stamp, "d": {"abcdefghijk": [[stamp, 100]]}}),
                encoding="utf-8",
            )
            radar.update_history_shards(history_dir, set(), {}, {}, stamp + 3600000)
            history = json.loads(path.read_text(encoding="utf-8"))
        self.assertEqual(history["d"]["abcdefghijk"], [[stamp, 100]])

    def test_canonical_owned_video_sheet_fails_closed(self):
        with patch.object(radar.urllib.request, "urlopen", side_effect=OSError("offline")):
            with self.assertRaisesRegex(RuntimeError, "canonical dashboard video list"):
                radar.sheet_video_ids()

    def test_sheet_manifest_includes_every_video_tab_visible_in_the_dashboard(self):
        from openpyxl import Workbook

        workbook = Workbook()
        all_videos = workbook.active
        all_videos.title = "All Videos"
        all_videos["A2"] = '=IMAGE("https://i.ytimg.com/vi/abcdefghijk/mqdefault.jpg")'
        trends = workbook.create_sheet("Trends")
        trends["A2"] = '=HYPERLINK("https://youtu.be/zyxwvutsrqp","Thumb")'
        news = workbook.create_sheet("News")
        news["B2"] = '=HYPERLINK("https://www.youtube.com/shorts/mnopqrstuvw","News")'
        ours = workbook.create_sheet("Our Videos")
        ours["A2"] = "12345678901"
        ours["C2"] = datetime(2026, 8, 13)
        ours["F2"] = 0.125
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        class Response(io.BytesIO):
            def __enter__(self):
                return self

            def __exit__(self, *args):
                self.close()

        with patch.object(radar.urllib.request, "urlopen", return_value=Response(payload.read())):
            catalog = radar.sheet_video_catalog()
            ids = catalog["all"]
        self.assertEqual(
            ids,
            {"abcdefghijk", "zyxwvutsrqp", "mnopqrstuvw", "12345678901"},
        )
        self.assertEqual(catalog["ours"], {"12345678901"})
        self.assertEqual(
            catalog["ours_rows"]["12345678901"],
            {
                "vid": "12345678901",
                "pub": int(datetime(2026, 8, 13, tzinfo=timezone.utc).timestamp() * 1000),
                "durH": 0.125,
            },
        )

    def test_one_canonical_manifest_is_reused_by_all_shards(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            manifest_path = root / "artifacts" / "tracked.json"
            radar.write_snapshot(snapshot, {"videoMetricsT": 123, "d": {}})
            with patch.object(
                radar,
                "sheet_video_catalog",
                return_value={
                    "all": {"abcdefghijk", "zyxwvutsrqp"},
                    "ours": {"abcdefghijk"},
                    "ours_rows": {
                        "abcdefghijk": {
                            "vid": "abcdefghijk",
                            "pub": 1_700_000_000_000,
                            "durH": 1.0,
                        },
                    },
                },
            ), patch.object(radar, "tracked_ids", return_value=["abcdefghijk", "zyxwvutsrqp"]) as tracked:
                manifest = radar.write_tracked_manifest(snapshot, manifest_path)
            loaded = radar.read_tracked_manifest(manifest_path)
        tracked.assert_called_once()
        self.assertEqual(manifest["ids"], ["abcdefghijk", "zyxwvutsrqp"])
        self.assertEqual(manifest["ours_ids"], ["abcdefghijk"])
        self.assertEqual(manifest["ours_total"], 1)
        self.assertEqual(manifest["ours_metadata_total"], 1)
        self.assertEqual(manifest["ours_metadata"][0]["durH"], 1.0)
        self.assertEqual(loaded, ["abcdefghijk", "zyxwvutsrqp"])

    def test_empty_tracked_manifest_is_allowed_only_for_kids_scope(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            kids_manifest = root / "kids.json"
            standard_manifest = root / "standard.json"
            radar.write_snapshot(snapshot, {"d": {"kids": []}})
            kids = radar.write_tracked_manifest(snapshot, kids_manifest, "kids")
            self.assertEqual(kids["ids"], [])
            self.assertEqual(radar.read_tracked_manifest(kids_manifest, "kids"), [])
            standard_manifest.write_text(json.dumps({
                "version": 2, "scan_scope": "standard", "ids": [],
            }), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "empty tracked-video manifest"):
                radar.read_tracked_manifest(standard_manifest, "standard")

    def test_publicly_unavailable_ids_are_quarantined_but_kept_as_recovery_probes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            manifest_path = root / "artifacts" / "tracked.json"
            radar.write_snapshot(snapshot, {
                "videoMetricsT": 123,
                "videoMetrics": {"unavailable_ids": ["abcdefghijk"]},
                "d": {
                    "all": [
                        {"vid": "abcdefghijk", "title": "Unavailable"},
                        {"vid": "zyxwvutsrqp", "title": "Public"},
                    ],
                },
            })
            with patch.object(
                radar,
                "sheet_video_catalog",
                return_value={
                    "all": {"abcdefghijk", "zyxwvutsrqp"},
                    "ours": {"abcdefghijk", "zyxwvutsrqp"},
                    "ours_rows": {
                        "abcdefghijk": {
                            "vid": "abcdefghijk",
                            "pub": 1_700_000_000_000,
                            "durH": 1.0,
                        },
                        "zyxwvutsrqp": {
                            "vid": "zyxwvutsrqp",
                            "pub": 1_700_000_100_000,
                            "durH": 2.0,
                        },
                    },
                },
            ):
                manifest = radar.write_tracked_manifest(snapshot, manifest_path)
            active = radar.read_tracked_manifest(manifest_path)
            probes = radar.read_quarantine_manifest(manifest_path)
        self.assertEqual(manifest["ids"], ["zyxwvutsrqp"])
        self.assertEqual(manifest["ours_ids"], ["abcdefghijk", "zyxwvutsrqp"])
        self.assertEqual(manifest["ours_total"], 2)
        self.assertEqual(active, ["zyxwvutsrqp"])
        self.assertEqual(probes, ["abcdefghijk"])

    def test_canonical_manifest_rejects_truncated_sheet_metadata(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tracked.json"
            ids = ["abcdefghijk", "zyxwvutsrqp"]
            metadata = [
                {"vid": ids[0], "pub": 1_700_000_000_000, "durH": 0.01},
                {"vid": ids[1], "pub": 1_700_000_100_000, "durH": 1.0},
            ]
            path.write_text(json.dumps({
                "version": 2,
                "scan_scope": "standard",
                "ids": ids,
                "ours_ids": ids,
                "ours_total": 2,
                "ours_digest": radar.hashlib.sha256(
                    "\n".join(ids).encode("utf-8")
                ).hexdigest(),
                "ours_metadata": metadata[:1],
                "ours_metadata_total": 2,
                "ours_metadata_digest": radar.canonical_ours_metadata_digest(
                    metadata
                ),
            }), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "metadata proof"):
                radar.read_ours_manifest(path, "standard")

    def test_missing_subscriber_count_does_not_become_zero(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        row = radar.info_to_row(
            {
                "id": "abcdefghijk",
                "title": "Focus mix",
                "view_count": 100,
                "duration": 3600,
                "upload_date": "20260719",
            },
            now,
        )
        self.assertNotIn("subs", row)

    def test_video_info_preserves_channel_id_for_avatar_lookup(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        row = radar.info_to_row(
            {
                "id": "abcdefghijk",
                "title": "Focus mix",
                "view_count": 100,
                "duration": 3600,
                "upload_date": "20260719",
                "channel_id": "UC1234567890123456789012",
                "channel_url": "https://www.youtube.com/@FocusChannel",
                "comment_count": 17,
            },
            now,
        )
        self.assertEqual(row["channelId"], "UC1234567890123456789012")
        self.assertEqual(row["comments"], 17)

    def test_public_watch_duration_uses_real_length_seconds(self):
        self.assertEqual(
            radar.parse_watch_duration('{"videoDetails":{"lengthSeconds":"3661"}}'),
            3661.0,
        )
        self.assertIsNone(radar.parse_watch_duration('{"videoDetails":{}}'))

    def test_owned_genre_is_inferred_only_from_explicit_title_words(self):
        self.assertEqual(
            radar.owned_genre_from_title("summer lofi - chill beats to relax to"),
            "Lofi / chillhop",
        )
        self.assertEqual(radar.owned_genre_from_title("Deep Focus"), "")

    def test_update_row_backfills_genre_without_overwriting_curated_genre(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        empty = {"vid": "abcdefghijk"}
        radar.update_row(
            empty,
            {"genre": "Lofi / chillhop", "genreSource": "title_explicit"},
            now,
        )
        self.assertEqual(empty["genre"], "Lofi / chillhop")
        self.assertEqual(empty["genreSource"], "title_explicit")

        curated = {"vid": "abcdefghijk", "genre": "Ambient"}
        radar.update_row(
            curated,
            {"genre": "Lofi / chillhop", "genreSource": "title_explicit"},
            now,
        )
        self.assertEqual(curated["genre"], "Ambient")

    def test_update_row_keeps_precise_api_metadata_against_yt_dlp(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
        precise_pub = 1_700_000_000_123
        existing = {
            "vid": "abcdefghijk",
            "title": "Canonical API title",
            "views": 100,
            "pub": precise_pub,
            "metadataSource": radar.METADATA_SOURCE_API,
            "pubSource": radar.METADATA_SOURCE_API,
        }
        radar.update_row(existing, {
            "vid": "abcdefghijk",
            "title": "Less reliable discovery title",
            "views": 150,
            "pub": 1_700_006_400_000,
            "metadataSource": radar.METADATA_SOURCE_YTDLP,
            "pubSource": radar.METADATA_SOURCE_YTDLP,
        }, now)
        self.assertEqual(existing["views"], 150)
        self.assertEqual(existing["title"], "Canonical API title")
        self.assertEqual(existing["pub"], precise_pub)
        self.assertEqual(existing["pubSource"], radar.METADATA_SOURCE_API)

        legacy = {"vid": "zyxwvutsrqp", "pub": precise_pub, "views": 10}
        radar.update_row(legacy, {
            "vid": "zyxwvutsrqp",
            "title": "yt-dlp title must not replace legacy metadata",
            "views": 20,
            "pub": 1_700_006_400_000,
            "metadataSource": radar.METADATA_SOURCE_YTDLP,
            "pubSource": radar.METADATA_SOURCE_YTDLP,
        }, now)
        self.assertEqual(legacy["views"], 20)
        self.assertEqual(legacy["pub"], precise_pub)
        self.assertEqual(legacy["pubSource"], radar.METADATA_SOURCE_PRESERVED)
        self.assertEqual(
            legacy["title"],
            "yt-dlp title must not replace legacy metadata",
        )

    def test_card_validation_rejects_stale_views_or_derived_metrics(self):
        now = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
        pub = 1_700_000_000_000
        fresh = {"abcdefghijk": {"vid": "abcdefghijk", "views": 200, "pub": pub}}
        age = radar.age_months(pub, now)
        valid = {
            "all": [{
                "vid": "abcdefghijk",
                "views": 200,
                "pub": pub,
                "ageM": age,
                "vpm": 200 / age,
            }],
        }
        self.assertEqual(
            radar.validate_card_refresh(valid, fresh, {"abcdefghijk"}, now),
            (1, 1),
        )
        valid["all"][0]["vpm"] = 0
        with self.assertRaisesRegex(RuntimeError, "vpm"):
            radar.validate_card_refresh(valid, fresh, {"abcdefghijk"}, now)

    def test_sheet_duration_is_a_fallback_and_api_duration_remains_authoritative(self):
        now = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
        pub = 1_700_000_000_000
        sheet = {"vid": "abcdefghijk", "pub": pub, "durH": 0.01}

        fallback = {"vid": "abcdefghijk"}
        radar.merge_sheet_ours_metadata(fallback, sheet)
        radar.update_row(fallback, {
            "vid": "abcdefghijk", "views": 100, "pub": pub,
            "metadataSource": radar.METADATA_SOURCE_YTDLP,
            "pubSource": radar.METADATA_SOURCE_YTDLP,
        }, now)
        self.assertEqual(fallback["durH"], 0.01)
        self.assertEqual(fallback["durationSource"], radar.METADATA_SOURCE_SHEET)

        authoritative = {"vid": "abcdefghijk"}
        radar.merge_sheet_ours_metadata(authoritative, sheet)
        radar.update_row(authoritative, {
            "vid": "abcdefghijk", "views": 100, "pub": pub, "durH": 1.0,
            "metadataSource": radar.METADATA_SOURCE_API,
            "pubSource": radar.METADATA_SOURCE_API,
        }, now)
        self.assertEqual(authoritative["durH"], 1.0)
        self.assertEqual(authoritative["durationSource"], radar.METADATA_SOURCE_API)

    def test_official_upload_lookup_uses_the_channel_uploads_playlist(self):
        responses = iter([
            {"items": [{"contentDetails": {"relatedPlaylists": {"uploads": "UUofficial"}}}]},
            {"items": [{"contentDetails": {"videoId": "abcdefghijk"}}]},
        ])
        with patch.object(radar, "youtube_api_payload", side_effect=lambda *args: next(responses)) as api:
            ids = radar.fetch_owned_upload_ids("test-key")
        self.assertEqual(ids, ["abcdefghijk"])
        self.assertEqual(api.call_args_list[0].args[0], "channels")
        self.assertEqual(api.call_args_list[1].args[0], "playlistItems")

    def test_official_upload_fallback_uses_the_public_channel_videos_page(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)

        class Channel:
            def extract_info(self, url, download=False):
                self.url = url
                return {"entries": [{"id": "abcdefghijk"}]}

        channel = Channel()
        with patch.object(radar, "owned_ydl", return_value=channel), patch.object(
            radar,
            "fetch_one_video",
            return_value={
                "vid": "abcdefghijk",
                "title": "summer lofi - chill beats to relax to",
                "views": 100,
            },
        ), patch.object(radar, "fetch_public_duration", return_value=3661.0) as duration:
            rows = radar.fetch_owned_ydl_rows(now)
        self.assertEqual(channel.url, "https://www.youtube.com/@LofiGirl/videos")
        self.assertEqual(rows["abcdefghijk"]["source"], "Official Lofi Girl daily scan")
        self.assertEqual(rows["abcdefghijk"]["genre"], "Lofi / chillhop")
        self.assertAlmostEqual(rows["abcdefghijk"]["durH"], 3661 / 3600)
        duration.assert_called_once_with("abcdefghijk")

    def test_merge_inserts_official_upload_into_analysis_and_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            radar.write_snapshot(snapshot, {"t": 1, "d": {"all": [{"vid": "abcdefghijk", "views": 100, "pub": 1700000000000}], "trends": [], "news": [], "ours": [], "recos": [], "roadmap": []}})
            generated = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
            tracked = {"vid": "abcdefghijk", "title": "Tracked", "views": 101, "pub": 1700000000000}
            owned = {"vid": "zyxwvutsrqp", "title": "New Lofi Girl upload", "views": 200, "comments": 27, "pub": generated, "durH": 1.0, "source": "Official Lofi Girl daily scan"}
            artifact = {"version": 1, "generated_ms": generated, "shard": 0, "shards": 1, "tracked_total": 1, "tracked_ok": 1, "tracked_ids": ["abcdefghijk"], "tracked_fresh_ids": ["abcdefghijk"], "queries_total": 1, "queries_ok": 1, "queries_raw": 1, "queries_enriched": 1, "fresh": [tracked, owned], "owned_fresh": [owned], "candidates": []}
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            radar.merge_artifacts(snapshot, avatars, shards, 1)
            merged = radar.read_snapshot(snapshot)
            self.assertEqual(merged["d"]["ours"][0]["vid"], "zyxwvutsrqp")
            self.assertEqual(merged["d"]["ours"][0]["comments"], 27)
            history = json.loads((root / "video_history" / "7a.json").read_text(encoding="utf-8"))
            self.assertEqual(history["d"]["zyxwvutsrqp"], [[generated, 200]])

    def test_existing_owned_card_uses_tracked_fresh_not_owned_discovery_counter(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            precise_pub = 1_700_000_000_123
            radar.write_snapshot(snapshot, {
                "d": {
                    "all": [], "trends": [], "news": [], "kids": [],
                    "ours": [{
                        "vid": "abcdefghijk",
                        "title": "Canonical title",
                        "views": 100,
                        "pub": precise_pub,
                        "metadataSource": radar.METADATA_SOURCE_API,
                        "pubSource": radar.METADATA_SOURCE_API,
                    }],
                    "recos": [], "roadmap": [], "lives": [],
                },
            })
            tracked = {
                "vid": "abcdefghijk",
                "title": "Canonical title",
                "views": 250,
                "pub": precise_pub,
                "metadataSource": radar.METADATA_SOURCE_API,
                "pubSource": radar.METADATA_SOURCE_API,
            }
            discovery = {
                "vid": "abcdefghijk",
                "title": "Discovery title",
                "views": 175,
                "pub": precise_pub + 86_400_000,
                "metadataSource": radar.METADATA_SOURCE_YTDLP,
                "pubSource": radar.METADATA_SOURCE_YTDLP,
                "source": "Official Lofi Girl daily scan",
            }
            artifact = {
                "version": 1, "generated_ms": generated, "shard": 0, "shards": 1,
                "tracked_total": 1, "tracked_ok": 1,
                "tracked_ids": ["abcdefghijk"],
                "tracked_fresh_ids": ["abcdefghijk"],
                "tracked_failed_ids": [], "tracked_unavailable_ids": [],
                "tracked_recovered_ids": [], "queries_total": 0, "queries_ok": 0,
                "queries_raw": 0, "queries_enriched": 0, "owned_ok": True,
                "fresh": [tracked], "owned_fresh": [discovery], "candidates": [],
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            summary = radar.merge_artifacts(
                snapshot,
                avatars,
                shards,
                1,
                generate_recommendations=False,
            )
            merged = radar.read_snapshot(snapshot)
        card = merged["d"]["ours"][0]
        self.assertEqual(card["views"], 250)
        self.assertEqual(card["pub"], precise_pub)
        self.assertEqual(card["title"], "Canonical title")
        self.assertEqual(card["source"], "Official Lofi Girl daily scan")
        self.assertEqual(summary["card_rows_expected"], 1)
        self.assertEqual(summary["card_rows_updated"], 1)

    def test_all_83_sheet_owned_cards_are_materialized_from_fresh_beyond_discovery_cap(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            ids = [f"{index:011d}" for index in range(83)]
            digest = radar.hashlib.sha256("\n".join(ids).encode("utf-8")).hexdigest()
            radar.write_snapshot(snapshot, {
                "d": {
                    "all": [], "trends": [], "news": [], "kids": [],
                    "ours": [{"vid": video_id, "title": "Sheet title"} for video_id in ids[:50]],
                    "recos": [], "roadmap": [], "lives": [],
                },
            })
            fresh = [{
                "vid": video_id,
                "title": f"Canonical {index}",
                "views": 100_000 + index,
                "pub": 1_700_000_000_000 + index,
                "metadataSource": radar.METADATA_SOURCE_API,
                "pubSource": radar.METADATA_SOURCE_API,
            } for index, video_id in enumerate(ids)]
            sheet_metadata = [{
                "vid": video_id,
                "pub": 1_700_000_000_000 + index,
                "durH": 1.0,
            } for index, video_id in enumerate(ids)]
            artifact = {
                "version": 1, "scan_scope": "standard", "generated_ms": generated,
                "shard": 0, "shards": 1, "tracked_total": 83, "tracked_ok": 83,
                "tracked_ids": ids, "tracked_fresh_ids": ids,
                "tracked_failed_ids": [], "tracked_unavailable_ids": [],
                "tracked_recovered_ids": [], "queries_total": 0, "queries_ok": 0,
                "queries_raw": 0, "queries_enriched": 0, "owned_ok": True,
                "fresh": fresh, "owned_fresh": fresh[:50], "candidates": [],
                "canonical_ours_manifest": True,
                "canonical_ours_ids": ids,
                "canonical_ours_total": 83,
                "canonical_ours_digest": digest,
                "canonical_ours_metadata": sheet_metadata,
                "canonical_ours_metadata_total": 83,
                "canonical_ours_metadata_digest": radar.canonical_ours_metadata_digest(
                    sheet_metadata
                ),
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            summary = radar.merge_artifacts(
                snapshot, avatars, shards, 1,
                generate_recommendations=False,
                scan_scope="standard",
            )
            merged = radar.read_snapshot(snapshot)

        self.assertEqual(len(merged["d"]["ours"]), 83)
        by_id = {row["vid"]: row for row in merged["d"]["ours"]}
        self.assertEqual(by_id[ids[-1]]["views"], 100_082)
        self.assertEqual(by_id[ids[-1]]["pub"], 1_700_000_000_082)
        self.assertEqual(summary["sheet_ours_expected"], 83)
        self.assertEqual(summary["sheet_ours_updated"], 83)
        self.assertEqual(merged["videoMetrics"]["sheet_ours_expected"], 83)
        self.assertEqual(merged["videoMetrics"]["sheet_ours_updated"], 83)
        self.assertEqual(summary["analysis_rows_expected"], 83)
        self.assertEqual(summary["analysis_rows_updated"], 83)

    def test_visible_analysis_proof_matches_real_sheet_snapshot_merge_shape(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            sheet_ids = [f"s{index:010d}" for index in range(79)]
            extra_ids = [f"x{index:010d}" for index in range(14)]
            filler_ids = [f"f{index:010d}" for index in range(8)]
            all_ids = sorted(sheet_ids + extra_ids + filler_ids)
            digest = radar.hashlib.sha256(
                "\n".join(sorted(sheet_ids)).encode("utf-8")
            ).hexdigest()
            sheet_metadata = [{
                "vid": video_id,
                "pub": 1_700_000_000_000 + index,
                "durH": 0.01 if index < 7 else 1.0,
            } for index, video_id in enumerate(sorted(sheet_ids))]
            # Mirrors production: 79 unique Sheet rows, seven short rows hidden
            # by Analyse, plus twelve ordinary snapshot-only rows of which one
            # is short. Two more rows prove the final public exclusions: a
            # deferred phonk row and an unavailable row.
            existing_extras = [{
                "vid": video_id,
                "title": "Phonk must be pruned" if index == 12 else f"Snapshot-only {index}",
                "views": 90_000 + index,
                "pub": 1_700_000_100_000 + index,
                "durH": 0.025 if index == 0 else None,
            } for index, video_id in enumerate(extra_ids)]
            radar.write_snapshot(snapshot, {"d": {
                "all": [], "trends": [], "news": [], "kids": [],
                "ours": existing_extras,
                "recos": [], "roadmap": [], "lives": [],
            }})
            fresh = []
            for index, video_id in enumerate(sheet_ids):
                fresh.append({
                    "vid": video_id,
                    "title": f"Sheet {index}",
                    "views": 100_000 + index,
                    "pub": 1_700_000_000_000 + index,
                    "metadataSource": radar.METADATA_SOURCE_API,
                    "pubSource": radar.METADATA_SOURCE_API,
                })
            for index, row in enumerate(existing_extras):
                fresh.append({
                    **row,
                    "views": 110_000 + index,
                    "metadataSource": radar.METADATA_SOURCE_API,
                    "pubSource": radar.METADATA_SOURCE_API,
                })
            for index, video_id in enumerate(filler_ids):
                fresh.append({
                    "vid": video_id,
                    "title": f"Tracked filler {index}",
                    "views": 120_000 + index,
                    "pub": 1_700_000_200_000 + index,
                    "metadataSource": radar.METADATA_SOURCE_API,
                    "pubSource": radar.METADATA_SOURCE_API,
                })
            artifact = {
                "version": 1, "scan_scope": "standard", "generated_ms": generated,
                "shard": 0, "shards": 1, "tracked_total": 101, "tracked_ok": 100,
                "tracked_ids": all_ids,
                "tracked_fresh_ids": [video_id for video_id in all_ids if video_id != extra_ids[13]],
                "tracked_failed_ids": [extra_ids[13]],
                "tracked_unavailable_ids": [extra_ids[13]],
                "tracked_recovered_ids": [], "queries_total": 0, "queries_ok": 0,
                "queries_raw": 0, "queries_enriched": 0, "owned_ok": True,
                "fresh": [row for row in fresh if row["vid"] != extra_ids[13]],
                "owned_fresh": fresh[:50], "candidates": [],
                "canonical_ours_manifest": True,
                "canonical_ours_ids": sorted(sheet_ids),
                "canonical_ours_total": 79,
                "canonical_ours_digest": digest,
                "canonical_ours_metadata": sheet_metadata,
                "canonical_ours_metadata_total": 79,
                "canonical_ours_metadata_digest": radar.canonical_ours_metadata_digest(
                    sheet_metadata
                ),
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            summary = radar.merge_artifacts(
                snapshot, avatars, shards, 1,
                generate_recommendations=False,
                scan_scope="standard",
            )
            merged = radar.read_snapshot(snapshot)

        visible = [
            row for row in merged["d"]["ours"]
            if row["vid"] != extra_ids[13] and radar.is_analysis_card(row)
        ]
        self.assertNotIn(extra_ids[12], {row["vid"] for row in merged["d"]["ours"]})
        self.assertEqual(len({row["vid"] for row in merged["d"]["ours"]}), 92)
        self.assertEqual(len(visible), 83)
        self.assertEqual(summary["sheet_ours_expected"], 79)
        self.assertEqual(summary["sheet_ours_updated"], 79)
        self.assertEqual(summary["analysis_rows_expected"], 83)
        self.assertEqual(summary["analysis_rows_updated"], 83)
        by_id = {row["vid"]: row for row in merged["d"]["ours"]}
        self.assertEqual(by_id[sheet_ids[0]]["durH"], 0.01)
        self.assertEqual(by_id[sheet_ids[7]]["durH"], 1.0)

    def test_canonical_sheet_metadata_proof_rejects_a_truncated_shard(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            ids = ["abcdefghijk", "zyxwvutsrqp"]
            digest = radar.hashlib.sha256("\n".join(ids).encode("utf-8")).hexdigest()
            complete_metadata = [
                {"vid": ids[0], "pub": 1_700_000_000_000, "durH": 0.01},
                {"vid": ids[1], "pub": 1_700_000_100_000, "durH": 1.0},
            ]
            radar.write_snapshot(snapshot, {"d": {
                "all": [], "trends": [], "news": [], "kids": [], "ours": [],
                "recos": [], "roadmap": [], "lives": [],
            }})
            fresh = [{
                "vid": video_id,
                "views": 100_000 + index,
                "pub": 1_700_000_000_000 + index,
                "metadataSource": radar.METADATA_SOURCE_API,
                "pubSource": radar.METADATA_SOURCE_API,
            } for index, video_id in enumerate(ids)]
            artifact = {
                "version": 1, "scan_scope": "standard", "generated_ms": generated,
                "shard": 0, "shards": 1, "tracked_total": 2, "tracked_ok": 2,
                "tracked_ids": ids, "tracked_fresh_ids": ids,
                "tracked_failed_ids": [], "tracked_unavailable_ids": [],
                "tracked_recovered_ids": [], "queries_total": 0, "queries_ok": 0,
                "queries_raw": 0, "queries_enriched": 0, "owned_ok": True,
                "fresh": fresh, "owned_fresh": [], "candidates": [],
                "canonical_ours_manifest": True,
                "canonical_ours_ids": ids,
                "canonical_ours_total": 2,
                "canonical_ours_digest": digest,
                "canonical_ours_metadata": complete_metadata[:1],
                "canonical_ours_metadata_total": 2,
                "canonical_ours_metadata_digest": radar.canonical_ours_metadata_digest(
                    complete_metadata
                ),
            }
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            with self.assertRaisesRegex(RuntimeError, "metadata shard is truncated"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    scan_scope="standard",
                )

    def test_canonical_sheet_owned_proof_rejects_a_truncated_shard_union(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            ids = [f"{index:011d}" for index in range(83)]
            digest = radar.hashlib.sha256("\n".join(ids).encode("utf-8")).hexdigest()
            radar.write_snapshot(snapshot, {"d": {
                "all": [], "trends": [], "news": [], "kids": [], "ours": [],
                "recos": [], "roadmap": [], "lives": [],
            }})
            partial = ids[:-1]
            fresh = [{
                "vid": video_id, "views": 100_000 + index,
                "pub": 1_700_000_000_000 + index,
                "metadataSource": radar.METADATA_SOURCE_API,
                "pubSource": radar.METADATA_SOURCE_API,
            } for index, video_id in enumerate(partial)]
            artifact = {
                "version": 1, "scan_scope": "standard", "generated_ms": generated,
                "shard": 0, "shards": 1, "tracked_total": 82, "tracked_ok": 82,
                "tracked_ids": partial, "tracked_fresh_ids": partial,
                "tracked_failed_ids": [], "tracked_unavailable_ids": [],
                "tracked_recovered_ids": [], "queries_total": 0, "queries_ok": 0,
                "queries_raw": 0, "queries_enriched": 0, "owned_ok": True,
                "fresh": fresh, "owned_fresh": fresh[:50], "candidates": [],
                "canonical_ours_manifest": True,
                "canonical_ours_ids": partial,
                "canonical_ours_total": 83,
                "canonical_ours_digest": digest,
            }
            (shards / "youtube-shard-0.json").write_text(json.dumps(artifact), encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "truncated or inconsistent"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    scan_scope="standard",
                )

    def test_fallback_quarantines_only_after_two_consecutive_missing_scans(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            public_id = "abcdefghijk"
            missing_id = "zyxwvutsrqp"
            radar.write_snapshot(snapshot, {
                "t": 1,
                "d": {
                    "all": [
                        {"vid": public_id, "views": 100, "pub": 1700000000000},
                        {"vid": missing_id, "views": 200, "pub": 1700000000000},
                    ],
                    "trends": [],
                    "news": [],
                    "ours": [],
                    "recos": [],
                    "roadmap": [],
                },
            })

            def write_missing_artifact(generated, views):
                artifact = {
                    "version": 1,
                    "generated_ms": generated,
                    "shard": 0,
                    "shards": 1,
                    "tracked_total": 2,
                    "tracked_ok": 1,
                    "tracked_ids": [public_id, missing_id],
                    "tracked_fresh_ids": [public_id],
                    "tracked_failed_ids": [missing_id],
                    "tracked_unavailable_ids": [],
                    "tracked_recovered_ids": [],
                    "queries_total": 1,
                    "queries_ok": 1,
                    "queries_raw": 1,
                    "queries_enriched": 1,
                    "fresh": [{"vid": public_id, "views": views, "pub": 1700000000000}],
                    "owned_fresh": [],
                    "candidates": [],
                }
                (shards / "youtube-shard-0.json").write_text(
                    json.dumps(artifact), encoding="utf-8"
                )

            first = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
            write_missing_artifact(first, 101)
            with patch.object(radar, "MIN_PUBLISH_TRACK_RATIO", 0.0):
                radar.merge_artifacts(snapshot, avatars, shards, 1)
            first_metrics = radar.read_snapshot(snapshot)["videoMetrics"]
            self.assertTrue(first_metrics["partial"])
            self.assertEqual(first_metrics["tracked"], 2)
            self.assertEqual(first_metrics["updated"], 1)
            self.assertEqual(first_metrics["missing_ids"], [missing_id])
            self.assertEqual(first_metrics["unavailable_ids"], [])

            second = first + 3600000
            write_missing_artifact(second, 102)
            with patch.object(radar, "MIN_PUBLISH_TRACK_RATIO", 0.0):
                radar.merge_artifacts(snapshot, avatars, shards, 1)
            second_metrics = radar.read_snapshot(snapshot)["videoMetrics"]
            self.assertFalse(second_metrics["partial"])
            self.assertEqual(second_metrics["tracked"], 1)
            self.assertEqual(second_metrics["updated"], 1)
            self.assertEqual(second_metrics["missing_ids"], [])
            self.assertEqual(second_metrics["unavailable_ids"], [missing_id])

            third = second + 3600000
            recovery_artifact = {
                "version": 1,
                "generated_ms": third,
                "shard": 0,
                "shards": 1,
                "tracked_total": 1,
                "tracked_ok": 1,
                "tracked_ids": [public_id],
                "tracked_fresh_ids": [public_id],
                "tracked_failed_ids": [],
                "tracked_unavailable_ids": [],
                "tracked_recovered_ids": [missing_id],
                "queries_total": 1,
                "queries_ok": 1,
                "queries_raw": 1,
                "queries_enriched": 1,
                "fresh": [
                    {"vid": public_id, "views": 103, "pub": 1700000000000},
                    {"vid": missing_id, "views": 201, "pub": 1700000000000},
                ],
                "owned_fresh": [],
                "candidates": [],
            }
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(recovery_artifact), encoding="utf-8"
            )
            radar.merge_artifacts(snapshot, avatars, shards, 1)
            recovery_metrics = radar.read_snapshot(snapshot)["videoMetrics"]
            self.assertFalse(recovery_metrics["partial"])
            self.assertEqual(recovery_metrics["tracked"], 2)
            self.assertEqual(recovery_metrics["updated"], 2)
            self.assertEqual(recovery_metrics["unavailable_ids"], [])
            recovery_history = json.loads(
                (root / "video_history" / "7a.json").read_text(encoding="utf-8")
            )
            self.assertEqual(recovery_history["d"][missing_id][-1], [third, 201])

    def test_fallback_single_miss_then_success_never_quarantines(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            public_id, intermittent_id = "abcdefghijk", "zyxwvutsrqp"
            radar.write_snapshot(snapshot, {
                "d": {
                    "all": [
                        {"vid": public_id, "pub": 1700000000000},
                        {"vid": intermittent_id, "pub": 1700000000000},
                    ],
                    "trends": [], "news": [], "ours": [], "recos": [], "roadmap": [],
                }
            })

            def write_artifact(generated, fresh, failed):
                ids = [public_id, intermittent_id]
                artifact = {
                    "version": 1, "generated_ms": generated, "shard": 0, "shards": 1,
                    "tracked_total": 2, "tracked_ok": len(fresh), "tracked_ids": ids,
                    "tracked_fresh_ids": [row["vid"] for row in fresh],
                    "tracked_failed_ids": failed, "tracked_unavailable_ids": [],
                    "tracked_recovered_ids": [], "queries_total": 1, "queries_ok": 1,
                    "queries_raw": 1, "queries_enriched": 1, "fresh": fresh,
                    "owned_fresh": [], "candidates": [],
                }
                (shards / "youtube-shard-0.json").write_text(
                    json.dumps(artifact), encoding="utf-8"
                )

            first = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)
            write_artifact(first, [{"vid": public_id, "views": 100, "pub": 1700000000000}], [intermittent_id])
            with patch.object(radar, "MIN_PUBLISH_TRACK_RATIO", 0.0):
                radar.merge_artifacts(snapshot, avatars, shards, 1)
            self.assertEqual(
                radar.read_snapshot(snapshot)["videoMetrics"]["missing_ids"],
                [intermittent_id],
            )

            second = first + 3600000
            write_artifact(second, [
                {"vid": public_id, "views": 101, "pub": 1700000000000},
                {"vid": intermittent_id, "views": 201, "pub": 1700000000000},
            ], [])
            radar.merge_artifacts(snapshot, avatars, shards, 1)
            metrics = radar.read_snapshot(snapshot)["videoMetrics"]
            self.assertFalse(metrics["partial"])
            self.assertEqual(metrics["tracked"], 2)
            self.assertEqual(metrics["updated"], 2)
            self.assertEqual(metrics["missing_ids"], [])
            self.assertEqual(metrics["unavailable_ids"], [])

    def test_avatar_overlay_links_handle_to_channel_id_without_overwriting_atlas(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "avatars.js"
            count = radar.write_avatar_overlay(
                {
                    "d": {
                        "all": [{
                            "chUrl": "https://www.youtube.com/@FocusChannel",
                            "channelId": "UC1234567890123456789012",
                        }],
                        "trends": [],
                        "news": [],
                    }
                },
                output,
            )
            rendered = output.read_text(encoding="utf-8")
        self.assertEqual(count, 1)
        self.assertIn('"@FocusChannel":"UC1234567890123456789012"', rendered)
        self.assertIn("if(!atlas.channels[key])", rendered)

    def test_recent_search_uses_month_filter_and_enriches_results(self):
        now = int(datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000)

        class FlatSearch:
            url = None

            def extract_info(self, url, download=False):
                self.url = url
                return {"entries": [{"id": "abcdefghijk"}]}

        class FullVideo:
            def extract_info(self, url, download=False):
                return {
                    "id": "abcdefghijk",
                    "title": "Long focus mix",
                    "view_count": 20_000,
                    "duration": 3600,
                    "upload_date": "20260719",
                    "channel": "Focus Channel",
                }

        flat = FlatSearch()
        with patch.object(radar, "search_ydl", return_value=flat), patch.object(
            radar, "ydl", return_value=FullVideo()
        ):
            rows, raw, enriched = radar.fetch_search(
                {"query": "focus music", "genre": "Ambient", "cluster": "Focus"}, now
            )
        self.assertIn("sp=EgIIBA%3D%3D", flat.url)
        self.assertEqual((raw, enriched, len(rows)), (1, 1, 1))
        self.assertEqual(rows[0]["rank"], 1)
        self.assertEqual(rows[0]["added"], now)

    def test_news_view_floor_is_100k_and_prunes_legacy_rows(self):
        data = {
            "news": [
                {"vid": "abcdefghijk", "views": 99_999},
                {"vid": "zyxwvutsrqp", "views": 100_000},
                {"vid": "mnopqrstuvw", "views": 250_000},
            ]
        }
        removed = radar.prune_news_below_view_floor(data)
        self.assertEqual(radar.MIN_NEWS_VIEWS, 100_000)
        self.assertEqual(removed, 1)
        self.assertEqual(
            [row["vid"] for row in data["news"]],
            ["zyxwvutsrqp", "mnopqrstuvw"],
        )

    def test_merge_rejects_99999_views_and_accepts_100000(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "Lofi_Radar_data.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(
                datetime(2026, 7, 20, 8, tzinfo=timezone.utc).timestamp() * 1000
            )
            radar.write_snapshot(
                snapshot,
                {
                    "t": 1,
                    "d": {
                        "all": [{"vid": "abcdefghijk", "views": 1_000_000, "pub": 1700000000000}],
                        "trends": [],
                        "news": [],
                        "ours": [],
                        "recos": [],
                        "roadmap": [],
                    },
                },
            )
            artifact = {
                "version": 1,
                "generated_ms": generated,
                "shard": 0,
                "shards": 1,
                "tracked_total": 1,
                "tracked_ok": 1,
                "tracked_ids": ["abcdefghijk"],
                "tracked_fresh_ids": ["abcdefghijk"],
                "queries_total": 1,
                "queries_ok": 1,
                "queries_raw": 2,
                "queries_enriched": 2,
                "fresh": [{"vid": "abcdefghijk", "views": 1_000_001, "pub": 1700000000000}],
                "candidates": [
                    {
                        "vid": "zyxwvutsrqp",
                        "title": "Below floor",
                        "views": 99_999,
                        "ageM": 1,
                        "vpm": 99_999,
                        "kw": "focus music",
                    },
                    {
                        "vid": "mnopqrstuvw",
                        "title": "At floor",
                        "views": 100_000,
                        "ageM": 1,
                        "vpm": 100_000,
                        "kw": "focus music",
                    },
                ],
            }
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            radar.merge_artifacts(snapshot, avatars, shards, 1)
            merged = radar.read_snapshot(snapshot)
        self.assertEqual(
            [row["vid"] for row in merged["d"]["news"]],
            ["mnopqrstuvw"],
        )

    def test_watchdog_requires_today_history_and_publish_thresholds(self):
        with tempfile.TemporaryDirectory() as tmp:
            snapshot = Path(tmp) / "snapshot.js"
            stamp = int(datetime(2026, 7, 28, 8, tzinfo=timezone.utc).timestamp() * 1000)
            radar.write_snapshot(snapshot, {
                "videoMetricsT": stamp,
                "videoMetrics": {
                    "tracked": 1000,
                    "updated": 1000,
                    "keywords": 100,
                    "keywords_ok": 100,
                    "history_updated": 1000,
                    "card_rows_expected": 1000,
                    "card_rows_updated": 1000,
                    "sheet_ours_expected": 83,
                    "sheet_ours_updated": 83,
                    "analysis_rows_expected": 83,
                    "analysis_rows_updated": 83,
                    "history_day": "2026-07-28",
                    "day_timezone": "Europe/Paris",
                    "partial": False,
                },
                "d": {},
            })
            healthy = radar.snapshot_freshness(snapshot, stamp + 3600000)
            stale = radar.snapshot_freshness(snapshot, stamp + 24 * 3600000)
            partial_payload = radar.read_snapshot(snapshot)
            partial_payload["videoMetrics"].update({"updated": 999, "history_updated": 999, "partial": True})
            radar.write_snapshot(snapshot, partial_payload)
            partial = radar.snapshot_freshness(snapshot, stamp + 3600000)
        self.assertTrue(healthy["fresh"])
        self.assertFalse(stale["fresh"])
        self.assertFalse(partial["fresh"])

    def test_pages_verifier_requires_both_snapshot_and_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            history_dir = root / "video_history"
            history_dir.mkdir()
            stamp = int(datetime(2026, 7, 28, 8, tzinfo=timezone.utc).timestamp() * 1000)
            radar.write_snapshot(snapshot, {
                "videoMetricsT": stamp,
                "videoMetrics": {
                    "history_updated": 2,
                    "card_rows_expected": 2,
                    "card_rows_updated": 2,
                    "sheet_ours_expected": 1,
                    "sheet_ours_updated": 1,
                    "analysis_rows_expected": 1,
                    "analysis_rows_updated": 1,
                },
                "d": {},
            })
            pool = root / "Lofi_Radar_recommendation_pool.js"
            pool.write_text(
                radar.POOL_PREFIX + json.dumps({"sourceT": stamp, "items": [{}] * 1001}) + ";\n",
                encoding="utf-8",
            )
            (history_dir / "61.json").write_text(
                json.dumps({"version": 1, "updated": stamp, "d": {"abcdefghijk": [[stamp, 100]]}}),
                encoding="utf-8",
            )
            (history_dir / "62.json").write_text(
                json.dumps({"version": 1, "updated": stamp, "d": {"lmnopqrstuv": [[stamp, 200]]}}),
                encoding="utf-8",
            )

            class Response(io.BytesIO):
                def __enter__(self):
                    return self

                def __exit__(self, *args):
                    self.close()

            responses = iter([
                Response(snapshot.read_bytes()),
                Response(pool.read_bytes()),
                Response((history_dir / "61.json").read_bytes()),
                Response((history_dir / "62.json").read_bytes()),
            ])
            with patch.object(radar.urllib.request, "urlopen", side_effect=lambda *args, **kwargs: next(responses)):
                result = radar.verify_publication(
                    "https://example.test/radar/",
                    snapshot,
                    history_dir,
                    timeout_seconds=1,
                    interval_seconds=1,
                )
        self.assertTrue(result["published"])
        self.assertEqual(result["recommendations"], 1001)
        self.assertEqual(result["snapshot"], stamp)
        self.assertEqual(result["history_min"], stamp)
        self.assertEqual(result["history_shards"], 2)
        self.assertEqual(result["history_points"], 2)

    def test_core_pages_verifier_does_not_depend_on_recommendation_pool(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            history_dir = root / "video_history"
            history_dir.mkdir()
            stamp = int(datetime(2026, 8, 4, 8, tzinfo=timezone.utc).timestamp() * 1000)
            radar.write_snapshot(snapshot, {
                "videoMetricsT": stamp,
                "videoMetrics": {
                    "history_updated": 1,
                    "card_rows_expected": 1,
                    "card_rows_updated": 1,
                    "sheet_ours_expected": 1,
                    "sheet_ours_updated": 1,
                    "analysis_rows_expected": 1,
                    "analysis_rows_updated": 1,
                },
                "d": {},
            })
            shard = history_dir / "61.json"
            shard.write_text(
                json.dumps({"version": 1, "updated": stamp, "d": {"abcdefghijk": [[stamp, 100]]}}),
                encoding="utf-8",
            )

            class Response(io.BytesIO):
                def __enter__(self):
                    return self

                def __exit__(self, *args):
                    self.close()

            responses = iter([Response(snapshot.read_bytes()), Response(shard.read_bytes())])
            with patch.object(
                radar.urllib.request,
                "urlopen",
                side_effect=lambda *args, **kwargs: next(responses),
            ):
                result = radar.verify_publication(
                    "https://example.test/radar/",
                    snapshot,
                    history_dir,
                    timeout_seconds=1,
                    interval_seconds=1,
                    require_recommendation_pool=False,
                )
        self.assertTrue(result["published"])
        self.assertIsNone(result["recommendations"])
        self.assertEqual(result["snapshot"], stamp)
        self.assertEqual(result["history_min"], stamp)

    def test_pages_verifier_rejects_fresh_shard_metadata_with_missing_points(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            history_dir = root / "video_history"
            history_dir.mkdir()
            stamp = int(datetime(2026, 8, 4, 8, tzinfo=timezone.utc).timestamp() * 1000)
            radar.write_snapshot(snapshot, {
                "videoMetricsT": stamp,
                "videoMetrics": {
                    "history_updated": 1,
                    "card_rows_expected": 1,
                    "card_rows_updated": 1,
                    "sheet_ours_expected": 1,
                    "sheet_ours_updated": 1,
                    "analysis_rows_expected": 1,
                    "analysis_rows_updated": 1,
                },
                "d": {},
            })
            shard = history_dir / "61.json"
            shard.write_text(
                json.dumps({"version": 1, "updated": stamp, "d": {"abcdefghijk": [[stamp, 100]]}}),
                encoding="utf-8",
            )
            remote_shard = json.dumps({"version": 1, "updated": stamp, "d": {}}).encode()

            class Response(io.BytesIO):
                def __enter__(self):
                    return self

                def __exit__(self, *args):
                    self.close()

            responses = iter([Response(snapshot.read_bytes()), Response(remote_shard)])
            with patch.object(
                radar.urllib.request,
                "urlopen",
                side_effect=lambda *args, **kwargs: next(responses),
            ), patch.object(radar.time, "monotonic", side_effect=[0, 0, 2]), patch.object(
                radar.time, "sleep", return_value=None
            ):
                with self.assertRaisesRegex(RuntimeError, "stale history shards"):
                    radar.verify_publication(
                        "https://example.test/radar/",
                        snapshot,
                        history_dir,
                        timeout_seconds=1,
                        interval_seconds=1,
                        require_recommendation_pool=False,
                    )


class ScanScopeTests(unittest.TestCase):
    @staticmethod
    def _payload():
        return {
            "t": 111,
            "videoMetricsT": 222,
            "videoMetrics": {
                "tracked": 1,
                "updated": 1,
                "missing_ids": [],
                "unavailable_ids": [],
                "sentinel": {"keep": [1, 2, 3]},
            },
            "videoHistory": {"updated": 222, "sentinel": "keep"},
            "d": {
                "all": [{
                    "vid": "abcdefghijk",
                    "title": "Focus music",
                    "views": 100_000,
                    "pub": 1_700_000_000_000,
                    "kw": "focus music",
                    "genre": "Lofi / chillhop",
                    "cluster": "Study / focus / work",
                }],
                "trends": [{"vid": "abcdefghijk", "sentinel": "trend"}],
                "news": [{"vid": "abcdefghijk", "sentinel": "news"}],
                "ours": [{"vid": "abcdefghijk", "sentinel": "ours"}],
                "recos": [{"id": "recommendation", "sentinel": ["keep"]}],
                "roadmap": [{"id": "roadmap", "sentinel": {"keep": True}}],
                "lives": [{"vid": "livevideo01", "sentinel": "live"}],
                "kids": [],
            },
            "sentinel": {"top_level": "keep"},
        }

    @staticmethod
    def _artifact(scope, generated, **overrides):
        artifact = {
            "version": 1,
            "scan_scope": scope,
            "generated_ms": generated,
            "shard": 0,
            "shards": 1,
            "tracked_total": 1,
            "tracked_ok": 1,
            "tracked_ids": ["abcdefghijk"],
            "tracked_fresh_ids": ["abcdefghijk"],
            "tracked_failed_ids": [],
            "tracked_unavailable_ids": [],
            "tracked_recovered_ids": [],
            "queries_total": 0,
            "queries_ok": 0,
            "queries_raw": 0,
            "queries_enriched": 0,
            "kids_queries_total": 0,
            "kids_queries_ok": 0,
            "kids_results_examined": 0,
            "kids_candidates_kept": 0,
            "owned_ok": True,
            "owned_fresh": [],
            "live_audiences": {},
            "fresh": [{
                "vid": "abcdefghijk",
                "title": "Fresh focus music",
                "views": 100_001,
                "pub": 1_700_000_000_000,
            }],
            "candidates": [],
        }
        artifact.update(overrides)
        return artifact

    def test_query_specs_keeps_top_100_after_previous_kids_scan(self):
        payload = {
            "d": {"kids": []},
            "kidsMetrics": {"queries": len(radar.KIDS_QUERY_SPECS)},
            "videoMetrics": {"kids_queries": 0},
        }
        specs = [
            spec
            for spec in radar.query_specs(payload, include_kids=True)
            if spec.get("audience") == "kids"
        ]
        self.assertEqual(len(specs), len(radar.KIDS_QUERY_SPECS))
        self.assertTrue(all(spec["searchResults"] == 100 for spec in specs))
        self.assertTrue(all(len(spec["searchLanes"]) == 2 for spec in specs))

    def test_standard_shard_excludes_every_kids_operation(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            output = root / "youtube-shard-0.json"
            manifest = root / "tracked.json"
            payload = self._payload()
            payload["d"]["kids"] = [{
                "vid": "kidsvideo01",
                "kw": "baby sleep music instrumental",
                "madeForKids": True,
            }]
            radar.write_snapshot(snapshot, payload)
            manifest.write_text(json.dumps({
                "version": 1,
                "scan_scope": "standard",
                "ids": ["abcdefghijk", "kidsvideo01"],
                "quarantine_ids": [],
            }), encoding="utf-8")
            seen_specs = []

            def discovery(spec, now_ms, api_key):
                seen_specs.append(dict(spec))
                return [], 1, 1

            with patch.dict(radar.os.environ, {"YOUTUBE_API_KEY": ""}), patch.object(
                radar,
                "fetch_one_video",
                side_effect=lambda video_id, now_ms: {
                    "vid": video_id,
                    "views": 100_001,
                    "pub": 1_700_000_000_000,
                },
            ), patch.object(
                radar, "fetch_owned_ydl_rows", return_value={}
            ), patch.object(
                radar, "fetch_discovery_spec", side_effect=discovery
            ), patch.object(
                radar, "kids_dom_validator"
            ) as validator, patch.object(
                radar, "fetch_kids_search"
            ) as kids_api, patch.object(
                radar, "fetch_kids_search_ydl"
            ) as kids_public, patch.object(
                radar, "sheet_video_ids"
            ) as sheet:
                artifact = radar.run_shard(
                    snapshot,
                    output,
                    0,
                    1,
                    manifest,
                    scan_scope="standard",
                )
                output_scope = json.loads(
                    output.read_text(encoding="utf-8")
                )["scan_scope"]

        self.assertEqual(artifact["scan_scope"], "standard")
        self.assertEqual(artifact["tracked_ids"], ["abcdefghijk", "kidsvideo01"])
        self.assertEqual(artifact["tracked_ok"], 2)
        self.assertEqual(artifact["kids_queries_total"], 0)
        self.assertTrue(seen_specs)
        self.assertTrue(
            all(spec.get("audience") == "youtube" for spec in seen_specs)
        )
        validator.assert_not_called()
        kids_api.assert_not_called()
        kids_public.assert_not_called()
        sheet.assert_not_called()
        self.assertEqual(output_scope, "standard")

    def test_standard_manifest_includes_existing_kids_without_kids_discovery(self):
        payload = self._payload()
        payload["d"]["kids"] = [{"vid": "kidsvideo01"}]
        with patch.object(radar, "sheet_video_ids", return_value=set()):
            ids = radar.tracked_ids(payload, "standard")
        self.assertEqual(ids, ["abcdefghijk", "kidsvideo01"])

    def test_kids_shard_runs_only_kids_specs_without_owned_live_or_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            output = root / "youtube-shard-0.json"
            radar.write_snapshot(snapshot, self._payload())
            seen_specs = []

            def discovery(spec, now_ms, api_key):
                seen_specs.append(dict(spec))
                return [], 2, 1

            with patch.dict(
                radar.os.environ, {"YOUTUBE_API_KEY": "test-key"}
            ), patch.object(
                radar, "fetch_discovery_spec", side_effect=discovery
            ), patch.object(
                radar, "fetch_api_rows", return_value={}
            ) as api_rows, patch.object(
                radar, "fetch_owned_api_rows"
            ) as owned_api, patch.object(
                radar, "fetch_owned_ydl_rows"
            ) as owned_public, patch.object(
                radar, "kids_dom_validator"
            ) as validator, patch.object(
                radar, "sheet_video_ids"
            ) as sheet:
                artifact = radar.run_shard(
                    snapshot,
                    output,
                    0,
                    1,
                    scan_scope="kids",
                )
                output_scope = json.loads(
                    output.read_text(encoding="utf-8")
                )["scan_scope"]

        self.assertEqual(artifact["scan_scope"], "kids")
        self.assertEqual(artifact["tracked_total"], 0)
        self.assertEqual(
            artifact["queries_total"], len(radar.KIDS_QUERY_SPECS)
        )
        self.assertEqual(
            artifact["kids_queries_total"], len(radar.KIDS_QUERY_SPECS)
        )
        self.assertEqual(len(seen_specs), len(radar.KIDS_QUERY_SPECS))
        self.assertTrue(
            all(spec.get("audience") == "kids" for spec in seen_specs)
        )
        owned_api.assert_not_called()
        owned_public.assert_not_called()
        validator.assert_not_called()
        sheet.assert_not_called()
        api_rows.assert_called_once()
        self.assertEqual(api_rows.call_args.args[0], [])
        self.assertEqual(output_scope, "kids")

    def test_standard_merge_refreshes_preserved_kids_counters_and_metrics(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(
                datetime(2026, 8, 11, 8, tzinfo=timezone.utc).timestamp()
                * 1000
            )
            payload = self._payload()
            payload["kidsMetricsT"] = 987654321
            payload["kidsMetrics"] = {
                "queries": 40,
                "queries_ok": 40,
                "nested": {"preserve": [3, 2, 1]},
            }
            payload["d"]["kids"] = [{
                "vid": "kidsvideo01",
                "title": "Preserve this Kids cohort row",
                "views": 222_222,
                "pub": 1_700_000_000_000,
                "ageM": 10,
                "vpm": 22_222,
                "durH": 2,
                "madeForKids": True,
                "madeForKidsSource": "youtube_data_api_status",
                "custom": {"preserve": ["byte", "value"]},
            }]
            radar.write_snapshot(snapshot, payload)
            artifact = self._artifact(
                "standard",
                generated,
                tracked_total=2,
                tracked_ok=2,
                tracked_ids=["abcdefghijk", "kidsvideo01"],
                tracked_fresh_ids=["abcdefghijk", "kidsvideo01"],
                fresh=[
                    {
                        "vid": "abcdefghijk",
                        "title": "Fresh focus music",
                        "views": 100_001,
                        "pub": 1_700_000_000_000,
                    },
                    {
                        "vid": "kidsvideo01",
                        "title": "Fresh official Kids title",
                        "views": 333_333,
                        "pub": 1_700_000_000_000,
                        "metadataSource": radar.METADATA_SOURCE_API,
                        "pubSource": radar.METADATA_SOURCE_API,
                        "madeForKids": True,
                        "madeForKidsSource": "youtube_data_api_status",
                    },
                ],
            )
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact),
                encoding="utf-8",
            )

            radar.merge_artifacts(
                snapshot,
                avatars,
                shards,
                1,
                generate_recommendations=False,
                scan_scope="standard",
            )
            after = radar.read_snapshot(snapshot)

        self.assertEqual([row["vid"] for row in after["d"]["kids"]], ["kidsvideo01"])
        self.assertEqual(after["d"]["kids"][0]["views"], 333_333)
        self.assertEqual(after["d"]["kids"][0]["title"], "Fresh official Kids title")
        self.assertEqual(
            after["d"]["kids"][0]["custom"],
            {"preserve": ["byte", "value"]},
        )
        self.assertEqual(after["kidsMetricsT"], generated)
        self.assertEqual(after["kidsMetrics"]["queries"], 40)
        self.assertEqual(after["kidsMetrics"]["nested"], {"preserve": [3, 2, 1]})
        self.assertEqual(after["kidsMetrics"]["tracked"], 1)
        self.assertEqual(after["kidsMetrics"]["updated"], 1)
        self.assertEqual(after["kidsMetrics"]["history_updated"], 1)
        self.assertFalse(after["kidsMetrics"]["partial"])
        self.assertEqual(after["videoMetrics"]["card_rows_expected"], after["videoMetrics"]["card_rows_updated"])

    def test_kids_bootstrap_changes_only_dedicated_state_avatar_and_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            history = root / "history"
            shards.mkdir()
            generated = int(
                datetime(2026, 8, 11, 8, tzinfo=timezone.utc).timestamp()
                * 1000
            )
            radar.write_snapshot(snapshot, self._payload())
            before = radar.read_snapshot(snapshot)
            candidate = {
                "vid": "kidsvideo01",
                "title": "Baby sleep music instrumental - 2 hours",
                "views": 250_000,
                "pub": generated - 30 * 86400000,
                "ageM": 1,
                "vpm": 250_000,
                "durH": 2,
                "channel": "Calm Baby",
                "chUrl": (
                    "https://www.youtube.com/channel/"
                    "UC1234567890123456789012"
                ),
                "channelId": "UC1234567890123456789012",
                "genre": "Baby sleep",
                "cluster": "Relaxation / meditation",
                "kw": "baby sleep music instrumental",
                "audiences": ["kids"],
                "madeForKids": True,
                "madeForKidsSource": (
                    "youtube_public_player_restrictions"
                ),
                "instrumentalVerified": True,
                "liveStatus": "none",
            }
            artifact = self._artifact(
                "kids",
                generated,
                version=2,
                tracked_total=0,
                tracked_ok=0,
                tracked_ids=[],
                tracked_fresh_ids=[],
                fresh=[candidate],
                queries_total=len(radar.KIDS_QUERY_SPECS),
                queries_ok=len(radar.KIDS_QUERY_SPECS),
                queries_raw=4000,
                queries_enriched=1800,
                kids_queries_total=len(radar.KIDS_QUERY_SPECS),
                kids_queries_ok=len(radar.KIDS_QUERY_SPECS),
                kids_results_examined=4000,
                kids_candidates_kept=1,
                kids_funnel={
                    **radar.empty_kids_funnel(),
                    "raw": 4000,
                    "unique": 1800,
                    "enriched": 1800,
                    "rejected_made_for_kids": 1799,
                    "kept": 1,
                    "lane_calls_expected": 80,
                    "lane_calls_completed": 80,
                },
                candidates=[candidate],
            )
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )

            summary = radar.merge_artifacts(
                snapshot,
                avatars,
                shards,
                1,
                history_dir=history,
                generate_recommendations=False,
                require_kids=True,
                scan_scope="kids",
            )
            after = radar.read_snapshot(snapshot)
            history_payload = json.loads(
                (
                    history
                    / radar.history_shard_name(candidate["vid"])
                ).read_text(encoding="utf-8")
            )
            avatar_text = avatars.read_text(encoding="utf-8")

        for key in before:
            if key != "d":
                self.assertEqual(after[key], before[key], key)
        for bucket, rows in before["d"].items():
            if bucket != "kids":
                self.assertEqual(after["d"][bucket], rows, bucket)

        self.assertEqual(
            [row["vid"] for row in after["d"]["kids"]],
            [candidate["vid"]],
        )
        self.assertEqual(after["kidsMetricsT"], generated)
        self.assertEqual(
            after["kidsMetrics"]["queries"],
            len(radar.KIDS_QUERY_SPECS),
        )
        self.assertEqual(after["kidsMetrics"]["candidates_kept"], 1)
        self.assertEqual(summary["kids_added"], 1)
        self.assertEqual(
            history_payload["d"][candidate["vid"]],
            [[generated, 250_000]],
        )
        self.assertIn(candidate["channelId"], avatar_text)

    def test_require_kids_rejects_zero_verified_yield_and_preserves_funnel(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            generated = int(datetime(2026, 8, 13, 8, tzinfo=timezone.utc).timestamp() * 1000)
            radar.write_snapshot(snapshot, self._payload())
            funnel = radar.empty_kids_funnel()
            funnel.update({
                "raw": 2_000,
                "unique": 40,
                "enriched": 40,
                "rejected_made_for_kids": 40,
                "lane_calls_expected": 80,
                "lane_calls_completed": 80,
            })
            artifact = self._artifact(
                "kids", generated,
                version=2,
                tracked_total=0,
                tracked_ok=0,
                tracked_ids=[],
                tracked_fresh_ids=[],
                fresh=[],
                queries_total=40,
                queries_ok=40,
                queries_raw=2_000,
                queries_enriched=40,
                kids_queries_total=40,
                kids_queries_ok=40,
                kids_results_examined=2_000,
                kids_candidates_kept=0,
                kids_funnel=funnel,
            )
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(artifact), encoding="utf-8"
            )
            with self.assertRaisesRegex(RuntimeError, "coverage or verified candidates"):
                radar.merge_artifacts(
                    snapshot, avatars, shards, 1,
                    generate_recommendations=False,
                    require_kids=True,
                    scan_scope="kids",
                )

    def test_scope_only_history_preserves_out_of_scope_series_same_shard(self):
        with tempfile.TemporaryDirectory() as tmp:
            history = Path(tmp)
            in_scope = "abcdefghijk"
            out_of_scope = "a1234567890"
            old = int(
                datetime(2026, 8, 9, 8, tzinfo=timezone.utc).timestamp()
                * 1000
            )
            now = int(
                datetime(2026, 8, 11, 8, tzinfo=timezone.utc).timestamp()
                * 1000
            )
            outside_points = [
                [old + 3000, 9],
                [old, 7],
                [old, 7],
            ]
            path = history / radar.history_shard_name(in_scope)
            self.assertEqual(
                path.name,
                radar.history_shard_name(out_of_scope),
            )
            path.write_text(json.dumps({
                "version": 1,
                "updated": old,
                "d": {
                    in_scope: [[old, 100]],
                    out_of_scope: outside_points,
                },
            }), encoding="utf-8")

            radar.update_history_shards(
                history,
                {in_scope},
                {in_scope: {"views": 150}},
                {},
                now,
                scope_only=True,
            )
            updated = json.loads(path.read_text(encoding="utf-8"))

        self.assertEqual(updated["d"][out_of_scope], outside_points)
        self.assertEqual(updated["d"][in_scope][-1], [now, 150])

    def test_merge_rejects_scope_mismatch_and_duplicate_shards(self):
        generated = int(
            datetime(2026, 8, 11, 8, tzinfo=timezone.utc).timestamp()
            * 1000
        )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            radar.write_snapshot(snapshot, self._payload())
            (shards / "youtube-shard-0.json").write_text(
                json.dumps(self._artifact("kids", generated)),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(
                RuntimeError, "Artifact scope mismatch"
            ):
                radar.merge_artifacts(
                    snapshot,
                    avatars,
                    shards,
                    1,
                    generate_recommendations=False,
                    scan_scope="standard",
                )

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            snapshot = root / "snapshot.js"
            avatars = root / "avatars.js"
            shards = root / "shards"
            shards.mkdir()
            radar.write_snapshot(snapshot, self._payload())
            duplicate = self._artifact(
                "standard", generated, shard=0, shards=2
            )
            for suffix in ("a", "b"):
                (shards / f"youtube-shard-{suffix}.json").write_text(
                    json.dumps(duplicate), encoding="utf-8"
                )
            with self.assertRaisesRegex(
                RuntimeError, "Duplicate shard artifacts"
            ):
                radar.merge_artifacts(
                    snapshot,
                    avatars,
                    shards,
                    2,
                    generate_recommendations=False,
                    scan_scope="standard",
                )



if __name__ == "__main__":
    unittest.main()

import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

import refresh_youtube_live_snapshot as live


class RefreshYouTubeLiveSnapshotTests(unittest.TestCase):
    def workbook(self):
        workbook = openpyxl.Workbook()
        streams = workbook.active
        streams.title = "Live Streams"
        streams.append(["Video ID", "Channel", "Title", "URL", "Started", "Discovery keywords"])
        streams.append(["abcdefghijk", "Channel A", "Stream A 2026-08-11 12:00", "", "2025-01-01", "lofi"])
        streams.append(["zyxwvutsrqp", "Channel B", "Stream B", "", "2025-02-01", "jazz"])
        history = workbook.create_sheet("Live History")
        history.append(["Video ID", "Scan Date", "Concurrent Viewers"])
        history.append(["abcdefghijk", "2026-08-10", 100])
        history.append(["zyxwvutsrqp", "2026-08-10", 50])
        hourly = workbook.create_sheet("Live Hourly")
        hourly.append(["Video ID", "Scan Timestamp", "Concurrent Viewers"])
        hourly.append(["abcdefghijk", "2026-08-11T11:00:00Z", 60])
        hourly.append(["abcdefghijk", "2026-08-11T12:00:00Z", 42])
        hourly.append(["zyxwvutsrqp", "2026-08-11T02:00:00Z", 9])
        return workbook

    def test_builds_compact_summary_and_preserves_official_audience(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "Lofi_Radar_data.js"
            snapshot.write_text(
                'window.LOFI_DATA={"d":{"lives":[{"vid":"abcdefghijk",'
                '"madeForKids":false,"audiences":["youtube"]}]}};',
                encoding="utf-8",
            )
            payload = live.build_payload(self.workbook(), snapshot)
        self.assertEqual(payload["metrics"]["tracked"], 2)
        self.assertEqual(payload["metrics"]["active"], 1)
        self.assertNotIn("liveHourly", payload["d"])
        first = next(row for row in payload["d"]["lives"] if row["vid"] == "abcdefghijk")
        self.assertEqual(first["title"], "Stream A")
        self.assertIs(first["madeForKids"], False)
        self.assertEqual(first["audiences"], ["youtube"])
        self.assertEqual(payload["d"]["liveSummary"]["abcdefghijk"]["now"], 42)
        self.assertEqual(payload["d"]["liveSummary"]["abcdefghijk"]["peak24"], 60)
        self.assertIs(payload["d"]["liveSummary"]["abcdefghijk"]["active"], True)
        self.assertIs(payload["d"]["liveSummary"]["zyxwvutsrqp"]["active"], False)

    def test_discovers_official_streams_and_does_not_invent_historical_peaks(self):
        class Reader:
            def __init__(self, payloads):
                self.payloads = payloads

            def extract_info(self, url, download=False):
                return self.payloads[url]

        observed = 1_786_640_000_000
        ids = ("0muHFBSiybw", "LTiqKDrjqr4")
        flat = Reader({
            live.OFFICIAL_STREAMS_URL: {
                "entries": [
                    {"id": ids[0], "live_status": "is_live"},
                    {"id": ids[1], "live_status": "is_live"},
                    {"id": "abcdefghijk", "live_status": "was_live"},
                ]
            },
            live.OFFICIAL_UPLOADS_URL: {
                "entries": [{"id": ids[0], "live_status": "is_live"}]
            },
        })
        detail = Reader({
            f"https://www.youtube.com/watch?v={ids[0]}": {
                "id": ids[0], "channel_id": live.OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "title": "summer lofi radio ☀️ music to put you in a better mood 2026-08-13 19:08",
                "is_live": True, "live_status": "is_live", "concurrent_view_count": 3208,
                "release_timestamp": 1_786_546_921,
            },
            f"https://www.youtube.com/watch?v={ids[1]}": {
                "id": ids[1], "channel_id": live.OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "title": "24/7 deep sleep music 🌌 calm ambient to sleep & dream to",
                "is_live": True, "live_status": "is_live", "concurrent_view_count": 859,
                "release_timestamp": 1_785_344_581,
            },
            "https://www.youtube.com/watch?v=abcdefghijk": {
                "id": "abcdefghijk",
                "channel_id": live.OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "is_live": False,
                "live_status": "was_live",
            },
        })
        official = live.discover_official_live_streams(
            observed, flat_reader=flat, detail_reader=detail
        )
        payload = live.build_payload(self.workbook(), official_snapshot=official)
        by_id = {row["vid"]: row for row in payload["d"]["lives"]}
        self.assertEqual(set(ids) - set(by_id), set())
        self.assertEqual(by_id[ids[0]]["title"], "summer lofi radio ☀️ music to put you in a better mood")
        self.assertEqual(by_id[ids[0]]["started"], 1_786_546_921_000)
        self.assertEqual(payload["d"]["liveSummary"][ids[0]]["now"], 3208)
        self.assertIsNone(payload["d"]["liveSummary"][ids[0]]["peak24"])
        self.assertIsNone(payload["d"]["liveSummary"][ids[0]]["peakAll"])
        self.assertEqual(payload["metrics"]["officialExpected"], 2)
        self.assertEqual(payload["metrics"]["officialVerified"], 2)
        self.assertEqual(payload["metrics"]["officialAdded"], 2)
        self.assertEqual(payload["metrics"]["officialObservedT"], observed)
        self.assertEqual(official["metrics"]["streamsTabCandidates"], 2)
        self.assertEqual(official["metrics"]["streamsTabActive"], 2)
        self.assertEqual(official["metrics"]["listingConfirmedEnded"], 0)
        self.assertEqual(official["metrics"]["uploadsPlaylistActive"], 1)
        self.assertLess(payload["metrics"]["sheetSourceLatestT"], observed)

    def test_partial_streams_listing_recovers_previous_live_and_confirms_ended(self):
        class Reader:
            def __init__(self, payloads):
                self.payloads = payloads

            def extract_info(self, url, download=False):
                return self.payloads[url]

        listing_id = "0muHFBSiybw"
        recovered_id = "LTiqKDrjqr4"
        ended_id = "X4VbdwhkE10"
        flat = Reader({
            live.OFFICIAL_STREAMS_URL: {
                "entries": [
                    {"id": listing_id, "live_status": "is_live"},
                    {"id": "abcdefghijk", "live_status": "is_upcoming"},
                ]
            },
            live.OFFICIAL_UPLOADS_URL: {"entries": []},
        })
        active = lambda video_id, title: {
            "id": video_id,
            "channel_id": live.OFFICIAL_CHANNEL_ID,
            "availability": "public",
            "title": title,
            "is_live": True,
            "live_status": "is_live",
            "concurrent_view_count": 10,
            "release_timestamp": 1_785_344_581,
        }
        detail = Reader({
            f"https://www.youtube.com/watch?v={listing_id}": active(listing_id, "summer"),
            f"https://www.youtube.com/watch?v={recovered_id}": active(recovered_id, "sleep"),
            f"https://www.youtube.com/watch?v={ended_id}": {
                "id": ended_id,
                "channel_id": live.OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "is_live": False,
                "live_status": "was_live",
            },
            "https://www.youtube.com/watch?v=abcdefghijk": {
                "id": "abcdefghijk",
                "channel_id": live.OFFICIAL_CHANNEL_ID,
                "availability": "public",
                "is_live": False,
                "live_status": "is_upcoming",
            },
        })
        result = live.discover_official_live_streams(
            1_786_640_000_000,
            flat_reader=flat,
            detail_reader=detail,
            previous_active={
                recovered_id: {"trusted": True},
                ended_id: {"trusted": True},
            },
        )
        self.assertEqual({row["vid"] for row in result["rows"]}, {listing_id, recovered_id})
        self.assertEqual(result["metrics"]["listingActive"], 1)
        self.assertEqual(result["metrics"]["previousActive"], 2)
        self.assertEqual(result["metrics"]["missingFromListing"], 2)
        self.assertEqual(result["metrics"]["recoveredStillLive"], 1)
        self.assertEqual(result["metrics"]["confirmedEnded"], 1)
        self.assertEqual(result["metrics"]["listingConfirmedEnded"], 0)
        self.assertEqual(result["metrics"]["listingUpcoming"], 1)

    def test_flat_upcoming_rows_never_call_detail_or_watch_fallback(self):
        active_id = "0muHFBSiybw"
        streams_upcoming = "abcdefghijk"
        uploads_upcoming = "zyxwvutsrqp"

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [
                        {"id": active_id, "live_status": "is_live"},
                        {"id": streams_upcoming, "live_status": "is_upcoming"},
                    ]}
                return {"entries": [
                    {"id": uploads_upcoming, "live_status": "is_upcoming"}
                ]}

        class DetailReader:
            def __init__(self):
                self.calls = []

            def extract_info(self, url, download=False):
                video_id = url.rsplit("=", 1)[-1]
                self.calls.append(video_id)
                if video_id != active_id:
                    raise RuntimeError("Sign in to confirm you're not a bot")
                return {
                    "id": active_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "active radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 12,
                    "release_timestamp": 1_786_546_921,
                }

        class UnexpectedWatch:
            def extract_info(self, video_id):
                raise AssertionError("Explicit upcoming rows must not reach WatchPage")

        detail = DetailReader()
        result = live.discover_official_live_streams(
            1_786_640_000_000,
            flat_reader=FlatReader(),
            detail_reader=detail,
            watch_reader=UnexpectedWatch(),
        )
        self.assertEqual(detail.calls, [active_id])
        self.assertEqual([row["vid"] for row in result["rows"]], [active_id])
        self.assertEqual(result["metrics"]["streamsTabActive"], 1)
        self.assertEqual(result["metrics"]["streamsTabCandidates"], 1)
        self.assertEqual(result["metrics"]["listingUpcoming"], 2)
        self.assertEqual(result["metrics"]["watchPageFallbacks"], 0)

    def test_trusted_previous_live_with_ambiguous_identity_blocks_discovery(self):
        class Reader:
            def __init__(self, payload):
                self.payload = payload

            def extract_info(self, url, download=False):
                return self.payload

        video_id = "LTiqKDrjqr4"
        with self.assertRaisesRegex(RuntimeError, "conflicting identity"):
            live.discover_official_live_streams(
                1_786_640_000_000,
                flat_reader=Reader({"entries": []}),
                detail_reader=Reader({
                    "id": video_id,
                    "channel_id": "UCxxxxxxxxxxxxxxxxxxxxxx",
                }),
                previous_active={video_id: {"trusted": True}},
            )

    def test_trusted_zero_viewer_live_remains_in_previous_active_cohort(self):
        with tempfile.TemporaryDirectory() as directory:
            asset = Path(directory) / "Lofi_Radar_live_data.js"
            asset.write_text(
                'window.LOFI_LIVE_DATA={"d":{"lives":[{'
                '"vid":"0muHFBSiybw","channelId":"'
                + live.OFFICIAL_CHANNEL_ID
                + '","source":"Official Lofi Girl streams scan",'
                '"liveStatus":"is_live"}],"liveSummary":{'
                '"0muHFBSiybw":{"now":0,"active":false}}},"metrics":{'
                '"officialExpected":1,"officialVerified":1}};\n',
                encoding="utf-8",
            )
            cohort = live.load_previous_active_official_lives(asset)
        self.assertEqual(
            cohort,
            {"0muHFBSiybw": {"trusted": True, "started": None}},
        )

    @staticmethod
    def watch_payload(
        video_id="rFZHOHl-L8A",
        *,
        channel_id=None,
        title="lofi hip hop radio 📚 beats to relax/study to",
        micro_title=None,
        status="OK",
        reason=None,
        is_live=True,
        is_live_content=True,
        is_live_now=True,
        is_private=False,
        is_unlisted=False,
        is_crawlable=True,
        canonical_url=None,
        start="2026-08-19T07:43:43+00:00",
        end=None,
    ):
        channel_id = channel_id or live.OFFICIAL_CHANNEL_ID
        playability = {"status": status}
        if reason is not None:
            playability["reason"] = reason
        broadcast = {"isLiveNow": is_live_now, "startTimestamp": start}
        if end is not None:
            broadcast["endTimestamp"] = end
        return {
            "playabilityStatus": playability,
            "videoDetails": {
                "videoId": video_id,
                "channelId": channel_id,
                "author": "Lofi Girl",
                "title": title,
                "isLive": is_live,
                "isLiveContent": is_live_content,
                "isPrivate": is_private,
                "isCrawlable": is_crawlable,
            },
            "microformat": {
                "playerMicroformatRenderer": {
                    "externalVideoId": video_id,
                    "externalChannelId": channel_id,
                    "ownerChannelName": "Lofi Girl",
                    "title": {"simpleText": micro_title or title},
                    "canonicalUrl": canonical_url
                    or f"https://www.youtube.com/watch?v={video_id}",
                    "isUnlisted": is_unlisted,
                    "liveBroadcastDetails": broadcast,
                }
            },
        }

    def test_watch_page_proves_exact_public_live_and_ended_states(self):
        active = live.YouTubeWatchPageLiveProof._project(
            self.watch_payload(), "rFZHOHl-L8A"
        )
        self.assertIs(active["is_live"], True)
        self.assertEqual(active["release_timestamp"], 1_787_125_423)
        self.assertEqual(active["title"], "lofi hip hop radio 📚 beats to relax/study to")

        ended = live.YouTubeWatchPageLiveProof._project(
            self.watch_payload(
                video_id="X4VbdwhkE10",
                title="old official radio",
                status="UNPLAYABLE",
                reason=live.ENDED_UNPLAYABLE_REASON,
                is_live=False,
                is_live_now=False,
                start="2026-08-18T10:00:00Z",
                end="2026-08-19T03:09:13Z",
            ),
            "X4VbdwhkE10",
        )
        self.assertIs(ended["is_live"], False)
        self.assertEqual(ended["live_status"], "was_live")

        mutations = (
            {"channel_id": "UCxxxxxxxxxxxxxxxxxxxxxx"},
            {"status": "LOGIN_REQUIRED"},
            {"is_private": True},
            {"is_unlisted": True},
            {"is_crawlable": False},
            {"is_live_content": False},
            {"canonical_url": "https://www.youtube.com/watch?v=LTiqKDrjqr4"},
            {"micro_title": "another title"},
            {"start": "2026-08-19T07:43:43"},
        )
        for mutation in mutations:
            with self.subTest(mutation=mutation):
                with self.assertRaises(RuntimeError):
                    live.YouTubeWatchPageLiveProof._project(
                        self.watch_payload(**mutation), "rFZHOHl-L8A"
                    )
        with self.assertRaises(RuntimeError):
            live.YouTubeWatchPageLiveProof._project(
                self.watch_payload(
                    video_id="X4VbdwhkE10",
                    title="old official radio",
                    status="UNPLAYABLE",
                    reason="Video unavailable",
                    is_live=False,
                    is_live_now=False,
                    end="2026-08-19T03:09:13Z",
                ),
                "X4VbdwhkE10",
            )

    def test_watch_page_extracts_one_strict_player_assignment(self):
        payload = self.watch_payload()
        html = (
            "<script>var ytInitialPlayerResponse = "
            + json.dumps(payload, ensure_ascii=False)
            + ";</script>"
        ).encode("utf-8")
        self.assertEqual(
            live.YouTubeWatchPageLiveProof._player_response(html), payload
        )
        with self.assertRaisesRegex(RuntimeError, "ambiguous"):
            live.YouTubeWatchPageLiveProof._player_response(html + html)
        with self.assertRaisesRegex(RuntimeError, "malformed"):
            live.YouTubeWatchPageLiveProof._player_response(
                b'<script>ytInitialPlayerResponse={"a":1,"a":2};</script>'
            )

    @staticmethod
    def next_payload(
        video_id="0muHFBSiybw",
        *,
        channel_id=None,
        is_live=True,
        original="3208",
        displayed="3,208",
        started_text="Started streaming on Aug 12, 2026",
    ):
        channel_id = channel_id or live.OFFICIAL_CHANNEL_ID
        return {
            "currentVideoEndpoint": {
                "commandMetadata": {
                    "webCommandMetadata": {
                        "url": f"/watch?v={video_id}",
                        "webPageType": "WEB_PAGE_TYPE_WATCH",
                    }
                },
                "watchEndpoint": {"videoId": video_id},
            },
            "contents": {
                "twoColumnWatchNextResults": {
                    "results": {
                        "results": {
                            "contents": [
                                {
                                    "videoPrimaryInfoRenderer": {
                                        "title": {"runs": [{"text": "summer lofi radio"}]},
                                        "viewCount": {
                                            "videoViewCountRenderer": {
                                                "isLive": is_live,
                                                "originalViewCount": original,
                                                "viewCount": {
                                                    "runs": [
                                                        {"text": displayed},
                                                        {"text": " watching now"},
                                                    ]
                                                },
                                            }
                                        },
                                        "dateText": {"simpleText": started_text},
                                    }
                                },
                                {
                                    "videoSecondaryInfoRenderer": {
                                        "owner": {
                                            "videoOwnerRenderer": {
                                                "navigationEndpoint": {
                                                    "browseEndpoint": {"browseId": channel_id}
                                                },
                                                "title": {
                                                    "runs": [
                                                        {
                                                            "text": "Lofi Girl",
                                                            "navigationEndpoint": {
                                                                "browseEndpoint": {
                                                                    "browseId": channel_id
                                                                }
                                                            },
                                                        }
                                                    ]
                                                },
                                            }
                                        }
                                    }
                                },
                            ]
                        }
                    }
                }
            },
        }

    def test_next_fallback_projects_only_strict_current_live_facts(self):
        client = live.YouTubeNextLiveFallback(retries=0)
        info = client._project(
            self.next_payload(), "0muHFBSiybw", 1_786_546_921_000
        )
        self.assertEqual(info["concurrent_view_count"], 3208)
        self.assertEqual(info["release_timestamp"], 1_786_546_921)
        self.assertEqual(info["channel_id"], live.OFFICIAL_CHANNEL_ID)
        self.assertEqual(info["detail_source"], "youtube_next_previous_start")
        relative_observed = 1_786_546_921_000 + 8 * 3_600_000 + 30 * 60_000
        relative = client._project(
            self.next_payload(started_text="Started streaming 8 hours ago"),
            "0muHFBSiybw",
            1_786_546_921_000,
            start_source="youtube_watch_page",
            expected_title="summer lofi radio",
            observed_ms=relative_observed,
        )
        self.assertEqual(relative["detail_source"], "youtube_next_watch_page_start")
        trusted_relative = client._project(
            self.next_payload(started_text="Started streaming 8 hours ago"),
            "0muHFBSiybw",
            1_786_546_921_000,
            observed_ms=relative_observed,
        )
        self.assertEqual(
            trusted_relative["detail_source"], "youtube_next_previous_start"
        )
        with self.assertRaisesRegex(RuntimeError, "start date mismatch"):
            client._project(
                self.next_payload(started_text="Started streaming 8 hours ago"),
                "0muHFBSiybw",
                1_786_546_921_000,
                start_source="untrusted_previous_asset",
                observed_ms=relative_observed,
            )
        with self.assertRaisesRegex(RuntimeError, "start date mismatch"):
            client._project(
                self.next_payload(started_text="Started streaming 8 hours ago"),
                "0muHFBSiybw",
                1_786_546_921_000,
                observed_ms=1_786_546_921_000 + 10 * live.DAY_MS,
            )
        with self.assertRaises(RuntimeError):
            client._project(
                self.next_payload(),
                "0muHFBSiybw",
                1_786_546_921_000,
                expected_title="different title",
            )
        for payload in (
            self.next_payload(is_live=False),
            self.next_payload(channel_id="UCxxxxxxxxxxxxxxxxxxxxxx"),
            self.next_payload(original="3209"),
            self.next_payload(started_text="Started streaming on Jan 1, 2024"),
        ):
            with self.subTest(payload=payload):
                with self.assertRaises(RuntimeError):
                    client._project(payload, "0muHFBSiybw", 1_786_546_921_000)
        with self.assertRaisesRegex(RuntimeError, "no exact start time"):
            client._project(self.next_payload(), "0muHFBSiybw", None)

    def test_next_relative_age_uses_response_timestamp_not_scan_timestamp(self):
        started_ms = 1_786_546_921_000
        response_ms = started_ms + 8 * 3_600_000 + 15 * 60_000
        client = live.YouTubeNextLiveFallback(retries=0)
        response = (
            self.next_payload(started_text="Started streaming 8 hours ago"),
            live.NEXT_ENDPOINT,
            response_ms,
        )
        with mock.patch.object(client, "_fetch_once", return_value=response):
            info = client.extract_info("0muHFBSiybw", started_ms=started_ms)
        self.assertEqual(info["release_timestamp"], int(started_ms / 1000))

        stale_response = (response[0], response[1], started_ms + 10 * live.DAY_MS)
        with mock.patch.object(client, "_fetch_once", return_value=stale_response):
            with self.assertRaisesRegex(RuntimeError, "start date mismatch"):
                client.extract_info("0muHFBSiybw", started_ms=started_ms)

    def test_known_active_antibot_uses_prior_start_next_without_watch(self):
        video_id = "rFZHOHl-L8A"
        started_ms = 1_787_125_423_000

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": video_id, "live_status": "is_live"}]}
                return {"entries": []}

        class BlockedReader:
            def extract_info(self, url, download=False):
                raise RuntimeError("Sign in to confirm you're not a bot")

        class NextReader:
            def __init__(self):
                self.calls = []

            def extract_info(self, requested_id, *, started_ms):
                self.calls.append((requested_id, started_ms))
                return {
                    "id": requested_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "lofi hip hop radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 9876,
                    "release_timestamp": started_ms / 1000,
                }

        class UnexpectedWatch:
            def extract_info(self, requested_id):
                raise AssertionError("Trusted active Next proof must not call WatchPage")

        next_reader = NextReader()
        result = live.discover_official_live_streams(
            1_787_128_000_000,
            flat_reader=FlatReader(),
            detail_reader=BlockedReader(),
            watch_reader=UnexpectedWatch(),
            next_reader=next_reader,
            previous_active={
                video_id: {"trusted": True, "started": started_ms}
            },
        )
        self.assertEqual(next_reader.calls, [(video_id, started_ms)])
        self.assertEqual([row["vid"] for row in result["rows"]], [video_id])
        self.assertEqual(result["points"][video_id][0][1], 9876)
        self.assertEqual(result["metrics"]["nextFallbacks"], 1)
        self.assertEqual(result["metrics"]["watchPageFallbacks"], 0)

    def test_missing_previous_never_uses_trusted_next_fast_path(self):
        active_id = "rFZHOHl-L8A"
        missing_id = "X4VbdwhkE10"

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": active_id, "live_status": "is_live"}]}
                return {"entries": []}

        class DetailReader:
            def extract_info(self, url, download=False):
                video_id = url.rsplit("=", 1)[-1]
                if video_id == missing_id:
                    raise RuntimeError("Sign in to confirm you're not a bot")
                return {
                    "id": active_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "current radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 20,
                    "release_timestamp": 1_787_125_423,
                }

        class WatchReader:
            def __init__(self):
                self.calls = []

            def extract_info(self, video_id):
                self.calls.append(video_id)
                return {
                    "id": missing_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "ended radio",
                    "is_live": False,
                    "live_status": "was_live",
                    "release_timestamp": 1_780_555_998,
                    "end_timestamp": 1_787_108_953,
                }

        class UnexpectedNext:
            def extract_info(self, video_id, **kwargs):
                raise AssertionError("Missing previous IDs must not use trusted Next")

        watch_reader = WatchReader()
        result = live.discover_official_live_streams(
            1_787_128_000_000,
            flat_reader=FlatReader(),
            detail_reader=DetailReader(),
            watch_reader=watch_reader,
            next_reader=UnexpectedNext(),
            previous_active={
                missing_id: {"trusted": True, "started": 1_780_555_998_000}
            },
        )
        self.assertEqual(watch_reader.calls, [missing_id])
        self.assertEqual(result["endedIds"], [missing_id])
        self.assertEqual(result["metrics"]["nextFallbacks"], 0)

    def test_next_and_watch_failures_are_preserved_without_sensitive_text(self):
        video_id = "rFZHOHl-L8A"

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": video_id, "live_status": "is_live"}]}
                return {"entries": []}

        class BlockedReader:
            def extract_info(self, url, download=False):
                raise RuntimeError("Sign in to confirm you're not a bot")

        class BrokenNext:
            def extract_info(self, video_id, **kwargs):
                raise RuntimeError(
                    "YouTube next network failure for public ID: URLError SECRET_NEXT"
                )

        class BrokenWatch:
            def extract_info(self, video_id):
                raise RuntimeError(
                    "YouTube WatchPage evidence is incomplete SECRET_WATCH"
                )

        with self.assertRaisesRegex(RuntimeError, "Trusted YouTube next proof failed") as caught:
            live.discover_official_live_streams(
                flat_reader=FlatReader(),
                detail_reader=BlockedReader(),
                watch_reader=BrokenWatch(),
                next_reader=BrokenNext(),
                previous_active={
                    video_id: {"trusted": True, "started": 1_787_125_423_000}
                },
            )
        message = str(caught.exception)
        self.assertIn("network failure", message)
        self.assertIn("incomplete public evidence", message)
        self.assertNotIn("SECRET_NEXT", message)
        self.assertNotIn("SECRET_WATCH", message)
        self.assertIsNone(caught.exception.__cause__)
        self.assertIs(caught.exception.__suppress_context__, True)

    def test_antibot_cold_boot_combines_watch_proof_with_next_counter(self):
        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": "0muHFBSiybw", "live_status": "is_live"}]}
                return {"entries": []}

        class BlockedReader:
            def extract_info(self, url, download=False):
                raise RuntimeError("Sign in to confirm you’re not a bot")

        class WatchReader:
            def __init__(self):
                self.calls = []

            def extract_info(self, video_id):
                self.calls.append(video_id)
                return {
                    "id": video_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "summer lofi radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "release_timestamp": 1_786_546_921,
                }

        class NextReader:
            def __init__(self):
                self.calls = []

            def extract_info(
                self,
                video_id,
                *,
                started_ms,
                start_source,
                expected_title,
            ):
                self.calls.append(
                    (video_id, started_ms, start_source, expected_title)
                )
                return {
                    "id": video_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "summer lofi radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 3210,
                    "release_timestamp": int(started_ms / 1000),
                }

        next_reader = NextReader()
        watch_reader = WatchReader()
        result = live.discover_official_live_streams(
            1_786_640_000_000,
            flat_reader=FlatReader(),
            detail_reader=BlockedReader(),
            watch_reader=watch_reader,
            next_reader=next_reader,
        )
        self.assertEqual(watch_reader.calls, ["0muHFBSiybw"])
        self.assertEqual(
            next_reader.calls,
            [(
                "0muHFBSiybw",
                1_786_546_921_000,
                "youtube_watch_page",
                "summer lofi radio",
            )],
        )
        self.assertEqual(result["metrics"]["watchPageFallbacks"], 1)
        self.assertEqual(result["metrics"]["nextFallbacks"], 1)
        self.assertEqual(result["points"]["0muHFBSiybw"][0][1], 3210)

    def test_next_fallback_is_not_used_for_other_detail_failures(self):
        class FlatReader:
            def extract_info(self, url, download=False):
                return {
                    "entries": [{"id": "0muHFBSiybw", "live_status": "is_live"}]
                    if url == live.OFFICIAL_STREAMS_URL
                    else []
                }

        class BrokenReader:
            def extract_info(self, url, download=False):
                raise RuntimeError("unrelated extractor failure")

        class UnexpectedNext:
            def extract_info(self, video_id, *, started_ms):
                raise AssertionError("Next fallback must not run")

        with self.assertRaisesRegex(RuntimeError, "unrelated extractor failure"):
            live.discover_official_live_streams(
                flat_reader=FlatReader(),
                detail_reader=BrokenReader(),
                next_reader=UnexpectedNext(),
                previous_active={
                    "0muHFBSiybw": {
                        "trusted": True,
                        "started": 1_786_546_921_000,
                    }
                },
            )

    def test_antibot_watch_proof_failure_never_reaches_next(self):
        class FlatReader:
            def extract_info(self, url, download=False):
                return {
                    "entries": [{"id": "0muHFBSiybw", "live_status": "is_live"}]
                    if url == live.OFFICIAL_STREAMS_URL
                    else []
                }

        class BlockedReader:
            def extract_info(self, url, download=False):
                raise RuntimeError("Sign in to confirm you’re not a bot")

        class BrokenWatch:
            def extract_info(self, video_id):
                raise RuntimeError("strict WatchPage proof failed")

        class UnexpectedNext:
            def extract_info(self, video_id, **kwargs):
                raise AssertionError("Next must never run without WatchPage proof")

        with self.assertRaisesRegex(RuntimeError, "strict WatchPage proof failed"):
            live.discover_official_live_streams(
                flat_reader=FlatReader(),
                detail_reader=BlockedReader(),
                watch_reader=BrokenWatch(),
                next_reader=UnexpectedNext(),
            )

    def test_antibot_replacement_removes_ended_previous_radio(self):
        new_id = "rFZHOHl-L8A"
        old_id = "X4VbdwhkE10"

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": new_id, "live_status": "is_live"}]}
                return {"entries": []}

        class BlockedReader:
            def extract_info(self, url, download=False):
                video_id = url.rsplit("=", 1)[-1]
                if video_id == old_id:
                    raise RuntimeError(
                        f"ERROR: [youtube] {old_id}: "
                        f"{live.ENDED_UNPLAYABLE_REASON}"
                    )
                raise RuntimeError("Sign in to confirm you're not a bot")

        class WatchReader:
            def extract_info(self, video_id):
                if video_id == new_id:
                    return {
                        "id": new_id,
                        "channel_id": live.OFFICIAL_CHANNEL_ID,
                        "availability": "public",
                        "title": "lofi hip hop radio",
                        "is_live": True,
                        "live_status": "is_live",
                        "release_timestamp": 1_787_125_423,
                    }
                return {
                    "id": old_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "ended radio",
                    "is_live": False,
                    "live_status": "was_live",
                    "release_timestamp": 1_787_040_000,
                    "end_timestamp": 1_787_108_953,
                }

        class NextReader:
            def __init__(self):
                self.calls = []

            def extract_info(
                self,
                video_id,
                *,
                started_ms,
                start_source,
                expected_title,
            ):
                self.calls.append(video_id)
                return {
                    "id": video_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": expected_title,
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 4321,
                    "release_timestamp": started_ms / 1000,
                }

        next_reader = NextReader()
        result = live.discover_official_live_streams(
            1_787_128_000_000,
            flat_reader=FlatReader(),
            detail_reader=BlockedReader(),
            watch_reader=WatchReader(),
            next_reader=next_reader,
            previous_active={old_id: {"trusted": True, "started": 1_787_040_000_000}},
        )
        self.assertEqual([row["vid"] for row in result["rows"]], [new_id])
        self.assertEqual(result["endedIds"], [old_id])
        self.assertEqual(result["metrics"]["watchPageFallbacks"], 2)
        self.assertEqual(result["metrics"]["nextFallbacks"], 1)
        self.assertEqual(result["metrics"]["confirmedEnded"], 1)
        self.assertEqual(next_reader.calls, [new_id])

    def test_exact_ended_recording_error_uses_watch_proof_without_next(self):
        active_id = "rFZHOHl-L8A"
        ended_id = "X4VbdwhkE10"

        class FlatReader:
            def extract_info(self, url, download=False):
                if url == live.OFFICIAL_STREAMS_URL:
                    return {"entries": [{"id": active_id, "live_status": "is_live"}]}
                return {"entries": []}

        class DetailReader:
            def extract_info(self, url, download=False):
                video_id = url.rsplit("=", 1)[-1]
                if video_id == ended_id:
                    raise RuntimeError(
                        f"ERROR: [youtube] {ended_id}: "
                        f"{live.ENDED_UNPLAYABLE_REASON}"
                    )
                return {
                    "id": active_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "current radio",
                    "is_live": True,
                    "live_status": "is_live",
                    "concurrent_view_count": 20,
                    "release_timestamp": 1_787_125_423,
                }

        class WatchReader:
            def __init__(self):
                self.calls = []

            def extract_info(self, video_id):
                self.calls.append(video_id)
                return {
                    "id": ended_id,
                    "channel_id": live.OFFICIAL_CHANNEL_ID,
                    "availability": "public",
                    "title": "ended radio",
                    "is_live": False,
                    "live_status": "was_live",
                    "release_timestamp": 1_780_555_998,
                    "end_timestamp": 1_787_108_953,
                }

        class UnexpectedNext:
            def extract_info(self, video_id, **kwargs):
                raise AssertionError("An ended WatchPage proof must not call Next")

        watch_reader = WatchReader()
        result = live.discover_official_live_streams(
            1_787_128_000_000,
            flat_reader=FlatReader(),
            detail_reader=DetailReader(),
            watch_reader=watch_reader,
            next_reader=UnexpectedNext(),
            previous_active={ended_id: {"trusted": True}},
        )
        self.assertEqual(watch_reader.calls, [ended_id])
        self.assertEqual(result["endedIds"], [ended_id])
        self.assertEqual(result["metrics"]["confirmedEnded"], 1)
        self.assertEqual(result["metrics"]["watchPageFallbacks"], 1)
        self.assertEqual(result["metrics"]["nextFallbacks"], 0)
        self.assertFalse(
            live.is_youtube_ended_recording_error(
                RuntimeError(live.ENDED_UNPLAYABLE_REASON + " retry"), ended_id
            )
        )

    def test_streams_tab_entry_without_flat_status_is_hydrated_not_dropped(self):
        class Reader:
            def __init__(self, payloads):
                self.payloads = payloads

            def extract_info(self, url, download=False):
                return self.payloads[url]

        video_id = "0muHFBSiybw"
        flat = Reader({
            live.OFFICIAL_STREAMS_URL: {"entries": [{"id": video_id}]},
            live.OFFICIAL_UPLOADS_URL: {"entries": []},
        })
        detail = Reader({f"https://www.youtube.com/watch?v={video_id}": {
            "id": video_id,
            "channel_id": live.OFFICIAL_CHANNEL_ID,
            "availability": "public",
            "title": "summer lofi radio",
            "is_live": True,
            "live_status": "is_live",
            "concurrent_view_count": 25,
            "release_timestamp": 1_786_546_921,
        }})
        result = live.discover_official_live_streams(
            1_786_640_000_000, flat_reader=flat, detail_reader=detail
        )
        self.assertEqual([row["vid"] for row in result["rows"]], [video_id])
        self.assertEqual(result["metrics"]["streamsTabUnknown"], 1)

    def test_official_stream_discovery_rejects_wrong_channel_or_missing_counter(self):
        base = {
            "id": "0muHFBSiybw",
            "channel_id": live.OFFICIAL_CHANNEL_ID,
            "availability": "public",
            "title": "summer lofi radio",
            "is_live": True,
            "live_status": "is_live",
            "concurrent_view_count": 1,
            "release_timestamp": 1_786_546_921,
        }
        for mutation in (
            {"channel_id": "UCxxxxxxxxxxxxxxxxxxxxxx"},
            {"live_status": "is_upcoming"},
            {"is_live": False},
            {"availability": "private"},
            {"concurrent_view_count": None},
            {"concurrent_view_count": 1.5},
        ):
            with self.subTest(mutation=mutation):
                with self.assertRaises(RuntimeError):
                    live.official_live_row({**base, **mutation}, 1_786_640_000_000)

    def test_one_sheet_point_does_not_become_a_peak_or_mask_sheet_staleness(self):
        workbook = self.workbook()
        streams = live.sheet_by_name(workbook, "Live Streams")
        hourly = live.sheet_by_name(workbook, "Live Hourly")
        video_id = "0muHFBSiybw"
        streams.append([
            video_id, "Lofi Girl", "summer lofi radio", "", "2026-08-12", "official"
        ])
        hourly.append([video_id, "2026-08-11T12:00:00Z", 3200])
        observed = 1_786_640_000_000
        official = {
            "rows": [{
                "vid": video_id,
                "channel": "Lofi Girl",
                "channelId": live.OFFICIAL_CHANNEL_ID,
                "title": "summer lofi radio",
                "url": f"https://www.youtube.com/watch?v={video_id}",
                "started": 1_786_546_921_000,
                "liveStatus": "is_live",
            }],
            "points": {video_id: [(observed, 3250)]},
            "metrics": {"expected": 1, "verified": 1, "observedT": observed},
        }
        payload = live.build_payload(workbook, official_snapshot=official)
        summary = payload["d"]["liveSummary"][video_id]
        self.assertEqual(summary["now"], 3250)
        self.assertIsNone(summary["peak24"])
        self.assertIsNone(summary["peakAll"])
        self.assertEqual(
            payload["metrics"]["sheetSourceLatestT"],
            live.to_ms("2026-08-11T12:00:00Z"),
        )
        self.assertEqual(payload["metrics"]["sourceLatestT"], observed)
        with self.assertRaisesRegex(RuntimeError, "Livestream Sheet is"):
            live.ensure_sheet_freshness(payload, 8, now_ms=observed)

        official["points"][video_id] = [(observed + 1, 0)]
        zero_payload = live.build_payload(workbook, official_snapshot=official)
        self.assertIs(zero_payload["d"]["liveSummary"][video_id]["active"], True)

    def test_confirmed_ended_official_radio_cannot_remain_active_from_sheet_lag(self):
        observed = live.to_ms("2026-08-11T12:00:00Z")
        official_id = "0muHFBSiybw"
        official = {
            "rows": [{
                "vid": official_id,
                "channel": "Lofi Girl",
                "channelId": live.OFFICIAL_CHANNEL_ID,
                "title": "current radio",
                "url": f"https://www.youtube.com/watch?v={official_id}",
                "started": observed - 3_600_000,
                "liveStatus": "is_live",
            }],
            "points": {official_id: [(observed, 10)]},
            "endedIds": ["abcdefghijk"],
            "metrics": {"expected": 1, "verified": 1, "observedT": observed},
        }
        payload = live.build_payload(self.workbook(), official_snapshot=official)
        self.assertIs(payload["d"]["liveSummary"][official_id]["active"], True)
        self.assertIs(payload["d"]["liveSummary"]["abcdefghijk"]["active"], False)

    def test_pages_verification_requires_exact_asset_bytes(self):
        class Response:
            def __init__(self, request):
                self.request = request

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def geturl(self):
                return self.request.full_url

            def read(self, limit):
                self.asserted_limit = limit
                return b"window.LOFI_LIVE_DATA={};\n"

        with tempfile.TemporaryDirectory() as directory:
            asset = Path(directory) / "Lofi_Radar_live_data.js"
            asset.write_bytes(b"window.LOFI_LIVE_DATA={};\n")

            def open_response(request, timeout):
                self.assertEqual(timeout, 30)
                self.assertIn(
                    "/youtube-radar-kx9v2m/Lofi_Radar_live_data.js?live_verify=",
                    request.full_url,
                )
                return Response(request)

            with mock.patch.object(live.urllib.request, "urlopen", open_response):
                live.verify_published_asset(
                    asset,
                    "https://dim75017.github.io/youtube-radar-kx9v2m/",
                    timeout_seconds=1,
                    interval_seconds=0.1,
                )

    def test_daily_workflow_refreshes_asset_without_hourly_deploy_loop(self):
        workflow = Path('.github/workflows/refresh-instrumental-radar.yml').read_text(encoding='utf-8')
        push_paths = workflow.split('pull_request:', 1)[0]
        self.assertIn('python refresh_youtube_live_snapshot.py', workflow)
        self.assertIn('yt-dlp==2026.7.4 openpyxl', workflow)
        live_step = workflow.split('Refresh compact livestream first-paint snapshot', 1)[1].split(
            'Commit factual catalogue and daily analytics first', 1
        )[0]
        self.assertIn('id: live_snapshot', live_step)
        self.assertIn('continue-on-error: true', live_step)
        report_step = workflow.split('Report incomplete official livestream discovery', 1)[1]
        self.assertIn("steps.live_snapshot.outcome != 'success'", report_step)
        self.assertIn('exit 1', report_step)
        self.assertLess(
            workflow.index('Commit factual catalogue and daily analytics first'),
            workflow.index('Report incomplete official livestream discovery'),
        )
        self.assertIn('git add Lofi_Radar_data.js Lofi_Radar_live_data.js', workflow)
        self.assertNotIn("- 'Lofi_Radar_live_data.js'", push_paths)
        self.assertNotRegex(workflow, r"cron:\s*['\"]?[^\n]*\* \* \* \*['\"]?")


if __name__ == "__main__":
    unittest.main()

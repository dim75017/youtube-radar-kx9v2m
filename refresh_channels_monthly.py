"""Monthly public YouTube channel audit snapshot.

The source list and curation stay in the dedicated Google Sheet.  This job
refreshes volatile public metrics only and stores a monthly append-only history
in ``Lofi_Radar_chx.js``.  A YouTube Data API key enriches total views/video
counts; the yt-dlp fallback still refreshes subscribers and latest upload.
"""
from __future__ import annotations

import concurrent.futures
import io
import json
import os
import re
import threading
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SNAPSHOT = Path(os.environ.get("CHANNEL_SNAPSHOT", ROOT / "Lofi_Radar_chx.js"))
SHEET_EXPORT = (
    "https://docs.google.com/spreadsheets/d/"
    "1jDbcryjTDbRsW4Uw6OP_SYwgvbPoLQu_KcSWqC3Dfoc/export?format=xlsx"
)
WORKERS = int(os.environ.get("CHANNEL_WORKERS", "10"))
MIN_SUCCESS_RATIO = 0.80
MIN_LATEST_UPLOAD_RATIO = 0.70
THREAD = threading.local()


def now_ms() -> int:
    return int(datetime.now(timezone.utc).timestamp() * 1000)


def atomic_write(path: Path, text: str) -> None:
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def read_snapshot(path: Path) -> dict:
    raw = path.read_text(encoding="utf-8")
    return json.loads(re.sub(r"^window\.CHX=", "", raw).rstrip(";\n"))


def normalized_alias(value: str | None) -> str:
    return str(value or "").strip().lower().lstrip("@")


def load_targets() -> list[dict]:
    from openpyxl import load_workbook

    with urllib.request.urlopen(SHEET_EXPORT, timeout=45) as response:
        workbook = load_workbook(io.BytesIO(response.read()), read_only=True, data_only=True)
    title = next(name for name in workbook.sheetnames if "Audit" in name)
    sheet = workbook[title]
    header = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and str(row[0] or "").strip().lower() in {"chaîne", "chaine"}:
            header = row_number
            break
    if not header:
        raise RuntimeError("Channel audit header not found")
    out = []
    for row in sheet.iter_rows(min_row=header + 1, values_only=True):
        if len(row) < 3:
            continue
        name = str(row[0] or "").strip()
        url = str(row[2] or "").strip()
        if name and url.startswith("https://www.youtube.com/"):
            out.append({"name": name, "url": url})
    # Lofi Girl is displayed as an owned synthetic row in the UI.
    out.append({"name": "Lofi Girl", "url": "https://www.youtube.com/@LofiGirl"})
    deduped = {target["url"].rstrip("/").lower(): target for target in out}
    return list(deduped.values())


def channel_ydl():
    if not hasattr(THREAD, "channel_ydl"):
        import yt_dlp

        THREAD.channel_ydl = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "extract_flat": True,
                "playlistend": 1,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "cachedir": False,
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return THREAD.channel_ydl


def video_ydl():
    if not hasattr(THREAD, "video_ydl"):
        import yt_dlp

        THREAD.video_ydl = yt_dlp.YoutubeDL(
            {
                "quiet": True,
                "no_warnings": True,
                "skip_download": True,
                "socket_timeout": 15,
                "retries": 1,
                "extractor_retries": 1,
                "ignoreerrors": True,
                "ignore_no_formats_error": True,
                "cachedir": False,
                "extractor_args": {"youtube": {"lang": ["en"], "player_client": ["web"]}},
            }
        )
    return THREAD.video_ydl


def fetch_channel(target: dict) -> dict | None:
    url = target["url"].rstrip("/") + "/videos"
    info = channel_ydl().extract_info(url, download=False) or {}
    channel_id = str(info.get("channel_id") or info.get("id") or "")
    if not channel_id.startswith("UC"):
        return None
    handle = normalized_alias(info.get("uploader_id"))
    row = {
        "name": target["name"],
        "url": target["url"],
        # Preserve the case-sensitive ID for YouTube Data API requests.  We
        # normalize aliases only when writing lookup keys to the snapshot.
        "channel_id": channel_id,
        "handle": handle,
    }
    followers = info.get("channel_follower_count")
    if isinstance(followers, (int, float)) and followers > 0:
        row["s"] = int(followers)
    entries = info.get("entries") or []
    if entries and entries[0] and entries[0].get("id"):
        latest = video_ydl().extract_info(
            f"https://www.youtube.com/watch?v={entries[0]['id']}", download=False
        ) or {}
        upload_date = str(latest.get("upload_date") or "")
        if re.match(r"^\d{8}$", upload_date):
            row["lu"] = f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:]}"
    return row


def enrich_from_api(rows: list[dict], key: str) -> int:
    by_id = {row["channel_id"]: row for row in rows}
    ids = list(by_id)
    matched = 0
    for start in range(0, len(ids), 50):
        batch = ids[start : start + 50]
        query = urllib.parse.urlencode(
            {"part": "snippet,statistics", "id": ",".join(batch), "key": key}
        )
        with urllib.request.urlopen(
            "https://www.googleapis.com/youtube/v3/channels?" + query, timeout=30
        ) as response:
            payload = json.load(response)
        for item in payload.get("items") or []:
            row = by_id.get(str(item.get("id") or ""))
            if not row:
                continue
            matched += 1
            stats = item.get("statistics") or {}
            snippet = item.get("snippet") or {}
            for source, target in (("subscriberCount", "s"), ("viewCount", "v"), ("videoCount", "n")):
                try:
                    row[target] = int(stats[source])
                except (KeyError, TypeError, ValueError):
                    pass
            published = str(snippet.get("publishedAt") or "")
            if re.match(r"^\d{4}-\d{2}-\d{2}", published):
                row["cr"] = published[:10]
    return matched


MONTH_DAYS = 30.4375


def _positive_int(value: object) -> int | None:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def append_monthly(points: list, timestamp: int, subscribers: int | None, views: int | None) -> list:
    clean = []
    for point in points or []:
        if isinstance(point, list) and len(point) >= 3:
            try:
                clean.append(
                    [
                        int(point[0]),
                        _positive_int(point[1]),
                        _positive_int(point[2]),
                    ]
                )
            except (TypeError, ValueError):
                pass
    month = datetime.fromtimestamp(timestamp / 1000, timezone.utc).strftime("%Y-%m")
    previous_same_month = next(
        (
            point
            for point in reversed(clean)
            if datetime.fromtimestamp(point[0] / 1000, timezone.utc).strftime("%Y-%m") == month
        ),
        None,
    )
    # A partial rerun must never replace a valid monthly subscriber/view point
    # with 0/None.  This is especially important for rounded YouTube counters.
    incoming_subscribers = _positive_int(subscribers)
    incoming_views = _positive_int(views)
    if previous_same_month:
        incoming_subscribers = incoming_subscribers or previous_same_month[1]
        incoming_views = incoming_views or previous_same_month[2]
    clean = [
        point
        for point in clean
        if datetime.fromtimestamp(point[0] / 1000, timezone.utc).strftime("%Y-%m") != month
    ]
    clean.append([timestamp, incoming_subscribers, incoming_views])
    clean.sort(key=lambda point: point[0])
    return clean[-36:]


def smoothed_monthly_subscriber_growth(points: list, recent_days: float = 366.0) -> int | None:
    """Return a non-zero monthly subscriber estimate.

    The preferred estimate is the average change across the most recent twelve
    months.  When rounded public counters make that window look flat, the
    function expands to the complete retained history.  A genuinely flat or
    insufficient series is reported as unknown instead of the misleading 0.
    """

    by_month: dict[str, tuple[int, int]] = {}
    for point in points or []:
        if not isinstance(point, list) or len(point) < 2:
            continue
        try:
            timestamp = int(point[0])
        except (TypeError, ValueError):
            continue
        subscribers = _positive_int(point[1])
        if subscribers is None:
            continue
        month = datetime.fromtimestamp(timestamp / 1000, timezone.utc).strftime("%Y-%m")
        previous = by_month.get(month)
        if previous is None or timestamp >= previous[0]:
            by_month[month] = (timestamp, subscribers)
    series = sorted(by_month.values())
    if len(series) < 2:
        return None

    def estimate(window_days: float | None) -> int | None:
        selected = series
        if window_days is not None:
            cutoff = series[-1][0] - int(window_days * 86_400_000)
            selected = [point for point in series if point[0] >= cutoff]
        if len(selected) < 2:
            return None
        first, last = selected[0], selected[-1]
        span_days = (last[0] - first[0]) / 86_400_000
        if span_days < 20:
            return None
        value = round((last[1] - first[1]) / (span_days / MONTH_DAYS))
        return int(value) if value != 0 else None

    return estimate(recent_days) or estimate(None)


def main() -> None:
    targets = load_targets()
    rows = []
    failed = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=WORKERS) as pool:
        future_to_target = {pool.submit(fetch_channel, target): target for target in targets}
        for future in concurrent.futures.as_completed(future_to_target):
            target = future_to_target[future]
            try:
                row = future.result()
                if row:
                    rows.append(row)
                else:
                    failed += 1
            except Exception as exc:
                failed += 1
                print(f"WARN {target['name']}: {type(exc).__name__}: {exc}")
    if not targets or len(rows) / len(targets) < MIN_SUCCESS_RATIO:
        raise RuntimeError(f"Channel refresh rejected: {len(rows)}/{len(targets)} resolved")

    subscriber_updated = sum(
        1 for row in rows if isinstance(row.get("s"), int) and row["s"] > 0
    )
    latest_upload_updated = sum(1 for row in rows if row.get("lu"))
    if subscriber_updated / len(targets) < MIN_SUCCESS_RATIO:
        raise RuntimeError(
            f"Channel refresh rejected: {subscriber_updated}/{len(targets)} subscriber totals"
        )
    if latest_upload_updated / len(targets) < MIN_LATEST_UPLOAD_RATIO:
        raise RuntimeError(
            f"Channel refresh rejected: {latest_upload_updated}/{len(targets)} latest uploads"
        )

    api_key = os.environ.get("YOUTUBE_API_KEY", "").strip()
    api_updated = 0
    if api_key:
        api_updated = enrich_from_api(rows, api_key)
        if api_updated / len(rows) < MIN_SUCCESS_RATIO:
            raise RuntimeError(
                f"Official channel metrics rejected: {api_updated}/{len(rows)} API matches"
            )

    snapshot = read_snapshot(SNAPSHOT)
    data = snapshot.setdefault("d", {})
    history = snapshot.setdefault("hist", {})
    timestamp = now_ms()
    for row in rows:
        channel_alias = normalized_alias(row["channel_id"])
        aliases = {channel_alias}
        if row.get("handle"):
            aliases.add(normalized_alias(row["handle"]))
        base = dict(data.get(channel_alias) or {})
        for key in ("s", "v", "n", "lu", "cr"):
            if row.get(key) not in (None, "", 0):
                base[key] = row[key]
        for alias in aliases:
            existing = dict(data.get(alias) or {})
            existing.update(base)
            # Never stamp an old cached total-view count as a fresh monthly
            # observation when the official API secret is unavailable.
            updated_history = append_monthly(
                history.get(alias) or [],
                timestamp,
                row.get("s"),
                row.get("v"),
            )
            history[alias] = updated_history
            smoothed_growth = smoothed_monthly_subscriber_growth(updated_history)
            if smoothed_growth is not None:
                existing["sm"] = smoothed_growth
            elif existing.get("sm") == 0:
                existing.pop("sm", None)
            data[alias] = existing
        if row["name"] == "Lofi Girl":
            snapshot["lg"] = dict(snapshot.get("lg") or {})
            for key in ("s", "v", "n", "lu", "cr"):
                if base.get(key) is not None:
                    snapshot["lg"][key] = base[key]

    snapshot["t"] = timestamp
    snapshot["coverage"] = {
        "targets": len(targets),
        "resolved": len(rows),
        "updated": subscriber_updated,
        "subscribers_updated": subscriber_updated,
        "latest_upload_updated": latest_upload_updated,
        "failed": failed,
        "official_api": bool(api_key),
        "official_api_updated": api_updated,
        "growth_estimates": sum(
            1
            for value in data.values()
            if isinstance(value, dict) and isinstance(value.get("sm"), int) and value["sm"] != 0
        ),
    }
    atomic_write(
        SNAPSHOT,
        "window.CHX=" + json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")) + ";",
    )
    print(json.dumps(snapshot["coverage"]))


if __name__ == "__main__":
    main()
